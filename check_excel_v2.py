import openpyxl
import os

excel_path = r'C:\Users\selcu\OneDrive\Masaüstü\Kaynakhane.xlsx'
if not os.path.exists(excel_path):
    print("Excel not found")
    exit()

wb = openpyxl.load_workbook(excel_path, data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n--- Sheet: {sheet}, Rows: {ws.max_row} ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] and ('6416' in str(row[0]) or '2062' in str(row[0])):
            print(f"Row {i+1}: {row}")
