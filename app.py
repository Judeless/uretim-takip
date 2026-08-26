from flask import (Flask, request, jsonify, render_template, send_file, g, session,
                   redirect, make_response)
from flask_cors import CORS
from datetime import datetime, date, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
import re
import secrets
import traceback # For debugging

from database import get_db as db_connect, init_db
from oee import hesapla_oee, hesapla_oee_ozet
from import_excel import import_data, durus_sebepleri_yukle, import_tum, export_referans_cycle_times, import_tk1
from export_excel import export_arsiv
# Tel proses mantığı — TK1_ROBOT_NOLARI (aşağıda) TEL_HATLARI'ndan türediği için
# import DOSYA BAŞINDA olmak zorunda. Ayrıntılı açıklama tanımın yanında.
from tel_proses import (TEL_ADIMLARI, TEL_ADIM_SIRA, TEL_HATLARI,      # noqa: F401
                        TEL_GIZLI_HATLAR,
                        TEL_ADIM_EKI, tel_hat_adimi, tel_referans_kodu,
                        tel_ek_ayikla, tel_adim_etiketi, tel_koddan_adim)

# ODS dosyası yolu
FIKSTUR_ODS_YOLU = r'C:\Users\selcu\OneDrive\Masaüstü\KAYNAKHANE FİKSTÜR RAF LİSTESİ.ods'
UYUM_ODS_YOLU = r'C:\Users\selcu\OneDrive\Masaüstü\RAF ve PROGRAM..ods'

def _ods_uyum_oku():
    """ODS dosyasındaki referans-robot uyum bilgilerini döndürür. (ref, robot_no, ist_no) listesi"""
    try:
        from odf.opendocument import load as ods_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        
        doc = ods_load(UYUM_ODS_YOLU)
        sh = doc.spreadsheet.getElementsByType(Table)[0]
        rows = sh.getElementsByType(TableRow)
        
        col_map = {}
        for r_idx in range(1, 15):
            base = 2 + (r_idx - 1) * 2
            col_map[base] = (f'ABB-{r_idx}', 1)
            col_map[base + 1] = (f'ABB-{r_idx}', 2)
            
        kayitlar = []
        for idx, row in enumerate(rows):
            if idx < 2: continue # basliklari atla
            cells = row.getElementsByType(TableCell)
            vals = []
            for c in cells:
                ps = c.getElementsByType(P)
                text = str(ps[0]).strip() if ps else ''
                rep = c.getAttribute('numbercolumnsrepeated')
                cnt = min(int(rep) if rep else 1, 35 - len(vals))
                vals.extend([text] * cnt)
                if len(vals) >= 35: break
            while len(vals) < 35: vals.append('')
            
            ref = vals[0].strip()
            if not ref or len(ref) < 3 or ref == 'ROBOT' or ref == 'İSTASYON': continue
            
            for col_idx, (r_no, ist) in col_map.items():
                if col_idx < len(vals):
                    val = vals[col_idx].strip()
                    if val and len(val) > 0:
                        kayitlar.append((ref, r_no, ist))
        return kayitlar
    except Exception as e:
        print('ODS Uyum Okuma Hatası:', e)
        return []

def _ods_kayitlari_oku():
    """ODS dosyasından tüm fikstür kayıtlarını {referans_kodu: raf_adresi} dict olarak döndürür."""
    try:
        from odf.opendocument import load as ods_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P

        doc = ods_load(FIKSTUR_ODS_YOLU)
        sh = doc.spreadsheet.getElementsByType(Table)[0]
        rows = sh.getElementsByType(TableRow)
        kayitlar = {}
        for row in rows:
            cells = row.getElementsByType(TableCell)
            vals = []
            for c in cells:
                ps = c.getElementsByType(P)
                text = str(ps[0]).strip() if ps else ''
                rep = c.getAttribute('numbercolumnsrepeated')
                cnt = min(int(rep) if rep else 1, 12 - len(vals))
                vals.extend([text] * cnt)
                if len(vals) >= 12:
                    break
            while len(vals) < 12:
                vals.append('')
            # Her 3 sütun grubunda 0=ref, 1=raf
            skip = {'A RAFI','B RAFI','C RAFI','D RAFI','REFERANS','RAF',''}
            for start in [0, 3, 6, 9]:
                ref = vals[start].strip()
                raf = vals[start + 1].strip() if start + 1 < 12 else ''
                if ref and ref not in skip and len(ref) > 2:
                    kayitlar[ref] = raf
        return kayitlar
    except Exception:
        return {}


def _ods_guncelle(tum_kayitlar):
    """
    tum_kayitlar: list of (referans_kodu, raf_adresi, notlar) tuples
    ODS dosyasını tamamen yeniden yazar.
    """
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableColumnProperties

        doc = OpenDocumentSpreadsheet()
        sheet = Table(name='Fikstür Listesi')
        doc.spreadsheet.addElement(sheet)

        # Başlık satırı
        hrow = TableRow()
        sheet.addElement(hrow)
        for baslik in ['Referans Kodu', 'Raf Adresi', 'Notlar']:
            tc = TableCell(valuetype='string')
            tc.addElement(P(text=baslik))
            hrow.addElement(tc)

        # Veri satırları
        for ref, raf, notlar in tum_kayitlar:
            row = TableRow()
            sheet.addElement(row)
            for val in [ref, raf, notlar]:
                tc = TableCell(valuetype='string')
                tc.addElement(P(text=str(val) if val else ''))
                row.addElement(tc)

        doc.save(FIKSTUR_ODS_YOLU)
    except Exception as e:
        print(f'ODS güncelleme hatası: {e}')

app = Flask(__name__)

# ── Oturum (session) gizli anahtarı ──
# Panel girişi imzalı cookie ile tutulur. Anahtar RESTART'lar arasında SABİT
# olmalı (yoksa her restart tüm oturumları düşürür). Öncelik: COFLE_SECRET_KEY
# env; yoksa proje kökünde tek seferlik üretilip saklanan .flask_secret dosyası
# (gitignore'da). Dosya yazılamazsa (salt-okunur FS) süreç-ömürlü rastgele anahtar.
def _secret_key_yukle():
    env_key = os.environ.get('COFLE_SECRET_KEY')
    if env_key:
        return env_key
    yol = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.flask_secret')
    try:
        if os.path.exists(yol):
            with open(yol, 'r', encoding='utf-8') as f:
                k = f.read().strip()
                if k:
                    return k
        k = secrets.token_hex(32)
        with open(yol, 'w', encoding='utf-8') as f:
            f.write(k)
        return k
    except Exception:
        return secrets.token_hex(32)

app.secret_key = _secret_key_yukle()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

# ── Panel yetki modeli ──
# İzin verilebilen sayfa anahtarları (dashboard_v2 sidebar data-sayfa değerleri).
# 'kullanicilar' burada YOK — o admin'e özeldir (rol ile korunur, izinle değil).
PANEL_SAYFALAR = [
    'ozet', 'bolum', 'kayitlar', 'is-yonetimi', 'fikstur', 'referanslar',
    'operatorler', 'saha-cihazlari', 'sinyal-analizi', 'andon-ayarlari', 'raporlar',
    'as400-teyit', 'kaynak-plan',
    # 2026-08-21: operatör performansı KİŞİ SIRALAR — yeni sayfa olarak eklendi,
    # mevcut kullanıcıların izin listesinde YOK (yalnız admin görür, yönetici
    # tek tek yetki verir). 'analiz' de aynı gerekçeyle ayrı.
    'operator-performans', 'analiz',
    # 2026-08-21 zaman çizelgesi: duruş sebebini DEĞİŞTİRİP yeniden atayabiliyor
    # (rapor/OEE'yi doğrudan etkiler) → aynı gerekçeyle izin listesine kendisi
    # eklenmez; yönetici tek tek yetki verir.
    'zaman-cizelgesi',
]

def panel_kullanici():
    """Aktif oturumdaki panel kullanıcısını DB'den taze çeker (izin/aktiflik anlık).
    Oturum yoksa veya kullanıcı pasif/silinmişse None."""
    kid = session.get('panel_uid')
    if not kid:
        return None
    try:
        conn = get_db()
        row = conn.execute(
            "SELECT id, kullanici_adi, ad_soyad, rol, izinler, aktif, sifre_gecici "
            "FROM panel_kullanicilari WHERE id=?", (kid,)
        ).fetchone()
    except Exception:
        return None
    if not row or not row['aktif']:
        return None
    try:
        izinler = json.loads(row['izinler'] or '[]')
        if not isinstance(izinler, list):
            izinler = []
    except Exception:
        izinler = []
    admin = (row['rol'] == 'admin')
    return {
        'id': row['id'],
        'kullanici_adi': row['kullanici_adi'],
        'ad_soyad': row['ad_soyad'] or '',
        'rol': row['rol'],
        'admin': admin,
        # admin tüm sayfaları görür; kullanıcı yalnız izin listesi
        'izinler': PANEL_SAYFALAR[:] if admin else [s for s in izinler if s in PANEL_SAYFALAR],
        'sifre_gecici': bool(row['sifre_gecici']),
    }

def panel_gerekli(izin=None, admin=False):
    """Endpoint koruma decorator'ı.
      panel_gerekli()                → sadece giriş yapılmış olmalı
      panel_gerekli(izin='referanslar') → o sayfa izni (admin daima geçer)
      panel_gerekli(admin=True)      → yalnız admin
    401 = giriş yok, 403 = yetki yok. Operatör mobil/andon uç noktalarına
    UYGULANMAZ (onlar oturumsuz çalışır)."""
    def dekorator(fn):
        @wraps(fn)
        def sarmal(*args, **kwargs):
            ku = panel_kullanici()
            if not ku:
                return jsonify({'hata': 'Oturum gerekli', 'giris_gerekli': True}), 401
            if admin and not ku['admin']:
                return jsonify({'hata': 'Bu işlem için yönetici yetkisi gerekli'}), 403
            if izin and not ku['admin'] and izin not in ku['izinler']:
                return jsonify({'hata': f'Bu sayfa için yetkiniz yok ({izin})'}), 403
            g.panel_ku = ku
            return fn(*args, **kwargs)
        return sarmal
    return dekorator

# Şablon hot-reload — debug'dan BAĞIMSIZ. debug=False (güvenlik) iken Flask normalde
# şablonları önbelleğe alır ve .html değişiklikleri ancak restart'ta yansır. Bunu açık
# tutarak HTML değişikliklerini RESTART OLMADAN canlıya alıyoruz (ufak mtime kontrolü,
# ihmal edilebilir maliyet) — güvenlik debug'ı kapalı kalır.
app.config['TEMPLATES_AUTO_RELOAD'] = True

@app.after_request
def _html_no_cache(resp):
    """HTML sayfaları her açılışta TAZE gelsin — PWA/ana ekran kısayolu bayat
    sürüm göstermesin (tarayıcı her seferinde sunucuyla doğrulasın)."""
    try:
        if resp.mimetype == 'text/html':
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            resp.headers['Pragma'] = 'no-cache'
    except Exception:
        pass
    return resp

# GUVENLIK: wildcard CORS YOK. Tum erisim same-origin oldugu icin varsayilan: cross-origin kapali.
# Frontend ayri domaine bolunurse: COFLE_CORS_ORIGINS=https://alan.adi (virgulle coklu) ile ac.
_cors_origins = [o.strip() for o in os.environ.get('COFLE_CORS_ORIGINS', '').split(',') if o.strip()]
if _cors_origins:
    CORS(app, origins=_cors_origins)

# ── Veritabanı Bağlantı Yönetimi ──
def get_db():
    if 'db' not in g:
        g.db = db_connect()
    return g.db

@app.teardown_appcontext
def teardown_db(exception):
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass


@app.errorhandler(Exception)
def api_hata_json(e):
    """/api/* yollarında yakalanmamış istisna HTML 500 yerine JSON döner.
    Mobil/dashboard r.json() bekler — HTML gelince operatör anlamsız 'Bağlantı hatası'
    görür ve gerçek sebep gizlenir (2026-07-06 'closed database' arızası dersi).
    Sayfa route'ları (HTML) etkilenmez. HTTPException'lar (404/400 vb.) aynen geçer."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        # /api/* yollarında 400/404/405 vb. de JSON — mobil r.json() her zaman parse edebilsin
        if request.path.startswith('/api/'):
            return jsonify({'hata': e.description or e.name, 'kod': e.code}), e.code
        return e
    if request.path.startswith('/api/'):
        print(f"HATA ({request.method} {request.path}): {traceback.format_exc()}")
        return jsonify({'hata': f'Sunucu hatası: {type(e).__name__}: {e}'}), 500
    raise e

# ─────────────────────────────────────────────────────────────
# OPERATÖR YETKİLENDİRME (PIN tabanlı)
# Operatör mobile'dan gelen istekler X-Operator + X-Operator-Pin
# header'larini icermeli. Dashboard/yonetici cagrilarinda bu
# header'lar yoktur ve serbest erisim verilir (geri uyumluluk).
# ─────────────────────────────────────────────────────────────
def operator_required(f):
    """Header'da X-Operator + X-Operator-Pin varsa PIN doğrular.
    Yoksa (dashboard/yönetici) izin verir ama g.operator_adi None olur.

    X-Operator header'i frontend'de encodeURIComponent ile encode edilir
    (HTTP header'lar ISO-8859-1 olmali, Turkce karakterler bozulur).
    """
    from urllib.parse import unquote
    @wraps(f)
    def wrapper(*args, **kwargs):
        op_raw = (request.headers.get('X-Operator', '') or '').strip()
        op = unquote(op_raw) if op_raw else ''
        pin = (request.headers.get('X-Operator-Pin', '') or '').strip()
        g.operator_adi = None
        g.operator_lokasyon = None
        if op and pin:
            # Lokasyon (varsa, fetch wrapper query'e ekler) ile daralt — composite UNIQUE
            # sonrası aynı ad iki tesiste olabilir. PIN'i tutan SATIRIN lokasyonu g'ye yazılır
            # (sahiplik kontrolü tesis karşılaştırması yapabilsin).
            lok = (request.args.get('lokasyon') or '').strip()
            # DIKKAT: get_db() flask.g'de REQUEST-PAYLASIMLI baglanti doner — BURADA KAPATMA!
            # (Decorator istegin BASINDA calisir; kapatirsak handler ayni kapali baglantiyi
            # alir → "Cannot operate on a closed database" 500. Kapatma teardown_db'de.)
            conn = get_db()
            if lok:
                rows = conn.execute("SELECT pin, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE ad=? AND COALESCE(lokasyon,'TK2')=?", (op, lok)).fetchall()
            else:
                rows = conn.execute("SELECT pin, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE ad=?", (op,)).fetchall()
            if not rows:
                return jsonify({'hata': f'Operatör bulunamadı: {op}'}), 403
            eslesen = next((r for r in rows if (r['pin'] or '0000') == pin), None)
            if eslesen is None:
                return jsonify({'hata': 'PIN yanlış'}), 403
            g.operator_adi = op
            g.operator_lokasyon = eslesen['lokasyon'] or 'TK2'
        elif op or pin:
            return jsonify({'hata': 'Operatör adı ve PIN birlikte gönderilmeli'}), 403
        return f(*args, **kwargs)
    return wrapper


def _vardiya_sahibi_kontrol(vardiya_id):
    """g.operator_adi varsa vardiya o operatöre ait mi kontrol eder.
    g.operator_adi None ise (dashboard) izin verir.
    Returns: (ok: bool, hata_mesaji: str|None)
    """
    if not getattr(g, 'operator_adi', None):
        return True, None  # dashboard modu — operatör yetkisi gerekmez
    if g.operator_adi == ADMIN_ADI:
        return True, None  # yönetici — tüm vardiyalara erişir
    conn = get_db()
    row = conn.execute("SELECT operator_adi, COALESCE(lokasyon,'TK2') as lokasyon FROM vardiyalar WHERE id=?", (vardiya_id,)).fetchone()
    if not row:
        return False, f'Vardiya bulunamadı (id={vardiya_id})'
    if row['operator_adi'] != g.operator_adi:
        return False, f'Bu vardiya başka operatöre ait: {row["operator_adi"]}'
    # Tesis karşılaştırması: composite UNIQUE sonrası iki tesiste aynı ad olabilir —
    # TK1'deki 'X', TK2'deki aynı adlı 'X'in vardiyasına dokunamasın.
    op_lok = getattr(g, 'operator_lokasyon', None)
    if op_lok and row['lokasyon'] != op_lok:
        return False, f'Bu vardiya başka tesise ait ({row["lokasyon"]})'
    return True, None


def _uretim_vardiya_bul(uretim_id):
    """uretim_kayit id'den vardiya_id getirir (yoksa None)."""
    conn = get_db()
    row = conn.execute('SELECT vardiya_id FROM uretim_kayitlari WHERE id=?', (uretim_id,)).fetchone()
    return row['vardiya_id'] if row else None


def _durus_vardiya_bul(durus_id):
    """durus id'den vardiya_id getirir (yoksa None)."""
    conn = get_db()
    row = conn.execute('SELECT vardiya_id FROM duruslar WHERE id=?', (durus_id,)).fetchone()
    return row['vardiya_id'] if row else None


# ─────────────────────────────────────────────────────────────
# SAYAÇ ENTEGRASYONU — pilot sensör pulse'larını üretim kaydına bağlar
# ─────────────────────────────────────────────────────────────
# Sadece sahada kurulu + doğrulanmış cihazlarda OK adet sensörden otomatik beslenir.
# Yeni cihaz devreye alınıp test edildikçe robot_no buraya eklenir.
# Kaynak ABB1..ABB9: tüm robotlar devreye alındı (2026-06). NOT: 2 robotun fiziksel sayaç
#   montajı 2026-06-26'da tamamlanacak — o robotlarda cihaz sayana kadar operatör ADET'i ELLE
#   girmeli (auto mod sensörü 0 okur → boş bırakılırsa üretim 0 görünür).
# Montaj M1..M12: tüm hatlara sayaç modülü kuruldu (2026-06) — hepsi sensör destekli.
# Metal: 300T + 400T + 550T. 300T [PULSE] testinden geçti, sahada sayıyor (2026-07-16).
# NOT: andon "makine çalışıyor" + sayaç rozeti bu allowlist'ten BAĞIMSIZ (saha_cihazlari
# robot_calisiyor'u sinyalden türetir); allowlist SADECE operatör mobil otomatik sayacını açar.
SAYAC_AUTO_CIHAZLAR = (
    {f'ABB{i}' for i in range(1, 10)} | {'300T', '400T', '550T'} | {f'M{i}' for i in range(1, 13)}
    | {'YF1', 'LF-LFP'}
    # 2026-07-23 yeni sayaç modülleri: abkant (röle) + eksantrik pres → bölüm 'pres';
    # plastik enjeksiyon (320T/407T, metal mantığı) → bölüm 'plastik' (TK1).
    # 2026-07-27 saha rollout: yapıştırma (kısa röle pulse, plastik hattının makinesi)
    # → bölüm 'plastik' (TK1). robot_no ASCII 'Yapistirma' (firmware ile birebir).
    # 2026-07-31: pres bölümünün TÜM makineleri makinenin KENDİ SAYACINA takılan röleden
    # sinyal alır (eski iki-el buton modeli bırakıldı) — 5 eksantrik + 1 hidrolik + 1 broş.
    # robot_no ASCII 'Bros' (firmware/klasör adı ASCII olmalı — 'Yapistirma' kalıbı).
    | {'Abkant 1', 'Abkant 2', 'Abkant 3'} | {f'Pres {i}' for i in range(1, 6)}
    | {'Hidrolik Pres', 'Bros'}
    | {'320T', '407T', 'Yapistirma'}
    # 2026-07-31: TK1 montaj masaları (7 adet, buton+buzzer). TK2'deki M1..M12
    # modelinin aynısı — MASA = hat; operatör mobilde masasını seçer, sayaç
    # doğrudan eşleşir (ek hat→cihaz eşlemesi gerekmez). YF1 ayrı, LF-LFP'ye bağlı.
    | {f'TK1-M{i}' for i in range(1, 5)}   # 2026-08-19: buton olanlar (M5/M6/M7 değil)
    # 2026-08-07: TK1 kapama presleri (12 adet, röle). Presler ÜÇERLİ ORTAK
    # PANOLARDA → 4 ESP32 modülü 3'er presi okur, ama her kanal KENDİ makine adıyla
    # (robot_no='Kapama N', istasyon=1..3) sinyal gönderir. Yani sayaç eşleşmesi
    # tek makineli cihazlarla AYNI: operatör hattını seçer, doğrudan eşleşir.
    # DİKKAT: burası CİHAZ (firmware robot_no) adlarıdır — 2026-08-18'den beri
    # operatörün gördüğü HAT ADI saha kodudur ('Kapama 5' hattı → 'Kapama 1' cihazı).
    # Çeviri HAT_SAYAC_CIHAZI'nda yapılır; bu küme firmware tarafında kalır.
    | {f'Kapama {i}' for i in range(1, 13)}
    # 2026-08-19: TK1 otomatik hazırlık makinesi (tek röle, 5-15 sn/parça).
    # robot_no ASCII — hat adı Türkçe, çeviri HAT_SAYAC_CIHAZI'nda.
    | {'Otomatik Hazirlik'}
)
# YF1: TK1 (yan tesis) montaj deneme modülü — MONTAJ-YF1 cihazı robot_no='YF1' ile flash'lı
# (sahada zaten yüklü, yeniden flash YOK). pilot.db'de bolum=montaj/robot_no=YF1; saha_cihazlari
# BEKLENEN'inde lokasyon=TK1 olarak izlenir. robot_no global benzersiz → firmware'de lokasyon gerekmez.

# TK1 (yan tesis) cihazlarının robot_no'ları — pilot.db'de lokasyon kolonu olmadığı için
# saha_cihazlari/sinyal_kalitesi'ni lokasyona göre filtrelerken robot_no ile eşleriz.
# YF1 = sahadaki deneme modülü; diğerleri TK1 hatları (ileride cihaz takılırsa kapsanır).
#   Pull / Push-Pull: TEL ÜRÜNÜ adları, hat değil (kullanıcı 2026-08-04) — montaj
#   hat listesinden çıkarıldı, ama GEÇMİŞ vardiyaları bu isimlerle kayıtlı olduğu
#   için burada KALIR (yoksa eski kayıtlar TK1 filtrelerinden düşerdi).
TK1_ROBOT_NOLARI = ({'YF1', 'Pull', 'Push-Pull', 'Iveco', 'LF-LFP',
                     '320T', '407T', 'Yapistirma', 'Sizdirmazlik Test',  # 2026-07-27 plastik enj + yapıştırma (TK1)
                     'Plastik Enjeksiyon',   # 2026-08-18 plastik SABİT HAT (makine artık üretim kaydında)
                     'TK1 Montaj', 'Tel Üretimi'}   # 2026-08-18 TK1 montaj + tel SABİT HATLARI
                    | {f'TK1-M{i}' for i in range(1, 8)}   # 2026-07-31 TK1 montaj masaları
                    # (M5/M6/M7 2026-08-19'da buton listesinden çıktı ama GEÇMİŞ
                    #  kayıtları TK1 filtrelerinden düşmesin diye bu kümede KALIYOR)
                    # Tel üretimi proses hatları — TEL_HATLARI'ndan TÜRETİLİR (elle
                    # yazılmaz). Eskiden burada 13 hat elle sayılıydı; kapama 4'ten
                    # 12'ye çıkınca (2026-08-07) yeni presler bu kümeye eklenmeyi
                    # unutulsaydı TK1 filtrelerinden sessizce düşerlerdi.
                    | set(TEL_HATLARI))

# Hat → sayaç cihazının robot_no eşlemesi: operatör hattı seçer (vardiya.robot_no='LF-LFP')
# ama saha modülü pulse'ları robot_no='YF1' ile pilot.db'ye düşer. Sayım yapılırken hat
# adı cihazın robot_no'suna çevrilir. (Yeniden flash gerektirmeden eşleştirme.)
# ── KAPAMA PRESLERİ: SAHA KODU → FIRMWARE robot_no (kullanıcı 2026-08-18) ──
# Modüller 'Kapama 1..12' olarak flash'lı ama sahadaki preslerin kodları başka
# (sistemdeki 'Kapama 1' modülü saha kodu 5 olan prese bağlı). Operatör sahada
# gördüğü kodu seçmeli → HAT ADI saha kodu oldu, cihaz adı firmware'deki gibi
# KALDI. Yeniden flash GEREKMEZ; eşleme burada (LF-LFP→YF1 ile aynı kalıp).
# DİKKAT: 'Kapama 12' hem HAT adı (saha kodu 12 → cihaz 'Kapama 2') hem CİHAZ adı
# (modül 4 kanal 3 → hat 'Kapama 26'). Karışmaz çünkü arama tek yönlü ve tek adımdır:
# ileri yön hat adıyla (HAT_SAYAC_CIHAZI['Kapama 12'] = 'Kapama 2'), geri yön
# DEĞERLERDE arayarak (cihaz 'Kapama 12' → hat 'Kapama 26').
# SOL = OPERATÖRÜN GÖRDÜĞÜ HAT ADI (2026-08-26'dan beri 'Hidrolik Pres N'),
# SAĞ = FIRMWARE'DEKİ CİHAZ ADI ('Kapama N') — cihaz adları DEĞİŞMEDİ, yeniden
# flash gerekmez. Adlar artık birbirine benzemediği için eski "'Kapama 12' hem
# hat hem cihaz" karışıklığı da kalmadı.
_KAPAMA_SAHA_CIHAZ = {
    # Modül 1-2 (2026-08-18)
    'Hidrolik Pres 5':  'Kapama 1',   'Hidrolik Pres 12': 'Kapama 2',
    'Hidrolik Pres 4':  'Kapama 3',   'Hidrolik Pres 30': 'Kapama 4',
    'Hidrolik Pres 28': 'Kapama 5',   'Hidrolik Pres 29': 'Kapama 6',
    # Modül 3-4 (2026-08-19) — 12 presin tamamı adlandırıldı, gizli hat kalmadı
    'Hidrolik Pres 21': 'Kapama 7',   'Hidrolik Pres 19': 'Kapama 8',
    'Hidrolik Pres 20': 'Kapama 9',   'Hidrolik Pres 27': 'Kapama 10',
    'Hidrolik Pres 25': 'Kapama 11',  'Hidrolik Pres 26': 'Kapama 12',
}
# TK1 BUTON MODÜLLERİ: buton etiketi → firmware robot_no (kullanıcı 2026-08-19).
# Modüller 'TK1-M1..M7' / 'YF1' olarak flash'lı; sahadaki butonlar MONTAJ - 1..5
# diye etiketlendi. Yeniden flash GEREKMEZ, eşleme burada.
# 2026-08-20 SAHA DÜZENİ — 5 modül İKİ HATTA bölündü:
#     MONTAJ hattı      : buton 1 · 2 · 3   (TK1-M1 · TK1-M2 · TK1-M3)
#     TEL SON MONTAJ    : buton 4 · 5       (TK1-M4 · YF1)
# ⚠ Buton↔modül sırası VARSAYIM (kullanıcı numaraları verdi, modül seri no'sunu
# değil). Sahada her butona bir kez basıp sayacın doğru hatta arttığı
# DOĞRULANMALI; tutmuyorsa yalnız bu tablo düzeltilir.
_TK1_MONTAJ_BUTON_CIHAZ = {
    # MONTAJ bölümünün kullandığı 3 buton: 1 · 2 · 3 (kullanıcı 2026-08-20)
    'MONTAJ - 1': 'TK1-M1',
    'MONTAJ - 2': 'TK1-M2',
    'MONTAJ - 3': 'TK1-M3',
    # 4 ve 5 numaralı slotlar montaj listesinden ÇIKTI — bu iki modül artık tel
    # SON MONTAJ hattında. Eşlemeler burada KALIR: 2026-08-19'dan beri yazılmış
    # eski montaj kayıtlarının sayacı hâlâ bunlarla çözülüyor.
    'MONTAJ - 4': 'TK1-M4',
    'MONTAJ - 5': 'YF1',
    # TEL SON MONTAJ (kullanıcı 2026-08-20): "son montaj hattındaki cihazlar
    # MONTAJ - 4 ve MONTAJ - 5 olmalı … son montaj hattında 2 buton modülü olacak
    # (buton 4 - buton 5)". Modüller ETİKETLERİYLE BİRLİKTE taşındı → hat adı
    # sahadaki buton numarasını taşır, yeniden flash GEREKMEZ:
    #     buton 4 → MONTAJ-TK1-M4 modülü      buton 5 → MONTAJ-YF1 modülü (eski YF)
    # ⚠ Sahada BİR KEZ doğrula: butona bas, sayaç 'Son Montaj 4/5'te mi artıyor.
    #   Tutmuyorsa yalnız bu iki satır yer değiştirir (başka yer düzeltilmez).
    'Son Montaj 4': 'TK1-M4',
    'Son Montaj 5': 'YF1',
}

# Bu butonlar montaj bölümü listesinden ÇIKARILDI (tel son montajda kullanılıyor)
# ama TK1_MONTAJ_HATLARI'nda POZİSYONLARI DURUR — liste kısalsaydı istasyon
# numaraları kayar, 2026-08-19'dan beri yazılmış TK1 montaj kayıtları başka hattı
# gösterirdi. Operatör listesinde gizlenirler (bkz. /api/kayit_hatlari).
GIZLI_MONTAJ_HATLARI = frozenset({'MONTAJ - 4', 'MONTAJ - 5'})
# Operatör listesinde GÖRÜNMEYEN tüm hat adları (emekli buton + kodu bekleyen slot).
# İki yerde kullanılır: /api/kayit_hatlari (listeden düşür) ve /api/saha_cihazlari'nın
# CİHAZ→HAT ters aramasında (taşınmış modül eski adıyla değil, BUGÜN çalıştığı hatla
# etiketlenmeli — yoksa panelde 'MONTAJ - 5' yazar, cihaz aslında son montajdadır).
GIZLI_HATLAR = TEL_GIZLI_HATLAR | GIZLI_MONTAJ_HATLARI
# 'LF-LFP': 'YF1' GEÇMİŞ vardiyalar için kalıyor (hat listesinden çıktı ama eski
# kayıtların sayacı hâlâ bu eşlemeyle çözülüyor). Aynı cihaza iki ad işaret
# ediyor; ters arama listede BULUNAN adı seçtiği için karışmaz.
HAT_SAYAC_CIHAZI = {'LF-LFP': 'YF1',
                    # Hat adı Türkçe (operatör görür) ↔ cihaz robot_no ASCII
                    # (firmware/klasör adı kuralı) — 2026-08-19 otomatik hazırlık.
                    'Otomatik Hazırlık': 'Otomatik Hazirlik',
                    **_TK1_MONTAJ_BUTON_CIHAZ, **_KAPAMA_SAHA_CIHAZ}

# ── ROBOT KAYNAK MAKİNELERİ (kullanıcı 2026-08-21: Punta Kaynak eklendi) ──
# Kaynak bölümünde hat VARDİYADA seçilir (robot_no) — tel/pres'in aksine üretim
# kaydında değil. Liste eskiden /api/robotlar içinde satır-içi üretiliyordu
# (f'ABB{i}'); punta kaynak eklenince tek kaynak buraya alındı.
#
# İSTASYON SAYISI MAKİNENİN ÖZELLİĞİ: ABB robotlarında operatör bir istasyonda
# söktak yaparken robot diğerinde kaynak yapar (bkz. oee.pair_cycle_hesapla) —
# 2 istasyon. PUNTA KAYNAK TEK İSTASYONLU (kullanıcı): orada 'İst. 2' seçeneği
# operatöre gösterilmez, yoksa var olmayan bir istasyona kayıt açılır ve o kayıt
# raporda hiçbir yere oturmaz.
KAYNAK_ROBOTLARI = [f'ABB{i}' for i in range(1, 10)] + ['Punta Kaynak']
TEK_ISTASYONLU_KAYNAK = frozenset({'Punta Kaynak'})


def kaynak_istasyon_sayisi(robot_no):
    """Kaynak makinesinin istasyon sayısı (varsayılan 2, punta 1).
    Kaynak dışı bölümlerde anlamsızdır — çağıran bölümü kontrol eder."""
    return 1 if str(robot_no or '').strip() in TEK_ISTASYONLU_KAYNAK else 2


# ── PRES (Pres Abkant) MAKİNE MODELİ (kullanıcı 2026-07-28) ──
# Operatör gün içinde makine değiştirerek çalışıyor → makine VARDİYADA DEĞİL, her
# ÜRETİM KAYDINDA seçilir (lazer modelinin aynısı). Vardiya hattı tek ve sabittir:
# PRES_SABIT_HAT; makine, uretim_kayitlari.istasyon kolonunda 1..10 olarak durur.
# Sıra ÖNEMLİ: istasyon no = bu listedeki sıra (1=Abkant 1 ... 10=Bros); makine
# adları pilot.db cihaz robot_no'larıyla (firmware) BİREBİR aynı olmalı.
# YENİ MAKİNE HEP SONA EKLENİR — araya girmek eski üretim kayıtlarının istasyon
# numaralarını başka makineye kaydırır (2026-07-31: 9=Hidrolik Pres, 10=Bros).
PRES_SABIT_HAT  = 'Pres Abkant'
PRES_MAKINELERI = ['Abkant 1', 'Abkant 2', 'Abkant 3',
                   'Pres 1', 'Pres 2', 'Pres 3', 'Pres 4', 'Pres 5',
                   'Hidrolik Pres', 'Bros']


# ── PLASTİK ENJEKSİYON MAKİNE MODELİ (kullanıcı 2026-08-18) ──
# Pres/abkant ile AYNI model: makine artık VARDİYADA değil her ÜRETİM KAYDINDA
# seçilir (operatör referansı seçtikten sonra makinesini işaretler). Vardiya hattı
# tek ve sabittir: PLASTIK_SABIT_HAT.
# SIRA ÖNEMLİ: istasyon no = bu listedeki sıra (1=320T ... 4=Sizdirmazlik Test).
# Adlar pilot.db cihaz robot_no'larıyla (firmware, generate.py DEVICES['plastik'])
# BİREBİR aynı — ASCII yazım zorunlu. YENİ MAKİNE HEP SONA EKLENİR; araya girmek
# eski üretim kayıtlarının istasyon numaralarını başka makineye kaydırır.
PLASTIK_SABIT_HAT  = 'Plastik Enjeksiyon'
PLASTIK_MAKINELERI = ['320T', '407T', 'Yapistirma', 'Sizdirmazlik Test']

# ── TK1: HAT DA ÜRETİM KAYDINDA (kullanıcı 2026-08-18) ──
# "TK1'de bütün üretim mantığında operatör referansı seçtikten SONRA hattı seçsin —
# bir operatör bir referansı yaparken birden fazla hatta çalışabilir." Somut vaka:
# aynı kişi bir ürünün halat kapamasını yapıp ardından son montajını da yapıyor;
# eskiden bunun için İKİ AYRI VARDİYA açmak gerekiyordu.
# TK2 montaj DEĞİŞMEDİ — orada hat vardiyada seçilmeye devam eder.
TK1_MONTAJ_SABIT_HAT = 'TK1 Montaj'
# HAT ADLARI = SAHADAKİ BUTON ETİKETLERİ (kullanıcı 2026-08-19): "5 adet buton
# kullanıyoruz, butonları MONTAJ - 1 … MONTAJ - 5 diye isimlendirdim." Operatör
# ekranda, masanın üstünde yazan adı görmeli — kapama preslerinde olduğu gibi.
# LF-LFP ve Iveco ÇIKARILDI (kullanıcı): artık montaj hattı değiller. Geçmiş
# vardiyaları bu adlarla kayıtlı olduğundan TK1_ROBOT_NOLARI'nda ve /api/robotlar
# geçmiş listesinde KALIRLAR — dashboard filtreleri bozulmaz.
# SIRA ÖNEMLİ (istasyon no = sıra); yeni hat HEP SONA eklenir.
TK1_MONTAJ_HATLARI = [f'MONTAJ - {i}' for i in range(1, 6)]
TEL_SABIT_HAT = 'Tel Üretimi'

# Hattı/makinesi ÜRETİM KAYDINDA (istasyon kolonunda) seçilen bölümler.
#   KAYITTA_HAT:    (lokasyon, bolum) → vardiyada seçilen TEK sabit hat adı
#   HAT_MAKINELERI: sabit hat adı → o hattın makine/alt-hat listesi
# Anahtar SABİT HAT ADI: adlar global benzersiz olduğundan çözümleme fonksiyonlarına
# lokasyon taşımak gerekmez (vardiyanın robot_no'su zaten sabit hattın kendisidir).
# Lazer BURADA DEĞİL: makinesi (Lazer 1..6) kayıtta seçilir ama sahada sayaç cihazı
# yok — istasyon yalnız rapor kırılımı, robot_no→cihaz eşlemesine girmez.
KAYITTA_HAT = {
    ('TK2', 'pres'):    PRES_SABIT_HAT,
    ('TK1', 'plastik'): PLASTIK_SABIT_HAT,
    ('TK1', 'montaj'):  TK1_MONTAJ_SABIT_HAT,
    ('TK1', 'tel'):     TEL_SABIT_HAT,
}
HAT_MAKINELERI = {
    PRES_SABIT_HAT:       PRES_MAKINELERI,
    PLASTIK_SABIT_HAT:    PLASTIK_MAKINELERI,
    TK1_MONTAJ_SABIT_HAT: TK1_MONTAJ_HATLARI,
    TEL_SABIT_HAT:        list(TEL_HATLARI),
}


# Kayıtlar sayfasında makine filtresinin ETİKETİ — operatörün ekranda gördüğü
# kavramın adı. Kullanıcı 2026-08-25: "operatör ekranındaki isim ile aynı olsun."
# TK1 montajda operatör bir MASA (buton) seçiyor, tel/plastik/preste bir MAKİNE.
KAYIT_MAKINE_ETIKETI = {
    TK1_MONTAJ_SABIT_HAT: 'Masa',
    TEL_SABIT_HAT:        'Makine',
    PLASTIK_SABIT_HAT:    'Makine',
    PRES_SABIT_HAT:       'Makine',
}


def kayit_makine_etiketi(lokasyon, bolum, hatlar=None):
    """Bölümün makine filtresi etiketi; hattı vardiyada seçilen bölümlerde ''.

    LOKASYON BOŞ GELEBİLİR: /api/ozet lokasyonsuz da çağrılabiliyor. O hâlde
    sabit_hat_adi TK2'ye düşer ve TK1'e özgü bölümler (tel, plastik, TK1 montaj)
    sessizce 'hat modu'na kayardı — panel yine eski hat adlarını gösterirdi.
    Lokasyonsuz istekte karar TAHMİNLE değil VERİYLE verilir: çağıran, dönen
    kayıtların vardiya hatlarını 'hatlar' ile geçirir; içlerinden biri sabit hat
    ise (KAYIT_MAKINE_ETIKETI'nde varsa) o bölüm makine modundadır. 'montaj' gibi
    iki tesiste de olan bir bölümde lokasyon tahmini yanlış tarafı seçebilirdi."""
    lok = (lokasyon or '').strip().upper()
    if lok:
        return KAYIT_MAKINE_ETIKETI.get(sabit_hat_adi(lok, bolum), '')
    for h in (hatlar or ()):
        e = KAYIT_MAKINE_ETIKETI.get(str(h or '').strip())
        if e:
            return e
    return ''


def sabit_hat_adi(lokasyon, bolum):
    """(lokasyon, bolum) → vardiyada seçilecek sabit hat adı; yoksa '' (hat vardiyada seçilir)."""
    return KAYITTA_HAT.get(((lokasyon or 'TK2').upper(), (bolum or '').strip()), '')


def kayit_makine_ad(robot_no, istasyon):
    """(vardiya hattı, istasyon no) → kaydın makine/alt-hat adı.
    Sabit hat değilse ya da istasyon 0/geçersizse '' (eski kayıt: hat vardiyadaydı)."""
    makineler = HAT_MAKINELERI.get(str(robot_no or '').strip())
    if not makineler:
        return ''
    try:
        i = int(istasyon or 0)
    except (TypeError, ValueError):
        return ''
    return makineler[i - 1] if 1 <= i <= len(makineler) else ''


def kayit_ist_no(robot_no, makine_ad):
    """(vardiya hattı, makine adı) → istasyon no. Listede yoksa 0."""
    makineler = HAT_MAKINELERI.get(str(robot_no or '').strip()) or []
    try:
        return makineler.index(str(makine_ad or '').strip()) + 1
    except ValueError:
        return 0


def kayit_hatti(robot_no, istasyon):
    """Kaydın GERÇEKTE çalışıldığı hat: sabit hatlı bölümlerde kaydın makinesi,
    diğerlerinde (ve eski kayıtlarda) vardiyanın hattı. Tel'de kod eki, sayaç
    eşlemesi ve rapor adımı BU değerden türer."""
    return kayit_makine_ad(robot_no, istasyon) or str(robot_no or '').strip()


# ── TEST CİHAZI (SVP) EŞLEMESİ (kullanıcı 2026-08-18) ────────────────────────
# "Test cihazı ile çalışıyorum" kutusu HER KAYITTA çıkmasın: yalnız gerçekten SVP
# test cihazından sayılan hat/makinelerde görünsün, orada da SADECE o hattın
# cihazları listelensin (operatör 40 cihazlık listeden doğru olanı aramasın).
#
# Anahtar: (lokasyon, bolum, hat) — hat None ise o bölümün TÜM hatları.
#   'hat' = vardiyanın robot_no'su; makinesi ÜRETİM KAYDINDA seçilen bölümlerde
#   (plastik/tel/TK1 montaj) KAYDIN hat adı (bkz. KAYITTA_HAT).
# Değer: SVP device_id listesi. TEK cihaz varsa mobil onu KENDİLİĞİNDEN seçer.
#
# GERİYE UYUM — BİLİNÇLİ: bir (lokasyon, bolum) için HİÇ kayıt yoksa eski davranış
# sürer (kutu görünür, TÜM cihazlar listelenir). TK2 montajda cihaz 31/35/53 hâlâ
# sahada kullanılıyor (M3/M7/M11, son kayıt 2026-07-23); kullanıcı TK2 eşlemesini
# vermeden burayı kapatmak canlı bir özelliği sessizce keserdi.
# Bölümün EN AZ BİR kaydı varsa eşlenmemiş hatlarda kutu GİZLENİR (TK1 plastikte
# yalnız Sizdirmazlik Test, tel'de yalnız iki yarı otomat hattı gibi).
#
# device_id'ler beta.coflesvp.com/device/<id>/test-logs adresindeki id'dir.
TEST_CIHAZ_ESLEME = {
    # TK1 montaj: hatlarda ayrıca buton/sayaç modülü de var (TK1-M1..M7) — operatör
    # test cihazını işaretlemezse ESP32 sayar, işaretlerse SVP sayar.
    ('TK1', 'montaj',  None):               [22, 65, 57],   # JKP · JKP YENI · JKW
    # TK1 plastik: sızdırmazlık testinin sayacı SVP 'Tank'. Yapistirma/320T/407T'nin
    # kendi ESP32 sayacı var → onlarda kutu HİÇ görünmez.
    ('TK1', 'plastik', 'Sizdirmazlik Test'): [12],          # Tank
    # TK1 tel — bu makinelere saha modülü TAKILMAYACAK (kullanıcı), adet doğrudan
    # SVP'den gelir. Yarı otomat ikilisi 2026-08-20'de sahada YER DEĞİŞTİRDİ
    # (kullanıcı) — hangi makinenin hangi SVP cihazına bağlı olduğu ancak sahada
    # görülüyor, bu yüzden eşleme kodda tek satırda tutuluyor.
    ('TK1', 'tel',     'Yarı Otomatik 1'):   [48],          # PP LINE 7 (Otomatik 1)
    ('TK1', 'tel',     'Yarı Otomatik 2'):   [43],          # PP LINE 8 (Otomatik 2)
    # HAT ADI 2026-08-21'de 'Tam Otomatik' → 'Otomatik Kesim Makinesi' oldu
    # (tel_proses.TEL_HATLARI). Bu sözlük ADIMLA değil HAT ADIYLA aranıyor
    # (test_cihaz_eslemesi → kayit_hatti sonucu) — anahtar güncellenmediği için
    # o makinede SVP cihazı otomatik seçili GELMİYORDU. Eski anahtar ulaşılamaz
    # hale geldi (kayit_makine_ad her zaman GÜNCEL listeden okur), o yüzden
    # bırakılmadı. HAT ADI DEĞİŞTİRİLİRSE BURASI DA DEĞİŞMELİ.
    ('TK1', 'tel',     'Otomatik Pres Makinesi'): [41],    # PP OTOMATIK SPIRAL PRESLEME
}


def test_cihaz_eslemesi(lokasyon, bolum, hat):
    """(lokasyon, bolum, hat) → (device_id listesi, bolum_eslenmis_mi).

    Dönen liste boş + bolum_eslenmis True  → bu hatta test cihazı YOK (kutu gizlenir).
    Dönen liste boş + bolum_eslenmis False → eşleme tanımlanmamış (eski davranış:
                                             tüm cihazlar listelenir).
    """
    lok = (lokasyon or 'TK2').upper()
    bol = (bolum or '').strip()
    hat_ad = (hat or '').strip()
    bolum_eslenmis = any(k[0] == lok and k[1] == bol for k in TEST_CIHAZ_ESLEME)
    if not bolum_eslenmis:
        return [], False
    return (list(TEST_CIHAZ_ESLEME.get((lok, bol, hat_ad))
                 or TEST_CIHAZ_ESLEME.get((lok, bol, None))
                 or []), True)


# CİHAZIN FIRMWARE'DE BİLDİRDİĞİ BÖLÜM (kullanıcı 2026-08-20).
# YF1 ve TK1-M4 modülleri montaj olarak flash'lı ve pilot.db'ye bolum='montaj'
# yazıyor; ama artık TEL'in son montaj hattında kullanılıyorlar (butonlar 4 ve 5).
# Sayım (bolum, robot_no) ikilisiyle yapıldığı için kaydın bölümü ('tel') ile
# sorgulanırsa HİÇ PULSE BULUNAMAZ ve sayaç sessizce 0 okurdu.
# Yeniden flash GEREKMESİN diye çeviri burada.
_CIHAZ_FW_BOLUM = {'YF1': 'montaj', 'TK1-M4': 'montaj'}


def _fw_bolum(bolum, cihaz):
    """Pilot DB'de bu cihazın sinyalleri hangi bölüm adıyla duruyor?"""
    return _CIHAZ_FW_BOLUM.get(cihaz, bolum)


def _sayac_cihazi(bolum, robot_no, istasyon=0):
    """Vardiya hattı + üretim kaydı istasyonundan pilot.db CIHAZ robot_no'sunu bulur.
    - pres / plastik: hat sabit, MAKİNE istasyondan gelir (pres 1=Abkant 1 ... 10=Bros;
      plastik 1=320T ... 4=Sizdirmazlik Test). istasyon 0 ise (eski kayıtlar: makine
      vardiyada seçiliyordu) hat adı aynen kullanılır → geriye dönük uyum korunur.
    - diğer bölümler: hat→cihaz eşlemesi (LF-LFP→YF1), yoksa hattın kendisi."""
    ad = kayit_makine_ad(robot_no, istasyon)
    if ad:
        # Hat→cihaz eşlemesi kayıt hattına da uygulanır (TK1 montajda LF-LFP→YF1)
        return HAT_SAYAC_CIHAZI.get(ad, ad)
    return HAT_SAYAC_CIHAZI.get(robot_no, robot_no)


def _sayac_destekli(bolum, robot_no, istasyon=0):
    """KAYIT düzeyi: bu üretim kaydı için otomatik sensör sayımı var mı (allowlist)?
    KATI — pres'te makine (istasyon) çözülemiyorsa False döner. Aksi halde makinesiz
    kayıt 'auto' işaretlenir, sayım hep 0 okur ve gerçek ok_adet'i EZERDİ."""
    return _sayac_cihazi(bolum, robot_no, istasyon) in SAYAC_AUTO_CIHAZLAR


def _sayac_destekli_bolum(bolum, robot_no):
    """VARDİYA düzeyi UI ipucu: bu vardiyada sensör sayımı mümkün mü? pres'te makine
    üretim kaydında seçildiğinden bölümdeki herhangi bir sayaçlı makine yeterlidir
    (mobil sensör rozeti/sıfırla butonu vardiya açılışında da görünsün)."""
    makineler = HAT_MAKINELERI.get(str(robot_no or '').strip())
    if makineler:
        return any(HAT_SAYAC_CIHAZI.get(m, m) in SAYAC_AUTO_CIHAZLAR for m in makineler)
    return _sayac_destekli(bolum, robot_no, 0)

# Yönetici kullanıcı — operatorler tablosunda 'Admin' (PIN varsayılan 9999, panelden
# değiştirilebilir). Bu adla PIN girişi yapan TÜM vardiyalara erişir (sahiplik kontrolü
# bypass, mobilde kart kilidi yok) ve başka operatör adına vardiya açabilir.
ADMIN_ADI = 'Admin'

# Geçerli bölüm anahtarları — TEK kaynak. isleme/lazer/pres (2026-07): TK2'ye eklenen
# yeni bölümler; sayaç/andon/iş takibi YOK, sadece vardiya + üretim girişi + rapor.
# Yeni bölüm eklerken: import_excel.BOLUM_SAYFA + mobil/dashboard/rapor template
# toggle'ları da güncellenmeli (anahtarlar ASCII slug — URL/CSS class/JS map'lerde kullanılır).
GECERLI_BOLUMLER = ('kaynak', 'montaj', 'metal', 'isleme', 'lazer', 'pres', 'plastik', 'tel')
BOLUM_AD = {
    'kaynak': 'Robot Kaynak',
    'montaj': 'Montaj',
    'metal':  'Metal Enjeksiyon',
    'isleme': 'İşleme',
    'lazer':  'Lazer Kesim',
    'pres':   'Pres Abkant',
    'plastik': 'Plastik Enjeksiyon',
    'tel':    'Tel Üretimi',
}

# ── TEL ÜRETİMİ PROSES ADIMLARI (TK1, 2026-08-04) ──────────────────────────
# Bir tel referansı birden çok HATTAN geçer; her adımı ayrı kaydedilir.
# 2026-08-19: adım artık vardiyanın değil ÜRETİM KAYDININ hattından gelir
# (bkz. KAYITTA_HAT) — aynı operatör aynı vardiyada birden çok adım kaydedebilir.
#
# HİÇBİR ADIM ZORUNLU DEĞİL (kullanıcı 2026-08-19: "her üründe otomatik hazırlık
# yapılmak zorunda değil"). Referans bazlı proses tanımı (referans_listesi.
# tel_adimlar) 2026-08-04'te TERK EDİLDİ: kapaması/son montajı dışarıda yapılacak
# ürünler sabit değil. Yapılmayan adımın kaydı hiç açılmaz, raporda '—' görünür.
#
# ⚠ ESKİ KURAL DEĞİL: "üretim yalnız SON adımda sayılır" kuralı KALDIRILDI. Çoklu
# sayım artık KOD EKİYLE çözülüyor (her adım '… KAPAMA' gibi ayrı kodda toplanır)
# ve rapor ürün bazlı adım kırılımı veriyor — adımlar tek toplamda birleştirilmez.
# Yarı/tam otomatik AYNI SIRADA iki alternatif hattır (ürün hangisine düşerse).
# Tanımlar tel_proses.py'de — mail_raporu.py ve oee.py de aynı kuralı kullanıyor,
# app.py'yi import edemedikleri için (Flask'ı ayağa kaldırır) ortak modüle taşındı.
# (import DOSYA BAŞINDA — TK1_ROBOT_NOLARI yukarıda TEL_HATLARI'nı kullanıyor.)

# Push bildirim alıcıları — bölüme yeni referans/launch eklendiğinde kimin
# telefonuna bildirim gitsin. İsimler operatorler tablosundaki adlarla
# (harf/Türkçe karakter duyarsız) eşleştirilir. Yeni alıcı: buraya ekle.
BILDIRIM_ALICILARI = {
    'kaynak': ['Cihan Zilci', 'Mevlüt Bora'],
    'montaj': ['Mustafa Kuş'],
}

# Metal "makine calisiyor" durumu — her sinyal aslinda makine-calisiyor rolesi
# darbesidir (400T firmware o ~22sn aktif roleyi izler). Yesil SADECE role aktifken
# yansin diye esik role-aktif penceresine yakin tutulur: sinyal gelince yesil, ~20sn
# sonra (role birakinca / sinyal kesilince) soner. Boylece andon makinenin nabzini
# gosterir. NOT: andon bu durumu 3sn'de bir hizli poll eder (ana 30sn refresh'i
# beklemeden) — yoksa 20sn'lik pencere 30sn poll arasinda kacardi.
METAL_CALISIYOR_ESIK_SN = 20


def _cop_key(uretim_tarihi, kod):
    """COP dedup anahtari: 'uretim_tarihi|NORMALIZE(kod)'.

    NORMALIZE = tum bosluklar silinir + BUYUK harf — kod tabaninin her yerinde
    kullanilan kanonik karsilastirmanin aynisi. NEDEN (2026-08-17): panel
    "COP verildi"yi bu anahtarla ariyor, log satirini ise robota GONDERILEN kod
    yaziyor. Yalniz upper() yapmak yetmedi; kod ERP'den bosluk dolgulu ('92.10.4312
    B') ya da operatorun yazdigi gibi gelebiliyor. Anahtarin iki tarafi ayni
    normalizasyondan gecmezse gonderim "zaten girilmis" derken panel satiri
    "bekliyor" gostermeye devam ediyor ve satir KISIR DONGUYE giriyor:
    listede duruyor ama gonderilince atlaniyor. Frontend de ayni normalizasyonu
    uygular (bosluk sil + upper).
    """
    return f"{uretim_tarihi}|" + re.sub(r'\s+', '', str(kod or '')).upper()


def _pilot_db_yolu():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pilot', 'pilot.db')


def _pilot_pulse_say(bolum, robot_no, istasyon, basla_ts, biti_ts=None):
    """Pilot DB'de (bolum, robot_no) için basla_ts ile biti_ts arasındaki pulse sayısı.
    istasyon > 0 ise o istasyona filtrele; 0/None ise tüm istasyonlar (montaj/metal
    tek istasyon, firmware istasyon=1 gönderir → filtre yok hepsini sayar)."""
    # Hat → cihaz robot_no eşlemesi (LF-LFP→YF1; pres'te istasyon→makine)
    cihaz = _sayac_cihazi(bolum, robot_no, istasyon)
    # PRES/PLASTİK: istasyon MAKİNE seçimidir, cihaz-içi istasyon değil — makineye
    # çevrildiyse istasyon filtresi KALKAR (abkant/pres/plastik firmware'i montaj gibi
    # istasyon=1 gönderir; filtre bırakılırsa istasyon=3 aranır ve hiç pulse bulunamazdı).
    if kayit_makine_ad(robot_no, istasyon):
        istasyon = 0
    bolum = _fw_bolum(bolum, cihaz)   # cihaz başka bölüm adıyla yazıyor olabilir
    robot_no = cihaz
    pilot_db = _pilot_db_yolu()
    if not basla_ts:
        return 0
    if not os.path.exists(pilot_db):
        return None   # pilot.db yok = OKUNAMADI (hata); gercek 0 ile karistirma
    import sqlite3
    try:
        pc = sqlite3.connect(pilot_db, timeout=5.0)
        try:
            pc.execute('PRAGMA busy_timeout=5000')   # kilitliyse 5sn bekle, hemen hata atma
            q = 'SELECT COUNT(*) FROM sayac_olaylari WHERE bolum=? AND robot_no=? AND ts >= ?'
            params = [bolum, robot_no, basla_ts]
            if biti_ts:
                q += ' AND ts < ?'
                params.append(biti_ts)
            if istasyon and istasyon > 0:
                q += ' AND istasyon=?'
                params.append(istasyon)
            row = pc.execute(q, params).fetchone()
            return row[0] if row else 0
        finally:
            pc.close()
    except Exception:
        return None   # KILIT/HATA: 0 DONDURME — caller ok_adet'i sifirlamasin


# ─────────────────────────────────────────────────────────────
# SAYAÇ PAYLAŞIMI — bir makinede birden fazla operatör (kullanıcı 2026-08-20)
# ─────────────────────────────────────────────────────────────
# "Bir tel referansını iki operatör birden seçebilsin, ikisi de aynı presi
#  seçsin, tek presten sayı alınsın. Üretim sonunda adedi ikiye bölmeye
#  çalışmasınlar; sinyal üretildiğinde her operatör için yarım yarım saysın —
#  fazla üretilmiş gibi yanılgıya düşülmesin."
#
# Eski davranış: iki kayıt da presin TÜM sinyallerini sayardı → aynı 100 parça
# panelde 200 görünürdü. Yeni davranış: bir sinyal üretildiği ANDA o referansta
# kaç kayıt canlıysa sinyal onlara eşit bölünür.
#
# ZAMAN DİLİMİ ÖNEMLİ: ikinci operatör işe sonradan katılabilir. Kaydın kendi
# penceresini körü körüne 2'ye bölmek, ilk operatörün YALNIZ çalıştığı süreyi de
# yarıya indirirdi (üretim kaybolurdu). Bu yüzden pencere, paylaşanların
# başlangıçlarıyla DİLİMLERE ayrılır ve her dilim o dilimde canlı olan kayıt
# sayısına bölünür.
#
# TAM SAYI'ya çevirirken KALAN DAĞITILIR (en büyük kesir yöntemi): 7 sinyal / 2
# kişi → 4 + 3 = 7. Herkes aşağı yuvarlansaydı 3 + 3 = 6 olur, her tur bir parça
# buharlaşırdı; yukarı yuvarlansaydı 8 olur, kullanıcının kaçınmak istediği
# "fazla üretim" yanılgısı geri gelirdi.


def _ref_normal(ref):
    """Referans karşılaştırma anahtarı — SQL'deki UPPER(REPLACE(kod,' ','')) ile aynı."""
    return str(ref or '').replace(' ', '').upper()


def _sayac_anahtari(bolum, robot_no, istasyon):
    """Kaydın beslendiği SİNYAL AKIŞININ kimliği: (pilot_bolum, cihaz, istasyon_filtresi).

    _pilot_pulse_say'in sorgu filtresiyle BİREBİR aynı üçlü — aynı anahtara sahip
    iki kayıt fiziksel olarak AYNI sensörden sayar. Buradaki çözüm sırası
    değişirse _pilot_pulse_say de değişmeli, yoksa paylaşım yanlış kayıtları
    eşleştirir."""
    cihaz = _sayac_cihazi(bolum, robot_no, istasyon)
    ist = 0 if kayit_makine_ad(robot_no, istasyon) else int(istasyon or 0)
    return (_fw_bolum(bolum, cihaz), cihaz, ist)


def _canli_auto_kayitlar(conn, vardiya_id):
    """Bu vardiyanın GÜNÜNDEKİ tüm ESP32 sayaç kayıtları — DONMUŞ OLANLAR DAHİL.

    DONMUŞ KAYITLAR DA ALINIR (kullanıcı 2026-08-26 hata bildirimi): eskiden
    yalnız sayac_otomatik=1 olanlar alınıyordu. İki operatör aynı makinede aynı
    referansı paylaşırken BİRİ DURUNCA (vardiyası kapandı / başka referansa geçti)
    o kayıt donuyor ve paylaşım kümesinden düşüyordu; kalan kaydın penceresi kendi
    başlangıcından itibaren TEK BAŞINA sayılmaya başlıyordu, yani birlikte
    çalıştıkları dönem İKİNCİ KEZ ona yazılıyordu. Sonuç: bir kayıt gerçeğin iki
    katına fırlıyor, ikisinin toplamı makinenin ürettiğini aşıyordu.
    Artık pencere bitişi (bitis_ts) de taşınır; paylaşım o ana kadar geçerli
    sayılır — donmuş kayıt payını korur, canlı olan fazlasını almaz.

    Test cihazı kayıtları hariç: onların sayacı pilot.db değil test_sonuclari.
    canli=1 ise kayıt hâlâ sayıyor (bitişi YOK sayılır)."""
    try:
        return conn.execute(
            "SELECT u.id, u.istasyon, u.sayac_baslangic_ts, u.bitis_ts, "
            "       u.referans_kodu, COALESCE(u.sayac_otomatik,0) AS canli, "
            "       COALESCE(v.bolum,'kaynak') AS bolum, v.robot_no "
            "FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id "
            "WHERE v.tarih = (SELECT tarih FROM vardiyalar WHERE id=?) "
            "  AND u.sayac_baslangic_ts IS NOT NULL "
            "  AND u.test_cihaz_id IS NULL",
            (vardiya_id,)).fetchall()
    except Exception as e:
        print(f'[_canli_auto_kayitlar] hata: {e}')
        return []


def _paylasanlar(canli, bolum, robot_no, istasyon, ref):
    """Aynı sensörü + aynı referansı paylaşan canlı kayıtlar: [(bas_ts, id), ...] artan.

    Referans şartı bilinçli: paylaşım YALNIZ 'iki kişi aynı işi yapıyor' halinde
    açılmalı. Aynı makinede farklı referansta bir kayıt varsa (unutulmuş açık
    vardiya gibi) bu kaydın adedi yarıya inmemeli.

    DÖNÜŞ: [(bas_ts, bitis_ts|None, id), ...] — 'bitis_ts' None ise kayıt hâlâ
    sayıyor. Bitişi taşımak ŞART: yoksa duran ortağın payı kalan kayda ikinci kez
    yazılır (bkz. _canli_auto_kayitlar)."""
    anahtar = _sayac_anahtari(bolum, robot_no, istasyon)
    rn = _ref_normal(ref)
    if not rn:
        return []
    grup = []
    for r in canli:
        if _ref_normal(r['referans_kodu']) != rn:
            continue
        if _sayac_anahtari(r['bolum'], r['robot_no'], r['istasyon'] or 0) != anahtar:
            continue
        _k = r.keys()
        _canli = r['canli'] if 'canli' in _k else 1
        _bit = (r['bitis_ts'] if 'bitis_ts' in _k else None)
        # Hâlâ sayan kaydın bitişi YOK sayılır: 'tamamlandı'dan geri alınmış bir
        # kayıtta eski damga kalmış olabilir, pencereyi erken kapatmasın.
        grup.append((r['sayac_baslangic_ts'], (None if _canli else _bit), r['id']))
    grup.sort(key=lambda g: (g[0], g[2]))
    return grup


def _paylasimli_pulse(bolum, robot_no, istasyon, kayit_id, grup, biti_ts=None):
    """Paylaşımlı sinyal sayımı — bu kaydın TAM SAYI payı.

    grup: _paylasanlar() çıktısı (en az 2 üye). Tek üye varsa çağıran zaten
    _pilot_pulse_say'i doğrudan kullanır (ek sorgu yok, eski davranış aynen).
    pilot.db okunamazsa None döner → çağıran ok_adet'i EZMEZ."""
    # DİLİM SINIRLARI: başlangıçlar + BİTİŞLER (2026-08-26). Bitişler olmadan,
    # duran ortağın ayrıldığı an bir sınır oluşturmuyordu; o âna kadarki dönem
    # kalan kayda bütün olarak yeniden yazılıyordu ("adet 2 katı" şikâyeti).
    sinirlar = sorted({g[0] for g in grup} | {g[1] for g in grup if g[1]})
    paylar = {g[2]: 0.0 for g in grup}
    toplam = 0
    for i, bas in enumerate(sinirlar):
        son = sinirlar[i + 1] if i + 1 < len(sinirlar) else biti_ts
        adet = _pilot_pulse_say(bolum, robot_no, istasyon, bas, son)
        if adet is None:
            return None                      # okunamadı — hiçbir payı yazma
        # Bu dilimde AKTİF olanlar: başlamış VE henüz bitmemiş kayıtlar.
        aktif = [kid for bts, bit, kid in grup if bts <= bas and (not bit or bit > bas)]
        if not aktif:
            continue
        toplam += adet
        for kid in aktif:
            paylar[kid] += adet / len(aktif)
    # Kalan dağıtımı: tam sayıya inerken toplam KORUNUR (4+3=7, 3+3=6 değil).
    # Sıralama (kesir, id) — deterministik: aynı veriyle her senkronda aynı sonuç.
    tam = {kid: int(p) for kid, p in paylar.items()}
    kalan = toplam - sum(tam.values())
    if kalan > 0:
        for kid in sorted(paylar, key=lambda k: (-(paylar[k] - int(paylar[k])), k))[:kalan]:
            tam[kid] += 1
    return tam.get(kayit_id, 0)


def _devir_ekle(cnt, devir):
    """Sayaçtan gelen adede DEVİR ADEDİNİ ekler (kaldığı yerden devam).

    Devir, kayıt 'tamamlandı'dan geri alınıp yeniden sayılmaya başlandığında
    o ana kadar birikmiş adettir (bkz. uretim_tamamlandi_guncelle). PARÇA
    cinsindendir → bölen/çarpan uygulandıktan SONRA eklenir.
    cnt None ise (pilot.db okunamadı) None döner: çağıran ok_adet'i EZMEZ."""
    if cnt is None:
        return None
    try:
        return cnt + max(0, int(devir or 0))
    except (TypeError, ValueError):
        return cnt


def _sayac_oku(conn, vardiya_id, kayit_id, bolum, robot_no, istasyon, basla_ts, ref,
               biti_ts=None, canli=None):
    """Bir üretim kaydının sensör adedi — paylaşım varsa payı, yoksa tam sayım.

    canli: _canli_auto_kayitlar() önceden çekilmişse (toplu senkron) geçilir;
    yoksa burada çekilir. Paylaşan yoksa TEK sorgu yapılır — yaygın durumda
    (herkes kendi makinesinde) ek maliyet yok."""
    try:
        if canli is None:
            canli = _canli_auto_kayitlar(conn, vardiya_id)
        grup = _paylasanlar(canli, bolum, robot_no, istasyon, ref)
    except Exception as e:
        print(f'[_sayac_oku] paylasim hatasi: {e}')
        grup = []
    # Kaydın kendisi grupta değilse (beklenmez; canlı küme ile kayıt arasında bir
    # tutarsızlık olursa) paylaşıma GİRME — _paylasimli_pulse tanımadığı id için 0
    # döner ve gerçek üretimi sıfırlardı.
    if len(grup) < 2 or kayit_id not in {g[2] for g in grup}:
        return _pilot_pulse_say(bolum, robot_no, istasyon, basla_ts, biti_ts)
    return _paylasimli_pulse(bolum, robot_no, istasyon, kayit_id, grup, biti_ts)


def _pilot_son_pulse_dk(bolum, robot_no, gun=None, istasyon=0):
    """Bu cihazdan en son pulse'tan bu yana geçen dakika. Hiç pulse yoksa None.
    10 dk duruş uyarısı için kullanılır. pres'te istasyon = operatörün o an
    çalıştığı makine (üretim kaydından gelir)."""
    robot_no = _sayac_cihazi(bolum, robot_no, istasyon)   # LF-LFP→YF1; pres: istasyon→makine
    bolum = _fw_bolum(bolum, robot_no)   # cihaz başka bölüm adıyla yazıyor olabilir
    pilot_db = _pilot_db_yolu()
    if not os.path.exists(pilot_db):
        return None
    import sqlite3
    from datetime import datetime as _dt
    gun = gun or date.today().isoformat()
    try:
        pc = sqlite3.connect(pilot_db)
        try:
            row = pc.execute(
                "SELECT MAX(ts) FROM sayac_olaylari WHERE bolum=? AND robot_no=? AND date(ts)=?",
                (bolum, robot_no, gun)
            ).fetchone()
            if not row or not row[0]:
                return None
            try:
                son = _dt.fromisoformat(row[0])
            except Exception:
                return None
            return max(0, int((_dt.now() - son).total_seconds() // 60))
        finally:
            pc.close()
    except Exception:
        return None


def _test_ref_norm(referans_kodu):
    """SVP test cihazı eşleşmesi için referans anahtarı (boşluksuz, BÜYÜK harf).

    TEL ADIM EKİ ATILIR (2026-08-19 saha hatası): tel kaydında kod adım ekiyle durur
    ('93.00 3062 YARI OTOMAT') ama test cihazı ürünü BASE koduyla bildirir
    ('93.00 3062' → test_conf_name). Ek atılmazsa referans_norm eşleşmesi HİÇ tutmaz,
    sayaç sessizce 0 okur ve operatör "test cihazı sayamıyor" der.
    Tel dışı kodlarda no-op (yalnız bilinen adım ekleri ayıklanır)."""
    return tel_ek_ayikla(referans_kodu).strip().upper().replace(' ', '')


def _test_basari_say(conn, cihaz_id, referans_kodu, basla_ts, biti_ts=None):
    """Cofle test cihazından (cofle_test poller'ı doldurur) basla_ts ile biti_ts
    arasında o referansa ait BAŞARILI test sayısı. ESP32 pulse sayımının test-cihazı
    karşılığı: 1 başarılı test = 1 pulse. Kaynak: test_sonuclari (ana DB).
    basla_ts/biti_ts yerel '%Y-%m-%d %H:%M:%S' string; epoch'a çevrilir. Hata → None."""
    if not cihaz_id or not basla_ts:
        return None
    try:
        bas_epoch = int(datetime.strptime(basla_ts, '%Y-%m-%d %H:%M:%S').timestamp())
    except Exception:
        return None
    norm = _test_ref_norm(referans_kodu)
    # UPPER(referans_norm): eşleşme HARF DUYARSIZ (kullanıcı 2026-08-19). Poller
    # kaydı zaten büyütüyor ama SVP'de kod küçük harfle tanımlanırsa ya da eski
    # kayıtlar farklı yazımla düştüyse sayaç sessizce 0 okurdu — SQLite '=' harfe
    # duyarlıdır, bu yüzden karşılaştırma açıkça normalize edilir.
    q = ("SELECT COUNT(*) FROM test_sonuclari "
         "WHERE cihaz_id=? AND basarili=1 AND ts_epoch>=? "
         "AND (?='' OR UPPER(referans_norm)=?)")
    params = [cihaz_id, bas_epoch, norm, norm]
    if biti_ts:
        try:
            params.append(int(datetime.strptime(biti_ts, '%Y-%m-%d %H:%M:%S').timestamp()))
            q += " AND ts_epoch < ?"
        except Exception:
            pass
    try:
        row = conn.execute(q, params).fetchone()
        return row[0] if row else 0
    except Exception:
        return None   # tablo yok/hata: ok_adet'i SIFIRLAMA


def _test_son_aktivite_dk(conn, vardiya_id):
    """Vardiyanın AKTİF test-cihazı kayıtları (sayac_otomatik=1 + test_cihaz_id) için
    son BAŞARILI testten beri geçen dakika. Pilot pulse'ın test-cihazı karşılığı:
    operatör test cihazıyla çalışırken ESP32'ye pulse düşmez — '10dk sinyal yok'
    duruş uyarısı test aktivitesine bakmazsa YANLIŞ duruş sayar.
    Henüz hiç test yoksa referans başlangıcından beri sayar (o da gerçek duruştur).
    Test-cihazı kaydı yoksa None (pilot mantığı tek başına geçerli)."""
    try:
        rows = conn.execute(
            "SELECT test_cihaz_id, referans_kodu, sayac_baslangic_ts FROM uretim_kayitlari "
            "WHERE vardiya_id=? AND sayac_otomatik=1 AND test_cihaz_id IS NOT NULL "
            "AND sayac_baslangic_ts IS NOT NULL",
            (vardiya_id,)
        ).fetchall()
        if not rows:
            return None
        en_yeni_epoch = None
        for r in rows:
            try:
                bas_epoch = int(datetime.strptime(r['sayac_baslangic_ts'], '%Y-%m-%d %H:%M:%S').timestamp())
            except Exception:
                continue
            norm = _test_ref_norm(r['referans_kodu'])
            row = conn.execute(
                "SELECT MAX(ts_epoch) FROM test_sonuclari "
                "WHERE cihaz_id=? AND basarili=1 AND ts_epoch>=? "
                "AND (?='' OR UPPER(referans_norm)=?)",
                (r['test_cihaz_id'], bas_epoch, norm, norm)
            ).fetchone()
            aktivite = (row[0] if row and row[0] else bas_epoch)  # hiç test yok → başlangıç anı
            if en_yeni_epoch is None or aktivite > en_yeni_epoch:
                en_yeni_epoch = aktivite
        if en_yeni_epoch is None:
            return None
        return max(0, int((datetime.now().timestamp() - en_yeni_epoch) // 60))
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────
# MAKİNE BAŞINA SİNYAL BÖLENİ (kullanıcı 2026-08-21)
# ─────────────────────────────────────────────────────────────
# Bazı makineler BİR PARÇA için birden fazla röle sinyali üretir. Bu, referansın
# değil MAKİNENİN özelliğidir — o makinedeki her referans için geçerlidir, o
# yüzden referans bazlı bukum_operasyon ile çözülmez.
#
# BROŞ (saha ölçümü 2026-08-21): broş aşağı inip çekerken röle çekiyor, tekrar
# ilk konuma çıkarken BİR KEZ DAHA çekiyor → 1 parça = 2 sinyal.
# İki hareket AYNI SÜREDE olduğu için sinyaller birbirinden ayırt EDİLEMEZ:
# tanı verisinde 406 sayımın tamamı ~25 sn aralıkla, pulse'lar 6,5-6,9 sn —
# yani dağılım tek tepeli. Gerçek çevrim ~50 sn, biz yarısında bir sayıyoruz.
# Kullanıcı doğruladı: "görülen adetin yarısı kadar üretim yapılmış oluyor."
#
# NEDEN FIRMWARE'DE DEĞİL SUNUCUDA:
#   · Firmware'de "her ikinci sinyali gönder" demek PARİTE durumu tutmak demek;
#     tek bir sinyal kaçarsa parite ters döner ve o andan sonra HEP yanlış sayar.
#   · Sunucuda bölme ham TOPLAM üzerinden yapılır ve her senkronda yeniden
#     hesaplanır — kaçan sinyal en fazla yarım parçalık sapma yaratır, kalıcı
#     kaymaz. (Aynı gerekçe _bukum_bol'un taban bölmesinde de geçerli.)
#   · OTA gerekmez; andon "makine çalışıyor" göstergesi ham sinyalle beslendiği
#     için etkilenmez (makine gerçekten sinyal üretiyor).
#
# YENİ MAKİNE EKLERKEN: önce ÖLÇ (python pilot\tani_analiz.py "<makine>" →
# [6] SİNYAL ↔ ÜRETİM oranı), tahminle değer yazma.
MAKINE_SINYAL_BOLEN = {
    'Bros': 2,      # broş: iniş + çıkış = 2 sinyal, 1 parça
}


def _sinyal_bolen(robot_no, istasyon, bukum_op):
    """Ham sinyali parçaya çevirmek için TOPLAM bölen.

    İki bölen ÇARPILIR çünkü bağımsızdırlar: makine bir parça için 2 sinyal
    üretiyorsa ve referans 3 büküm istiyorsa, 1 parça = 6 sinyaldir."""
    makine = kayit_makine_ad(robot_no, istasyon) or robot_no
    return max(1, int(bukum_op or 1)) * MAKINE_SINYAL_BOLEN.get(makine, 1)


def _bukum_bol(pulse, bukum_op):
    """Ham pulse sayısını ÜRETİLEN PARÇAYA çevirir (Pres Abkant büküm modeli).

    Bükümde bir parçaya sırayla tüm büküm operasyonları uygulanır, sonra sıradaki
    parçaya geçilir. 3 operasyonlu bir referansta abkant 3 kez sinyal üretir ama
    ortaya 1 parça çıkar → üretim teyidi 3. sinyaldir.

    TABAN BÖLME bilinçli: 7 pulse / 3 op = 2 parça (7. pulse üçüncü parçanın ilk
    bükümü, parça henüz BİTMEDİ). Kalan, kayıt açık kaldığı sürece bir sonraki
    senkronda tamamlanır — ok_adet her seferinde ham pulse'tan yeniden hesaplandığı
    için kayıp olmaz.

    pulse None (pilot.db okunamadı) ise None döner: çağıran ok_adet'i EZMEZ.
    bukum_op ≤ 1 ise dokunmaz — tüm eski kayıtlar ve diğer bölümler etkilenmez."""
    if pulse is None:
        return None
    try:
        n = int(bukum_op or 1)
    except (TypeError, ValueError):
        n = 1
    if n <= 1:
        return pulse
    return pulse // n


def _bukum_excel_yaz(ref, adet):
    """Büküm operasyon sayısını 'Pres Abkant Referans' sayfasına yazar (D kolonu).
    Sayfa 3 kolonluydu (Kod | Açıklama | Süre) → D yeni. Başlık yoksa açılır.
    EN İYİ ÇABA: Excel biri tarafından açıksa PermissionError gelir, üretim kaydı
    bundan ETKİLENMEZ — hata yalnız loglanır (bkz. referans/sureler aynı kalıp)."""
    try:
        from import_excel import EXCEL_YOL, BOLUM_SAYFA
        import openpyxl, os
        if not os.path.exists(EXCEL_YOL):
            return
        wb = openpyxl.load_workbook(EXCEL_YOL)
        sayfa_adi = BOLUM_SAYFA['pres']['ref']
        if sayfa_adi not in wb.sheetnames:
            return
        ws = wb[sayfa_adi]
        if not ws.cell(row=1, column=4).value:
            ws.cell(row=1, column=4, value='Büküm Op.')
        hedef = str(ref).upper().replace(' ', '')
        for r in range(2, ws.max_row + 1):
            hucre = ws.cell(row=r, column=1).value
            if hucre and str(hucre).upper().replace(' ', '') == hedef:
                ws.cell(row=r, column=4, value=int(adet))
                break
        wb.save(EXCEL_YOL)
    except Exception as e:
        print(f'[bukum] Excel sync hatasi ({ref}): {e}')


def _paket_carp(adet, paket):
    """Sayılan adedi PAKET ADEDİYLE çarpar (TK1 montaj: 1 buton = 10 ürün).

    _bukum_bol'un TERSİ: o böler (N sinyal = 1 parça), bu çarpar (1 sinyal = N
    parça). TK1'de tel ürünleri 10'arlı montajlanıyor — operatör sepeti bitirince
    bir kez basıyor, sepette 10 adet var.
    adet None (kaynak okunamadı) ise None döner: çağıran ok_adet'i EZMEZ.
    paket <= 1 ise dokunmaz — tüm eski kayıtlar ve diğer bölümler etkilenmez."""
    if adet is None:
        return None
    try:
        n = int(paket or 1)
    except (TypeError, ValueError):
        n = 1
    return adet if n <= 1 else adet * n


def _paket_acik_mi(bolum, lokasyon, hat):
    """Paket (montaj adet) çarpanı bu kayıtta anlamlı mı?

    YALNIZ tel bölümünün 'Son Montaj' adımı (kullanıcı 2026-08-20: "tk1 montaj
    masalarında olmasın, sadece son montaj hatları için geçerli"). Ürün orada
    10'arlı sepetlerle toplanıyor: operatör sepeti bitirince bir kez basıyor.
    Başka hiçbir yerde açılmaz — TK1/TK2 montajda 1 basış = 1 parça, çarpan açık
    kalsa üretim 10 katına çıkardı."""
    if (lokasyon or 'TK2').upper() != 'TK1':
        return False
    return (bolum or '') == 'tel' and tel_hat_adimi(hat) == 'Son Montaj'


def _paket_coz(c, gelen, ref, bolum, lokasyon, hat=None):
    """Üretim kaydına yazılacak paket adedini belirler (_bukum_op_coz kalıbı).
    Kapsam dışında her zaman 1 → özellik kapalı, hiçbir şey değişmez."""
    if not _paket_acik_mi(bolum, lokasyon, hat):
        return 1
    kayitli = 1
    try:
        row = c.execute(
            "SELECT COALESCE(paket_adedi,1) AS p FROM referans_listesi "
            "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
            "AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=? LIMIT 1",
            (ref, bolum, lokasyon)
        ).fetchone()
        if row:
            kayitli = int(row['p'] or 1)
    except Exception:
        kayitli = 1

    if gelen in (None, '', 0, '0'):
        return max(1, kayitli)
    try:
        yeni = int(gelen)
    except (TypeError, ValueError):
        return max(1, kayitli)
    yeni = max(1, min(999, yeni))   # 1..999 — sepet adedi büyük olabilir
    if yeni != kayitli:
        try:
            c.execute(
                "UPDATE referans_listesi SET paket_adedi=? "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
                "AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=?",
                (yeni, ref, bolum, lokasyon)
            )
        except Exception as e:
            print(f'[paket] referans guncellenemedi ({ref}): {e}')
    return yeni


def _bukum_op_coz(c, gelen, ref, bolum, lokasyon, hat=None):
    """Üretim kaydına yazılacak "1 parça = kaç sinyal" sayısını belirler.

    İKİ YERDE AÇIK (2026-08-19):
      · pres (abkant büküm operasyonu)
      · tel KAPAMA hatları — kullanıcı: "bir ürün için birden fazla kez prese
        basabiliyorlar". Kapama dışındaki tel adımlarında (kesim/otomat/son montaj)
        KAPALI: orada 1 sinyal = 1 parça, açık bırakılsa referansın hatırlanan
        kapama değeri o adımın sayacını da bölerdi.
    - Diğer bölümlerde her zaman 1 → özellik kapalı, hiçbir şey değişmez.
    - İstemci değer gönderdiyse: referansa da YAZILIR (bir sonraki üretimde hazır
      gelsin) ve Excel'e senkronlanır — ama YALNIZ değer gerçekten değiştiyse,
      yoksa her üretim kaydında 700 satırlık Excel açılıp kaydedilirdi.
    - Göndermediyse: referansın hatırlanan değeri kullanılır (yoksa 1)."""
    _acik = ((bolum or '') == 'pres'
             or ((bolum or '') == 'tel' and tel_hat_adimi(hat) == 'Kapama'))
    if not _acik:
        return 1
    kayitli = 1
    try:
        row = c.execute(
            "SELECT COALESCE(bukum_operasyon,1) AS b FROM referans_listesi "
            "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
            "AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=? LIMIT 1",
            (ref, bolum, lokasyon)
        ).fetchone()
        if row:
            kayitli = int(row['b'] or 1)
    except Exception:
        kayitli = 1

    if gelen in (None, '', 0, '0'):
        return max(1, kayitli)
    try:
        yeni = int(gelen)
    except (TypeError, ValueError):
        return max(1, kayitli)
    yeni = max(1, min(99, yeni))   # 1..99 — sıfır/negatif bölme çöktürür, 99 üstü hatalı giriş

    if yeni != kayitli:
        try:
            c.execute(
                "UPDATE referans_listesi SET bukum_operasyon=? "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
                "AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=?",
                (yeni, ref, bolum, lokasyon)
            )
        except Exception as e:
            print(f'[bukum] referans guncellenemedi ({ref}): {e}')
        # Excel senkronu YALNIZ pres: hedef dosya Pres Abkant referans sayfası,
        # tel kodu oraya yazılırsa sayfaya ait olmayan satır açar.
        if (bolum or '') == 'pres':
            _bukum_excel_yaz(ref, yeni)
    return yeni


def _uretim_sayac_senkron(conn, vardiya_id):
    """Auto modlu (sayac_otomatik=1) üretim kayıtlarının ok_adet'ini sensör sayımıyla
    günceller. GET /api/vardiya ve OEE okumadan önce çağrılır, böylece liste/rapor/OEE
    hep güncel sensör sayısını gösterir. İki sayaç kaynağı:
      • test_cihaz_id dolu → Cofle test cihazı (test_sonuclari, başarılı test)
      • aksi + robot allowlist'te → ESP32 pulse (pilot.db)"""
    try:
        v = conn.execute(
            "SELECT COALESCE(bolum,'kaynak') as bolum, robot_no FROM vardiyalar WHERE id=?",
            (vardiya_id,)
        ).fetchone()
        if not v:
            return
        rows = conn.execute(
            "SELECT id, istasyon, sayac_baslangic_ts, test_cihaz_id, referans_kodu, "
            "       COALESCE(bukum_operasyon, 1) AS bukum_operasyon, "
            "       COALESCE(paket_adedi, 1) AS paket_adedi, "
            "       COALESCE(sayac_devir_adet, 0) AS sayac_devir_adet "
            "FROM uretim_kayitlari "
            "WHERE vardiya_id=? AND sayac_otomatik=1 AND sayac_baslangic_ts IS NOT NULL",
            (vardiya_id,)
        ).fetchall()
        # Paylaşım adayları BİR KEZ çekilir (satır başına değil): bu senkron
        # dashboard/andon her yenilendiğinde çalışır, satır başına sorgu maliyeti
        # gereksiz yere katlanırdı.
        canli = _canli_auto_kayitlar(conn, vardiya_id) if rows else []
        for r in rows:
            # Allowlist kontrolü SATIR bazlı: pres'te makine üretim kaydındadır
            # (istasyon), dolayısıyla desteklenirlik satırdan satıra değişebilir.
            robot_allow = bool(v['robot_no']) and _sayac_destekli(
                v['bolum'], v['robot_no'], r['istasyon'] or 0)
            if r['test_cihaz_id']:
                cnt = _test_basari_say(conn, r['test_cihaz_id'], r['referans_kodu'],
                                       r['sayac_baslangic_ts'])
            elif robot_allow:
                # Aynı makinede aynı referansta başka canlı kayıt varsa sinyal
                # PAYLAŞILIR (bkz. _sayac_oku) — yoksa tek sorgulu eski yol.
                cnt = _sayac_oku(conn, vardiya_id, r['id'], v['bolum'], v['robot_no'],
                                 r['istasyon'] or 0, r['sayac_baslangic_ts'],
                                 r['referans_kodu'], canli=canli)
                cnt = _bukum_bol(cnt, _sinyal_bolen(v['robot_no'], r['istasyon'] or 0,
                                                   r['bukum_operasyon']))
            else:
                cnt = None
            # PAKET ÇARPANI en sona: sinyal sayısı önce parçaya çevrilir
            # (büküm bölmesi), sonra sepet adediyle çarpılır. Test cihazı
            # sayımında da geçerli — 1 başarılı test = 1 sepet olabilir.
            cnt = _paket_carp(cnt, r['paket_adedi'])
            # DEVİR en sonda: 'tamamlandı'dan geri alınan kayıt kaldığı yerden devam eder
            cnt = _devir_ekle(cnt, r['sayac_devir_adet'])
            if cnt is not None:   # kaynak okunamadıysa ok_adet'i SIFIRLAMA (gerçek üretimi ezme)
                conn.execute("UPDATE uretim_kayitlari SET ok_adet=? WHERE id=?", (cnt, r['id']))
        conn.commit()
    except Exception as e:
        print(f"[_uretim_sayac_senkron] hata: {e}")


def _acik_auto_vardiyalari_senkronla(conn, bolum=None, lokasyon=None):
    """Bugünün AÇIK vardiyalarındaki TÜM auto sayaç kayıtlarını (ESP32 + test cihazı)
    sensör sayımıyla tazeler. Andon ve dashboard-özet okumaları ÖNCESİ çağrılır —
    yoksa canlı makine adetleri (pilot rozeti) akarken performans/OEE bayat ok_adet'le
    hesaplanır (operatör mobili açık değilse saatlerce güncellenmez). Yalnız AÇIK
    vardiyalar: kapalı vardiyanın sayacı kapanış değerinde donuk kalmalı."""
    try:
        sql = ("SELECT DISTINCT v.id FROM vardiyalar v "
               "JOIN uretim_kayitlari u ON u.vardiya_id = v.id "
               "WHERE v.tarih = date('now','localtime') "
               "AND COALESCE(v.durum,'acik') != 'kapali' AND u.sayac_otomatik = 1")
        params = []
        if bolum:
            sql += " AND COALESCE(v.bolum,'kaynak') = ?"
            params.append(bolum)
        if lokasyon:
            sql += " AND COALESCE(v.lokasyon,'TK2') = ?"
            params.append(lokasyon)
        for r in conn.execute(sql, params).fetchall():
            _uretim_sayac_senkron(conn, r['id'])
    except Exception as e:
        print(f"[auto-senkron] hata: {e}")


# ─────────────────────────────────────────────────────────────
# OTOMATIK MOLALAR — sabit saatlerde planlı duruş (çay + yemek)
# ─────────────────────────────────────────────────────────────
# Şirket politikası: 09:00/13:00/15:00 çay (10dk), 11:00 yemek (30dk). Vardiya o
# saatten ÖNCE açıldıysa ve saat GELDİYSE duruş otomatik işlenir — operatör elle
# girmesin. AYNI TÜRDEN molalar tek duruş kaydında BİRİKİR (süre üzerine eklenir);
# yeni satır sadece o türden kayıt yoksa açılır. cay_molasi_bayrak bitmask'i her
# molayı bir kez işler (operatör kaydı silse de geri gelmez). Sadece BUGÜN.
# Çay tüm bölümlerde aynı; YEMEK bölüme göre değişir:
#   metal → 'Maden Tkv. ve Yemek Molası' 40dk (döküm alanı: maden takviyesi + yemek)
#   diğer → 'Yemek Molası' 30dk
def _bolum_molalari(bolum):
    m = [
        (1, '09:00', 10, 'Çay Molası'),
        (2, '13:00', 10, 'Çay Molası'),
        (4, '15:00', 10, 'Çay Molası'),
    ]
    if (bolum or 'kaynak') == 'metal':
        m.append((8, '11:00', 40, 'Maden Tkv. ve Yemek Molası'))
    else:
        m.append((8, '11:00', 30, 'Yemek Molası'))
    return sorted(m, key=lambda x: x[1])

def _saat_to_time(s):
    """'HH:MM' / 'HH:MM:SS' metnini datetime.time'a çevir; bozuksa None."""
    s = (s or '').strip()
    try:
        return datetime.strptime(s[:5], '%H:%M').time()
    except ValueError:
        return None

_TR_FOLD = str.maketrans('çÇıİöÖşŞüÜğĞ', 'cciioossuugg')
def _mola_norm(s):
    """Sebep adını eşleştirme için normalize et: Türkçe harf katla + küçült.
    'Çay Molası' / 'Cay Molasi' / 'ÇAY MOLASI' → 'cay molasi'."""
    return (s or '').translate(_TR_FOLD).lower().strip()

def _otomatik_mola_uygula(conn, vardiya):
    """Saati gelmiş otomatik molaları bu vardiyaya işler. Aynı türden planlı kayıt
    varsa süresini ARTIRIR, yoksa yeni satır açar. Idempotent (bitmask). Caller
    commit eder. Sadece bugünkü vardiya; eksik kolon/bozuk veri sessizce geçilir."""
    try:
        if 'cay_molasi_bayrak' not in vardiya.keys():
            return                      # migration henüz çalışmadı
    except Exception:
        return
    if (vardiya['tarih'] or '') != datetime.now().strftime('%Y-%m-%d'):
        return                          # geçmiş ayrı (tek seferlik düzenleme yapıldı)
    bas_t = _saat_to_time(vardiya['baslangic_saati'])
    if bas_t is None:
        return
    # Etkin bitiş: vardiya açıksa ŞİMDİ, kapalıysa bitiş saati
    if (vardiya['durum'] or '') == 'kapali':
        etkin_bitis = _saat_to_time(vardiya['bitis_saati']) or datetime.now().time()
    else:
        etkin_bitis = datetime.now().time()
    try:
        vardiya_bolum = (vardiya['bolum'] if 'bolum' in vardiya.keys() else 'kaynak') or 'kaynak'
    except Exception:
        vardiya_bolum = 'kaynak'
    bayrak = vardiya['cay_molasi_bayrak'] or 0
    yeni = bayrak
    for bit, mola_saati, mola_dk, sebep in _bolum_molalari(vardiya_bolum):
        if bayrak & bit:
            continue                    # bu mola zaten islendi
        mola_t = _saat_to_time(mola_saati)
        # Vardiya moladan ÖNCE açılmış + mola saati gelmiş olmalı
        if not (mola_t and bas_t < mola_t <= etkin_bitis):
            continue
        # Aynı türden mevcut planlı kayıt var mı? (yazım varyantlarına dayanıklı)
        hedef_id = None
        for r in conn.execute(
                "SELECT id, durus_sebebi FROM duruslar WHERE vardiya_id=? AND durus_tipi='planli' ORDER BY id",
                (vardiya['id'],)).fetchall():
            if _mola_norm(r['durus_sebebi']) == _mola_norm(sebep):
                hedef_id = r['id']
                break
        if hedef_id is not None:
            # BİRİKTİR: süreyi üzerine ekle, açıklamaya saati not düş
            # Süre büyüdü → bitiş damgası da uzamalı; başlangıç damgası boşsa
            # (elle girilmiş eski satır) molanın saatinden doldurulur.
            _mb, _mbit = _durus_ts_hesapla(vardiya, mola_saati, mola_dk)
            conn.execute(
                "UPDATE duruslar SET sure_dk = sure_dk + ?, "
                "aciklama = CASE WHEN COALESCE(aciklama,'')='' THEN ? ELSE aciklama || ' + ' || ? END, "
                "baslangic_ts = COALESCE(baslangic_ts, ?), "
                "bitis_ts = datetime(COALESCE(baslangic_ts, ?), "
                "                    '+' || CAST(sure_dk + ? AS TEXT) || ' minutes') "
                "WHERE id=?",
                (mola_dk, f'Otomatik: {mola_saati} ({mola_dk}dk)', f'{mola_saati} ({mola_dk}dk)',
                 _mb, _mb, mola_dk, hedef_id)
            )
        else:
            # Damga: molanın saati BELLİ (mola_saati) → doğrudan türetilir.
            # Zaman çizelgesinde çay/yemek molası da bar olarak görünür.
            _mb, _mbit = _durus_ts_hesapla(vardiya, mola_saati, mola_dk)
            conn.execute(
                "INSERT INTO duruslar (vardiya_id, durus_sebebi, aciklama, sure_dk, baslangic_saati, durus_tipi, baslangic_ts, bitis_ts) "
                "VALUES (?, ?, ?, ?, ?, 'planli', ?, ?)",
                (vardiya['id'], sebep, f'Otomatik: {mola_saati} ({mola_dk}dk)', mola_dk, mola_saati,
                 _mb, _mbit)
            )
        yeni |= bit
    if yeni != bayrak:
        conn.execute("UPDATE vardiyalar SET cay_molasi_bayrak=? WHERE id=?", (yeni, vardiya['id']))


# ─────────────────────────────────────────────────────────────
# WEB PUSH BİLDİRİM — operatör telefonuna (VAPID)
# ─────────────────────────────────────────────────────────────
# Operatör mobil sayfada PIN ile girince tarayıcı push aboneliği kaydedilir
# (/api/push/abone). Launch eklenince BILDIRIM_ALICILARI'ndaki operatörlerin
# tüm cihazlarına bildirim gider. Anahtarlar ilk kullanımda üretilip
# genel_ayarlar'da saklanır. pywebpush yoksa özellik sessizce devre dışı.
_vapid_cache = {}

def _vapid_keys():
    """VAPID anahtar çifti (yoksa üret, genel_ayarlar'a yaz). {'priv','pub'} | None."""
    if _vapid_cache:
        return _vapid_cache
    try:
        import base64
        from py_vapid import Vapid
        from cryptography.hazmat.primitives import serialization
        conn = db_connect()
        try:
            r_priv = conn.execute("SELECT deger FROM genel_ayarlar WHERE anahtar='push_vapid_priv'").fetchone()
            r_pub  = conn.execute("SELECT deger FROM genel_ayarlar WHERE anahtar='push_vapid_pub'").fetchone()
            if r_priv and r_pub:
                try:
                    Vapid.from_string(r_priv['deger'])   # format dogrulamasi (self-heal)
                    _vapid_cache.update(priv=r_priv['deger'], pub=r_pub['deger'])
                    return _vapid_cache
                except Exception:
                    print('[PUSH] Kayitli VAPID anahtari gecersiz formatta — yeniden uretiliyor')
            v = Vapid()
            v.generate_keys()
            # pywebpush, vapid_private_key olarak 32-byte HAM degerin base64url'unu bekler
            # (Vapid.from_string -> from_raw). PEM DEGIL.
            priv_raw = v.private_key.private_numbers().private_value.to_bytes(32, 'big')
            priv_b64 = base64.urlsafe_b64encode(priv_raw).rstrip(b'=').decode()
            pub_raw = v.public_key.public_bytes(
                serialization.Encoding.X962, serialization.PublicFormat.UncompressedPoint)
            pub_b64 = base64.urlsafe_b64encode(pub_raw).rstrip(b'=').decode()
            conn.execute("INSERT OR REPLACE INTO genel_ayarlar (anahtar, deger) VALUES ('push_vapid_priv', ?)", (priv_b64,))
            conn.execute("INSERT OR REPLACE INTO genel_ayarlar (anahtar, deger) VALUES ('push_vapid_pub', ?)", (pub_b64,))
            conn.commit()
            _vapid_cache.update(priv=priv_b64, pub=pub_b64)
            return _vapid_cache
        finally:
            conn.close()
    except Exception as e:
        print(f'[PUSH] VAPID anahtar hatası: {e}')
        return None

def _push_gonder(alici_adlari, baslik, govde, url='/'):
    """Adı geçen operatörlerin TÜM cihazlarına push gönder (senkron).
    Ölü abonelikler (404/410) otomatik silinir. Hata bildirimi düşürmez."""
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        print('[PUSH] pywebpush kurulu değil — bildirim atlandı')
        return
    keys = _vapid_keys()
    if not keys:
        return
    hedef = {_mola_norm(a) for a in alici_adlari}
    payload = json.dumps({'title': baslik, 'body': govde, 'url': url})
    try:
        conn = db_connect()
    except Exception as e:
        print(f'[PUSH] DB: {e}')
        return
    try:
        subs = [s for s in conn.execute('SELECT * FROM push_abonelikler').fetchall()
                if _mola_norm(s['operator_adi']) in hedef]
        for s in subs:
            try:
                webpush(
                    subscription_info={'endpoint': s['endpoint'],
                                       'keys': {'p256dh': s['p256dh'], 'auth': s['auth']}},
                    data=payload,
                    vapid_private_key=keys['priv'],
                    vapid_claims={'sub': 'mailto:bildirim@coflemanage.online'},
                    timeout=10,
                )
            except WebPushException as e:
                kod = getattr(getattr(e, 'response', None), 'status_code', None)
                if kod in (404, 410):   # abonelik ölmüş (uygulama silinmiş vb.)
                    conn.execute('DELETE FROM push_abonelikler WHERE id=?', (s['id'],))
                    conn.commit()
                else:
                    print(f'[PUSH] gönderim hatası ({s["operator_adi"]}): {e}')
            except Exception as e:
                print(f'[PUSH] {e}')
    finally:
        conn.close()

def _push_gonder_async(alici_adlari, baslik, govde, url='/'):
    """Push'u arka plan thread'inde gönder — HTTP yanıtını bekletmez."""
    import threading
    threading.Thread(target=_push_gonder, args=(list(alici_adlari), baslik, govde, url),
                     daemon=True).start()


@app.route('/sw.js')
def service_worker_js():
    """Service worker — kök kapsam (push almak için / scope şart)."""
    resp = send_file(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'sw.js'),
                     mimetype='application/javascript')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/api/push/vapid', methods=['GET'])
def push_vapid_public():
    keys = _vapid_keys()
    if not keys:
        return jsonify({'hata': 'push devre dışı'}), 503
    return jsonify({'public_key': keys['pub']})


@app.route('/api/push/abone', methods=['POST'])
@operator_required
def push_abone():
    """Operatör mobil cihazının push aboneliğini kaydet (PIN doğrulamalı)."""
    if not getattr(g, 'operator_adi', None):
        return jsonify({'hata': 'Operatör girişi gerekli'}), 400
    sub = (request.get_json() or {}).get('subscription') or {}
    endpoint = (sub.get('endpoint') or '').strip()
    anahtarlar = sub.get('keys') or {}
    p256dh, auth = anahtarlar.get('p256dh'), anahtarlar.get('auth')
    if not (endpoint and p256dh and auth):
        return jsonify({'hata': 'Geçersiz abonelik'}), 400
    conn = get_db()
    # endpoint UNIQUE: aynı cihaz tekrar abone olursa operatör adı güncellenir
    conn.execute('''
        INSERT INTO push_abonelikler (operator_adi, endpoint, p256dh, auth)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(endpoint) DO UPDATE SET
            operator_adi = excluded.operator_adi,
            p256dh = excluded.p256dh,
            auth = excluded.auth
    ''', (g.operator_adi, endpoint, p256dh, auth))
    conn.commit()
    return jsonify({'basarili': True})


def _vardiya_efektif_dk(vardiya):
    """Vardiyanın EFEKTİF süresi (dk) — canlı performans/kullanılabilirlik paydası.
    Kapalı: toplam_sure_dk (gerçek açılış→kapanış; kapatırken yazılır).
    Açık (bugün): başlangıçtan ŞU ANA kadar geçen dakika — süre ilerledikçe artar,
    performans buna göre canlı değişir. Açık ama eski günde unutulmuş: kayıtlı süre."""
    try:
        durum = (vardiya['durum'] or 'kapali')
        if durum == 'kapali':
            return vardiya['toplam_sure_dk'] or 0
        if (vardiya['tarih'] or '') == datetime.now().strftime('%Y-%m-%d'):
            bas = _saat_to_time(vardiya['baslangic_saati'])
            if bas is None:
                return vardiya['toplam_sure_dk'] or 0
            simdi = datetime.now().time()
            gecen = (simdi.hour * 60 + simdi.minute) - (bas.hour * 60 + bas.minute)
            if gecen < 0:
                gecen += 1440   # gece yarısını geçen vardiya
            return max(0, gecen)
        return vardiya['toplam_sure_dk'] or 0
    except Exception:
        try:
            return vardiya['toplam_sure_dk'] or 0
        except Exception:
            return 0


# ─────────────────────────────────────────────────────────────
# SAYFA ROUTES
# ─────────────────────────────────────────────────────────────

# ── CİHAZ TESİS SABİTLEME (kullanıcı 2026-08-18) ────────────────────────────
# Sorun: TK1 operatörü tarayıcıda /tk1'i açıp "ana ekrana/masaüstüne ekle" dediğinde
# kısayol sadece "coflemanage.online" olarak kaydediliyor (yol düşüyor) → uygulama
# '/' açılıyor, lokasyon TK2'ye dönüyor ve operatör TK2 ekranını görüyor.
# manifest.json'daki start_url='/tk1' bunu ÇÖZMÜYOR: iOS "Ana Ekrana Ekle" ve
# masaüstü "kısayol oluştur" akışları manifest'i dikkate almıyor, o anki adresi
# (bazen yalnız kök adresi) kaydediyor.
# Çözüm kısayoldan BAĞIMSIZ: /tk1 bir kez açıldığında cihaza uzun ömürlü bir çerez
# yazılır; '/' bu çerezi görürse /tk1'e yönlendirir. Böylece kısayol nasıl kaydedilmiş
# olursa olsun TK1 telefonu/bilgisayarı hep TK1 ekranını açar.
# ÇIKIŞ KAPISI: /tk2 (ya da /?tesis=TK2) çerezi siler — aynı cihaz TK2'ye dönebilir.
LOKASYON_COOKIE = 'cofle_lokasyon'
LOKASYON_COOKIE_OMUR = 60 * 60 * 24 * 365 * 2   # 2 yıl (fabrika cihazı kalıcı sayılır)


def _lokasyon_cerezi_yaz(resp, lokasyon):
    """Cihazı bir tesise sabitler. lokasyon=None → sabitlemeyi kaldırır.
    secure=False BİLİNÇLİ: sistem fabrika içinden düz http (192.168.x) ile de
    açılıyor; secure çerez orada hiç yazılmaz ve düzeltme çalışmazdı."""
    if lokasyon:
        resp.set_cookie(LOKASYON_COOKIE, lokasyon, max_age=LOKASYON_COOKIE_OMUR,
                        path='/', samesite='Lax')
    else:
        resp.set_cookie(LOKASYON_COOKIE, '', max_age=0, path='/', samesite='Lax')
    return resp


@app.route('/')
def operator_sayfasi():
    """Operatör mobil giriş sayfası (v2). Varsayılan lokasyon: TK2 (mevcut fabrika).

    Cihaz TK1'e sabitlenmişse (bkz. LOKASYON_COOKIE) /tk1'e yönlendirilir —
    kısayolun yolu düşse bile TK1 operatörü TK2 ekranını görmez.
    ?tesis=TK2 sabitlemeyi kaldırır (ortak/deneme cihazları için çıkış kapısı).
    """
    if (request.args.get('tesis') or '').upper() == 'TK2':
        return _lokasyon_cerezi_yaz(
            make_response(render_template('mobile_v2.html', lokasyon='TK2')), None)
    if request.cookies.get(LOKASYON_COOKIE) == 'TK1':
        return redirect('/tk1')
    return render_template('mobile_v2.html', lokasyon='TK2')


@app.route('/tk2')
def operator_tk2_sayfasi():
    """TK2 operatör sayfası — AÇIK adres. '/' ile aynı ekran, ek olarak cihazın
    TK1 sabitlemesini KALDIRIR: TK1'de denenmiş bir cihaz buradan TK2'ye döner."""
    return _lokasyon_cerezi_yaz(
        make_response(render_template('mobile_v2.html', lokasyon='TK2')), None)


@app.route('/mobile_v2')
def operator_v2_preview():
    """Operatör v2 önizlemesi (canlı ile aynı şablon — direkt link için saklanıyor)."""
    return render_template('mobile_v2.html', lokasyon='TK2')


@app.route('/tk1')
def operator_tk1_sayfasi():
    """TK1 (yan tesis) operatör mobil veri giriş sayfası. Lokasyon URL'den SABİT (TK1);
    operatör ekranda TK1/TK2 seçmez — bu adres TK1 operatörlerinin telefonu içindir.

    Bu sayfayı bir kez açan cihaz TK1'e SABİTLENİR (çerez): kısayol kök adrese
    kaydedilmiş olsa bile '/' buradan devam eder. Geri alma: /tk2
    """
    return _lokasyon_cerezi_yaz(
        make_response(render_template('mobile_v2.html', lokasyon='TK1')), 'TK1')


@app.route('/onizleme')
def operator_onizleme_sayfasi():
    """UI v3 "ENDÜSTRİYEL" ÖNİZLEMESİ (2026-07-29, design_handoff_ui_v3) —
    operatör ekranı adayı; canlı / ve /tk1'e DOKUNMAZ. OK/OEE sayaç halkası +
    ink/amber/kırmızı-sınır CTA reçetesi. Beğenilirse mobile_onizleme.html
    içeriği mobile_v2.html'e taşınır."""
    return render_template('mobile_onizleme.html', lokasyon='TK2')


@app.route('/mobile_legacy')
def operator_legacy_sayfasi():
    """Eski operatör tasarımı (geri dönüş için saklanıyor)."""
    return render_template('mobile.html')


@app.route('/dashboard')
def dashboard_sayfasi():
    """Yönetici dashboard — YENİ tasarım (v2: önce genel bakış, gruplu menü, i18n).
    Oturum yoksa giriş ekranı sunulur. Kullanıcının rol/izin bilgisi şablona
    enjekte edilir (ekran titremeden yetkisiz sayfalar gizlenir)."""
    ku = panel_kullanici()
    if not ku:
        return render_template('panel_giris.html')
    return render_template('dashboard_v2.html', panel_ku=json.dumps({
        'kullanici_adi': ku['kullanici_adi'],
        'ad_soyad': ku['ad_soyad'],
        'rol': ku['rol'],
        'admin': ku['admin'],
        'izinler': ku['izinler'],
        'sifre_gecici': ku['sifre_gecici'],
    }, ensure_ascii=False))


@app.route('/dashboard/onizleme')
def dashboard_onizleme_sayfasi():
    """UI v3 "ENDÜSTRİYEL" ÖNİZLEMESİ (2026-07-29, design_handoff_ui_v3) —
    canlı /dashboard'a DOKUNMAZ. Nötr gri + indigo + ink chrome; düz yüzeyler,
    gölgesiz, mono rakam. Beğenilirse dashboard_onizleme.html içeriği
    dashboard_v2.html'e taşınır. Auth/panel_ku canlı rotayla birebir aynı."""
    ku = panel_kullanici()
    if not ku:
        return render_template('panel_giris.html')
    return render_template('dashboard_onizleme.html', panel_ku=json.dumps({
        'kullanici_adi': ku['kullanici_adi'],
        'ad_soyad': ku['ad_soyad'],
        'rol': ku['rol'],
        'admin': ku['admin'],
        'izinler': ku['izinler'],
        'sifre_gecici': ku['sifre_gecici'],
    }, ensure_ascii=False))


@app.route('/dashboard/tezgah')
def dashboard_tezgah_sayfasi():
    """TASARIM ADAYI 2 — "Tezgah / Sıcak Nötr" (2026-07-28). Marka turuncusu
    (logo #EC6736) birincil kimlik, ılık nötr zemin, yoğun tablo, süs katmanı yok.
    /dashboard/onizleme (aday 1) ve /dashboard (canlı) ile yan yana karşılaştırılır.
    Canlıya DOKUNMAZ."""
    ku = panel_kullanici()
    if not ku:
        return render_template('panel_giris.html')
    return render_template('dashboard_tezgah.html', panel_ku=json.dumps({
        'kullanici_adi': ku['kullanici_adi'],
        'ad_soyad': ku['ad_soyad'],
        'rol': ku['rol'],
        'admin': ku['admin'],
        'izinler': ku['izinler'],
        'sifre_gecici': ku['sifre_gecici'],
    }, ensure_ascii=False))


@app.route('/dashboard_eski')
def dashboard_eski_sayfasi():
    """Eski yönetim paneli (v2 canlıya alınınca yedek olarak saklanıyor).
    Bu panel sayfa-izni filtresi TAŞIMAZ (her şeyi gösterir) — bu yüzden yalnız
    admin erişebilir; kısıtlı kullanıcı buradan kısıtlamayı aşamaz."""
    ku = panel_kullanici()
    if not ku:
        return render_template('panel_giris.html')
    if not ku['admin']:
        return redirect('/dashboard')
    return render_template('dashboard.html')


@app.route('/dashboard_legacy')
def dashboard_legacy_sayfasi():
    """Eski dashboard tasarımı (geri dönüş için saklanıyor). Admin-only (yukarıdaki
    gerekçeyle — izin filtresi yok)."""
    ku = panel_kullanici()
    if not ku:
        return render_template('panel_giris.html')
    if not ku['admin']:
        return redirect('/dashboard')
    return render_template('dashboard_legacy.html')


# ═══════════════════════ PANEL GİRİŞ / OTURUM ═══════════════════════
@app.route('/api/panel/giris', methods=['POST'])
def panel_giris():
    """Panel (dashboard) giriş. Body: {kullanici_adi, sifre}.
    Başarılı → imzalı oturum cookie'si kurulur. 401 = hatalı/pasif."""
    data = request.get_json(silent=True) or {}
    ka = (data.get('kullanici_adi') or '').strip().lower()
    sifre = (data.get('sifre') or '')
    if not ka or not sifre:
        return jsonify({'hata': 'Kullanıcı adı ve şifre zorunlu'}), 400
    conn = get_db()
    row = conn.execute(
        "SELECT id, sifre_hash, aktif FROM panel_kullanicilari WHERE lower(kullanici_adi)=?",
        (ka,)
    ).fetchone()
    if not row or not check_password_hash(row['sifre_hash'], sifre):
        return jsonify({'hata': 'Kullanıcı adı veya şifre hatalı'}), 401
    if not row['aktif']:
        return jsonify({'hata': 'Bu hesap pasif durumda. Yöneticinize başvurun.'}), 403
    conn.execute("UPDATE panel_kullanicilari SET son_giris=datetime('now','localtime') WHERE id=?", (row['id'],))
    conn.commit()
    session.permanent = True
    session['panel_uid'] = row['id']
    ku = panel_kullanici()
    return jsonify({'basarili': True, 'kullanici': {
        'kullanici_adi': ku['kullanici_adi'],
        'ad_soyad': ku['ad_soyad'],
        'rol': ku['rol'],
        'admin': ku['admin'],
        'izinler': ku['izinler'],
        'sifre_gecici': ku['sifre_gecici'],
    }}), 200


@app.route('/api/panel/cikis', methods=['POST'])
def panel_cikis():
    """Oturumu kapatır."""
    session.pop('panel_uid', None)
    session.clear()
    return jsonify({'basarili': True}), 200


@app.route('/api/panel/oturum', methods=['GET'])
def panel_oturum():
    """Aktif oturum bilgisi (frontend yetki senkronu için)."""
    ku = panel_kullanici()
    if not ku:
        return jsonify({'oturum': False}), 200
    return jsonify({'oturum': True, 'kullanici': {
        'kullanici_adi': ku['kullanici_adi'],
        'ad_soyad': ku['ad_soyad'],
        'rol': ku['rol'],
        'admin': ku['admin'],
        'izinler': ku['izinler'],
        'sifre_gecici': ku['sifre_gecici'],
    }}), 200


@app.route('/api/panel/sifre_degistir', methods=['POST'])
def panel_sifre_degistir():
    """Giriş yapmış kullanıcı kendi şifresini değiştirir.
    Body: {eski_sifre, yeni_sifre}. Geçici şifre bayrağı temizlenir."""
    ku = panel_kullanici()
    if not ku:
        return jsonify({'hata': 'Oturum gerekli', 'giris_gerekli': True}), 401
    data = request.get_json(silent=True) or {}
    eski = (data.get('eski_sifre') or '')
    yeni = (data.get('yeni_sifre') or '')
    if len(yeni) < 6:
        return jsonify({'hata': 'Yeni şifre en az 6 karakter olmalı'}), 400
    conn = get_db()
    row = conn.execute("SELECT sifre_hash FROM panel_kullanicilari WHERE id=?", (ku['id'],)).fetchone()
    if not row or not check_password_hash(row['sifre_hash'], eski):
        return jsonify({'hata': 'Mevcut şifre hatalı'}), 403
    conn.execute(
        "UPDATE panel_kullanicilari SET sifre_hash=?, sifre_gecici=0 WHERE id=?",
        (generate_password_hash(yeni), ku['id'])
    )
    conn.commit()
    return jsonify({'basarili': True}), 200


# ═══════════════════════ PANEL KULLANICI YÖNETİMİ (admin) ═══════════════════════
def _panel_ku_json(row):
    try:
        izinler = json.loads(row['izinler'] or '[]')
        if not isinstance(izinler, list):
            izinler = []
    except Exception:
        izinler = []
    admin = (row['rol'] == 'admin')
    return {
        'id': row['id'],
        'kullanici_adi': row['kullanici_adi'],
        'ad_soyad': row['ad_soyad'] or '',
        'rol': row['rol'],
        'izinler': PANEL_SAYFALAR[:] if admin else [s for s in izinler if s in PANEL_SAYFALAR],
        'aktif': bool(row['aktif']),
        'sifre_gecici': bool(row['sifre_gecici']),
        'son_giris': row['son_giris'] if 'son_giris' in row.keys() else None,
    }


def _temiz_izinler(deger):
    """Gelen izin listesini geçerli sayfa anahtarlarına süzer."""
    if not isinstance(deger, list):
        return []
    return [s for s in deger if s in PANEL_SAYFALAR]


@app.route('/api/panel/kullanicilar', methods=['GET'])
@panel_gerekli(admin=True)
def panel_kullanicilar_liste():
    """Tüm panel kullanıcılarını listeler (admin)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT id, kullanici_adi, ad_soyad, rol, izinler, aktif, sifre_gecici, son_giris "
        "FROM panel_kullanicilari ORDER BY (rol='admin') DESC, kullanici_adi"
    ).fetchall()
    return jsonify({'sayfalar': PANEL_SAYFALAR, 'kullanicilar': [_panel_ku_json(r) for r in rows]}), 200


@app.route('/api/panel/kullanicilar', methods=['POST'])
@panel_gerekli(admin=True)
def panel_kullanici_ekle():
    """Yeni panel kullanıcısı (admin). Body:
    {kullanici_adi, ad_soyad, sifre, rol, izinler[]}. sifre_gecici=1 verilir."""
    data = request.get_json(silent=True) or {}
    ka = (data.get('kullanici_adi') or '').strip().lower()
    ad_soyad = (data.get('ad_soyad') or '').strip()
    sifre = (data.get('sifre') or '')
    rol = 'admin' if (data.get('rol') == 'admin') else 'kullanici'
    izinler = _temiz_izinler(data.get('izinler'))
    if not ka:
        return jsonify({'hata': 'Kullanıcı adı zorunlu'}), 400
    if len(sifre) < 6:
        return jsonify({'hata': 'Şifre en az 6 karakter olmalı'}), 400
    conn = get_db()
    var = conn.execute("SELECT 1 FROM panel_kullanicilari WHERE lower(kullanici_adi)=?", (ka,)).fetchone()
    if var:
        return jsonify({'hata': f'Bu kullanıcı adı zaten var: {ka}'}), 409
    conn.execute(
        "INSERT INTO panel_kullanicilari (kullanici_adi, ad_soyad, sifre_hash, rol, izinler, aktif, sifre_gecici) "
        "VALUES (?, ?, ?, ?, ?, 1, 1)",
        (ka, ad_soyad, generate_password_hash(sifre), rol, json.dumps(izinler))
    )
    conn.commit()
    return jsonify({'basarili': True, 'kullanici_adi': ka}), 201


@app.route('/api/panel/kullanicilar/<int:uid>', methods=['PATCH'])
@panel_gerekli(admin=True)
def panel_kullanici_guncelle(uid):
    """Kullanıcı ad_soyad / rol / izinler / aktif günceller (admin).
    Kilitlenme koruması: admin kendi rolünü düşüremez/kendini pasifleştiremez;
    son aktif admin'in rolü düşürülemez."""
    data = request.get_json(silent=True) or {}
    conn = get_db()
    row = conn.execute("SELECT id, rol, aktif FROM panel_kullanicilari WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'hata': 'Kullanıcı bulunamadı'}), 404
    ben = g.panel_ku
    yeni_rol = row['rol']
    if 'rol' in data:
        yeni_rol = 'admin' if (data.get('rol') == 'admin') else 'kullanici'
    yeni_aktif = row['aktif']
    if 'aktif' in data:
        yeni_aktif = 1 if data.get('aktif') else 0

    # Kendini kilitleme koruması
    if uid == ben['id']:
        if yeni_rol != 'admin':
            return jsonify({'hata': 'Kendi yönetici rolünüzü kaldıramazsınız'}), 400
        if not yeni_aktif:
            return jsonify({'hata': 'Kendi hesabınızı pasifleştiremezsiniz'}), 400
    # Son admin koruması
    if row['rol'] == 'admin' and (yeni_rol != 'admin' or not yeni_aktif):
        aktif_admin = conn.execute(
            "SELECT COUNT(*) c FROM panel_kullanicilari WHERE rol='admin' AND aktif=1"
        ).fetchone()['c']
        if aktif_admin <= 1:
            return jsonify({'hata': 'Sistemde en az bir aktif yönetici kalmalı'}), 400

    ad_soyad = data.get('ad_soyad')
    if ad_soyad is None:
        ad_soyad = conn.execute("SELECT ad_soyad FROM panel_kullanicilari WHERE id=?", (uid,)).fetchone()['ad_soyad']
    else:
        ad_soyad = ad_soyad.strip()
    izinler = _temiz_izinler(data['izinler']) if 'izinler' in data else None
    if izinler is None:
        izinler_json = conn.execute("SELECT izinler FROM panel_kullanicilari WHERE id=?", (uid,)).fetchone()['izinler']
    else:
        izinler_json = json.dumps(izinler)
    conn.execute(
        "UPDATE panel_kullanicilari SET ad_soyad=?, rol=?, izinler=?, aktif=? WHERE id=?",
        (ad_soyad, yeni_rol, izinler_json, yeni_aktif, uid)
    )
    conn.commit()
    return jsonify({'basarili': True}), 200


@app.route('/api/panel/kullanicilar/<int:uid>/sifre', methods=['POST'])
@panel_gerekli(admin=True)
def panel_kullanici_sifre_sifirla(uid):
    """Admin bir kullanıcının şifresini sıfırlar. Body: {sifre}.
    sifre_gecici=1 → kullanıcı ilk girişte değiştirmeye zorlanır."""
    data = request.get_json(silent=True) or {}
    sifre = (data.get('sifre') or '')
    if len(sifre) < 6:
        return jsonify({'hata': 'Şifre en az 6 karakter olmalı'}), 400
    conn = get_db()
    row = conn.execute("SELECT id FROM panel_kullanicilari WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'hata': 'Kullanıcı bulunamadı'}), 404
    conn.execute(
        "UPDATE panel_kullanicilari SET sifre_hash=?, sifre_gecici=1 WHERE id=?",
        (generate_password_hash(sifre), uid)
    )
    conn.commit()
    return jsonify({'basarili': True}), 200


@app.route('/api/panel/kullanicilar/<int:uid>', methods=['DELETE'])
@panel_gerekli(admin=True)
def panel_kullanici_sil(uid):
    """Kullanıcı siler (admin). Kendini veya son aktif admin'i silemez."""
    ben = g.panel_ku
    if uid == ben['id']:
        return jsonify({'hata': 'Kendi hesabınızı silemezsiniz'}), 400
    conn = get_db()
    row = conn.execute("SELECT rol, aktif FROM panel_kullanicilari WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'hata': 'Kullanıcı bulunamadı'}), 404
    if row['rol'] == 'admin' and row['aktif']:
        aktif_admin = conn.execute(
            "SELECT COUNT(*) c FROM panel_kullanicilari WHERE rol='admin' AND aktif=1"
        ).fetchone()['c']
        if aktif_admin <= 1:
            return jsonify({'hata': 'Sistemde en az bir aktif yönetici kalmalı'}), 400
    conn.execute("DELETE FROM panel_kullanicilari WHERE id=?", (uid,))
    conn.commit()
    return jsonify({'basarili': True}), 200


@app.route('/logo')
def logo_serve():
    """Marka logosu (COFLE FORGE — 2026-07-29 kimlik yenilemesi).

    ?tema=koyu → static/logo_koyu.png   (wordmark BEYAZ; koyu zeminli ekranlar:
                                         andon panoları, giriş ekranı)
    ?tip=mark  → static/logo_mark.png   (SEMBOL tek başına, 512×512 kare)
                 Sekme ikonu/PWA için: 1600×480 lockup 16×16'ya sıkışınca
                 okunamayan bir çizgiye dönüşüyordu.
    varsayılan → static/logo.png        (açık tema lockup)

    Tercih sırası: istenen varyant → static/logo.png → masaüstü override →
    static/logo.svg. static/logo.png BAŞTA: masaüstündeki eski 'LOGO_COFLE ONLY.png'
    varsa bile repodaki resmi logo kazanır (handoff README'si tersini söylüyor,
    o bilgi static/logo.svg'deki bayat yoruma dayanıyor — kod hep böyleydi)."""
    proje_dir = os.path.dirname(os.path.abspath(__file__))
    if (request.args.get('tip') or '').strip() == 'mark':
        mark = os.path.join(proje_dir, 'static', 'logo_mark.png')
        if os.path.exists(mark):
            return send_file(mark, mimetype='image/png')
        mark_svg = os.path.join(proje_dir, 'static', 'logo_mark.svg')
        if os.path.exists(mark_svg):
            return send_file(mark_svg, mimetype='image/svg+xml')
    # LOCKUP'TA SVG ÖNCELİKLİ (2026-07-30): logo.png / logo_koyu.png dosyalarının
    # SAĞ KENARI KIRPIK — wordmark'ın son harfi ("FORGE"un E'si) kesiliyor. Ölçüm:
    # 1600×480 dosyada solda 109px boşluk var, sağda 0px ve son 14 sütun boyunca
    # dolu piksel sayısı sabit (35) — yani harf kenarda kesilmiş, boşluk bitmemiş.
    # Kullanıcı bunu iki kez bildirdi; CSS ile düzelmez, kaynak dosya kırpık.
    # SVG vektörel ve metni viewBox'a sığıyor → her ölçekte tam görünür.
    if (request.args.get('tema') or '').strip() == 'koyu':
        for ad, mime in (('logo_koyu.svg', 'image/svg+xml'), ('logo_koyu.png', 'image/png')):
            yol = os.path.join(proje_dir, 'static', ad)
            if os.path.exists(yol):
                return send_file(yol, mimetype=mime)
    logo_yollar = [
        (os.path.join(proje_dir, 'static', 'logo.svg'), 'image/svg+xml'),
        (os.path.join(proje_dir, 'static', 'logo.png'), 'image/png'),
        (os.path.join(os.path.expanduser('~'), 'OneDrive', 'Masaüstü', 'LOGO_COFLE ONLY.png'), 'image/png'),
        (os.path.join(os.path.expanduser('~'), 'Desktop', 'LOGO_COFLE ONLY.png'), 'image/png'),
    ]
    for yol, mime in logo_yollar:
        if os.path.exists(yol):
            return send_file(yol, mimetype=mime)
    return '', 404


# ───── Favicon / iOS / Android home-screen ikonları ─────
# Telefonda "Ana ekrana ekle" denildiğinde Cofle logosu görünür.
@app.route('/favicon.ico')
@app.route('/favicon.png')
@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
def favicon_serve():
    """Tüm tarayıcı/iOS/Android ikon istekleri SEMBOLE yönlendirilir.

    Eskiden lockup (1600×480) dönülüyordu: 16×16 sekme ikonuna küçülünce
    okunamayan bir çizgi oluyordu. logo_mark.png kare (512×512) ve tek başına
    okunur. Kaynak: handoff README, madde 5."""
    proje_dir = os.path.dirname(os.path.abspath(__file__))
    mark = os.path.join(proje_dir, 'static', 'logo_mark.png')
    if os.path.exists(mark):
        return send_file(mark, mimetype='image/png')
    return logo_serve()


@app.route('/manifest.json')
def web_manifest():
    """Android/iOS PWA manifesti — ana ekrana eklendiğinde Cofle logosu + neon tema.

    ?lokasyon=TK1 → start_url '/tk1' (TK1 operatör telefonu ana ekran kısayolu TK1'den
    açılır; yoksa PWA '/' açıp localStorage lokasyonu TK2'ye ezerdi ve TK2 operatörleri
    listelenirdi). id alanı lokasyona göre farklı → iki tesis kısayolu aynı telefonda
    ayrı uygulama olarak yaşayabilir.
    """
    lokasyon = (request.args.get('lokasyon') or 'TK2').upper()
    tk1 = (lokasyon == 'TK1')
    return jsonify({
        "id": "/tk1" if tk1 else "/",
        "name": "Cofle Forge TK1" if tk1 else "Cofle Forge",
        "short_name": "Cofle TK1" if tk1 else "Cofle",
        "description": "Cofle Forge Üretim Takip Sistemi" + (" — TK1" if tk1 else ""),
        "start_url": "/tk1" if tk1 else "/",
        "scope": "/",
        "display": "standalone",
        "orientation": "any",
        # Açılış (splash) renkleri UYGULAMANIN GERÇEK zeminiyle eşleşmeli: eskiden
        # koyu mor (#060414/#1c1f3a) idi, operatör ekranı tezgah paletine geçince
        # koyu bir splash'tan açık bir uygulamaya düşüyordu (göz alıyordu).
        "background_color": "#ECECF1",
        "theme_color": "#ECECF1",
        "lang": "tr",
        # PWA ikonu ayrı asset: /logo lockup'ı (1600×480) kare ikon yuvasına
        # sıkıştırılıyordu. pwa_icon.png 512×512 kare (handoff README, madde 6).
        "icons": [
            {"src": "/static/pwa_icon.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "/static/pwa_icon.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/static/pwa_icon.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"}
        ]
    })

@app.route('/andon')
def andon_sayfasi():
    """Andon TV pano sayfasi — Robot Kaynak (v5)."""
    return render_template('andon_v5.html',
                           bolum='kaynak',
                           bolum_ad='Robot Kaynak',
                           bolum_ikon='🔧',
                           panel_baslik='Robot Durum Paneli')


@app.route('/andon_legacy')
def andon_legacy_sayfasi():
    """Eski andon tasarımı (geri dönüş için saklanıyor)."""
    return render_template('andon.html')

@app.route('/andon_v2')
def andon_v2_preview():
    """Andon V2 onizleme sayfasi."""
    return render_template('andon_v2_preview.html')


@app.route('/rapor')
def rapor_sayfasi():
    """Operatörler için günlük üretim raporu paylaşım sayfası."""
    return render_template('rapor.html')


@app.route('/andon_v4')
def andon_v4_preview():
    """Andon v4 Tema Önizleme."""
    return render_template('andon_v4.html')


@app.route('/andon_v5')
def andon_v5_preview():
    """Andon v5 — komuta merkezi tema önizlemesi (Robot Kaynak)."""
    return render_template('andon_v5.html')


@app.route('/andon_montaj')
def andon_montaj_sayfasi():
    """Montaj Andon ekranı (v5 tasarım)."""
    return render_template('andon_v5.html',
                           bolum='montaj',
                           bolum_ad='Montaj',
                           bolum_ikon='🔩',
                           panel_baslik='Hat Durum Paneli')


@app.route('/andon_metal')
def andon_metal_sayfasi():
    """Metal Enjeksiyon Andon ekranı (v5 tasarım)."""
    return render_template('andon_v5.html',
                           bolum='metal',
                           bolum_ad='Metal Enjeksiyon',
                           bolum_ikon='🏭',
                           panel_baslik='Makine Durum Paneli')


@app.route('/andon_tv')
def andon_tv_sayfasi():
    """Andon TV KABUĞU (2026-08-04) — tam ekranı kalıcı tutar.

    Andon sayfası bellek birikimini önlemek için ~30 dk'da bir kendini yeniliyor;
    Fullscreen API sayfa yenilenince çıktığı ve kullanıcı jesti olmadan geri
    alınamadığı için TV yarım saatte bir tam ekrandan düşüyordu. Kullanıcı TV'ye
    uzaktan bağlanamıyor (yalnız fare) → kiosk modu kurulamıyor.

    Bu kabuk ASLA yenilenmez: tam ekranı o tutar, andon içeriği iframe'de durur ve
    periyodik olarak yalnız iframe tazelenir (bellek temizliği korunur, tam ekran
    bozulmaz). TV'de bir kez açılıp fareyle "Tam Ekran"a basılması yeterlidir.

    Kullanım: /andon_tv?bolum=metal   (kaynak | montaj | metal | tk1)
    Hedef sunucuda BEYAZ LİSTEDEN seçilir — dışarıdan gelen serbest URL iframe'e
    verilmez (açık yönlendirme / yabancı içerik gömme riski).
    """
    HEDEFLER = {'kaynak': '/andon', 'montaj': '/andon_montaj',
                'metal': '/andon_metal', 'tk1': '/andon_tk1'}
    bolum = (request.args.get('bolum') or 'kaynak').strip().lower()
    return render_template('andon_tv.html',
                           hedef=HEDEFLER.get(bolum, HEDEFLER['kaynak']))


@app.route('/andon_tk1')
def andon_tk1_sayfasi():
    """TK1 (yan tesis) Andon ekranı — montaj mantığı, lokasyon=TK1 (v5 tasarım)."""
    return render_template('andon_v5.html',
                           bolum='montaj',
                           lokasyon='TK1',
                           bolum_ad='TK1 Montaj',
                           bolum_ikon='🏢',
                           panel_baslik='TK1 Hat Durum Paneli')


@app.route('/andon_montaj_legacy')
def andon_montaj_legacy_sayfasi():
    """Eski Montaj andon tasarımı (geri dönüş için saklanıyor)."""
    return render_template('andon_montaj.html')


@app.route('/andon_metal_legacy')
def andon_metal_legacy_sayfasi():
    """Eski Metal Enjeksiyon andon tasarımı (geri dönüş için saklanıyor)."""
    return render_template('andon_metal.html')


@app.route('/dashboard_v3')
def dashboard_v3_preview():
    """Dashboard v3 Tema Önizleme."""
    return render_template('dashboard_v3.html')


# ─────────────────────────────────────────────────────────────
# VARDİYA API
# ─────────────────────────────────────────────────────────────

@app.route('/api/vardiya', methods=['GET'])
def vardiya_listesi():
    """Vardiya listesini döndür (filtrelenebilir)."""
    conn = get_db()
    c = conn.cursor()

    tarih = request.args.get('tarih')
    robot = request.args.get('robot')
    vardiya_turu = request.args.get('vardiya_turu')
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    limit = int(request.args.get('limit', 100))

    query = 'SELECT * FROM vardiyalar WHERE 1=1'
    params = []

    if tarih:
        query += ' AND tarih = ?'
        params.append(tarih)
    if robot:
        query += ' AND robot_no = ?'
        params.append(robot)
    if vardiya_turu:
        query += ' AND vardiya_turu = ?'
        params.append(vardiya_turu)
    if bolum:
        query += " AND COALESCE(bolum, 'kaynak') = ?"
        params.append(bolum)
    if lokasyon:
        query += " AND COALESCE(lokasyon, 'TK2') = ?"
        params.append(lokasyon)

    query += ' ORDER BY tarih DESC, created_at DESC LIMIT ?'
    params.append(limit)

    rows = c.execute(query, params).fetchall()
    conn.close()

    return jsonify([dict(r) for r in rows])


@app.route('/api/vardiya', methods=['POST'])
@operator_required
def vardiya_ekle():
    """Yeni vardiya kaydı oluştur.
    Operatör mobile'dan gelirse X-Operator header'ı body'deki operator_adi ile eşleşmeli.
    """
    data = request.get_json()

    zorunlu = ['tarih', 'vardiya_turu', 'robot_no', 'operator_adi']
    for alan in zorunlu:
        if not data.get(alan):
            return jsonify({'hata': f'"{alan}" alanı zorunludur'}), 400

    # Operatör mobile'dan gelen istekte header'daki operator = body'deki operator olmalı
    # (başkasının adına vardiya açma engeli). Dashboard modunda g.operator_adi None.
    # Admin bu kısıttan muaf — başka operatör adına vardiya açabilir.
    if (getattr(g, 'operator_adi', None) and g.operator_adi != data['operator_adi']
            and g.operator_adi != ADMIN_ADI):
        return jsonify({'hata': f'Sadece kendi adınıza vardiya açabilirsiniz ({g.operator_adi})'}), 403

    # Başlangıç: verilmediyse ŞU AN (operatör vardiyayı açtığı saat — sabit 07:00 yok).
    # Bitiş: vardiya KAPATILINCA belli olur; açılışta boş kalabilir (toplam_sure_dk=0,
    # açık vardiyada performans zaten "o ana kadar geçen süre" üzerinden hesaplanır).
    bas_str = (data.get('baslangic_saati') or '').strip() or datetime.now().strftime('%H:%M')
    bit_str = (data.get('bitis_saati') or '').strip()
    try:
        bas = datetime.strptime(bas_str, '%H:%M')
    except ValueError:
        return jsonify({'hata': 'Başlangıç saat formatı HH:MM olmalıdır'}), 400
    fark_dk = 0
    if bit_str:
        try:
            bit = datetime.strptime(bit_str, '%H:%M')
        except ValueError:
            return jsonify({'hata': 'Bitiş saat formatı HH:MM olmalıdır'}), 400
        fark_dk = int((bit - bas).total_seconds() / 60)
        if fark_dk < 0:
            fark_dk += 1440  # Gece yarısı geçen vardiyalar için

    conn = get_db()
    c = conn.cursor()
    c.execute('''
        INSERT INTO vardiyalar (tarih, vardiya_turu, robot_no, operator_adi, baslangic_saati, bitis_saati, toplam_sure_dk, notlar, bolum, lokasyon)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['tarih'],
        data['vardiya_turu'],
        data['robot_no'],
        data['operator_adi'],
        bas_str,
        bit_str,
        fark_dk,
        data.get('notlar', ''),
        data.get('bolum', 'kaynak'),
        (data.get('lokasyon') or request.args.get('lokasyon') or 'TK2')
    ))
    vardiya_id = c.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'basarili': True, 'vardiya_id': vardiya_id, 'sure_dk': fark_dk}), 201


@app.route('/api/vardiya/<int:vid>', methods=['GET'])
def vardiya_detay(vid):
    """Tek vardiya detayı (üretim + duruşlar dahil)."""
    conn = get_db()
    c = conn.cursor()

    vardiya = c.execute('SELECT * FROM vardiyalar WHERE id = ?', (vid,)).fetchone()
    if not vardiya:
        conn.close()
        return jsonify({'hata': 'Vardiya bulunamadı'}), 404

    # Otomatik molalar (çay/yemek) — saati gelmişse bu vardiyaya işle (idempotent)
    _otomatik_mola_uygula(c, vardiya)
    conn.commit()

    # Auto sayaç kayıtlarının ok_adet'ini sensör pulse sayımıyla tazele
    _uretim_sayac_senkron(conn, vid)

    uretim = c.execute('SELECT * FROM uretim_kayitlari WHERE vardiya_id = ?', (vid,)).fetchall()
    duruslar = c.execute('SELECT * FROM duruslar WHERE vardiya_id = ?', (vid,)).fetchall()

    # 10dk duruş uyarısı: son AKTİVİTE'den beri geçen dakika. İki kaynak:
    # ESP32 pulse (pilot.db) VE test cihazı başarılı testleri (test_sonuclari) —
    # test cihazıyla çalışan hatta pulse düşmez, yalnız pilot'a bakmak YANLIŞ duruş sayar.
    # En yakın (en küçük dk) aktivite geçerli.
    vbolum = (vardiya['bolum'] if 'bolum' in vardiya.keys() else None) or 'kaynak'
    # pres/plastik: makine üretim kaydında → o an çalışılan makineyi son auto kayıttan al
    # (duruş uyarısı doğru cihazın pulse'ına bakmalı; makine yoksa 0 = hat adı).
    _mak_ist = 0
    if vardiya['robot_no'] in HAT_MAKINELERI:
        _pr = c.execute(
            "SELECT istasyon FROM uretim_kayitlari WHERE vardiya_id=? "
            "AND sayac_baslangic_ts IS NOT NULL ORDER BY id DESC LIMIT 1", (vid,)).fetchone()
        _mak_ist = int((_pr['istasyon'] if _pr else 0) or 0)
    son_pulse_dk = _pilot_son_pulse_dk(vbolum, vardiya['robot_no'], istasyon=_mak_ist)
    test_dk = _test_son_aktivite_dk(conn, vid)
    if test_dk is not None:
        son_pulse_dk = test_dk if son_pulse_dk is None else min(son_pulse_dk, test_dk)
    conn.close()

    # Kaynakta istasyon sayısı MAKİNEYE bağlı (punta kaynak tek istasyonlu).
    # Mobil bunu vardiya yükünden okur — ayrı istek atmaz, kendi listesini tutmaz.
    _v_dict = dict(vardiya)
    _v_dict['kaynak_istasyon_sayisi'] = kaynak_istasyon_sayisi(vardiya['robot_no'])

    return jsonify({
        'vardiya': _v_dict,
        'uretim': [dict(r) for r in uretim],
        'duruslar': [dict(r) for r in duruslar],
        'son_pulse_dk': son_pulse_dk,
        # Bu cihazda sensör otomatik sayım destekli mi (allowlist) — mobil UI buna göre
        # sensör rozeti/ipucu/sıfırla butonu gösterir
        'sayac_destekli': _sayac_destekli_bolum(vbolum, vardiya['robot_no']),
    })


@app.route('/api/vardiya/<int:vid>/sayac_log', methods=['GET'])
def vardiya_sayac_log(vid):
    """Operatör mobil 'Sinyal Geçmişi' — vardiyanın sayaç olay logu (en yeni üstte, son 100).
    İki kaynak birleşik: ESP32 pulse (pilot.db sayac_olaylari, vardiya günü + cihaz)
    ve test cihazı kayıtları (test_sonuclari, referans başlangıcından beri)."""
    conn = get_db()
    v = conn.execute(
        "SELECT robot_no, COALESCE(bolum,'kaynak') as bolum, tarih FROM vardiyalar WHERE id=?",
        (vid,)
    ).fetchone()
    if not v:
        return jsonify({'hata': 'Vardiya bulunamadı'}), 404
    loglar = []

    # ── ESP32 pulse'ları (hat → cihaz eşlemesi: LF-LFP→YF1) ──
    robot = HAT_SAYAC_CIHAZI.get(v['robot_no'], v['robot_no'])
    pilot_db = _pilot_db_yolu()
    if os.path.exists(pilot_db):
        try:
            import sqlite3 as _sq
            pc = _sq.connect(pilot_db, timeout=5.0)
            pc.row_factory = _sq.Row
            try:
                pc.execute('PRAGMA busy_timeout=5000')
                for r in pc.execute(
                    "SELECT ts, istasyon FROM sayac_olaylari "
                    "WHERE bolum=? AND robot_no=? AND date(ts)=? ORDER BY ts DESC LIMIT 100",
                    (v['bolum'], robot, v['tarih'])
                ).fetchall():
                    loglar.append({
                        'ts': r['ts'], 'kaynak': 'sensor',
                        'detay': (f"İst.{r['istasyon']}" if (r['istasyon'] or 0) > 0 else 'Sayım'),
                    })
            finally:
                pc.close()
        except Exception as e:
            print(f"[sayac_log] pilot okuma hatası: {e}")

    # ── Test cihazı kayıtları (başarısızlar da listelenir — operatör tam geçmişi görsün) ──
    try:
        for u in conn.execute(
            "SELECT test_cihaz_id, referans_kodu, sayac_baslangic_ts FROM uretim_kayitlari "
            "WHERE vardiya_id=? AND test_cihaz_id IS NOT NULL", (vid,)
        ).fetchall():
            norm = (u['referans_kodu'] or '').strip().upper().replace(' ', '')
            bas_epoch = 0
            if u['sayac_baslangic_ts']:
                try:
                    bas_epoch = int(datetime.strptime(u['sayac_baslangic_ts'], '%Y-%m-%d %H:%M:%S').timestamp())
                except Exception:
                    pass
            for t in conn.execute(
                "SELECT ts_epoch, sonuc, basarili, cihaz_adi FROM test_sonuclari "
                "WHERE cihaz_id=? AND ts_epoch>=? AND (?='' OR referans_norm=?) "
                "ORDER BY ts_epoch DESC LIMIT 100",
                (u['test_cihaz_id'], bas_epoch, norm, norm)
            ).fetchall():
                loglar.append({
                    'ts': datetime.fromtimestamp(t['ts_epoch']).strftime('%Y-%m-%d %H:%M:%S'),
                    'kaynak': 'test',
                    'basarili': int(t['basarili'] or 0),
                    'detay': (t['sonuc'] or ('BAŞARILI' if t['basarili'] else 'BAŞARISIZ')),
                })
    except Exception as e:
        print(f"[sayac_log] test okuma hatası: {e}")

    loglar.sort(key=lambda x: x['ts'], reverse=True)
    return jsonify({'robot_no': v['robot_no'], 'loglar': loglar[:100]})


@app.route('/api/vardiya/<int:vid>', methods=['DELETE'])
@operator_required
def vardiya_sil(vid):
    """Vardiyayi sil (uretim ve duruslarla birlikte).
    Operatör sadece kendi vardiyasını silebilir.
    """
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    conn = get_db()
    conn.execute('DELETE FROM vardiyalar WHERE id = ?', (vid,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/vardiya/bugun', methods=['GET'])
def vardiya_bugun():
    """Bugune ait tum vardiylari getirir (operatorn secim ekrani icin). ?bolum= ve ?lokasyon= ile filtrelenebilir."""
    from datetime import date as _date
    bugun = _date.today().isoformat()
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    q = 'SELECT * FROM vardiyalar WHERE tarih = ?'
    params = [bugun]
    if bolum:
        q += " AND COALESCE(bolum, 'kaynak') = ?"
        params.append(bolum)
    if lokasyon:
        q += " AND COALESCE(lokasyon, 'TK2') = ?"
        params.append(lokasyon)
    q += ' ORDER BY baslangic_saati DESC'
    conn = get_db()
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/vardiya/<int:vid>/kapat', methods=['PATCH'])
@operator_required
def vardiya_kapat(vid):
    """Vardiyayi kapat ve bitis saatini guncelle.
    Operatör sadece kendi vardiyasını kapatabilir.
    """
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    data = request.get_json() or {}
    bitis = data.get('bitis_saati', datetime.now().strftime('%H:%M'))
    conn = get_db()
    c = conn.cursor()
    vardiya = c.execute('SELECT * FROM vardiyalar WHERE id = ?', (vid,)).fetchone()
    if not vardiya:
        conn.close()
        return jsonify({'hata': 'Vardiya bulunamadi'}), 404
    # Kapanmadan önce: vardiya boyunca geçen otomatik molaları kesinleştir (rapor tam olsun)
    _otomatik_mola_uygula(c, vardiya)
    # Saat parse KORUMALI: DB'de bozuk baslangic_saati varsa (eski/elle düzenlenmiş kayıt)
    # veya body'de bozuk bitis gelirse vardiya SONSUZA dek kapatılamaz hale gelmesin —
    # 'HH:MM:SS' toleransı + hata durumunda 400 JSON.
    try:
        bas = datetime.strptime(str(vardiya['baslangic_saati'])[:5], '%H:%M')
        bit = datetime.strptime(str(bitis)[:5], '%H:%M')
    except (ValueError, TypeError):
        conn.close()
        return jsonify({'hata': f"Saat formatı geçersiz (başlangıç='{vardiya['baslangic_saati']}', bitiş='{bitis}') — HH:MM olmalı"}), 400
    fark_dk = int((bit - bas).total_seconds() / 60)
    if fark_dk < 0:
        fark_dk += 1440
    # Auto sayaç kayıtlarını KAPANIŞ DEĞERİNDE DONDUR: son sensör/test sayımını yaz,
    # otomatik modu kapat. Yoksa kayıt sayac_otomatik=1 kaldığı için sonraki okumalar
    # (mobil GET/OEE) vardiya KAPANDIKTAN sonraki pulse'ları da bu kayda yazar —
    # yeni vardiyayla ÇİFT SAYIM olur.
    _uretim_sayac_senkron(conn, vid)
    c.execute("UPDATE uretim_kayitlari SET sayac_otomatik=0 WHERE vardiya_id=? AND sayac_otomatik=1", (vid,))
    # Hâlâ açık üretim kayıtlarının penceresi VARDİYANIN BİTİŞİNDE kapanır
    # (kullanıcı 2026-08-21). 'now' DEĞİL: vardiya geriye dönük kapatılıyorsa
    # (operatör unuttu, yönetici akşam kapattı) kayıtlar gerçekte olmadıkları
    # saatlere kadar uzardı. Gece vardiyasında bitiş ertesi güne taşınır.
    _kapanis_dt = _ts_parse(f"{vardiya['tarih']} {str(bitis)[:5]}")
    if _kapanis_dt and _ts_parse(f"{vardiya['tarih']} {str(vardiya['baslangic_saati'])[:5]}") \
            and _kapanis_dt < _ts_parse(f"{vardiya['tarih']} {str(vardiya['baslangic_saati'])[:5]}"):
        _kapanis_dt += timedelta(days=1)
    c.execute("UPDATE uretim_kayitlari SET bitis_ts=? WHERE vardiya_id=? AND bitis_ts IS NULL",
              (_ts_metin(_kapanis_dt) or datetime.now().strftime('%Y-%m-%d %H:%M:%S'), vid))
    c.execute(
        "UPDATE vardiyalar SET durum='kapali', bitis_saati=?, toplam_sure_dk=? WHERE id=?",
        (bitis, fark_dk, vid)
    )
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'sure_dk': fark_dk})


@app.route('/api/vardiya/<int:vid>/robotla_calisiyor', methods=['PATCH'])
@operator_required
def vardiya_robotla_calisiyor(vid):
    """Vardiya 'robotla çalışıyor / tam otomasyon' bayrağını aç/kapat.
    Metal enjeksiyonda makine+robot otomatik çalışırken kullanılır.
    Body: { robotla_calisiyor: 0|1 }
    """
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    data = request.get_json() or {}
    yeni = 1 if data.get('robotla_calisiyor') else 0
    conn = get_db()
    c = conn.cursor()
    v = c.execute('SELECT id FROM vardiyalar WHERE id=?', (vid,)).fetchone()
    if not v:
        conn.close()
        return jsonify({'hata': 'Vardiya bulunamadı'}), 404
    c.execute('UPDATE vardiyalar SET robotla_calisiyor=? WHERE id=?', (yeni, vid))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'robotla_calisiyor': yeni})


@app.route('/api/vardiya/<int:vid>/ac', methods=['PATCH'])
@operator_required
def vardiya_ac(vid):
    """Kapanmış vardiyayı yeniden aç."""
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    conn = get_db()
    c = conn.cursor()
    vardiya = c.execute('SELECT * FROM vardiyalar WHERE id = ?', (vid,)).fetchone()
    if not vardiya:
        conn.close()
        return jsonify({'hata': 'Vardiya bulunamadi'}), 404
    if vardiya['durum'] != 'kapali':
        conn.close()
        return jsonify({'hata': 'Vardiya zaten açık'}), 400
    # Vardiya yeniden açılıyor → KAPANIŞTA atılan üretim bitiş damgalarını geri al
    # (2026-08-21). Yoksa kayıtlar 'kapanmış' kalır ve zaman çizelgesinde barları
    # eski kapanış saatinde donardı. YALNIZ kapanışta atılanlar temizlenir:
    #   · tamamlandi_ts dolu   → operatör bitirdi, damga KORUNUR
    #   · bitis_ts ≠ kapanış anı → referans değişiminde donduruldu, KORUNUR
    _eski_kapanis = _ts_metin(_ts_parse(f"{vardiya['tarih']} {str(vardiya['bitis_saati'] or '')[:5]}"))
    _eski_kapanis_ertesi = None
    if _eski_kapanis:
        _b = _ts_parse(_eski_kapanis)
        _eski_kapanis_ertesi = _ts_metin(_b + timedelta(days=1))   # gece vardiyası karşılığı
    c.execute(
        "UPDATE uretim_kayitlari SET bitis_ts=NULL "
        "WHERE vardiya_id=? AND tamamlandi_ts IS NULL "
        "  AND bitis_ts IS NOT NULL AND bitis_ts IN (?, ?)",
        (vid, _eski_kapanis, _eski_kapanis_ertesi)
    )
    c.execute(
        "UPDATE vardiyalar SET durum='aktif', bitis_saati=NULL, toplam_sure_dk=NULL WHERE id=?",
        (vid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/vardiya/<int:vid>/hat', methods=['PATCH'])
@operator_required
def vardiya_hat_degistir(vid):
    """Açık vardiyanın hattını (robot_no) değiştir — operatör masa değiştirince
    vardiyayı kapatmadan geçtiği masanın sayacına bağlanır.

    Eski masadaki AKTİF sayaç kayıtları o ana kadarki değerde DONDURULUR (yeni
    masanın pulse'ları eski referansa karışmasın); operatör yeni masada '+ Ekle'
    ile sayımı sıfırdan başlatır. Bölüm değişmez (montaj içinde masa değişimi).
    Body: { robot_no }
    """
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    data = request.get_json() or {}
    yeni_robot = (data.get('robot_no') or '').strip()
    if not yeni_robot:
        return jsonify({'hata': 'Hat (robot_no) zorunludur'}), 400

    conn = get_db()
    try:
        c = conn.cursor()
        v = c.execute(
            "SELECT COALESCE(bolum,'kaynak') as bolum, robot_no, durum FROM vardiyalar WHERE id=?",
            (vid,)
        ).fetchone()
        if not v:
            return jsonify({'hata': 'Vardiya bulunamadı'}), 404
        if v['durum'] == 'kapali':
            return jsonify({'hata': 'Kapalı vardiyanın hattı değiştirilemez'}), 400

        eski_robot = v['robot_no']
        if yeni_robot == eski_robot:
            return jsonify({'basarili': True, 'degisti': False, 'robot_no': yeni_robot,
                            'sayac_destekli': (yeni_robot in SAYAC_AUTO_CIHAZLAR)})

        # Eski masadaki aktif sayaç kayıtlarını O ANA kadar dondur (eski robot_no ile say).
        # uretim_ekle'deki referans-değişimi dondurma mantığıyla aynı.
        simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        aktif = c.execute(
            "SELECT id, istasyon, sayac_baslangic_ts, referans_kodu, "
            "       COALESCE(bukum_operasyon,1) AS bukum_operasyon, "
            "       COALESCE(paket_adedi,1) AS paket_adedi "
            "FROM uretim_kayitlari "
            "WHERE vardiya_id=? AND sayac_otomatik=1 AND sayac_baslangic_ts IS NOT NULL",
            (vid,)
        ).fetchall()
        for o in aktif:
            fcnt = _sayac_oku(conn, vid, o['id'], v['bolum'], eski_robot,
                              o['istasyon'] or 0, o['sayac_baslangic_ts'],
                              o['referans_kodu'], biti_ts=simdi)
            # Dondurma da bölünmeli — yoksa kayıt kapanınca adet ham pulse'a fırlar
            fcnt = _paket_carp(_bukum_bol(fcnt, _sinyal_bolen(eski_robot, o['istasyon'] or 0,
                                                              o['bukum_operasyon'])),
                               o['paket_adedi'])
            # BİTİŞ DAMGASI ŞART (2026-08-26): paylaşım hesabı pencereleri artık
            # bitis_ts ile sınırlıyor. Damgasız donan kayıt "hâlâ paylaşıyor"
            # sayılır ve ortağının payını gereksiz yere küçültürdü.
            if fcnt is not None:
                c.execute("UPDATE uretim_kayitlari SET ok_adet=?, sayac_otomatik=0, "
                          "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?",
                          (fcnt, simdi, o['id']))
            else:
                # pilot.db okunamadı — ok_adet'i SIFIRLAMA; mevcut değerle manuel'e al
                c.execute("UPDATE uretim_kayitlari SET sayac_otomatik=0, "
                          "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?", (simdi, o['id']))

        c.execute("UPDATE vardiyalar SET robot_no=? WHERE id=?", (yeni_robot, vid))
        conn.commit()
        return jsonify({'basarili': True, 'degisti': True, 'robot_no': yeni_robot,
                        'eski_robot': eski_robot, 'donduruldu': len(aktif),
                        'sayac_destekli': (yeni_robot in SAYAC_AUTO_CIHAZLAR)})
    finally:
        conn.close()


@app.route('/api/vardiya/<int:vid>', methods=['PUT'])
@operator_required
def vardiya_guncelle(vid):
    """Vardiya saatlerini ve detaylarını güncelle."""
    ok, hata = _vardiya_sahibi_kontrol(vid)
    if not ok:
        return jsonify({'hata': hata}), 403
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    mevcut = c.execute('SELECT * FROM vardiyalar WHERE id = ?', (vid,)).fetchone()
    if not mevcut:
        conn.close()
        return jsonify({'hata': 'Vardiya bulunamadi'}), 404
        
    bas = data.get('baslangic_saati', mevcut['baslangic_saati'])
    bit = data.get('bitis_saati', mevcut['bitis_saati'])
    fark_dk = data.get('toplam_sure_dk')
    operator_adi = data.get('operator_adi', mevcut['operator_adi'])
    tarih = data.get('tarih', mevcut['tarih'])
    vardiya_turu = data.get('vardiya_turu', mevcut['vardiya_turu'])

    # Saat FORMAT DOĞRULAMA: bozuk değer DB'ye yazılırsa vardiya_kapat her denemede
    # patlar ve vardiya kapatılamaz hale gelir — bozuk formatı 400 ile reddet.
    # (Boş bitis_saati meşru: açık vardiya. 'HH:MM:SS' ilk 5 karaktere kırpılır.)
    def _saat_ok(s):
        if s is None or str(s).strip() == '':
            return ''
        try:
            datetime.strptime(str(s).strip()[:5], '%H:%M')
            return str(s).strip()[:5]
        except ValueError:
            return None
    bas = _saat_ok(bas)
    bit = _saat_ok(bit)
    if bas is None or bit is None or bas == '':
        conn.close()
        return jsonify({'hata': 'Saat formatı HH:MM olmalıdır (örn: 07:30)'}), 400

    if fark_dk is None:
        try:
            b1 = datetime.strptime(bas, '%H:%M')
            b2 = datetime.strptime(bit, '%H:%M')
            fark_dk = int((b2 - b1).total_seconds() / 60)
            if fark_dk < 0:
                fark_dk += 1440
        except Exception:
            fark_dk = mevcut['toplam_sure_dk']
            
    c.execute(
        "UPDATE vardiyalar SET baslangic_saati=?, bitis_saati=?, toplam_sure_dk=?, operator_adi=?, tarih=?, vardiya_turu=? WHERE id=?",
        (bas, bit, fark_dk, operator_adi, tarih, vardiya_turu, vid)
    )
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'sure_dk': fark_dk})



# ─────────────────────────────────────────────────────────────
# ÜRETİM API
# ─────────────────────────────────────────────────────────────

@app.route('/api/uretim', methods=['POST'])
@operator_required
def uretim_ekle():
    """Üretim kaydı ekle."""
    data = request.get_json() or {}

    if not data.get('vardiya_id') or not data.get('referans_kodu'):
        return jsonify({'hata': 'vardiya_id ve referans_kodu zorunludur'}), 400
    try:
        vardiya_id_int = int(data['vardiya_id'])
    except (ValueError, TypeError):
        return jsonify({'hata': f"vardiya_id sayısal olmalı: {data.get('vardiya_id')!r}"}), 400

    # Operatör yetki kontrolü: header'da operator+pin geldiyse vardiya sahibi mi?
    ok, hata = _vardiya_sahibi_kontrol(vardiya_id_int)
    if not ok:
        return jsonify({'hata': hata}), 403

    conn = None
    try:
        conn = get_db()
        c = conn.cursor()

        # Vardiyanın bölümünü + robot_no'sunu + lokasyonunu öğren
        v_row = c.execute(
            "SELECT COALESCE(bolum, 'kaynak') as bolum, robot_no, COALESCE(lokasyon, 'TK2') as lokasyon FROM vardiyalar WHERE id = ?",
            (data['vardiya_id'],)
        ).fetchone()
        vardiya_bolum = v_row['bolum'] if v_row else 'kaynak'
        vardiya_robot = v_row['robot_no'] if v_row else ''
        vardiya_lokasyon = v_row['lokasyon'] if v_row else 'TK2'

        # Birden fazla kayıt gelebilir
        satirlar = data.get('satirlar', [data])
        tek_satir = (len(satirlar) == 1)
        simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        eklenen = 0
        for satir in satirlar:
            ref = (satir.get('referans_kodu') or data.get('referans_kodu') or '').strip()
            # ── PROSES ADIMI EKİ (kullanıcı 2026-08-04) ───────────────────────
            # KAYDIN HATTI — tel/TK1 montaj/plastik/pres'te hat ÜRETİM KAYDINDA seçilir
            # (kullanıcı 2026-08-18). Kod eki, kapama kısıtı ve sayaç eşlemesi artık
            # vardiyanın hattından DEĞİL bu değerden türer. istasyon 0 (eski kayıt ya
            # da hattı vardiyada seçilen bölüm) → vardiyanın hattına düşer.
            ist_val = int(satir.get('istasyon', data.get('istasyon', 0)) or 0)
            kayit_hat = kayit_hatti(vardiya_robot, ist_val)

            # Tel üretiminde HER adım kendi ekiyle kaydedilir ('93.TK.464 KAPAMA').
            # Böylece adımlar ayrı kodlarda toplanır: ne çoklu sayım olur ne de
            # kimin ne iş yaptığı kaybolur. "Son adım base kod alır" kuralı YOK —
            # kapaması/son montajı dışarıda yapılacak ürünler sabit değil (üretim
            # yoğunluğu + tedarikçi durumuna göre değişiyor), önceden tanımlanamaz.
            # OTOMATİK yapılır (operatör elle yazmaz): yazım hatası referansı
            # bambaşka bir koda çevirir ve kayıt kaybolurdu. Elle yazılmışsa da
            # tekrarlanmaz — tel_referans_kodu önce mevcut eki ayıklar.
            if vardiya_bolum == 'tel' and ref:
                ref = tel_referans_kodu(ref, kayit_hat)
            # ── KAPAMA: AYNI REFERANS = AYNI PRES (kullanıcı 2026-08-20) ───────
            # ESKİ KURAL (2026-08-04): "kapamada bir referansı tek operatör yapar"
            # → ikinci operatör 409 ile reddediliyordu.
            # YENİ KURAL: "Bir tel referansını iki operatör birden seçebilsin, fakat
            # birisi hangi presi seçtiyse diğeri de AYNI presi seçsin. Tek presten
            # sayı alınsın." Yani kısıt kalkmadı, YER DEĞİŞTİRDİ: artık engellenen
            # şey ikinci operatör değil, AYNI referansın BAŞKA bir preste açılması.
            # Sebep: sayaç payı sensör bazlı bölünüyor (bkz. _sayac_oku); iki ayrı
            # preste açılan aynı referans iki ayrı sinyal akışından TAM sayardı ve
            # üretim yine iki katı görünürdü.
            # Kesim/otomat/son montaj serbest — kısıt YALNIZ kapamaya özel.
            if vardiya_bolum == 'tel' and tel_hat_adimi(kayit_hat) == 'Kapama' and ref:
                # LIMIT YOK: aynı referansın o günkü TÜM tel kayıtları çekilir ve
                # kapama olanlar Python'da süzülür. (LIMIT 1 ile ilk satır kesim/otomat
                # olabiliyor, kapama filtresi boşa düşüyor ve kısıt hiç çalışmıyordu.)
                _cak = c.execute("""
                    SELECT v.operator_adi, v.robot_no, u.istasyon FROM uretim_kayitlari u
                    JOIN vardiyalar v ON v.id = u.vardiya_id
                    WHERE COALESCE(v.bolum,'')='tel'
                      AND v.tarih = (SELECT tarih FROM vardiyalar WHERE id=?)
                      AND UPPER(REPLACE(u.referans_kodu,' ','')) = UPPER(REPLACE(?,' ',''))
                      AND u.vardiya_id != ?
                """, (vardiya_id_int, ref, vardiya_id_int)).fetchall()
                # Karşı kaydın hattı da KAYITTAN çözülür (vardiya hattı artık sabit)
                _cak_hat = {id(r): kayit_hatti(r['robot_no'], r['istasyon']) for r in _cak}
                # BAŞKA PRESTE olanlar engellenir; AYNI preste olan serbesttir
                # (iki operatör tek presi paylaşıyor → sayaç ikiye bölünür).
                _baska = [r for r in _cak
                          if tel_hat_adimi(_cak_hat[id(r)]) == 'Kapama'
                          and _cak_hat[id(r)] != kayit_hat]
                if _baska:
                    return jsonify({'hata':
                        f"Bu referansın kapaması bugün {_baska[0]['operator_adi']} tarafından "
                        f"{_cak_hat[id(_baska[0])]} hattında girilmiş. Aynı referans iki ayrı "
                        f"preste kapanamaz — birlikte çalışıyorsanız siz de "
                        f"{_cak_hat[id(_baska[0])]} hattını seçin; sayaç ikinize eşit bölünür."}), 409
            ct_in = float(satir.get('cycle_time_sn', 0) or 0)
            # cycle_time gönderilmediyse ya da 0 ise referanslardan otomatik çek.
            # LOKASYON filtresi ŞART: composite UNIQUE sonrası aynı kod TK1+TK2'de
            # iki satır olabilir — vardiyanın tesisinin cycle'ı çekilmeli.
            # BÖLÜM TERCİHİ: aynı kod birden fazla bölümde farklı cycle ile olabilir —
            # önce vardiyanın bölümündeki satır; yoksa (eski davranış) herhangi biri.
            if ct_in <= 0 and ref:
                ref_row = c.execute(
                    "SELECT hedef_cycle_time_sn FROM referans_listesi "
                    "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) AND COALESCE(lokasyon,'TK2')=? "
                    "ORDER BY (COALESCE(bolum,'kaynak') = ?) DESC LIMIT 1",
                    (ref, vardiya_lokasyon, vardiya_bolum)
                ).fetchone()
                if ref_row:
                    ct_in = float(ref_row['hedef_cycle_time_sn'] or 0)

            ok_val  = int(satir.get('ok_adet', 0) or 0)

            # Test cihazı (Cofle/wicow) seçildiyse sayaç kaynağı o cihaz olur (ESP32 yerine)
            tc_raw = satir.get('test_cihaz_id', data.get('test_cihaz_id'))
            try:
                tc_id = int(tc_raw) if tc_raw not in (None, '', 0, '0') else None
            except (ValueError, TypeError):
                tc_id = None

            # Sayaç otomatik mod: tek satır + OK girilmemiş (0) + (cihaz allowlist'te VEYA
            # test cihazı seçilmiş). Operatör referansa YENİ başlıyor → sayaç o andan sıfırlanır.
            # (Toplu/geçmiş giriş, OK elle yazılmış veya desteklenmeyen kaynak → manuel mod.)
            # pres: allowlist kontrolü SATIRIN makinesine (istasyon) göre — makine
            # seçilmemişse otomatik sayaç YOK (manuel giriş; 0-ezme tuzağını önler).
            auto = 1 if (tek_satir and ok_val == 0
                         and (_sayac_destekli(vardiya_bolum, vardiya_robot, ist_val) or tc_id)) else 0

            if auto and tc_id:
                # Aynı test cihazında önceki auto kayıt(lar)ı dondur (referans değişti)
                onceki = c.execute(
                    "SELECT id, sayac_baslangic_ts, referans_kodu, "
                    "       COALESCE(sayac_devir_adet,0) AS sayac_devir_adet "
                    "FROM uretim_kayitlari "
                    "WHERE vardiya_id=? AND test_cihaz_id=? AND sayac_otomatik=1",
                    (data['vardiya_id'], tc_id)
                ).fetchall()
                for o in onceki:
                    fcnt = _test_basari_say(conn, tc_id, o['referans_kodu'],
                                            o['sayac_baslangic_ts'], biti_ts=simdi)
                    fcnt = _devir_ekle(fcnt, o['sayac_devir_adet'])
                    if fcnt is not None:
                        c.execute("UPDATE uretim_kayitlari SET ok_adet=?, sayac_otomatik=0, "
                                  "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?",
                                  (fcnt, simdi, o['id']))
                    else:
                        c.execute("UPDATE uretim_kayitlari SET sayac_otomatik=0, "
                                  "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?", (simdi, o['id']))
            elif auto:
                # Aynı vardiya+istasyon önceki ESP32 auto kayıt(lar)ı dondur (test-cihazı hariç)
                onceki = c.execute(
                    "SELECT id, istasyon, sayac_baslangic_ts, referans_kodu, "
                    "       COALESCE(bukum_operasyon,1) AS bukum_operasyon, "
                    "       COALESCE(paket_adedi,1) AS paket_adedi, "
                    "       COALESCE(sayac_devir_adet,0) AS sayac_devir_adet "
                    "FROM uretim_kayitlari "
                    "WHERE vardiya_id=? AND istasyon=? AND sayac_otomatik=1 AND test_cihaz_id IS NULL",
                    (data['vardiya_id'], ist_val)
                ).fetchall()
                for o in onceki:
                    # Dondurma da PAYLAŞIMLI okunur: kayıt kapanırken adet ham
                    # sinyale fırlarsa paylaşım anlamını yitirirdi.
                    fcnt = _sayac_oku(conn, data['vardiya_id'], o['id'], vardiya_bolum,
                                      vardiya_robot, ist_val, o['sayac_baslangic_ts'],
                                      o['referans_kodu'], biti_ts=simdi)
                    # Dondurma da bölünüp çarpılmalı (bkz. _bukum_bol / _paket_carp)
                    fcnt = _paket_carp(_bukum_bol(fcnt, _sinyal_bolen(vardiya_robot, ist_val,
                                                                     o['bukum_operasyon'])),
                                       o['paket_adedi'])
                    fcnt = _devir_ekle(fcnt, o['sayac_devir_adet'])
                    # DONDURMA ANI = O REFERANSIN BİTİŞİDİR (2026-08-21): operatör
                    # aynı makinede yeni referansa geçti, öncekinin üretim penceresi
                    # burada kapanır. bitis_ts zaten doluysa (tamamlandı işaretlenmiş)
                    # KORUNUR — ilk kapanış anı doğru olandır.
                    if fcnt is not None:
                        c.execute("UPDATE uretim_kayitlari SET ok_adet=?, sayac_otomatik=0, "
                                  "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?",
                                  (fcnt, simdi, o['id']))
                    else:
                        # pilot.db okunamadi — ok_adet'i SIFIRLAMA; mevcut degerle dondur (manuel'e al)
                        c.execute("UPDATE uretim_kayitlari SET sayac_otomatik=0, "
                                  "bitis_ts=COALESCE(bitis_ts, ?) WHERE id=?", (simdi, o['id']))

            # Referansı listeye otomatik ekle — VARDIYANIN bölümü VE lokasyonuyla etiketle
            # (montaj operatörü tanımsız bir kod girince montaj'a düşsün, kaynak'a değil;
            #  TK1 operatörünün girdiği kod TK1'e düşsün, TK2'ye karışmasın)
            # NORMALIZE kontrol: UNIQUE ham kod üzerinde — operatör '94. LTK. 341/10' gibi
            # boşluklu/farklı yazarsa INSERT OR IGNORE yeni VARYANT satır açıyordu; bu
            # duplike satırlar JOIN'lerde iş emirlerini çoğaltıyordu (fan-out). Normalize
            # eş (aynı lokasyonda) varsa yeni satır AÇMA.
            # SIRA: bu blok üretim INSERT'inden ÖNCE olmalı — _bukum_op_coz büküm
            # sayısını referans satırına yazıyor, satır henüz yoksa değer kaybolurdu.
            ref_var = c.execute(
                "SELECT 1 FROM referans_listesi "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
                "AND COALESCE(lokasyon,'TK2')=? LIMIT 1",
                (ref, vardiya_lokasyon)
            ).fetchone()
            if not ref_var:
                c.execute(
                    'INSERT OR IGNORE INTO referans_listesi (referans_kodu, bolum, lokasyon) VALUES (?, ?, ?)',
                    (ref, vardiya_bolum, vardiya_lokasyon)
                )

            # BÜKÜM OPERASYON SAYISI — yalnız pres (abkant) için anlamlı.
            # İstemci göndermediyse referansın hatırlanan değerine düşer; o da yoksa 1
            # (bölme yok = eski davranış). Kayda KOPYALANIR: referansın varsayılanı
            # sonradan değişse bile bu kaydın adedi retroaktif kaymasın.
            bop = _bukum_op_coz(c, satir.get('bukum_operasyon', data.get('bukum_operasyon')),
                                ref, vardiya_bolum, vardiya_lokasyon, kayit_hat)
            # Paket (montaj adet) çarpanı — TK1 montaj / tel Son Montaj dışında 1
            pak = _paket_coz(c, satir.get('paket_adedi', data.get('paket_adedi')),
                             ref, vardiya_bolum, vardiya_lokasyon, kayit_hat)

            c.execute('''
                INSERT INTO uretim_kayitlari (vardiya_id, referans_kodu, ok_adet, nok_adet, tamir_adet, hedef_adet, cycle_time_sn, istasyon, launch_adet, aciklama, sayac_baslangic_ts, sayac_otomatik, test_cihaz_id, bukum_operasyon, paket_adedi, baslangic_ts)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['vardiya_id'],
                ref,
                ok_val,
                int(satir.get('nok_adet', 0) or 0),
                int(satir.get('tamir_adet', 0) or 0),
                int(satir.get('hedef_adet', 0) or 0),
                ct_in,
                ist_val,
                int(satir.get('launch_adet', data.get('launch_adet', 0)) or 0),
                (satir.get('aciklama') or data.get('aciklama') or '').strip(),
                simdi if auto else None,
                auto,
                tc_id,
                bop,
                pak,
                # baslangic_ts: SAYAÇTAN BAĞIMSIZ (kullanıcı 2026-08-21). sayac_baslangic_ts
                # yalnız otomatik sayaçlı kayıtlarda dolu; elle giren hatların tamamı
                # NULL kalıyordu ve zaman çizelgesinde hiç görünmezlerdi.
                simdi
            ))
            eklenen += 1

        conn.commit()
        return jsonify({'basarili': True, 'eklenen': eklenen}), 201
    except Exception as e:
        print(f"HATA (uretim_ekle): {traceback.format_exc()}")
        return jsonify({'hata': f'Üretim eklenemedi: {str(e)}'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/uretim/<int:uid>/sayac_sifirla', methods=['POST'])
@operator_required
def uretim_sayac_sifirla(uid):
    """Operatör emniyet reseti — bu üretim kaydının sayacını ŞİMDİDEN sıfırlar.
    Deneme/test sinyalleri karışıklık yaratırsa operatör sayacı 0'lar.
    baseline=now → bu andan sonraki pulse'lar sayılır; otomatik mod yeniden aktif."""
    vid = _uretim_vardiya_bul(uid)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    conn = get_db()
    try:
        simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # DEVİR DE SIFIRLANIR: bu buton "her şeyi sıfırla"dır; devir kalsaydı
        # operatör 0 beklerken eski adet geri gelirdi.
        cur = conn.execute(
            "UPDATE uretim_kayitlari SET sayac_baslangic_ts=?, ok_adet=0, "
            "sayac_otomatik=1, sayac_devir_adet=0 WHERE id=?",
            (simdi, uid)
        )
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'hata': 'Kayit bulunamadi'}), 404
        return jsonify({'basarili': True, 'sayac_baslangic_ts': simdi})
    except Exception as e:
        # DB kilidi/hata: mobil JSON bekler — HTML 500 yerine anlaşılır hata
        return jsonify({'hata': f'Sayaç sıfırlanamadı: {e}'}), 500
    finally:
        conn.close()


@app.route('/api/uretim/<int:uid>', methods=['DELETE'])
@operator_required
def uretim_sil(uid):
    vid = _uretim_vardiya_bul(uid)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    conn = get_db()
    conn.execute('DELETE FROM uretim_kayitlari WHERE id = ?', (uid,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/uretim/<int:uid>', methods=['PUT'])
@operator_required
def uretim_guncelle(uid):
    """Uretim kaydini guncelle."""
    vid = _uretim_vardiya_bul(uid)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    data = request.get_json() or {}
    conn = None
    try:
        conn = get_db()
        c = conn.cursor()
        mevcut = c.execute(
            'SELECT id, vardiya_id, istasyon, sayac_otomatik, sayac_baslangic_ts '
            'FROM uretim_kayitlari WHERE id = ?', (uid,)
        ).fetchone()
        if not mevcut:
            return jsonify({'hata': 'Kayit bulunamadi'}), 404

        ref = (data.get('referans_kodu') or '').strip()
        ct = float(data.get('cycle_time_sn', 0) or 0)
        # Gelen CT 0 veya gönderilmediyse referanslardan çek — vardiyanın LOKASYONUYLA
        # (composite UNIQUE sonrası aynı kod iki tesiste olabilir; yanlış tesisin ct'si çekilmesin)
        if ct <= 0 and ref:
            v_lok_row = c.execute(
                "SELECT COALESCE(lokasyon,'TK2') as lok FROM vardiyalar WHERE id = ?",
                (mevcut['vardiya_id'],)
            ).fetchone()
            v_lok = v_lok_row['lok'] if v_lok_row else 'TK2'
            ref_row = c.execute(
                "SELECT hedef_cycle_time_sn FROM referans_listesi "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) AND COALESCE(lokasyon,'TK2')=?",
                (ref, v_lok)
            ).fetchone()
            if ref_row:
                ct = float(ref_row['hedef_cycle_time_sn'] or 0)

        submitted_ok = int(data.get('ok_adet', 0) or 0)
        # Sayaç otomatik modu: mobil 'ok_elle_degisti' bayrağı gönderirse (operatör
        # OK alanına dokundu) otomatik kapatılır (değer dondurulur). Sadece NOK/açıklama
        # değişip OK'e dokunulmadıysa bayrak gelmez → auto devam eder.
        yeni_otomatik = mevcut['sayac_otomatik'] or 0
        if data.get('ok_elle_degisti'):
            yeni_otomatik = 0

        # BÜKÜM OPERASYON SAYISI — operatör kaydı düzenlerken de değiştirebilir.
        # Vardiyanın bölüm/lokasyonuyla çözülür; pres dışında her zaman 1 döner.
        v_bl = c.execute(
            "SELECT COALESCE(bolum,'kaynak') AS bolum, COALESCE(lokasyon,'TK2') AS lokasyon, "
            "       robot_no "
            "FROM vardiyalar WHERE id=?", (mevcut['vardiya_id'],)
        ).fetchone()
        _duz_hat = kayit_hatti(v_bl['robot_no'] if v_bl else '',
                               int(data.get('istasyon', 0) or 0))
        bop = _bukum_op_coz(c, data.get('bukum_operasyon'), ref,
                            v_bl['bolum'] if v_bl else 'kaynak',
                            v_bl['lokasyon'] if v_bl else 'TK2', _duz_hat)
        pak = _paket_coz(c, data.get('paket_adedi'), ref,
                         v_bl['bolum'] if v_bl else 'kaynak',
                         v_bl['lokasyon'] if v_bl else 'TK2', _duz_hat)

        # TEL: kod eki KAYDIN hattından türer. Hat artık üretim kaydında seçildiği
        # için DÜZENLEMEDE de uygulanmalı — yoksa 'Kapama'dan 'Son Montaj'a alınan
        # kaydın kodu KAPAMA ekiyle kalır ve iş yanlış adımda toplanırdı.
        # tel_referans_kodu önce mevcut eki ayıklar → ek iki kez yazılmaz.
        yeni_ist = int(data.get('istasyon', 0) or 0)
        if v_bl and v_bl['bolum'] == 'tel' and ref:
            ref = tel_referans_kodu(ref, kayit_hatti(v_bl['robot_no'], yeni_ist))

        c.execute('''
            UPDATE uretim_kayitlari
            SET referans_kodu=?, ok_adet=?, nok_adet=?, tamir_adet=?, hedef_adet=?, cycle_time_sn=?, istasyon=?, launch_adet=?, aciklama=?, sayac_otomatik=?, bukum_operasyon=?, paket_adedi=?
            WHERE id=?
        ''', (
            ref,
            submitted_ok,
            int(data.get('nok_adet', 0) or 0),
            int(data.get('tamir_adet', 0) or 0),
            int(data.get('hedef_adet', 0) or 0),
            ct,
            yeni_ist,
            int(data.get('launch_adet', 0) or 0),
            (data.get('aciklama') or '').strip(),
            yeni_otomatik,
            bop,
            pak,
            uid
        ))
        conn.commit()
        return jsonify({'basarili': True})
    except Exception as e:
        print(f"HATA (uretim_guncelle): {traceback.format_exc()}")
        return jsonify({'hata': f'Üretim güncellenemedi: {str(e)}'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/uretim/ct_guncelle', methods=['POST'])
def uretim_ct_toplu_guncelle():
    """Tum uretim kayitlarinda cycle_time_sn eksik olanlari referans listesinden toplu guncelle."""
    conn = get_db()
    # Referans eşleşmesi kaydın VARDİYASININ lokasyonuna kapalı — composite UNIQUE
    # sonrası aynı kod TK1+TK2'de olabilir; yanlış tesisin cycle'ı yazılmasın.
    cur = conn.execute("""
        UPDATE uretim_kayitlari
        SET cycle_time_sn = (
            SELECT r.hedef_cycle_time_sn
            FROM referans_listesi r
            WHERE UPPER(REPLACE(r.referans_kodu,' ','')) = UPPER(REPLACE(uretim_kayitlari.referans_kodu,' ',''))
              AND COALESCE(r.lokasyon,'TK2') = (
                    SELECT COALESCE(v.lokasyon,'TK2') FROM vardiyalar v WHERE v.id = uretim_kayitlari.vardiya_id)
              AND r.hedef_cycle_time_sn > 0
            LIMIT 1
        )
        WHERE (cycle_time_sn IS NULL OR cycle_time_sn = 0)
          AND EXISTS (
            SELECT 1 FROM referans_listesi r
            WHERE UPPER(REPLACE(r.referans_kodu,' ','')) = UPPER(REPLACE(uretim_kayitlari.referans_kodu,' ',''))
              AND COALESCE(r.lokasyon,'TK2') = (
                    SELECT COALESCE(v.lokasyon,'TK2') FROM vardiyalar v WHERE v.id = uretim_kayitlari.vardiya_id)
              AND r.hedef_cycle_time_sn > 0
          )
    """)
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'guncellenen': cur.rowcount})


# ─────────────────────────────────────────────────────────────
# ZAMAN DAMGASI TÜRETME (kullanıcı 2026-08-21)
# ─────────────────────────────────────────────────────────────
# Duruş ve üretim kayıtlarına TAM damga ('YYYY-MM-DD HH:MM:SS') yazılır; zaman
# çizelgesi sayfası bunlardan çizilir. Kurallar database.py'deki geriye-dönük
# doldurma SQL'i ile BİREBİR aynı olmalı — biri değişirse eski ve yeni kayıtlar
# aynı çizelgede farklı mantıkla yerleşir ve fark sessizce görünmez kalır.
TS_FMT = '%Y-%m-%d %H:%M:%S'


def _ts_metin(dt):
    return dt.strftime(TS_FMT) if dt else None


def _ts_parse(deger):
    """'YYYY-MM-DD HH:MM[:SS]' (ya da ISO 'T' ayraçlı) metni datetime'a çevirir.
    Bozuk/boş girdi None döner — çağıran 400 verir, HTML 500'e düşmez."""
    t = str(deger or '').strip().replace('T', ' ')
    if not t:
        return None
    for f in (TS_FMT, '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(t[:len('2026-01-01 00:00:00') if f == TS_FMT else 16], f)
        except ValueError:
            continue
    return None


def _vardiya_pencere(vardiya):
    """Vardiyanın (baslangic_dt, bitis_dt) penceresi.

    Vardiya AÇIKSA bitiş = ŞİMDİ (çizelgede bar 'şu ana kadar' uzar).
    GECE VARDİYASI: bitiş saati başlangıçtan küçükse ertesi güne taşınır —
    22:00-06:00 vardiyası tek günde 'negatif uzunlukta' görünmesin."""
    try:
        tarih = str(vardiya['tarih'] or '')
        bas = _ts_parse(f"{tarih} {str(vardiya['baslangic_saati'] or '')[:5]}")
    except Exception:
        bas = None
    if bas is None:
        return None, None
    if (vardiya['durum'] or '') == 'kapali' and vardiya['bitis_saati']:
        bit = _ts_parse(f"{tarih} {str(vardiya['bitis_saati'])[:5]}")
        if bit and bit < bas:
            bit += timedelta(days=1)
    else:
        bit = datetime.now()
        if bit < bas:            # gelecek tarihli/hatalı vardiya — bar ters dönmesin
            bit = bas
    return bas, bit


def _durus_ts_hesapla(vardiya, baslangic_saati, sure_dk, anchor_ts=None):
    """Duruşun (baslangic_ts, bitis_ts) damgasını türetir.

    İKİ KAYNAK, İKİ KURAL:
      · Operatör SAAT yazdıysa  → başlangıç odur (vardiyanın gününe oturtulur;
        vardiya başlangıcından önceki saat ERTESİ gündür — gece vardiyası).
      · Saat YOKSA              → duruş girildiği ANDA bitmiştir kabul edilir:
        bitiş = anchor (kaydın created_at'i / şimdi), başlangıç = bitiş − süre.
        Kayıtların çoğunda saat boş (bkz. /api/sinyal_analiz notu) — bu kural
        olmasaydı o duruşlar çizelgede hiç görünmezdi.
    Vardiya penceresi çözülemezse (bozuk saat) (None, None) döner; kayıt yine
    yazılır, çizelgede 'zamanı belirsiz' listesine düşer."""
    try:
        sure = max(0, int(sure_dk or 0))
    except (TypeError, ValueError):
        sure = 0
    saat = str(baslangic_saati or '').strip()[:5]
    bas = None
    if len(saat) >= 4:
        try:
            v_bas = str(vardiya['baslangic_saati'] or '')[:5]
            bas = _ts_parse(f"{vardiya['tarih']} {saat}")
            if bas and v_bas and saat < v_bas:
                bas += timedelta(days=1)      # gece vardiyası: ertesi güne taşı
        except Exception:
            bas = None
    if bas is None:
        bit_anchor = _ts_parse(anchor_ts) or datetime.now()
        bas = bit_anchor - timedelta(minutes=sure)
    return _ts_metin(bas), _ts_metin(bas + timedelta(minutes=sure))


def _vardiya_satiri(conn, vardiya_id):
    """Damga türetmek için gereken vardiya alanları (yoksa None)."""
    return conn.execute(
        "SELECT id, tarih, baslangic_saati, bitis_saati, durum FROM vardiyalar WHERE id=?",
        (vardiya_id,)
    ).fetchone()


# ─────────────────────────────────────────────────────────────
# DURUŞ API
# ─────────────────────────────────────────────────────────────

@app.route('/api/uretim_kayit/<int:uid>/tamamlandi', methods=['PATCH'])
@operator_required
def uretim_tamamlandi_guncelle(uid):
    """Üretim kaydını tamamlandı / tamamlanmadı olarak işaretle."""
    vid = _uretim_vardiya_bul(uid)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    data = request.get_json() or {}
    deger = 1 if data.get('tamamlandi') else 0
    conn = get_db()
    # tamamlandi_ts = ERKEN TEYİT sayacının başlangıcı (2026-08-18). İşaret
    # kaldırılırsa NULL'a döner ki bekleyen erken teyit de İPTAL olsun.
    # NOT: conn.close() KALDIRILDI — get_db() request-paylaşımlı (flask.g cache);
    # burada kapatmak aynı istekte sonra kullanan kodu düşürür (bkz. daha önce
    # decorator'da yaşanan prod 500).
    # bitis_ts de burada kapanır (2026-08-21): "tamamlandı" işareti operatörün
    # o referansı bitirdiğini söylediği andır. İşaret GERİ ALINIRSA damga da
    # NULL'a döner — kayıt yeniden 'açık' sayılır ve çizelgede bar uzamaya devam
    # eder (tamamlandi_ts ile aynı ömür, aynı gerekçe).
    conn.execute(
        "UPDATE uretim_kayitlari SET tamamlandi=?, "
        "tamamlandi_ts = CASE WHEN ?=1 THEN datetime('now','localtime') ELSE NULL END, "
        "bitis_ts      = CASE WHEN ?=1 THEN datetime('now','localtime') ELSE NULL END "
        "WHERE id=?", (deger, deger, deger, uid))

    # ── İŞARET GERİ ALINDI → SAYAÇ KALDIĞI YERDEN DEVAM ETSİN ────────────
    # (kullanıcı 2026-08-24: "tamamlandı işaretleyip tekrar devam etmek için
    #  açınca sayaç saymaya devam etmiyor")
    # Operatör referansı bitirip başka referansa geçtiğinde bu kayıt DONMUŞ olur
    # (sayac_otomatik=0 — uretim_ekle'deki dondurma). Geri dönüp işareti
    # kaldırmak onu çözmüyordu. Artık çözülüyor:
    #   · devir   = o ana kadarki ok_adet (kaybolmasın)
    #   · pencere = ŞİMDİ (aradaki pulse'lar başka referansa aitti, sayılmamalı)
    #   · auto    = 1
    # SADECE sayaçla beslenebilen kayıtlarda: makinesi allowlist'te değilse ya da
    # test cihazı yoksa kayıt ELLE giriştir; otomatiğe çevirmek adedi bozardı.
    sonuc = {'basarili': True, 'sayac_surduruldu': False}
    if deger == 0:
        u = conn.execute(
            "SELECT u.id, u.istasyon, u.ok_adet, u.sayac_otomatik, u.sayac_baslangic_ts, "
            "       u.test_cihaz_id, COALESCE(v.bolum,'kaynak') AS bolum, v.robot_no, "
            "       COALESCE(v.durum,'acik') AS durum "
            "FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id "
            "WHERE u.id=?", (uid,)).fetchone()
        if (u and u['durum'] != 'kapali' and not u['sayac_otomatik']
                and u['sayac_baslangic_ts']       # baştan sayaçlı açılmış bir kayıt
                and (u['test_cihaz_id']
                     or _sayac_destekli(u['bolum'], u['robot_no'], u['istasyon'] or 0))):
            simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            conn.execute(
                "UPDATE uretim_kayitlari SET sayac_otomatik=1, sayac_baslangic_ts=?, "
                "sayac_devir_adet=? WHERE id=?",
                (simdi, max(0, int(u['ok_adet'] or 0)), uid))
            sonuc['sayac_surduruldu'] = True
    conn.commit()
    return jsonify(sonuc)

@app.route('/api/durus', methods=['POST'])
@operator_required
def durus_ekle():
    """Duruş kaydı ekle — HER GİRİŞ KENDİ SATIRI (kullanıcı 2026-08-21).

    ESKİDEN aynı vardiya + aynı sebep TEK SATIRDA toplanıyordu. Toplam süre
    doğruydu ama olayın NE ZAMAN olduğu kayboluyordu: sabah 09:10'daki 15dk'lık
    arıza ile öğleden sonraki 40dk'lık arıza tek satırda '55dk' oluyordu. Zaman
    çizelgesi tam olarak bu iki olayı ayrı barlar olarak göstermek için var —
    birleştirme sürdükçe çizelge YANILTICI olurdu (kod içinde de böyle not
    edilmişti: bkz. /api/sinyal_analiz 'duruslar tablosu ... TOPLAR' notu).

    Toplam süreye bakan yerler etkilenmez: hepsi SUM(sure_dk) ile okuyor.
    Duruş listesi artık aynı sebepten birden çok satır gösterebilir — istenen bu."""
    data = request.get_json() or {}

    if not data.get('vardiya_id') or not data.get('durus_sebebi'):
        return jsonify({'hata': 'vardiya_id ve durus_sebebi zorunludur'}), 400
    try:
        durus_vid_int = int(data['vardiya_id'])
    except (ValueError, TypeError):
        return jsonify({'hata': f"vardiya_id sayısal olmalı: {data.get('vardiya_id')!r}"}), 400

    # Operatör yetki kontrolü: vardiya sahibi mi?
    ok, hata = _vardiya_sahibi_kontrol(durus_vid_int)
    if not ok:
        return jsonify({'hata': hata}), 403

    # Birden fazla duruş gelebilir
    satirlar = data.get('satirlar', [data])

    conn = None
    try:
        conn = get_db()
        c = conn.cursor()
        eklenen = 0
        birlestirilen = 0          # geriye uyum: istemci bu alanı okuyor, artık hep 0
        # Damga türetmek için vardiyanın tarihi/başlangıç saati gerekir (gece vardiyası)
        vardiya_row = _vardiya_satiri(conn, durus_vid_int)
        for satir in satirlar:
            vid    = data['vardiya_id']
            sebep  = (satir.get('durus_sebebi') or data.get('durus_sebebi') or '').strip()
            if not sebep:
                continue
            sure   = int(satir.get('sure_dk', 0) or 0)
            saat   = satir.get('baslangic_saati', '') or ''
            acikl  = (satir.get('aciklama') or '').strip()
            tipi   = satir.get('durus_tipi', data.get('durus_tipi', 'plansiz'))

            # Zaman damgası: istemci açıkça gönderdiyse o, yoksa saat+süreden türet.
            # (Zaman çizelgesi sayfası düzenlerken tam damga gönderebilir.)
            bas_ts = _ts_metin(_ts_parse(satir.get('baslangic_ts', data.get('baslangic_ts'))))
            bit_ts = _ts_metin(_ts_parse(satir.get('bitis_ts', data.get('bitis_ts'))))
            if not bas_ts:
                bas_ts, _tur_bit = _durus_ts_hesapla(vardiya_row, saat, sure)
                if not bit_ts:
                    bit_ts = _tur_bit
            elif not bit_ts:
                _b = _ts_parse(bas_ts)
                bit_ts = _ts_metin(_b + timedelta(minutes=sure)) if _b else None

            c.execute('''
                INSERT INTO duruslar (vardiya_id, durus_sebebi, aciklama, sure_dk, baslangic_saati, durus_tipi, baslangic_ts, bitis_ts)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (vid, sebep, acikl, sure, saat, tipi, bas_ts, bit_ts))
            eklenen += 1

        conn.commit()
        return jsonify({'basarili': True, 'eklenen': eklenen, 'birlestirilen': birlestirilen}), 201
    except Exception as e:
        print(f"HATA (durus_ekle): {traceback.format_exc()}")
        return jsonify({'hata': f'Duruş eklenemedi: {str(e)}'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/durus/<int:did>', methods=['DELETE'])
@operator_required
def durus_sil(did):
    vid = _durus_vardiya_bul(did)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    conn = get_db()
    conn.execute('DELETE FROM duruslar WHERE id = ?', (did,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/durus/<int:did>', methods=['PUT'])
@operator_required
def durus_guncelle(did):
    """Durus kaydini guncelle."""
    vid = _durus_vardiya_bul(did)
    if vid:
        ok, hata = _vardiya_sahibi_kontrol(vid)
        if not ok:
            return jsonify({'hata': hata}), 403
    data = request.get_json(silent=True) or {}
    # sure_dk güvenli parse (boş input '' → ValueError → HTML 500 olmasın; durus_ekle kalıbı)
    try:
        sure_dk_val = int(data.get('sure_dk', 0) or 0)
    except (ValueError, TypeError):
        return jsonify({'hata': f"Süre sayısal olmalı: {data.get('sure_dk')!r}"}), 400
    conn = get_db()
    c = conn.cursor()
    mevcut = c.execute(
        'SELECT id, vardiya_id, created_at, baslangic_ts FROM duruslar WHERE id = ?', (did,)
    ).fetchone()
    if not mevcut:
        conn.close()
        return jsonify({'hata': 'Kayit bulunamadi'}), 404

    saat_val = data.get('baslangic_saati', '')
    # Damga: istemci TAM damga gönderdiyse (zaman çizelgesi sayfası) o kullanılır;
    # yoksa saat+süreden yeniden türetilir. Süre veya saat değişince damganın da
    # değişmesi ŞART — yoksa çizelgedeki bar listedeki süreyle tutmaz.
    bas_ts = _ts_metin(_ts_parse(data.get('baslangic_ts')))
    bit_ts = _ts_metin(_ts_parse(data.get('bitis_ts')))
    if not bas_ts:
        v_row = _vardiya_satiri(conn, mevcut['vardiya_id'])
        # Saat boşsa eski davranışa düşmeden ÖNCE mevcut damgayı çapa al: kullanıcı
        # yalnız sebebi değiştirdiğinde duruş çizelgede yerinden OYNAMASIN.
        capa = mevcut['baslangic_ts'] or mevcut['created_at']
        if str(saat_val or '').strip()[:5] and len(str(saat_val).strip()) >= 4:
            bas_ts, _tb = _durus_ts_hesapla(v_row, saat_val, sure_dk_val)
        elif mevcut['baslangic_ts']:
            bas_ts, _tb = mevcut['baslangic_ts'], None
        else:
            bas_ts, _tb = _durus_ts_hesapla(v_row, '', sure_dk_val, anchor_ts=capa)
        if not bit_ts:
            _b = _ts_parse(bas_ts)
            bit_ts = _ts_metin(_b + timedelta(minutes=sure_dk_val)) if _b else _tb
    elif not bit_ts:
        _b = _ts_parse(bas_ts)
        bit_ts = _ts_metin(_b + timedelta(minutes=sure_dk_val)) if _b else None

    c.execute('''
        UPDATE duruslar
        SET durus_sebebi=?, aciklama=?, sure_dk=?, baslangic_saati=?, durus_tipi=?,
            baslangic_ts=?, bitis_ts=?
        WHERE id=?
    ''', (
        data.get('durus_sebebi'),
        data.get('aciklama', ''),
        sure_dk_val,
        saat_val,
        data.get('durus_tipi', 'plansiz'),
        bas_ts,
        bit_ts,
        did
    ))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'baslangic_ts': bas_ts, 'bitis_ts': bit_ts})



# ─────────────────────────────────────────────────────────────
# ZAMAN ÇİZELGESİ (kullanıcı 2026-08-21)
# ─────────────────────────────────────────────────────────────
# "dashboardda ayrı sayfa oluşturup bir zaman çizelgesi üzerinde bar üzerinden
#  duruşları görebilelim. vardiya barı üzerinden 10 dk dan uzun duruş barları,
#  ve zaman aralıkları olsun. tıklayınca duruş zamanı bölünebilsin ve duruş
#  değiştirilip atanabilsin."
#
# Bu uç nokta çizelgenin TEK veri kaynağıdır: vardiya barı + üzerindeki duruş
# ve üretim-referansı dilimleri. Süreler DAMGADAN (baslangic_ts/bitis_ts) gelir,
# sure_dk'dan değil — bölme/yeniden atama sonrası ikisi ayrışırsa çizelge ile
# rapor birbirini tutmaz; damga tek doğru kabul edilir ve sure_dk ona göre
# yeniden yazılır (bkz. durus_bol).


@app.route('/api/zaman_cizelgesi', methods=['GET'])
def zaman_cizelgesi_api():
    """Bir günün vardiya/duruş/üretim penceresi — zaman çizelgesi sayfası için.

    ?tarih=YYYY-MM-DD (varsayılan bugün) &bolum= &lokasyon=TK2 &hat= &operator=
    Dönen: {'vardiyalar': [...], 'sebepler': [...], 'zamansiz': [...]}

    KISA duruşlar da döner ('uzun' bayrağıyla): eşik yalnız GÖRSEL bir ayrımdır.
    Filtrelenip atılsalardı 40dk'lık bir duruş 35+5 diye bölündüğünde 5dk'lık
    parça ekrandan kaybolur, kullanıcı da onu bir daha düzenleyemezdi.
    """
    tarih    = (request.args.get('tarih') or datetime.now().strftime('%Y-%m-%d')).strip()
    bolum    = (request.args.get('bolum') or '').strip()
    lokasyon = (request.args.get('lokasyon') or '').strip()
    hat      = (request.args.get('hat') or '').strip()
    operator = (request.args.get('operator') or '').strip()
    try:
        esik_dk = max(0, int(request.args.get('min_durus_dk', 10) or 10))
    except (TypeError, ValueError):
        esik_dk = 10

    if bolum and bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f'Geçersiz bölüm: {bolum}'}), 400

    conn = get_db()
    kosul = ["v.tarih = ?"]
    parm  = [tarih]
    if bolum:
        kosul.append("COALESCE(v.bolum,'kaynak') = ?"); parm.append(bolum)
    if lokasyon:
        kosul.append("COALESCE(v.lokasyon,'TK2') = ?"); parm.append(lokasyon)
    if hat:
        kosul.append("v.robot_no = ?"); parm.append(hat)
    if operator:
        kosul.append("v.operator_adi = ?"); parm.append(operator)

    vardiyalar = conn.execute(
        "SELECT * FROM vardiyalar v WHERE " + " AND ".join(kosul) +
        " ORDER BY v.baslangic_saati, v.robot_no, v.id", parm
    ).fetchall()

    # AÇIK vardiyaların otomatik molalarını da işle: çizelge açılır açılmaz
    # çay/yemek barları görünsün (vardiya kapanmayı beklemesin).
    for v in vardiyalar:
        try:
            _otomatik_mola_uygula(conn, v)
        except Exception as _e:
            print(f'[zaman_cizelgesi] mola uygula hata (vardiya {v["id"]}): {_e}')
    conn.commit()

    vids = [v['id'] for v in vardiyalar]
    duruslar_map, uretim_map = {}, {}
    zamansiz = []
    if vids:
        ph = ','.join('?' * len(vids))
        for d in conn.execute(
                f"SELECT * FROM duruslar WHERE vardiya_id IN ({ph}) "
                f"ORDER BY COALESCE(baslangic_ts,'9999'), id", vids).fetchall():
            _bas, _bit = d['baslangic_ts'], d['bitis_ts']
            kayit = {
                'id':        d['id'],
                'vardiya_id': d['vardiya_id'],
                'sebep':     d['durus_sebebi'] or '',
                'tip':       d['durus_tipi'] or 'plansiz',
                'aciklama':  d['aciklama'] or '',
                'sure_dk':   int(d['sure_dk'] or 0),
                'bas_ts':    _bas,
                'bit_ts':    _bit,
                'uzun':      int(d['sure_dk'] or 0) >= esik_dk,
            }
            if not _bas:
                # Damgası türetilememiş kayıt (bozuk vardiya saati). Ekranda ayrı
                # listede gösterilir ki kullanıcı zamanını elle atayabilsin.
                zamansiz.append(kayit)
            else:
                duruslar_map.setdefault(d['vardiya_id'], []).append(kayit)

        for u in conn.execute(
                f"SELECT * FROM uretim_kayitlari WHERE vardiya_id IN ({ph}) "
                f"ORDER BY COALESCE(baslangic_ts, created_at), id", vids).fetchall():
            uretim_map.setdefault(u['vardiya_id'], []).append(u)

    cikti = []
    for v in vardiyalar:
        bas, bit = _vardiya_pencere(v)
        v_bolum = (v['bolum'] if 'bolum' in v.keys() else 'kaynak') or 'kaynak'
        satirlar = []
        for u in uretim_map.get(v['id'], []):
            u_bas = u['baslangic_ts'] or u['sayac_baslangic_ts'] or u['created_at']
            satirlar.append({
                'id':            u['id'],
                'referans_kodu': u['referans_kodu'],
                'makine':        kayit_makine_ad(v['robot_no'], u['istasyon'] or 0),
                'ok_adet':       int(u['ok_adet'] or 0),
                'nok_adet':      int(u['nok_adet'] or 0),
                'bas_ts':        u_bas,
                'bit_ts':        u['bitis_ts'],
                'acik':          not u['bitis_ts'],
            })
        cikti.append({
            'id':           v['id'],
            'operator_adi': v['operator_adi'],
            'hat':          v['robot_no'],
            'bolum':        v_bolum,
            'bolum_ad':     BOLUM_AD.get(v_bolum, v_bolum),
            'lokasyon':     (v['lokasyon'] if 'lokasyon' in v.keys() else 'TK2') or 'TK2',
            'vardiya_turu': v['vardiya_turu'],
            'durum':        v['durum'] or 'aktif',
            'bas_ts':       _ts_metin(bas),
            'bit_ts':       _ts_metin(bit),
            'sure_dk':      int((bit - bas).total_seconds() // 60) if (bas and bit) else 0,
            'duruslar':     duruslar_map.get(v['id'], []),
            'uretimler':    satirlar,
        })

    # Yeniden atama listesi: sayfada bölüm seçilmişse o bölümün sebepleri.
    # Bölüm seçilmemişse çizelgedeki vardiyaların bölümlerinden birleştirilir —
    # kullanıcı bölüm filtresi koymadan da sebep değiştirebilsin.
    sebepler, gorulen = [], set()
    for b, lk in ({(bolum, lokasyon or 'TK2')} if bolum
                  else {(c['bolum'], c['lokasyon']) for c in cikti}):
        try:
            liste = durus_sebepleri_yukle(b, lk) + list(_ek_durus_sebepleri(b, lk))
        except Exception as _e:
            print(f'[zaman_cizelgesi] sebep listesi hata ({b}/{lk}): {_e}')
            continue
        for x in liste:
            ad = str(x.get('sebep', '')).strip()
            if ad and ad.upper() not in gorulen:
                gorulen.add(ad.upper())
                sebepler.append({'sebep': ad, 'tip': x.get('tip', 'plansiz')})

    return jsonify({
        'tarih':        tarih,
        'min_durus_dk': esik_dk,
        'vardiyalar':   cikti,
        'zamansiz':     zamansiz,
        'sebepler':     sebepler,
    })


@app.route('/api/durus/<int:did>/bol', methods=['POST'])
@operator_required
def durus_bol(did):
    """Bir duruşu verilen ANDA ikiye böler; ikinci parçaya istenirse başka sebep atar.

    Body: {'bolme_ts': 'YYYY-MM-DD HH:MM', 'yeni_sebep': ..., 'yeni_tip': ...,
           'yeni_aciklama': ...}

    NEDEN: operatör 40dk'lık tek bir "Arıza" giriyor ama gerçekte 10dk kalıp
    değişimi + 30dk elektrik arızasıydı. Çizelgede bara tıklanıp bölünür ve her
    parçaya doğru sebep atanır — OEE ve duruş Pareto'su gerçeği gösterir.

    SÜRE DAMGADAN yeniden hesaplanır (sure_dk'ya güvenilmez): bölme noktası
    kullanıcının çizelgede gördüğü BAR üzerinden geliyor, bar da damgadan
    çiziliyor. İkisi ayrışırsa ekranda gördüğüyle kaydedilen tutmazdı."""
    ok, hata = _vardiya_sahibi_kontrol(_durus_vardiya_bul(did) or 0)
    if not ok:
        return jsonify({'hata': hata}), 403

    data = request.get_json(silent=True) or {}
    conn = get_db()
    d = conn.execute('SELECT * FROM duruslar WHERE id=?', (did,)).fetchone()
    if not d:
        return jsonify({'hata': 'Duruş bulunamadı'}), 404

    bas = _ts_parse(d['baslangic_ts'])
    bit = _ts_parse(d['bitis_ts'])
    if bas and not bit:
        bit = bas + timedelta(minutes=int(d['sure_dk'] or 0))
    if not bas or not bit:
        return jsonify({'hata': 'Bu duruşun zamanı belirsiz — önce başlangıç/bitiş atayın'}), 400

    kesim = _ts_parse(data.get('bolme_ts'))
    if not kesim:
        return jsonify({'hata': "bolme_ts 'YYYY-MM-DD HH:MM' biçiminde olmalı"}), 400
    # Saniyeler atılır: çizelge dakika çözünürlüğünde, iki parçanın toplamı
    # orijinal süreyi TAM tutmalı (yuvarlama farkı dakika kaçırırdı).
    kesim = kesim.replace(second=0, microsecond=0)
    if not (bas < kesim < bit):
        return jsonify({'hata': 'Bölme noktası duruşun İÇİNDE olmalı'}), 400

    # TOPLAM SÜRE DEĞİŞMEZ. Bölme bir RAPORLAMA düzeltmesidir: aynı duruşun
    # sebebini ayrıştırır, süresini değil. p2 = toplam − p1 diye hesaplanır;
    # iki parçayı damgadan AYRI AYRI yuvarlamak toplamdan dakika kaçırıyordu
    # (10:53:51 başlayan 30dk'lık duruş 14+15=29 oluyordu) ve OEE her bölmede
    # sessizce iyileşirdi.
    toplam_dk = int(d['sure_dk'] or 0) or int(round((bit - bas).total_seconds() / 60))
    if toplam_dk < 2:
        return jsonify({'hata': f'{toplam_dk} dakikalık duruş bölünemez (en az 2 dk gerekir)'}), 400
    p1_dk = int(round((kesim - bas).total_seconds() / 60))
    p1_dk = max(1, min(toplam_dk - 1, p1_dk))
    p2_dk = toplam_dk - p1_dk

    yeni_sebep = (data.get('yeni_sebep') or d['durus_sebebi'] or '').strip()
    yeni_tip   = (data.get('yeni_tip') or d['durus_tipi'] or 'plansiz').strip()
    yeni_acikl = (data.get('yeni_aciklama') or '').strip()

    conn.execute(
        "UPDATE duruslar SET sure_dk=?, bitis_ts=? WHERE id=?",
        (p1_dk, _ts_metin(kesim), did)
    )
    cur = conn.execute(
        "INSERT INTO duruslar (vardiya_id, durus_sebebi, aciklama, sure_dk, "
        "                      baslangic_saati, durus_tipi, baslangic_ts, bitis_ts) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (d['vardiya_id'], yeni_sebep, yeni_acikl, p2_dk, kesim.strftime('%H:%M'),
         yeni_tip, _ts_metin(kesim), _ts_metin(bit))
    )
    conn.commit()
    return jsonify({
        'basarili': True,
        'parca1': {'id': did, 'sure_dk': p1_dk,
                   'bas_ts': _ts_metin(bas), 'bit_ts': _ts_metin(kesim)},
        'parca2': {'id': cur.lastrowid, 'sure_dk': p2_dk, 'sebep': yeni_sebep,
                   'bas_ts': _ts_metin(kesim), 'bit_ts': _ts_metin(bit)},
    }), 201


# ─────────────────────────────────────────────────────────────
# OEE API
# ─────────────────────────────────────────────────────────────

@app.route('/api/oee/<int:vid>', methods=['GET'])
def oee_vardiya(vid):
    """Tek vardiya OEE hesapla."""
    # OEE'nin güncel sensör sayımını kullanması için auto kayıtları tazele.
    # NOT: get_db() g-paylaşımlı — BURADA KAPATMA (teardown kapatır); erken close
    # sonrası get_db kullanan kod eklenirse 'closed database' 500'e döner.
    try:
        _uretim_sayac_senkron(get_db(), vid)
    except Exception:
        pass
    sonuc = hesapla_oee(vid)
    if not sonuc:
        return jsonify({'hata': 'Vardiya bulunamadı'}), 404
    return jsonify(sonuc)


@app.route('/api/oee', methods=['GET'])
def oee_ozet():
    """Tarih aralığı, robot ve bölüm için OEE özeti."""
    tarih_bas = request.args.get('tarih_bas')
    tarih_bit = request.args.get('tarih_bit')
    robot = request.args.get('robot')
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')   # TK1/TK2 montaj OEE'si karışmasın

    # Varsayılan: bugün
    if not tarih_bas and not tarih_bit:
        bugun = date.today().isoformat()
        tarih_bas = bugun
        tarih_bit = bugun

    # Aralık bugünü kapsıyorsa auto sayaçları tazele — OEE özeti canlı adetlerle hesaplansın
    _bugun_o = date.today().isoformat()
    if (not tarih_bas or tarih_bas <= _bugun_o) and (not tarih_bit or tarih_bit >= _bugun_o):
        _acik_auto_vardiyalari_senkronla(get_db(), bolum or None, lokasyon or None)

    sonuc = hesapla_oee_ozet(tarih_bas, tarih_bit, robot, bolum, lokasyon)
    return jsonify(sonuc)


# ─────────────────────────────────────────────────────────────
# REFERANS & ROBOT API
# ─────────────────────────────────────────────────────────────

@app.route('/api/referanslar', methods=['GET'])
def referans_listesi():
    """Referans autocomplete listesi. ?bolum= ile filtrelenebilir.
    Kaynak için kaynak_suresi_sn ve soktak_suresi_sn alanlarını da döner."""
    q = request.args.get('q', '')
    bolum = request.args.get('bolum', '')
    lokasyon = request.args.get('lokasyon', '')
    conn = get_db()
    base = ("SELECT referans_kodu, aciklama, hedef_cycle_time_sn, kaynak_suresi_sn, soktak_suresi_sn, "
            "sure_teyit, sure_teyit_tarihi, COALESCE(bukum_operasyon,1) AS bukum_operasyon, "
            "COALESCE(paket_adedi,1) AS paket_adedi, "
            "COALESCE(depo_kodu,'') AS depo_kodu, "
            "COALESCE(karsi_depo_kodu,'') AS karsi_depo_kodu "
            "FROM referans_listesi")
    sql = base + " WHERE REPLACE(referans_kodu, ' ', '') LIKE REPLACE(?, ' ', '')"
    params = [f'%{q}%']
    if bolum:
        sql += " AND bolum = ?"
        params.append(bolum)
    if lokasyon:
        sql += " AND COALESCE(lokasyon, 'TK2') = ?"
        params.append(lokasyon)
    sql += " ORDER BY referans_kodu"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/referanslar', methods=['POST'])
@panel_gerekli(izin='referanslar')
def referans_ekle():
    """Yeni referans tanımla veya güncelle (bolum opsiyonel — varsayılan 'kaynak').
    Cycle time tanımlandığında Excel'in ilgili bölüm sayfası otomatik
    güncellenir (data/uretim_verileri.xlsx).
    """
    data = request.get_json()
    ref = data.get('referans_kodu', '').strip()
    if not ref:
        return jsonify({'hata': 'referans_kodu zorunludur'}), 400

    ct = float(data.get('hedef_cycle_time_sn', 0))
    desc = data.get('aciklama', '')
    bolum = (data.get('bolum') or 'kaynak').strip()
    if bolum not in GECERLI_BOLUMLER:
        bolum = 'kaynak'
    lokasyon = (data.get('lokasyon') or request.args.get('lokasyon') or 'TK2').strip() or 'TK2'

    conn = get_db()
    try:
        # BÖLÜM+LOKASYONA KAPALI UPSERT: aynı (referans_kodu, bolum, lokasyon) varsa GÜNCELLE,
        # yoksa EKLE. (INSERT OR REPLACE kullanılmıyor → sure_teyit/kaynak/söktak gibi diğer
        # kolonlar korunur; diğer fabrikanın/bölümün aynı kodlu satırına ASLA dokunulmaz —
        # aynı kod birden fazla bölümde farklı cycle ile yaşayabilir.)
        mevcut = conn.execute(
            "SELECT id FROM referans_listesi "
            "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
            "AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=?",
            (ref, bolum, lokasyon)
        ).fetchone()
        # AS400 depo kodları — GÖNDERİLMEDİYSE DOKUNULMAZ (None), gönderildiyse
        # (boş string dahil) yazılır. Kod alanı serbest metin değil ERP kodu:
        # büyük harfe çevrilip boşluklar atılır, 5 karakterle sınırlanır (ekran alanı).
        def _depo(x):
            if x is None:
                return None
            return str(x).strip().upper().replace(' ', '')[:5]
        d_kod = _depo(data.get('depo_kodu'))
        k_kod = _depo(data.get('karsi_depo_kodu'))

        if mevcut:
            conn.execute(
                "UPDATE referans_listesi SET aciklama=?, hedef_cycle_time_sn=?, "
                "depo_kodu=COALESCE(?, depo_kodu), "
                "karsi_depo_kodu=COALESCE(?, karsi_depo_kodu) WHERE id=?",
                (desc, ct, d_kod, k_kod, mevcut['id'])
            )
        else:
            conn.execute(
                "INSERT INTO referans_listesi (referans_kodu, aciklama, hedef_cycle_time_sn, "
                "bolum, lokasyon, depo_kodu, karsi_depo_kodu) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (ref, desc, ct, bolum, lokasyon, d_kod or '', k_kod or '')
            )

        # Geriye dönük: aynı referansın geçmiş üretim kayıtlarının cycle'ını güncelle —
        # AMA sadece bu lokasyonun + bu bölümün vardiyalarına ait olanları
        # (diğer fabrikayı/bölümü ezme — aynı kodun bölüm başına cycle'ı farklı olabilir).
        if ct > 0:
            conn.execute(
                "UPDATE uretim_kayitlari SET cycle_time_sn = ? "
                "WHERE UPPER(REPLACE(referans_kodu, ' ', '')) = UPPER(REPLACE(?, ' ', '')) "
                "AND vardiya_id IN (SELECT id FROM vardiyalar WHERE COALESCE(lokasyon,'TK2')=? AND COALESCE(bolum,'kaynak')=?)",
                (ct, ref, lokasyon, bolum)
            )

        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'hata': str(e)}), 400
    conn.close()

    # Excel auto-sync: SADECE TK2 (uretim_verileri.xlsx). TK1 referansları cycle time
    # kullanmaz ve ayrı Excel'dedir → TK1'de export atlanır (TK1 verisi TK2 dosyasına sızmaz).
    try:
        export_referans_cycle_times(bolum=bolum, lokasyon=lokasyon)
    except Exception as e:
        print(f'[referans_ekle] Excel auto-sync hatası ({bolum}/{lokasyon}): {e}')

    return jsonify({'basarili': True}), 201


# ── Referans KODU değiştirme (rename) — TÜM ilgili tablolarda ──
# Referans kodu 11 tablo.kolonda serbest string (FK yok); rename her yeri günceller.
# as400_teyit_log (audit) DOKUNULMAZ.
_REFERANS_KOD_TABLOLARI = [
    ('referans_listesi', 'referans_kodu'), ('uretim_kayitlari', 'referans_kodu'),
    ('referans_takip', 'referans_kodu'), ('referans_robot_uyumu', 'referans_kodu'),
    ('robot_programlari', 'referans_kodu'), ('robot_programlari_eski', 'referans_kodu'),
    ('robot_is_atamalari', 'referans_kodu'), ('fikstur_adresleri', 'referans_kodu'),
    ('fikstur_raf', 'referans_kodu'), ('test_sonuclari', 'referans_kodu'),
    ('test_sonuclari', 'referans_norm'),
]


def _guvenli_kod_replace(s, eski, yeni):
    """s içindeki 'eski' kodunu 'yeni' ile değiştirir AMA 'eski' daha büyük bir
    kodun parçasıysa (öncesi/sonrası alfasayısal veya - / ) o eşleşmeyi KORUR.
    Örn: eski='10.300.4081' → '10.300.4081W' DOKUNULMAZ (W devam), ama
    '10.300.4081 1. op' değişir (boşluk sınırı). Op etiketi hep korunur."""
    if not s or eski not in s:
        return s
    out = []
    i = 0
    n = len(eski)
    while True:
        j = s.find(eski, i)
        if j < 0:
            out.append(s[i:])
            break
        son = j + n
        onceki = s[j - 1] if j > 0 else ''
        sonraki = s[son] if son < len(s) else ''
        # DIKKAT: `sonraki in '-/'` KULLANMA — bos string ('' in '-/') True doner,
        # string SONUNDA biten tam eslesmeler asla degistirilemezdi (2026-07-22 bug).
        if onceki.isalnum() or sonraki.isalnum() or sonraki in ('-', '/'):
            out.append(s[i:son])          # daha büyük kodun parçası → dokunma
        else:
            out.append(s[i:j])
            out.append(yeni)
        i = son
    return ''.join(out)


def _referans_kodu_rename(conn, eski, yeni, uygula):
    """Tüm ilgili tablolarda kodu değiştirir. uygula=False → sadece önizleme (say + örnek).
    MASTER BİRLEŞTİRME (2026-07-20, 4081 dersi): referans_listesi'nde yeni kod aynı
    bölüm+lokasyonda ZATEN varsa (yazım hatası → doğru koda düzeltme senaryosu),
    hatalı master satırı SİLİNİR, hedef kayıt korunur (süreleri elle taşınabilir).
    Döner: (ozet {tablo.kolon: adet}, ornekler [{tablo, eski, yeni}])."""
    eski = (eski or '').strip()
    yeni = (yeni or '').strip()
    ozet = {}
    ornekler = []
    for t, k in _REFERANS_KOD_TABLOLARI:
        try:
            rows = conn.execute(f"SELECT rowid, {k} FROM {t} WHERE {k} LIKE ?", (f'%{eski}%',)).fetchall()
        except Exception:
            continue   # tablo yoksa atla
        say = 0
        for r in rows:
            rid, deger = r[0], r[1]
            yd = _guvenli_kod_replace(deger, eski, yeni)
            if yd != deger:
                if t == 'referans_listesi':
                    cakisan = conn.execute(
                        "SELECT rowid FROM referans_listesi WHERE rowid != ? "
                        "AND UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
                        "AND COALESCE(bolum,'kaynak')=(SELECT COALESCE(bolum,'kaynak') FROM referans_listesi WHERE rowid=?) "
                        "AND COALESCE(lokasyon,'TK2')=(SELECT COALESCE(lokasyon,'TK2') FROM referans_listesi WHERE rowid=?)",
                        (rid, yd, rid, rid)).fetchone()
                    if cakisan:
                        say += 1
                        if len(ornekler) < 15:
                            ornekler.append({'tablo': t, 'eski': deger, 'yeni': yd + ' (BİRLEŞTİRME: hedef kayıt zaten var, hatalı tanım silinir)'})
                        if uygula:
                            conn.execute("DELETE FROM referans_listesi WHERE rowid=?", (rid,))
                        continue
                say += 1
                if len(ornekler) < 15:
                    ornekler.append({'tablo': t, 'eski': deger, 'yeni': yd})
                if uygula:
                    conn.execute(f"UPDATE {t} SET {k}=? WHERE rowid=?", (yd, rid))
        if say:
            ozet[f'{t}.{k}'] = say
    return ozet, ornekler


@app.route('/api/referans/kod_degistir', methods=['POST'])
@panel_gerekli(izin='referanslar')
def referans_kod_degistir():
    """Bir referans kodunu TÜM ilgili tablolarda değiştirir (rename).
    Body: {eski, yeni, onizleme?:bool}. onizleme=true → uygulamadan say+örnek döndürür."""
    data = request.get_json(silent=True) or {}
    eski = (data.get('eski') or '').strip()
    yeni = (data.get('yeni') or '').strip()
    onizleme = bool(data.get('onizleme'))
    if not eski or not yeni:
        return jsonify({'hata': 'eski ve yeni kod zorunlu'}), 400
    if eski == yeni:
        return jsonify({'hata': 'eski ve yeni kod aynı'}), 400
    conn = get_db()
    try:
        ozet, ornekler = _referans_kodu_rename(conn, eski, yeni, uygula=not onizleme)
        if onizleme:
            return jsonify({'onizleme': True, 'eski': eski, 'yeni': yeni,
                            'toplam': sum(ozet.values()), 'ozet': ozet, 'ornekler': ornekler})
        conn.commit()
        return jsonify({'basarili': True, 'eski': eski, 'yeni': yeni,
                        'toplam': sum(ozet.values()), 'ozet': ozet})
    except Exception as e:
        conn.rollback()
        msg = str(e)
        if 'UNIQUE' in msg.upper():
            return jsonify({'hata': f'Hedef kod aynı bölüm/lokasyonda zaten var: {msg}'}), 409
        return jsonify({'hata': msg}), 400


@app.route('/api/referanslar/eksik', methods=['GET'])
def referans_eksik_listesi():
    """Süresi tanımlanmamış (0 veya NULL) referansları getirir. ?bolum= ve ?lokasyon= ile filtrelenebilir."""
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    conn = get_db()
    c = conn.cursor()
    sql = ("SELECT referans_kodu FROM referans_listesi "
           "WHERE (hedef_cycle_time_sn IS NULL OR hedef_cycle_time_sn = 0)")
    params = []
    if bolum:
        sql += " AND COALESCE(bolum, 'kaynak') = ?"
        params.append(bolum)
    if lokasyon:
        sql += " AND COALESCE(lokasyon, 'TK2') = ?"
        params.append(lokasyon)
    sql += " ORDER BY referans_kodu"
    rows = c.execute(sql, params).fetchall()
    conn.close()
    return jsonify([r['referans_kodu'] for r in rows])


@app.route('/api/referanslar/indir_excel', methods=['GET'])
@panel_gerekli(izin='referanslar')
def referans_excel_indir():
    """Bölümün referanslarını SÜRELERİYLE birlikte .xlsx olarak İNDİRİR.

    Amaç (kullanıcı 2026-08-04): mevcut referansları planlama bölümüyle paylaşmak.
    'Excel'e Yaz'dan FARKI: o, sunucudaki kaynak dosyayı (data/uretim_verileri.xlsx)
    günceller; bu ise tarayıcıya indirilen BAĞIMSIZ bir dosya üretir — paylaşmak
    için kaynak dosyaya dokunmak gerekmez.

    SÜRESİ TANIMSIZ OLANLAR LİSTEYE GİRMEZ (hedef_cycle_time_sn <= 0): planlama
    süresiz satırla hesap yapamaz; boş süre gönderilirse 0 sanılıp yanlış kapasite
    çıkarılır. Kaç satırın bu yüzden atlandığı dosyanın altında yazar ki eksik
    olduğu görünsün.

    Query: ?bolum=kaynak&lokasyon=TK2
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import io as _io

    bolum = (request.args.get('bolum') or 'kaynak').strip()
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip().upper()
    if bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f'Geçersiz bölüm: {bolum}'}), 400

    conn = get_db()
    satirlar = conn.execute(
        "SELECT referans_kodu, COALESCE(aciklama,'') aciklama, "
        "       COALESCE(hedef_cycle_time_sn,0) ct, "
        "       COALESCE(kaynak_suresi_sn,0) ks, COALESCE(soktak_suresi_sn,0) ss, "
        "       COALESCE(sure_teyit,0) teyit, COALESCE(sure_teyit_tarihi,'') teyit_tarih "
        "FROM referans_listesi "
        "WHERE COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=? "
        "ORDER BY referans_kodu", (bolum, lokasyon)).fetchall()
    dolu = [r for r in satirlar if (r['ct'] or 0) > 0]
    atlanan = len(satirlar) - len(dolu)

    wb = Workbook()
    ws = wb.active
    ws.title = (BOLUM_AD.get(bolum, bolum))[:31]      # Excel sayfa adı en fazla 31 karakter
    # Kaynak bölümünde süre kaynak+söktak olarak ayrışır — planlama ikisini de ister.
    kaynak_mi = (bolum == 'kaynak')
    basliklar = ['Referans Kodu', 'Açıklama', 'Cycle Time (sn)']
    if kaynak_mi:
        basliklar += ['Kaynak Süresi (sn)', 'Söktak Süresi (sn)']
    basliklar += ['Süre Teyitli', 'Teyit Tarihi']
    ws.append(basliklar)
    _bas_font = Font(bold=True, color='FFFFFF')
    _bas_dolgu = PatternFill('solid', fgColor='6D28D9')
    for h in ws[1]:
        h.font, h.fill = _bas_font, _bas_dolgu
        h.alignment = Alignment(horizontal='center', vertical='center')
    for r in dolu:
        satir = [r['referans_kodu'], r['aciklama'], round(float(r['ct']), 1)]
        if kaynak_mi:
            satir += [round(float(r['ks']), 1), round(float(r['ss']), 1)]
        satir += ['Evet' if r['teyit'] else 'Hayır', r['teyit_tarih']]
        ws.append(satir)
    # Kolon genişlikleri: içeriğe göre (planlama dosyayı açar açmaz okusun)
    for i, _b in enumerate(basliklar, start=1):
        en = max([len(str(_b))] + [len(str(c.value or '')) for c in ws[chr(64 + i)]][:400])
        ws.column_dimensions[chr(64 + i)].width = min(max(en + 3, 12), 45)
    ws.freeze_panes = 'A2'
    # Atlananları dosyanın SONUNA not düş — "liste eksik mi?" sorusu doğsun
    if atlanan:
        ws.append([])
        ws.append([f'NOT: süresi tanımlı olmadığı için {atlanan} referans bu listeye '
                   f'alınmadı (toplam {len(satirlar)} referans var).'])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color='9D174D')

    bellek = _io.BytesIO()
    wb.save(bellek)
    bellek.seek(0)
    _ad = f"referanslar_{bolum}_{lokasyon}_{date.today().isoformat()}.xlsx"
    return send_file(bellek, as_attachment=True, download_name=_ad,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def kayit_grubu(bolum, robot_no, istasyon):
    """Üretim kaydının GRUP etiketi — Kayıtlar sayfasındaki kırılım ve Excel için.

    Kullanıcı 2026-08-25: "tk1'de tel üretimi kısmında kapama makineleri
    üretimlerini bir arada gruplanmış şekilde görelim ve dilersek filtre ile
    sadece kapama bölümünün üretim verilerini excel ile indirebilelim."

    TEL'de grup = PROSES ADIMI ('Kapama 5' → 'Kapama'): 12 kapama presi tek
    başlıkta toplanır, kullanıcı "kapama ne üretti" sorusunu tek bakışta görür.
    Diğer bölümlerde grup = MAKİNE (pres/plastik/lazer'de makine üretim
    kaydındadır); makinesiz bölümlerde (kaynak/montaj) vardiyanın hattı.
    """
    if bolum == 'tel':
        adim = tel_adim_etiketi(kayit_hatti(robot_no, istasyon))
        if adim:
            return adim
    return kayit_makine_ad(robot_no, istasyon) or str(robot_no or '').strip() or '—'


@app.route('/api/kayitlar/uretim_excel', methods=['GET'])
@panel_gerekli(izin='kayitlar')
def kayitlar_uretim_excel():
    """Kayıtlar sayfasındaki üretim listesini .xlsx olarak indirir.

    Query: ?tarih_bas&tarih_bit&bolum&lokasyon&robot&grup
    'grup' verilirse YALNIZ o grup (tel'de proses adımı, diğerlerinde makine)
    dışa aktarılır — ekrandaki filtre ile birebir aynı küme iner.

    İki sayfa: 'Üretim' (satır satır, grup sütunlu) + 'Özet' (grup × gün).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import io as _io

    # ÇOKLU (2026-08-25): 'bolum' virgüllü liste olabilir; 'makine'/'grup' de
    # çoğul karşılıklarıyla ('makineler'/'gruplar') gelebilir. Tekil biçimler
    # KORUNDU — eski bağlantılar ve dış çağrılar bozulmasın.
    bolum = (request.args.get('bolum') or 'kaynak').strip()
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip().upper()
    bolumler = [b.strip() for b in bolum.split(',') if b.strip()] or ['kaynak']
    gecersiz = [b for b in bolumler if b not in GECERLI_BOLUMLER]
    if gecersiz:
        return jsonify({'hata': f'Geçersiz bölüm: {", ".join(gecersiz)}'}), 400
    bugun = date.today().isoformat()
    t_bas = (request.args.get('tarih_bas') or bugun).strip()
    t_bit = (request.args.get('tarih_bit') or bugun).strip()
    robot = (request.args.get('robot') or '').strip()
    # 'makine' = ekrandaki makine/masa filtresi (2026-08-25). 'robot' vardiyanın
    # hattına bakar; hattı üretim kaydında seçilen bölümlerde o değer sabit
    # olduğundan süzmez — ekranla aynı küme insin diye ayrı parametre.
    # ÇOĞUL BİÇİM: ekranda birden fazla makine/adım seçilebiliyor. Virgülle
    # ayrılır; tekil 'grup'/'makine' de kabul edilir (eski bağlantılar).
    def _csv(*adlar):
        out = []
        for a in adlar:
            for p in (request.args.get(a) or '').split(','):
                p = p.strip()
                if p and p not in out:
                    out.append(p)
        return out

    grup_f = _csv('gruplar', 'grup')
    makine_f = _csv('makineler', 'makine')

    def _eslesir(secim, satir_bolum, deger):
        """Seçim boşsa her şey geçer. Seçim 'bolum|ad' biçiminde de olabilir —
        panel bölümlü biçimi gönderir ki iki bölümde aynı adlı makine varsa
        biri diğerinin kayıtlarını da indirmesin. Düz ad da kabul edilir."""
        if not secim:
            return True
        for p in secim:
            if '|' in p:
                b, d = p.split('|', 1)
                if b == satir_bolum and d == deger:
                    return True
            elif p == deger:
                return True
        return False

    kosul = ["v.tarih BETWEEN ? AND ?",
             "COALESCE(v.bolum,'kaynak') IN (%s)" % ','.join('?' * len(bolumler)),
             "COALESCE(v.lokasyon,'TK2')=?"]
    parm = [t_bas, t_bit] + bolumler + [lokasyon]
    if robot:
        kosul.append("v.robot_no=?")
        parm.append(robot)

    conn = get_db()
    rows = conn.execute(
        "SELECT v.tarih, v.vardiya_turu, v.robot_no, "
        "       COALESCE(v.bolum,'kaynak') bolum, v.operator_adi, "
        "       COALESCE(u.istasyon,0) istasyon, u.referans_kodu, "
        "       COALESCE(u.ok_adet,0) ok_adet, COALESCE(u.nok_adet,0) nok_adet, "
        "       COALESCE(u.cycle_time_sn,0) ct, COALESCE(u.aciklama,'') aciklama "
        "FROM uretim_kayitlari u JOIN vardiyalar v ON v.id=u.vardiya_id "
        "WHERE " + " AND ".join(kosul) +
        " ORDER BY v.tarih, u.referans_kodu", parm).fetchall()

    satirlar = []
    for r in rows:
        g = kayit_grubu(r['bolum'], r['robot_no'], r['istasyon'])
        if not _eslesir(grup_f, r['bolum'], g):
            continue
        mk = kayit_hatti(r['robot_no'], r['istasyon'])
        if not _eslesir(makine_f, r['bolum'], mk):
            continue
        satirlar.append({
            'tarih': r['tarih'], 'vardiya': r['vardiya_turu'],
            'bolum': BOLUM_AD.get(r['bolum'], r['bolum']), 'grup': g,
            'hat': mk,
            'operator': r['operator_adi'], 'ref': r['referans_kodu'],
            'ok': int(r['ok_adet']), 'nok': int(r['nok_adet']),
            'ct': round(float(r['ct'] or 0), 1), 'aciklama': r['aciklama'],
        })

    wb = Workbook()
    ws = wb.active
    ws.title = 'Üretim'
    basliklar = ['Tarih', 'Vardiya', 'Bölüm', 'Grup', 'Makine / Hat', 'Operatör',
                 'Referans', 'OK', 'NOK', 'CT (sn)', 'Açıklama']
    ws.append(basliklar)
    _bf, _bd = Font(bold=True, color='FFFFFF'), PatternFill('solid', fgColor='6D28D9')
    for h in ws[1]:
        h.font, h.fill = _bf, _bd
        h.alignment = Alignment(horizontal='center', vertical='center')
    for s in satirlar:
        ws.append([s['tarih'], s['vardiya'], s['bolum'], s['grup'], s['hat'],
                   s['operator'], s['ref'], s['ok'], s['nok'], s['ct'], s['aciklama']])
    for i, _b in enumerate(basliklar, start=1):
        en = max([len(str(_b))] + [len(str(c.value or '')) for c in ws[chr(64 + i)]][:500])
        ws.column_dimensions[chr(64 + i)].width = min(max(en + 3, 10), 42)
    ws.freeze_panes = 'A2'

    # ── ÖZET: grup × gün (kullanıcının asıl istediği "bir arada gruplanmış" görünüm)
    # ÇOKLU BÖLÜMDE anahtar 'Bölüm · Grup': 'Kapama' telde proses adımı, başka
    # bölümde makine adı olabilir — tek adla toplansalar iki bölümün üretimi
    # aynı satıra yığılırdı.
    ws2 = wb.create_sheet('Özet')
    _cok_bolum = len({s['bolum'] for s in satirlar}) > 1
    for s in satirlar:
        s['ozet_ad'] = (f"{s['bolum']} · {s['grup']}" if _cok_bolum else s['grup'])
    gunler = sorted({s['tarih'] for s in satirlar})
    gruplar = sorted({s['ozet_ad'] for s in satirlar})
    ws2.append(['Bölüm · Grup' if _cok_bolum else 'Grup'] + gunler
               + ['TOPLAM OK', 'TOPLAM NOK'])
    for h in ws2[1]:
        h.font, h.fill = _bf, _bd
        h.alignment = Alignment(horizontal='center', vertical='center')
    for g in gruplar:
        satir = [g]
        for gun in gunler:
            satir.append(sum(s['ok'] for s in satirlar
                             if s['ozet_ad'] == g and s['tarih'] == gun))
        satir.append(sum(s['ok'] for s in satirlar if s['ozet_ad'] == g))
        satir.append(sum(s['nok'] for s in satirlar if s['ozet_ad'] == g))
        ws2.append(satir)
    if gruplar:
        gt = ['GENEL TOPLAM']
        for gun in gunler:
            gt.append(sum(s['ok'] for s in satirlar if s['tarih'] == gun))
        gt.append(sum(s['ok'] for s in satirlar))
        gt.append(sum(s['nok'] for s in satirlar))
        ws2.append(gt)
        for c in ws2[ws2.max_row]:
            c.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 34
    ws2.freeze_panes = 'B2'

    bellek = _io.BytesIO()
    wb.save(bellek)
    bellek.seek(0)
    # Dosya adı: tek seçimde adı yazılır, çoklu seçimde sayısı — 12 presin adı
    # dosya adına sığmaz ve Windows'ta ad uzunluğu sınırına takılır.
    def _ekle(deger, cogul_ad):
        if not deger:
            return ''
        if len(deger) == 1:
            # 'bolum|ad' geldiyse dosya adına yalnız ad girer: '|' Windows'ta
            # geçersiz karakter, indirme adı bozulurdu.
            ad = deger[0].split('|')[-1]
            return '_' + ad.replace(' ', '_')
        return f'_{len(deger)}{cogul_ad}'

    _ek = _ekle(grup_f, 'adim') + _ekle(makine_f, 'makine')
    # Çoklu bölümde adlar yerine sayı: virgüllü uzun ad hem çirkin hem de
    # bölüm sayısı arttıkça Windows'un dosya adı sınırına yaklaşıyor.
    _bad = bolumler[0] if len(bolumler) == 1 else f'{len(bolumler)}bolum'
    _ad = f"uretim_{lokasyon}_{_bad}{_ek}_{t_bas}_{t_bit}.xlsx"
    return send_file(bellek, as_attachment=True, download_name=_ad,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/referanslar/export_excel', methods=['POST'])
@panel_gerekli(izin='referanslar')
def referans_excel_export():
    """Veritabanındaki referans listesini data/uretim_verileri.xlsx'in
    aktif bölüm sayfasına yazar.

    Body: { bolum: 'kaynak' | 'montaj' | 'metal' } — yoksa kaynak.
    """
    # Sayfa adları import_excel.BOLUM_SAYFA'dan (tek kaynak) — yeni bölümler dahil.
    from import_excel import BOLUM_SAYFA as _IE_SAYFA
    BOLUM_SAYFA = {b: v['ref'] for b, v in _IE_SAYFA.items()}
    body = request.get_json(silent=True) or {}
    bolum = body.get('bolum', 'kaynak')
    lokasyon = (body.get('lokasyon') or request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    if bolum not in BOLUM_SAYFA:
        return jsonify({'hata': f"Geçersiz bölüm: {bolum}"}), 400
    # TK1 referansları cycle time kullanmaz ve ayrı dosyadadır (tek kolon) → TK2 Excel'ine yazma
    if lokasyon == 'TK1':
        return jsonify({'basarili': True, 'mesaj': 'TK1 referansları için cycle time Excel aktarımı yok.'}), 200
    # TEK implementasyon: import_excel.export_referans_cycle_times — kaynak sayfasında
    # 4+1 kolon (Kaynak/Söktak/Toplam/Süre Teyit), montaj/metal'de 2 kolon yazar.
    # (Eski yerel kopya kaynak'ta B kolonuna toplam cycle yazıp Kaynak Süresi'ni eziyordu.)
    try:
        sonuc = export_referans_cycle_times(bolum=bolum, lokasyon=lokasyon)
        if not sonuc.get('basarili'):
            return jsonify({'hata': sonuc.get('hata', 'Excel yazılamadı')}), 500
        return jsonify({'basarili': True, 'mesaj': f"{sonuc.get('yazilan', 0)} referans Excel'e yazıldı."})
    except Exception as e:
        return jsonify({'hata': str(e)}), 500


@app.route('/api/referanslar/<string:kod>', methods=['DELETE'])
@panel_gerekli(izin='referanslar')
def referans_sil(kod):
    """Referansı listeden sil — SADECE aktif lokasyonun satırı (diğer fabrikanın aynı kodlu
    referansına dokunma). lokasyon query/body'den; verilmezse TK2.
    ?bolum= verilirse yalnız o bölümün satırı silinir (aynı kod birden fazla bölümde
    yaşayabilir); verilmezse eski davranış (koddaki tüm TK2 satırları)."""
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    bolum = (request.args.get('bolum') or '').strip()
    conn = get_db()
    try:
        if bolum:
            conn.execute(
                "DELETE FROM referans_listesi WHERE referans_kodu = ? AND COALESCE(bolum,'kaynak') = ? AND COALESCE(lokasyon,'TK2') = ?",
                (kod, bolum, lokasyon)
            )
        else:
            conn.execute(
                "DELETE FROM referans_listesi WHERE referans_kodu = ? AND COALESCE(lokasyon,'TK2') = ?",
                (kod, lokasyon)
            )
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'hata': str(e)}), 400
    conn.close()
    return jsonify({'basarili': True}), 200


@app.route('/api/robotlar', methods=['GET'])
def robot_listesi():
    """Bölüm bazlı robot/hat/makine listesi.
    - kaynak: ABB1..ABB9 (sabit)
    - metal: 300T, 400T, 550T, Şerit Testere (sabit)
    - lazer: tek sabit hat 'Lazer' — makine (Lazer 1..6) vardiyada DEĞİL üretim kaydında
      seçilir (istasyon kolonu 1..6; operatör aynı anda birden fazla makine çalıştırır)
    - pres: tek sabit hat 'Pres Abkant' (2026-07-28) — makine (Abkant 1-3 / Pres 1-5 /
      Hidrolik Pres / Bros) vardiyada DEĞİL üretim kaydında seçilir (istasyon 1..10);
      operatör gün içinde
      makine değiştirebildiği için lazer modeline geçildi. Eski vardiyaların makine
      adları (Abkant 2, Pres...) listede KALIR → dashboard geçmiş filtreleri çalışsın.
    - montaj/isleme: vardiyalardan distinct robot_no — operatörler girdikçe dinamik büyür
    """
    bolum = request.args.get('bolum', 'kaynak')
    lokasyon = request.args.get('lokasyon', 'TK2')
    # Plastik enjeksiyon (TK1, 2026-07-23): sabit 2 makine + operatörün eklediği ekstralar.
    # TK1 montaj-hat kısayolundan ÖNCE — aksi halde plastik montaj hatlarıyla ezilirdi.
    if bolum == 'plastik':
        # 2026-08-18: makine artık VARDİYADA DEĞİL ÜRETİM KAYDINDA seçilir (pres modeli
        # — kullanıcı: "makine seçimini referansı seçtikten sonra yapalım, abkant gibi").
        # Vardiya hattı tek ve sabit: PLASTIK_SABIT_HAT. Mobil bu listeyi kullanmaz
        # (hat gizli/sabit); dashboard GEÇMİŞ filtreleri kullanır → eski vardiyaların
        # makine adları ('320T', 'Yapistirma' …) listede KALIR.
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum,'')='plastik' "
            "AND robot_no IS NOT NULL AND robot_no != '' ORDER BY robot_no").fetchall()
        conn.close()
        return jsonify([PLASTIK_SABIT_HAT]
                       + [r['robot_no'] for r in rows if r['robot_no'] != PLASTIK_SABIT_HAT])
    # Tel üretimi (TK1, 2026-08-04): hat = PROSES ADIMI. Operatör vardiyasında
    # hangi adımda çalıştığını seçer; aynı referans gün içinde farklı adımlarda
    # farklı operatörler tarafından kaydedilir (üretim yalnız SON adımda sayılır).
    if bolum == 'tel':
        # 2026-08-18: proses adımı (hat) artık VARDİYADA DEĞİL ÜRETİM KAYDINDA seçilir —
        # bir operatör aynı gün aynı referansın hem kapamasını hem son montajını
        # yapabiliyor. Vardiya hattı tek ve sabit: TEL_SABIT_HAT.
        # Mobil bu listeyi kullanmaz (hat gizli/sabit); dashboard GEÇMİŞ filtreleri
        # kullanır → eski vardiyaların hat adları listede KALIR.
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum,'')='tel' "
            "AND robot_no IS NOT NULL AND robot_no != '' ORDER BY robot_no").fetchall()
        conn.close()
        return jsonify([TEL_SABIT_HAT]
                       + [r['robot_no'] for r in rows if r['robot_no'] != TEL_SABIT_HAT])
    if lokasyon == 'TK1' and bolum == 'montaj':
        # 2026-08-18: TK1 montajda da hat ÜRETİM KAYDINDA seçilir (operatör gün içinde
        # masa/hat değiştirebiliyor). Vardiya hattı tek ve sabit: TK1_MONTAJ_SABIT_HAT.
        # TK2 montaj DEĞİŞMEDİ — orada hat vardiyada seçilir.
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum,'')='montaj' "
            "AND COALESCE(lokasyon,'TK2')='TK1' AND robot_no IS NOT NULL AND robot_no != '' "
            "ORDER BY robot_no").fetchall()
        conn.close()
        return jsonify([TK1_MONTAJ_SABIT_HAT]
                       + [r['robot_no'] for r in rows if r['robot_no'] != TK1_MONTAJ_SABIT_HAT])
    if lokasyon == 'TK1':
        # TK1 (yan tesis) sabit hatları — montaj mantığı.
        # 2026-07-31: 7 masaya sayaç modülü takıldı (TK1-M1..TK1-M7). TK2'deki
        # M1..M12 modeli: MASA = hat → operatör masasını seçince otomatik sayaç
        # çalışır. Eski 4 hat adı KALDI (geçmiş vardiyalar + hat bazlı çalışma).
        # Pull / Push-Pull ÇIKARILDI (kullanıcı 2026-08-04): bunlar TEL ÜRÜNÜ adları,
        # montaj hattı değil — telin kendisi 'tel' bölümünde proses adımlarında üretilir.
        # Geçmiş vardiyalar bu isimlerle kayıtlı kalır (TK1_ROBOT_NOLARI'nda duruyorlar).
        return jsonify(['LF-LFP', 'Iveco']
                       + [f'TK1-M{i}' for i in range(1, 8)])
    # Sabit makine tabanı olan bölümler (liste her zaman tam görünür; DISTINCT ekstraları eklenir)
    # pres BURADA DEĞİL: hattı sabit (aşağıda), makineleri üretim kaydında seçilir.
    # plastik BURADA DEĞİL (2026-08-18): pres gibi hattı sabit, makineleri üretim
    # kaydında seçilir — yukarıdaki erken dönüşte karşılanır.
    SABIT_TABAN = {}
    # Dinamik bölümlerde hiç vardiya yokken gösterilecek varsayılan makine/hat adı
    DINAMIK_VARSAYILAN = {'montaj': 'HAT 1', 'isleme': 'Tezgah 1'}
    if bolum == 'metal':
        robotlar = ['300T', '400T', '550T', 'Şerit Testere']
    elif bolum == 'lazer':
        robotlar = ['Lazer']
    elif bolum == 'pres':
        # Sabit hat + GEÇMİŞ vardiyaların makine adları (eski model: makine vardiyadaydı).
        # Mobil bu listeyi kullanmaz (hat gizli/sabit); dashboard geçmiş filtreleri kullanır.
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum,'')='pres' "
            "AND robot_no IS NOT NULL AND robot_no != '' ORDER BY robot_no").fetchall()
        conn.close()
        robotlar = [PRES_SABIT_HAT] + [r['robot_no'] for r in rows if r['robot_no'] != PRES_SABIT_HAT]
    elif bolum in SABIT_TABAN or bolum in DINAMIK_VARSAYILAN:
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum, 'kaynak') = ? "
            "AND COALESCE(lokasyon, 'TK2') = ? AND robot_no IS NOT NULL AND robot_no != '' ORDER BY robot_no",
            (bolum, lokasyon or 'TK2')
        ).fetchall()
        conn.close()
        kullanilan = [r['robot_no'] for r in rows]
        taban = SABIT_TABAN.get(bolum, [])
        robotlar = taban + [r for r in kullanilan if r not in taban]
        # Hiç kayıt yoksa varsayılanı göster — operatör formundan ilk girişte yenileri otomatik eklenir
        if not robotlar:
            robotlar = [DINAMIK_VARSAYILAN[bolum]]
    else:
        # kaynak — sabit liste + o bölümde GEÇMİŞTE kullanılmış adlar (eski kayıtlar
        # filtrelerden düşmesin). Sabit liste app.KAYNAK_ROBOTLARI'dır: punta kaynak
        # gibi yeni makineler tek yerden eklenir.
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT robot_no FROM vardiyalar WHERE COALESCE(bolum,'kaynak')='kaynak' "
            "AND COALESCE(lokasyon,'TK2')=? AND robot_no IS NOT NULL AND robot_no != '' "
            "ORDER BY robot_no", (lokasyon or 'TK2',)
        ).fetchall()
        conn.close()
        robotlar = KAYNAK_ROBOTLARI + [r['robot_no'] for r in rows
                                       if r['robot_no'] not in KAYNAK_ROBOTLARI]
    return jsonify(robotlar)


# ─────────────────────────────────────────────────────────────
# EOQ (AS400 sipariş miktarı) — kullanıcı 2026-08-19
# AS400 07/10/01 "SITUAZIONE ARTICOLO" ekranında her ürün için tanımlı EOQ var
# ve kullanıcıya göre çoğu ESKİ. Hepsini bir arada görebilmek için referans
# başına saklanır (referans_listesi.eoq / eoq_tarih / eoq_kaynak).
# Değer İKİ yoldan gelir: AS400 okuyucusu (as400/eoq_kesif.js ile keşfedilen
# akış) ya da panelden elle giriş. eoq_tarih = son güncelleme anı → listede
# hangi bilginin bayat olduğu görünür.
# ─────────────────────────────────────────────────────────────

@app.route('/kaynak_eoq')
@panel_gerekli()
def kaynak_eoq_sayfasi():
    """Kaynak referanslarının EOQ listesi (panel girişi gerekir)."""
    return render_template('kaynak_eoq.html')


@app.route('/api/kaynak_eoq', methods=['GET'])
def kaynak_eoq_listesi():
    """Bir bölümün referansları + EOQ değerleri. Varsayılan: kaynak / TK2."""
    bolum = (request.args.get('bolum') or 'kaynak').strip()
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    conn = get_db()
    rows = conn.execute(
        "SELECT referans_kodu, aciklama, hedef_cycle_time_sn, eoq, eoq_tarih, eoq_kaynak "
        "FROM referans_listesi "
        "WHERE COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=? "
        "ORDER BY referans_kodu",
        (bolum, lokasyon)
    ).fetchall()
    liste = [dict(r) for r in rows]
    resp = jsonify({
        'bolum': bolum, 'lokasyon': lokasyon,
        'referanslar': liste,
        'ozet': {
            'toplam':  len(liste),
            'eoq_var': sum(1 for r in liste if (r['eoq'] or 0) > 0),
            'eoq_yok': sum(1 for r in liste if not (r['eoq'] or 0)),
        },
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/kaynak_eoq', methods=['POST'])
def kaynak_eoq_kaydet():
    """EOQ yaz. Tek kayıt (panelden elle) ya da TOPLU (AS400 okuyucusu).

    Gövde: {'kaynak': 'as400'|'elle', 'lokasyon': 'TK2',
            'kayitlar': [{'referans_kodu': '10.300.3059W', 'eoq': 850}, ...]}
    Tek kayıt için 'referans_kodu' + 'eoq' doğrudan gövdede de olabilir.
    Eşleşme NORMALİZE koda göre (boşluk/harf farkı): AS400 kodu '10.300.3059 W'
    yazsa bile aynı referansa düşer. Bulunamayan kod 'eslesmeyen'de döner —
    sessizce yutulmaz, hangi kodun sistemde olmadığı görülür.
    """
    data = request.get_json(silent=True) or {}
    kayitlar = data.get('kayitlar')
    if not kayitlar:
        if data.get('referans_kodu') is None:
            return jsonify({'hata': 'kayitlar veya referans_kodu zorunlu'}), 400
        kayitlar = [{'referans_kodu': data.get('referans_kodu'), 'eoq': data.get('eoq')}]
    kaynak = (data.get('kaynak') or 'elle').strip()[:20]
    lokasyon = (data.get('lokasyon') or 'TK2').strip() or 'TK2'
    simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    guncellenen, eslesmeyen = 0, []
    conn = get_db()
    try:
        c = conn.cursor()
        for k in kayitlar:
            kod = (k.get('referans_kodu') or '').strip()
            if not kod:
                continue
            try:
                # Boş/None → EOQ temizle (elle silme mümkün olsun)
                deger = k.get('eoq')
                deger = None if deger in (None, '', '-') else int(deger)
            except (TypeError, ValueError):
                eslesmeyen.append(kod)
                continue
            cur = c.execute(
                "UPDATE referans_listesi SET eoq=?, eoq_tarih=?, eoq_kaynak=? "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
                "  AND COALESCE(lokasyon,'TK2')=?",
                (deger, simdi, kaynak, kod, lokasyon)
            )
            if cur.rowcount:
                guncellenen += cur.rowcount
            else:
                eslesmeyen.append(kod)
        conn.commit()
    except Exception as e:
        print(f'HATA (kaynak_eoq_kaydet): {traceback.format_exc()}')
        return jsonify({'hata': str(e)}), 500
    return jsonify({'basarili': True, 'guncellenen': guncellenen,
                    'eslesmeyen': eslesmeyen}), 200


@app.route('/api/kayit_hatlari', methods=['GET'])
def kayit_hatlari_api():
    """Hattı/makinesi ÜRETİM KAYDINDA seçilen bölümlerin hat listesi (mobil ekran D).

    ?lokasyon=TK1&bolum=tel →
      {'sabit_hat': 'Tel Üretimi', 'etiket': 'Proses Adımı',
       'hatlar': [{'no': 1, 'ad': 'Halat Kesme 1'}, ...]}
    sabit_hat '' ise bu bölümde hat VARDİYADA seçilir → 'hatlar' boş döner.
    'no' = uretim_kayitlari.istasyon değeri (listedeki sıra) — mobil bu numarayı
    gönderir, ad çözümü SUNUCUDA yapılır (tek kaynak: HAT_MAKINELERI).
    """
    lokasyon = request.args.get('lokasyon') or 'TK2'
    bolum = (request.args.get('bolum') or '').strip()
    sh = sabit_hat_adi(lokasyon, bolum)
    etiket = {'tel': 'Proses Adımı', 'montaj': 'Hat'}.get(bolum, 'Makine')
    # Kullanılmayan hatlar listede GÖSTERİLMEZ ama POZİSYONLARI listede DURUR
    # (istasyon numaraları kaymasın): kodu bekleyen slotlar (TEL_GIZLI_HATLAR) ve
    # tel son montaja devredilen montaj butonları (GIZLI_MONTAJ_HATLARI).
    hatlar = [{'no': i + 1, 'ad': ad, 'adim': tel_hat_adimi(ad) or ''}
              for i, ad in enumerate(HAT_MAKINELERI.get(sh, []))
              if ad not in GIZLI_HATLAR] if sh else []
    # TEL: LİSTE SIRASI ≠ POZİSYON (kullanıcı 2026-08-20). Yeni makine listenin
    # SONUNA eklenir (pozisyon = istasyon no, asla kaymamalı) ama operatör onu
    # kendi proses adımının yanında görmeli — 'Soyma' kesimin altında, dropdown'ın
    # dibinde değil. Çözüm: GÖNDERİM SIRASI proses sırasına göre, 'no' aynen kalır.
    # Kayda giden değer 'no' olduğu için sıralama tamamen görsel, risksiz.
    if bolum == 'tel':
        _sira = {a: i for i, a in enumerate(TEL_ADIMLARI)}
        hatlar.sort(key=lambda h: (_sira.get(h['adim'], len(_sira)), h['no']))
    resp = jsonify({'sabit_hat': sh, 'etiket': etiket, 'hatlar': hatlar})
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/test_cihazlari', methods=['GET'])
def test_cihazlari():
    """Cofle test cihazı listesi (operatör mobil dropdown'u için). Token sunucuda
    kalır — mobil API'ye doğrudan erişmez. Entegrasyon kapalıysa boş liste döner.

    ?lokasyon=&bolum=&hat= verilirse liste TEST_CIHAZ_ESLEME'ye göre SÜZÜLÜR:
    o hatta hangi cihaz(lar) kullanılıyorsa yalnız onlar döner. Tek cihaz varsa
    'otomatik' alanında o cihazın id'si gelir → mobil kendiliğinden seçer.
    'eslesme_var' False ise bölüm için eşleme tanımlanmamıştır (eski davranış).
    """
    try:
        import cofle_test
        if not cofle_test.etkin():
            return jsonify({'etkin': False, 'cihazlar': [], 'eslesme_var': False,
                            'otomatik': None})
        tum = cofle_test.cihaz_listesi()
        lokasyon = request.args.get('lokasyon')
        bolum = request.args.get('bolum')
        hat = request.args.get('hat')
        if not bolum:
            # Bağlam verilmediyse eski davranış (dashboard/tanı çağrıları)
            return jsonify({'etkin': True, 'cihazlar': tum, 'eslesme_var': False,
                            'otomatik': None})
        idler, eslesme_var = test_cihaz_eslemesi(lokasyon, bolum, hat)
        if not eslesme_var:
            return jsonify({'etkin': True, 'cihazlar': tum, 'eslesme_var': False,
                            'otomatik': None})
        # SVP listesi geçici olarak boş/bayat gelse bile eşlenmiş cihaz KAYBOLMASIN:
        # adı bulunamayanı id'siyle göster (kutuyu gizlemek, sayımı durdurmaktır).
        ad_map = {c['id']: c.get('ad') for c in tum}
        cihazlar = [{'id': i, 'ad': ad_map.get(i) or f'Cihaz {i}'} for i in idler]
        return jsonify({'etkin': True, 'cihazlar': cihazlar, 'eslesme_var': True,
                        'otomatik': (cihazlar[0]['id'] if len(cihazlar) == 1 else None)})
    except Exception as e:
        print(f"[test_cihazlari] hata: {e}")
        return jsonify({'etkin': False, 'cihazlar': [], 'eslesme_var': False,
                        'otomatik': None})


@app.route('/api/operatorler', methods=['GET'])
def operator_listesi():
    """Operatör listesini döndür. ?bolum= ile filtrelenebilir.
    PIN bilgisi bu endpoint'te VERİLMEZ (operatör mobile bunu çağırır).
    Yönetim için /api/operatorler/yonetim kullanın.
    """
    bolum = request.args.get('bolum', '')
    lokasyon = request.args.get('lokasyon', '')
    conn = get_db()
    sql = "SELECT id, ad, bolum, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE 1=1"
    params = []
    if bolum:
        # Admin her bölümün listesinde görünür (PIN modal + vardiya formu dropdown'u)
        sql += " AND (bolum = ? OR ad = ?)"
        params.extend([bolum, ADMIN_ADI])
    if lokasyon:
        sql += " AND COALESCE(lokasyon,'TK2') = ?"
        params.append(lokasyon)
    sql += " ORDER BY ad"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/operatorler/yonetim', methods=['GET'])
@panel_gerekli(izin='operatorler')
def operator_listesi_yonetim():
    """Yönetici için operatör listesi + PIN bilgisi. ?lokasyon= ile tesise göre filtrelenir.
    Dashboard'da PIN tanımlama panelinde kullanılır (aktif lokasyona göre).
    """
    lokasyon = request.args.get('lokasyon')
    conn = get_db()
    sql = "SELECT id, ad, bolum, pin, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE 1=1"
    params = []
    if lokasyon:
        sql += " AND COALESCE(lokasyon,'TK2') = ?"
        params.append(lokasyon)
    sql += " ORDER BY COALESCE(lokasyon,'TK2'), bolum, ad"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# Türkçe büyütme: Python'un upper()'ı 'i' harfini 'I' yapar ('YASIN'), oysa
# listedeki isimler 'YASİN'. İki özel harf elle çevrilir, gerisini upper() halleder.
def _ad_buyut(ad):
    return str(ad or '').replace('i', 'İ').replace('ı', 'I').upper()


def _ad_esitle(ad):
    """Ad karşılaştırma anahtarı — Türkçe harfler sadeleşir (import_excel._ad_normal'in eşi).
    'Ramazan Yasin Demir' ile 'RAMAZAN YASİN DEMİR' aynı kişidir; SQLite UNIQUE
    büyük/küçük harfe duyarlı olduğu için mükerrer satır bu kontrolle önlenir."""
    t = str(ad or '').strip().upper()
    for a, b in (('İ', 'I'), ('Ş', 'S'), ('Ğ', 'G'), ('Ü', 'U'), ('Ö', 'O'), ('Ç', 'C')):
        t = t.replace(a, b)
    return ' '.join(t.split())


@app.route('/api/operatorler', methods=['POST'])
@panel_gerekli(izin='operatorler')
def operator_ekle():
    """Yeni operatör ekler (dashboard Operatör panelinden).

    Body: {ad, bolum, pin?}  ·  lokasyon querystring'den gelir (dashboard fetch
    sarmalayıcısı aktif tesisi her /api/ isteğine ekler), yoksa body'den.

    NEDEN VAR (kullanıcı 2026-08-20): operatör listesi Excel'den içe aktarılıyordu
    ve sunucuda Excel yok — tek kişi eklemek için dosyayı laptopa çekip düzenleyip
    geri kopyalamak gerekiyordu. Bu uç nokta o zinciri kısaltır.

    EXCEL'İ BOZMAZ: içe aktarma INSERT OR IGNORE ile çalışır ve hiçbir operatörü
    SİLMEZ — buradan eklenen kişi bir sonraki Excel aktarımında yerinde kalır.
    Excel yine de listenin ana kaynağıdır; kalıcı olması için kişi oraya da
    yazılmalıdır (yoksa Excel ile veritabanı zamanla ayrışır)."""
    data = request.get_json() or {}
    ad = ' '.join(str(data.get('ad') or '').split())
    if len(ad) < 3:
        return jsonify({'hata': 'Ad soyad en az 3 karakter olmalı'}), 400
    ad = _ad_buyut(ad)
    bolum = (data.get('bolum') or '').strip()
    if bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f'Geçersiz bölüm: {bolum or "(boş)"}'}), 400
    lokasyon = (request.args.get('lokasyon') or data.get('lokasyon') or 'TK2').strip().upper()
    if lokasyon not in ('TK1', 'TK2'):
        return jsonify({'hata': f'Geçersiz lokasyon: {lokasyon}'}), 400
    pin = (data.get('pin') or '0000').strip()
    if len(pin) != 4 or not pin.isdigit():
        return jsonify({'hata': "PIN 4 haneli rakam olmalı (boş bırakılırsa 0000)"}), 400

    conn = get_db()
    # Mükerrer kontrolü Türkçe-duyarsız: UNIQUE(ad,bolum,lokasyon) kısıtı harf
    # büyüklüğüne duyarlı olduğundan 'Ali Şen' ile 'ALI SEN' iki satır olurdu.
    hedef = _ad_esitle(ad)
    for r in conn.execute(
            "SELECT ad, bolum FROM operatorler WHERE COALESCE(lokasyon,'TK2')=?",
            (lokasyon,)).fetchall():
        if _ad_esitle(r['ad']) == hedef and (r['bolum'] or '') == bolum:
            return jsonify({'hata': f"{r['ad']} zaten {lokasyon} · {bolum} listesinde kayıtlı"}), 409
    # Aynı kişi başka bölümde varsa PIN'i ORADAN devral — "bir kişi = tek PIN"
    # kuralı (operator_pin_guncelle ile aynı model); yeni satır 0000'a düşmesin.
    for r in conn.execute(
            "SELECT ad, COALESCE(pin,'0000') AS pin FROM operatorler "
            "WHERE COALESCE(lokasyon,'TK2')=?", (lokasyon,)).fetchall():
        if _ad_esitle(r['ad']) == hedef:
            ad = r['ad']                    # mevcut yazımı koru (liste tek isim göstersin)
            if pin == '0000':
                pin = r['pin']
            break
    import sqlite3   # modül düzeyinde import YOK (dosya kalıbı: yerel import)
    try:
        cur = conn.execute(
            "INSERT INTO operatorler (ad, bolum, pin, lokasyon) VALUES (?, ?, ?, ?)",
            (ad, bolum, pin, lokasyon))
        conn.commit()
    except sqlite3.IntegrityError as e:
        return jsonify({'hata': f'Eklenemedi (kayıt çakışması): {e}'}), 409

    # TK1 TUZAĞI: TK1 operatörünün bölümü Excel aktarımında İSİMDEN türetiliyor
    # (import_excel.tk1_operator_bolumu — montaj listesi SABİT, kalan herkes tel).
    # Listede olmayan bir isim 'montaj' olarak eklenirse bir sonraki aktarım onu
    # sessizce tel'e taşır. Engellemiyoruz ama söylüyoruz.
    uyari = ''
    if lokasyon == 'TK1' and bolum == 'montaj':
        try:
            from import_excel import tk1_operator_bolumu
            if tk1_operator_bolumu(ad) != 'montaj':
                uyari = (f'{ad} TK1 montaj listesinde tanımlı değil — bir sonraki '
                         f'Excel içe aktarımı bu kişiyi TEL bölümüne taşır. Kalıcı '
                         f'montaj için import_excel.TK1_MONTAJ_OPERATORLERI güncellenmeli.')
        except Exception:
            pass
    return jsonify({'basarili': True, 'id': cur.lastrowid, 'ad': ad,
                    'bolum': bolum, 'lokasyon': lokasyon, 'pin': pin,
                    'uyari': uyari}), 201


@app.route('/api/operator/oturum_ac', methods=['POST'])
def operator_oturum_ac():
    """Operatör mobile'da PIN ile oturum açma.
    Body: {operator_adi, pin}
    Return 200 ok / 403 hata
    """
    data = request.get_json() or {}
    op = (data.get('operator_adi') or '').strip()
    pin = (data.get('pin') or '').strip()
    lokasyon = (data.get('lokasyon') or request.args.get('lokasyon') or '').strip()
    if not op or not pin:
        return jsonify({'hata': 'operator_adi ve pin zorunlu'}), 400
    conn = get_db()
    # lokasyon verildiyse o lokasyondaki operatörü seç (ad artık lokasyonlar arası tek değil);
    # verilmezse ilk eşleşen (geriye dönük uyum).
    # ÇOKLU BÖLÜM: UNIQUE(ad, bolum, lokasyon) sonrası aynı kişinin bölüm başına satırı
    # olabilir — herhangi bir satırın PIN'i tutarsa giriş başarılı (satırlar tek PIN paylaşır,
    # import + PIN güncelleme bunu korur).
    if lokasyon:
        rows = conn.execute("SELECT pin, bolum, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE ad=? AND COALESCE(lokasyon,'TK2')=?", (op, lokasyon)).fetchall()
    else:
        rows = conn.execute("SELECT pin, bolum, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE ad=?", (op,)).fetchall()
    conn.close()
    if not rows:
        return jsonify({'hata': f'Operatör bulunamadı: {op}'}), 403
    eslesen = next((r for r in rows if (r['pin'] or '0000') == pin), None)
    if eslesen is None:
        return jsonify({'hata': 'PIN yanlış'}), 403
    return jsonify({
        'basarili':     True,
        'operator_adi': op,
        'bolum':        eslesen['bolum'] or '',
        'lokasyon':     eslesen['lokasyon'] or 'TK2',
    }), 200


@app.route('/api/operator/<int:oid>/pin', methods=['PATCH'])
@panel_gerekli(izin='operatorler')
def operator_pin_guncelle(oid):
    """Yönetici operatörün PIN'ini günceller. Body: {pin}
    PIN 4 haneli rakam olmalıdır.
    """
    data = request.get_json() or {}
    yeni_pin = (data.get('pin') or '').strip()
    if not yeni_pin or len(yeni_pin) != 4 or not yeni_pin.isdigit():
        return jsonify({'hata': "PIN 4 haneli rakam olmalı (örn: '1234')"}), 400
    conn = get_db()
    row = conn.execute("SELECT ad, COALESCE(lokasyon,'TK2') as lokasyon FROM operatorler WHERE id=?", (oid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'hata': 'Operatör bulunamadı'}), 404
    # Bir kişi = tek PIN: aynı kişinin tüm bölüm satırlarına yaz (aynı tesiste).
    # (Çoklu bölüm modeli: UNIQUE(ad, bolum, lokasyon) — kişi başına bölüm satırları var.)
    conn.execute(
        "UPDATE operatorler SET pin=? WHERE ad=? AND COALESCE(lokasyon,'TK2')=?",
        (yeni_pin, row['ad'], row['lokasyon'])
    )
    conn.commit()
    conn.close()
    return jsonify({'basarili': True, 'operator_adi': row['ad']}), 200


@app.route('/api/import_excel', methods=['POST'])
def excel_import():
    """Excel dosyalarından verileri çeker.
    ?bolum=kaynak|montaj|metal → sadece o bölümün dosyası;
    parametre yoksa: hepsi.
    """
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    # TK1: ayrı Excel + import_tk1 (mirror-sync YOK, TK2 verisini silmez)
    if lokasyon == 'TK1':
        try:
            return jsonify(import_tk1()), 200
        except Exception as e:
            return jsonify({'hata': str(e), 'basarili': False}), 500
    if bolum and bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f"Geçersiz bölüm: {bolum}", 'basarili': False}), 400
    try:
        sonuc = import_data(bolum=bolum)
        # Kaynak bölümü aktarımı ek sayfaları da kapsar: Robot Program Listesi + Fikstür
        # Raf (eski 'Toplu Veri Yönetimi' butonundan taşındı — tek buton, tek akış).
        if sonuc.get('basarili') and (bolum == 'kaynak' or not bolum):
            from import_excel import kaynak_ek_import
            sonuc.update(kaynak_ek_import())
        return jsonify(sonuc), 200
    except Exception as e:
        return jsonify({'hata': str(e), 'basarili': False}), 500


@app.route('/api/veri/import_tum', methods=['POST'])
def veri_import_tum():
    """Tüm Excel sayfalarını okuyup DB'yi günceller (tek tıkla import).
    Kapsam: referans + operatör (3 bölüm), robot program, fikstür raf.
    Duruş sebepleri her API isteğinde Excel'den okunur, import gerekmez.
    """
    # TK1: ayrı Excel + import_tk1 (TK2 dosyasını okumaz, TK2 verisini silmez)
    lokasyon = request.args.get('lokasyon')
    try:
        if lokasyon == 'TK1':
            return jsonify(import_tk1()), 200
        sonuc = import_tum()
        return jsonify(sonuc), 200
    except Exception as e:
        return jsonify({'hata': str(e), 'basarili': False}), 500


@app.route('/api/referans/teyit', methods=['PATCH'])
@panel_gerekli(izin='referanslar')
def referans_teyit_guncelle():
    """Bir referansın süre teyidi bayrağını aç/kapat.
    Body: {referans_kodu, sure_teyit: 0|1, lokasyon?} — lokasyon ile o fabrikanın
    satırı hedeflenir (diğer fabrikanın aynı kodlu referansının teyidi değişmez)."""
    data = request.get_json() or {}
    kod = (data.get('referans_kodu') or '').strip()
    teyit = 1 if data.get('sure_teyit') else 0
    lokasyon = (data.get('lokasyon') or request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    # Süre teyidi kaynak bölümüne özgü bir özellik — aynı kod başka bölümde de (lazer/pres)
    # yaşayabildiğinden güncelleme bölümle kapatılır (body'de bolum gelirse o, yoksa kaynak).
    bolum = (data.get('bolum') or 'kaynak').strip()
    if bolum not in GECERLI_BOLUMLER:
        bolum = 'kaynak'
    if not kod:
        return jsonify({'hata': 'referans_kodu zorunlu'}), 400
    conn = get_db()
    try:
        if teyit == 1:
            conn.execute(
                "UPDATE referans_listesi SET sure_teyit=1, sure_teyit_tarihi=datetime('now','localtime') "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=?",
                (kod, bolum, lokasyon)
            )
        else:
            conn.execute(
                "UPDATE referans_listesi SET sure_teyit=0, sure_teyit_tarihi=NULL "
                "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=?",
                (kod, bolum, lokasyon)
            )
        conn.commit()
        ref_row = {'bolum': bolum}
    finally:
        conn.close()
    # Excel auto-sync: teyit bilgisi artık Excel'de 'Süre Teyit' (E) kolonunda tutulur.
    # (Eskiden teyit yalnız DB'ye yazılıyor, kullanıcı Excel'de değişiklik göremiyordu.)
    excel_uyari = None
    if lokasyon != 'TK1' and ref_row:
        try:
            sonuc = export_referans_cycle_times(bolum=ref_row['bolum'], lokasyon=lokasyon)
            if not sonuc.get('basarili'):
                excel_uyari = sonuc.get('hata')
        except Exception as e:
            excel_uyari = str(e)
        if excel_uyari:
            print(f'[referans/teyit] Excel sync hatası: {excel_uyari}')
    return jsonify({'basarili': True, 'referans_kodu': kod, 'sure_teyit': teyit,
                    'excel_uyari': excel_uyari}), 200


@app.route('/api/saha_cihazlari', methods=['GET'])
def saha_cihazlari():
    """Beklenen tüm saha cihazlarının durumu (kaynak/montaj/metal).

    Beklenen cihaz listesi sabit:
      - 9 robot:    ABB1-IO ... ABB9-IO       (bolum: kaynak, TK2)
      - 12 montaj:  MONTAJ-M1 ... MONTAJ-M12  (bolum: montaj, TK2)
                    + MONTAJ-TK1-M1..M3 (TK1 montaj butonları)
                    (TK1-M4 ve YF1 butonları 2026-08-20'de 'tel' grubuna geçti)
      - 3 metal:    300T-IO, 400T-IO, 550T-IO (bolum: metal, TK2)
      - 10 pres:    ABKANT-A1/A2/A3, PRES-P1..P5,
                    PRES-HIDROLIK, PRES-BROS   (bolum: pres, TK2)
      - 3 plastik:  PLASTIK-320T/407T/YAPISTIRMA (bolum: plastik, TK1)

    Her cihaz için pilot.db'deki cihaz_kayitlari ile eşleştirilir.
    Eşleşme yoksa 'beklemede' (hiç bağlanmamış). ?lokasyon= ile tesise göre süzülür.
    """
    import os
    pilot_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pilot', 'pilot.db')

    # ?lokasyon= verilirse sadece o tesisin cihazları döner (saha cihazları + sinyal
    # analizi TK1/TK2'ye özel). Entry'de lokasyon yoksa 'TK2' varsayılır.
    lokasyon = request.args.get('lokasyon')
    BEKLENEN = {
        'kaynak': [{'cihaz_id': f'ABB{i}-IO', 'robot_no': f'ABB{i}'} for i in range(1, 10)],
        'montaj': [{'cihaz_id': f'MONTAJ-M{i}', 'robot_no': f'M{i}'} for i in range(1, 13)]
                  # TK1 montaj masaları (2026-07-31): cihaz_id/robot_no firmware ile
                  # birebir (generate.py DEVICES['montaj']).
                  # 2026-08-20: TK1 MONTAJ hattında 3 BUTON kaldı (MONTAJ - 1..3 =
                  # TK1-M1..M3). 4 ve 5 numaralı modüller tel SON MONTAJ hattına
                  # taşındı → aşağıda 'tel' grubunda izlenirler (cihaz, çalıştığı
                  # bölümün panelinde görünmeli).
                  # TK1-M5/M6/M7 buton DEĞİL → beklenenler listesinde YOK: sahada
                  # olmayan cihaz burada dururken panelde kalıcı "beklemede" görünür
                  # ve GERÇEK arızayı gizler.
                  + [{'cihaz_id': f'MONTAJ-TK1-M{i}', 'robot_no': f'TK1-M{i}',
                      'lokasyon': 'TK1'} for i in range(1, 4)],
        'metal':  [
            {'cihaz_id': '300T-IO', 'robot_no': '300T'},
            {'cihaz_id': '400T-IO', 'robot_no': '400T'},
            {'cihaz_id': '550T-IO', 'robot_no': '550T'},
        ],
        # 2026-07-27 saha rollout — abkant (pres bölümü, TK2) + plastik enjeksiyon &
        # yapıştırma (plastik bölümü, TK1). cihaz_id + robot_no firmware ile birebir
        # (generate.py DEVICES). 2026-07-31: pres bölümünün kalan makineleri de eklendi
        # (5 eksantrik + hidrolik + broş) — hepsi makinenin kendi sayacındaki röleyi izler.
        'pres':   [
            {'cihaz_id': 'ABKANT-A1', 'robot_no': 'Abkant 1'},
            {'cihaz_id': 'ABKANT-A2', 'robot_no': 'Abkant 2'},
            {'cihaz_id': 'ABKANT-A3', 'robot_no': 'Abkant 3'},
            {'cihaz_id': 'PRES-P1', 'robot_no': 'Pres 1'},
            {'cihaz_id': 'PRES-P2', 'robot_no': 'Pres 2'},
            {'cihaz_id': 'PRES-P3', 'robot_no': 'Pres 3'},
            {'cihaz_id': 'PRES-P4', 'robot_no': 'Pres 4'},
            {'cihaz_id': 'PRES-P5', 'robot_no': 'Pres 5'},
            {'cihaz_id': 'PRES-HIDROLIK', 'robot_no': 'Hidrolik Pres'},
            {'cihaz_id': 'PRES-BROS',     'robot_no': 'Bros'},
        ],
        'plastik': [
            {'cihaz_id': 'PLASTIK-320T',       'robot_no': '320T',       'lokasyon': 'TK1'},
            {'cihaz_id': 'PLASTIK-407T',       'robot_no': '407T',       'lokasyon': 'TK1'},
            {'cihaz_id': 'PLASTIK-YAPISTIRMA', 'robot_no': 'Yapistirma', 'lokasyon': 'TK1'},
        ],
        # 2026-08-07 TK1 kapama presleri — İLK ÇOK KANALLI KURULUM.
        # 12 pres üçerli ortak panolarda: 1 ESP32 = 3 makine (GPIO25/26/27).
        # Bu yüzden burada ÜÇ satır AYNI cihaz_id'yi paylaşır; onları ayıran
        # 'kanal' (= firmware'in gönderdiği istasyon 1..3). Sayım/reset kanala
        # göre süzülür, cihaz sağlığı (online/RSSI/firmware) üçünde de aynıdır —
        # 'modul' etiketi panoda hangi modülün baktığını gösterir.
        # cihaz_id + robot_no firmware ile BİREBİR (generate.py DEVICES['tel_kapama']).
        'tel': [
            # 2026-08-19 otomatik hazırlık makinesi — TEK kanallı (kapama modülleri
            # üçerli, bu değil). Modül montajı kullanıcıda; takılana kadar panelde
            # "beklemede" görünür.
            {'cihaz_id': 'TEL-HAZIRLIK', 'robot_no': 'Otomatik Hazirlik',
             'lokasyon': 'TK1'},
            # 2026-08-20 SON MONTAJ butonları — montaj hattından taşınan iki modül
            # (buton 4 = TK1-M4, buton 5 = YF1). cihaz_id/robot_no firmware'de
            # MONTAJ-* olarak KALDI (yeniden flash YOK) ve pilot.db'ye bolum='montaj'
            # yazarlar — çeviri _CIHAZ_FW_BOLUM'da. Panelde 'tel' altındalar çünkü
            # sayaç sıfırlama referansı (aktif vardiya / operatör baseline) TEL
            # vardiyasından gelmeli; montaj grubunda kalsalardı tel vardiyası açıkken
            # bile 0 sayarlardı.
            {'cihaz_id': 'MONTAJ-TK1-M4', 'robot_no': 'TK1-M4', 'lokasyon': 'TK1'},
            {'cihaz_id': 'MONTAJ-YF1',    'robot_no': 'YF1',    'lokasyon': 'TK1'},
        ] + [
            {'cihaz_id': f'TEL-KAPAMA-{m}',
             'robot_no': f'Kapama {(m - 1) * 3 + kn}',
             'lokasyon': 'TK1',
             'kanal':    kn,
             'modul':    f'Kapama {(m - 1) * 3 + 1}-{(m - 1) * 3 + 3}'}
            for m in range(1, 5) for kn in (1, 2, 3)
        ],
    }

    # Pilot DB'den mevcut cihaz_kayitlari + bugünün ist1/ist2 sayıları
    mevcut = {}
    reset_map = {}       # {(cihaz_id, istasyon): reset_ts}
    if os.path.exists(pilot_db):
        import sqlite3
        c = sqlite3.connect(pilot_db)
        c.row_factory = sqlite3.Row
        try:
            rows = c.execute('''
                SELECT *,
                       CAST((julianday('now','localtime') - julianday(son_heartbeat)) * 1440 AS INTEGER) as son_heartbeat_dk
                FROM cihaz_kayitlari
            ''').fetchall()
            for r in rows:
                d = dict(r)
                d['durum'] = 'offline' if (d.get('son_heartbeat_dk') or 0) > 2 else 'online'
                mevcut[d['cihaz_id']] = d

            # Manuel reset noktalari — dashboard'dan "sayaci sifirla" tıklayınca yazıldı
            # Sayım sırasında filtre olarak kullanılır (ts > reset_ts olanları say)
            try:
                for r in c.execute('SELECT cihaz_id, istasyon, reset_ts FROM sayac_reset_noktalari').fetchall():
                    reset_map[(r['cihaz_id'], int(r['istasyon'] or 0))] = r['reset_ts']
            except Exception:
                pass  # Tablo henüz yoksa (pilot_app.py yeniden başlatılmamışsa) — atla
        except Exception as e:
            print(f'[saha_cihazlari] DB hata: {e}')
        finally:
            c.close()

    # ─── Aktif referans bilgisi (durum='uretimde') ──────────────────
    # Her robot+istasyon icin: hangi referans uretimde + ne zamandir + simdiye
    # kadar kac pulse var. Operator yeni referans secince uretime_baslama_ts
    # yenilenir, sayim sifirdan baslamis gibi gorunur.
    aktif_ref_map = {}  # {(robot_no, istasyon): {referans_kodu, basla_ts}}
    # Aktif vardiyalar — sayac sifirlama icin ek baslangic referansi
    aktif_vardiya_map = {}  # {(bolum, robot_no): vardiya_basla_ts ISO}
    # Operator uretim baseline — yeni referans kaydi acinca sayac sifirlanir
    op_baseline_map = {}  # {(bolum, robot_no, istasyon): sayac_baslangic_ts}
    try:
        conn_main = get_db()
        for r in conn_main.execute('''
            SELECT robot_no, istasyon, referans_kodu,
                   COALESCE(uretime_baslama_ts, olusturma_tarihi) as basla_ts
            FROM referans_takip
            WHERE durum='uretimde' AND robot_no <> ''
        ''').fetchall():
            key = (r['robot_no'], int(r['istasyon'] or 0))
            # Eger ayni robot+istasyon icin birden fazla varsa en sonuncuyu al
            mevcut_ref = aktif_ref_map.get(key)
            if (not mevcut_ref) or ((r['basla_ts'] or '') > (mevcut_ref['basla_ts'] or '')):
                aktif_ref_map[key] = {'referans_kodu': r['referans_kodu'], 'basla_ts': r['basla_ts']}

        # Bugun acik vardiyalari cek — operator vardiya acince sayim o andan
        # itibaren olur (yeni referans secilmemis bile olsa sifirdan baslar)
        bugun_str = datetime.now().strftime('%Y-%m-%d')
        for r in conn_main.execute('''
            SELECT robot_no, COALESCE(bolum,'kaynak') as bolum, baslangic_saati, tarih
            FROM vardiyalar
            WHERE durum != 'kapali' AND tarih=?
        ''', (bugun_str,)).fetchall():
            bs = (r['baslangic_saati'] or '').strip()
            if bs and r['robot_no']:
                # baslangic_saati 'HH:MM' formatinda — tarih ile birlestir ISO yap
                try:
                    ts_iso = f"{r['tarih']} {bs}:00" if len(bs) == 5 else f"{r['tarih']} {bs}"
                    # Ayni (bolum, robot_no) icin birden fazla aktif vardiya varsa
                    # en sonra acilmis olani al (en buyuk ts)
                    key = (r['bolum'], r['robot_no'])
                    if (key not in aktif_vardiya_map) or (ts_iso > aktif_vardiya_map[key]):
                        aktif_vardiya_map[key] = ts_iso
                except Exception:
                    pass

        # Operatör üretim baseline — operatör yeni referans kaydı açınca sayac_baslangic_ts
        # set edilir; andon o andan sonrasını sayar (operatör aksiyonuyla sayaç sıfırlanır).
        # (bolum, robot_no, istasyon) -> en son baseline ts
        for r in conn_main.execute('''
            SELECT COALESCE(v.bolum,'kaynak') as bolum, v.robot_no, u.istasyon,
                   MAX(u.sayac_baslangic_ts) as bts
            FROM uretim_kayitlari u
            JOIN vardiyalar v ON u.vardiya_id = v.id
            WHERE v.durum != 'kapali' AND v.tarih=? AND u.sayac_baslangic_ts IS NOT NULL
            GROUP BY v.bolum, v.robot_no, u.istasyon
        ''', (bugun_str,)).fetchall():
            if r['robot_no'] and r['bts']:
                op_baseline_map[(r['bolum'], r['robot_no'], int(r['istasyon'] or 0))] = r['bts']
        # NOT: conn_main = get_db() g-paylaşımlı — istek ortasında KAPATMA (teardown kapatır).
        # Erken close, sonrasına get_db kullanan kod eklenirse 'closed database' 500 üretir.
    except Exception as e:
        print(f'[saha_cihazlari] referans_takip/vardiyalar hata: {e}')

    # ─── Pulse sayımı (pilot DB'den, baslangic_ts sonrası) ───
    def _aktif_pulse_say(cihaz_id, istasyon, basla_ts):
        """Verilen cihaz/istasyon icin basla_ts'den sonraki pulse sayisi."""
        if not basla_ts or not os.path.exists(pilot_db):
            return 0
        try:
            import sqlite3
            pc = sqlite3.connect(pilot_db)
            try:
                q = '''SELECT COUNT(*) FROM sayac_olaylari
                       WHERE cihaz_id=? AND ts >= ?'''
                params = [cihaz_id, basla_ts]
                if istasyon and istasyon > 0:
                    q += ' AND istasyon=?'
                    params.append(istasyon)
                row = pc.execute(q, params).fetchone()
                return row[0] if row else 0
            finally:
                pc.close()
        except Exception:
            return 0

    def _en_son_ts(*candidates):
        """Verilen ISO timestamp string'lerinden en buyugunu (en son) dondur.
        None/bos olanlar atlanir. Hicbiri yoksa None."""
        gecerli = [t for t in candidates if t]
        return max(gecerli) if gecerli else None

    sonuc = {}
    for bolum, cihazlar in BEKLENEN.items():
        liste = []
        for beklenen_cihaz in cihazlar:
            cid = beklenen_cihaz['cihaz_id']
            rno = beklenen_cihaz['robot_no']
            dev_lok = beklenen_cihaz.get('lokasyon', 'TK2')
            # ÇOK KANALLI MODÜL (kapama presleri): 3 makine 1 cihaz_id paylaşır.
            # kanal = firmware'in gönderdiği istasyon (1..3); sayım/reset buna göre
            # süzülür. Tek makineli cihazlarda 0 → filtre yok, eski davranış aynen.
            kanal = int(beklenen_cihaz.get('kanal') or 0)
            if lokasyon and dev_lok != lokasyon:
                continue   # bu tesisin cihazı değil — atla
            kayit = mevcut.get(cid)

            # PRES (2026-07-28) / PLASTIK (2026-08-18): makine artik VARDIYADA degil
            # URETIM KAYDINDA secilir. Cihaz 'Abkant 2' → vardiya hatti sabit hat,
            # baseline istasyonu = 2. ESKI kayitlar (vardiya.robot_no='Abkant 2' /
            # '320T', istasyon=0) icin fallback var.
            # Cihazın makine adı hangi sabit hattın listesinde? (LF-LFP→YF1 gibi
            # takma adlı cihazlar için ALIAS'a da bakılır: cihaz robot_no'su YF1 ama
            # operatörün seçtiği hat 'LF-LFP'dir.)
            v_rno, mak_ist = rno, 0
            _sabit_hat, _hat_ad = '', ''
            # ÖNCE alias: cihaz 'Kapama 4' hattı 'Kapama 30'dur; ham adla arasak
            # listedeki 'Kapama 4' HATTINI (cihaz 'Kapama 3') yanlışlıkla bulurduk.
            # EMEKLİ adlar ATLANIR (2026-08-20): TK1-M4/YF1 modülleri hem eski
            # 'MONTAJ - 4/5' hem yeni 'Son Montaj 4/5' adıyla eşleşir; gizli olanı
            # elemezsek panel modülü hâlâ montaj masasındaymış gibi etiketler ve
            # sayaç sıfırlama referansını YANLIŞ vardiyadan alır.
            _aliaslar = [k for k, v in HAT_SAYAC_CIHAZI.items()
                         if v == rno and k not in GIZLI_HATLAR] or [rno]
            for _sh in HAT_MAKINELERI:
                for _al in _aliaslar:
                    _pi = kayit_ist_no(_sh, _al)
                    if _pi:
                        v_rno, mak_ist, _sabit_hat, _hat_ad = _sh, _pi, _sh, _al
                        break
                if _sabit_hat:
                    break

            # Aktif vardiya baslangic ts (varsa) — sayac sifirlama referansi
            # (yeni model once; bulunamazsa eski makine-adli vardiyaya dus)
            vardiya_ts = aktif_vardiya_map.get((bolum, v_rno)) or aktif_vardiya_map.get((bolum, rno))
            # Manuel reset ts'leri (varsa) — dashboard'dan butonla sifirlanmis
            reset_ist1 = reset_map.get((cid, 1))
            reset_ist2 = reset_map.get((cid, 2))
            reset_all  = reset_map.get((cid, 0))  # tum istasyonlari kapsar
            # Cok kanalli modulde SADECE bu presin reset'i (kanal=3 ise ist1/ist2 degil)
            reset_kanal = reset_map.get((cid, kanal)) if kanal else None

            # Aktif referans tespiti — kaynakta her istasyon ayri olabilir,
            # montaj/metal'de tek istasyon (genelde 0 veya 1).
            # Sayim icin filtre = max(vardiya_ts, referans_ts, manuel_reset_ts):
            # hangisi daha sonraysa o andan itibaren sayim baslar.
            aktif_referanslar = []
            if bolum == 'kaynak':
                for ist in (1, 2):
                    a = aktif_ref_map.get((rno, ist))
                    if a:
                        reset_ist = reset_ist1 if ist == 1 else reset_ist2
                        op_bl = op_baseline_map.get((bolum, rno, ist))
                        filt_ts = _en_son_ts(vardiya_ts, a['basla_ts'], reset_ist, reset_all, op_bl)
                        aktif_referanslar.append({
                            'istasyon':      ist,
                            'referans_kodu': a['referans_kodu'],
                            'basla_ts':      filt_ts,
                            'pulse_sayisi':  _aktif_pulse_say(cid, ist, filt_ts),
                        })
            else:
                # Montaj/metal: istasyon belirtilmemis (0) veya 1 — her ikisini de yakala
                a = None
                for ist in (0, 1):
                    a = aktif_ref_map.get((rno, ist))
                    if a:
                        break
                # Montaj fallback: operator hat secmez, robot_no='MONTAJ' default (eski kayitlar)
                if not a and bolum == 'montaj':
                    for fallback_rno in ('MONTAJ', ''):
                        for ist in (0, 1):
                            a = aktif_ref_map.get((fallback_rno, ist))
                            if a:
                                break
                        if a:
                            break
                if a:
                    op_bl = op_baseline_map.get((bolum, rno, 0)) or op_baseline_map.get((bolum, rno, 1))
                    # Cok kanalli modulde ist1 reset'i DIGER presi ilgilendirmez —
                    # kanal varsa yalniz kendi reset'i (reset_kanal) hesaba katilir.
                    _r_ist = reset_kanal if kanal else reset_ist1
                    filt_ts = _en_son_ts(vardiya_ts, a['basla_ts'], reset_all, _r_ist, op_bl)
                    aktif_referanslar.append({
                        'istasyon':      kanal or None,
                        'referans_kodu': a['referans_kodu'],
                        'basla_ts':      filt_ts,
                        # kanal>0: SADECE bu presin pulse'lari (modulun ucu birden degil)
                        'pulse_sayisi':  _aktif_pulse_say(cid, kanal, filt_ts),
                    })

            # Sayac degerlerini hesapla — uretim kapisi = AKTIF VARDIYA (bkz. vardiya_acik):
            # - Aktif vardiya varsa: vardiya basindan (veya daha gec reset/baseline'dan) say
            # - Aktif vardiya yoksa: 0 (vardiya acilmadan gelen sinyal = hat paraziti, uretim degil)
            if kayit:
                # Istasyon basina baslangic referanslari (operator uretim baseline'i)
                op_bl_1 = op_baseline_map.get((bolum, rno, 1))
                op_bl_2 = op_baseline_map.get((bolum, rno, 2))
                op_bl_0 = op_baseline_map.get((bolum, rno, 0))
                # PRES/PLASTIK: baseline anahtarinda istasyon = MAKINE no (sabit hat + 1..N).
                # Cihaz tek makinedir → tek sayac (op_bl_0 gibi davranir); eski
                # makine-adli kayitlarin baseline'i (rno, 0) fallback olarak kalir.
                if mak_ist:
                    op_bl_0 = _en_son_ts(op_baseline_map.get((bolum, _sabit_hat, mak_ist)),
                                         op_bl_0)
                    op_bl_1 = op_bl_2 = None
                # ── URETIM KAPISI: AKTIF VARDIYA ──────────────────────────────
                # Sayim ancak ACIK bir vardiya (veya o vardiyada acilmis operator
                # uretim baseline'i) varsa yapilir. Aksi halde uretim baglami yoktur:
                # vardiya acilmadan gelen sinyaller (gece/hafta sonu hat paraziti)
                # uretim olarak SAYILMAZ.
                # NOT: gunler oncesinden kalan BAYAT manuel reset noktasi TEK BASINA
                # sayim baslatmaz; reset ancak acik vardiya icinde "su andan say" demektir.
                vardiya_acik = bool(vardiya_ts or op_bl_0 or op_bl_1 or op_bl_2)
                if not vardiya_acik:
                    ist1_v = 0
                    ist2_v = 0
                    toplam_v = 0
                elif kanal:
                    # COK KANALLI MODUL: bu satir modulun TEK bir presi. Sayim yalniz
                    # kendi kanalindan; modulun diger iki presi bu karta KARISMAZ
                    # (kanal filtresi olmasaydi uc presin toplami gorunurdu).
                    kanal_filt = _en_son_ts(vardiya_ts, reset_kanal, reset_all, op_bl_0)
                    ist1_v = _aktif_pulse_say(cid, kanal, kanal_filt) if kanal_filt else 0
                    ist2_v = 0
                    toplam_v = ist1_v
                else:
                    # En gec baslangic noktasindan say: vardiya / manuel reset / baseline
                    ist1_filt = _en_son_ts(vardiya_ts, reset_ist1, reset_all, op_bl_1)
                    ist2_filt = _en_son_ts(vardiya_ts, reset_ist2, reset_all, op_bl_2)
                    ist1_v = _aktif_pulse_say(cid, 1, ist1_filt) if ist1_filt else 0
                    ist2_v = _aktif_pulse_say(cid, 2, ist2_filt) if ist2_filt else 0
                    # toplam = cihazin TUM pulse'lari (istasyon farketmez).
                    # _aktif_pulse_say(cid, 0, ts) -> istasyon=0 falsy oldugu icin filtre
                    # UYGULAMAZ, tum pulse'lari sayar. ESKI BUG: ist1_v + ist2_v + toplam_0
                    # idi; montaj istasyon=1 gonderdigi icin o pulse hem ist1_v'de hem
                    # toplam_0'da sayiliyordu = 2x. Duzeltme: dogrudan toplam say.
                    toplam_ts = _en_son_ts(vardiya_ts, reset_all, op_bl_0)
                    toplam_v = _aktif_pulse_say(cid, 0, toplam_ts) if toplam_ts else (ist1_v + ist2_v)

                # ── "Makine calisiyor" durumu ──────────────────────────────────
                # Metal: sinyal akisindan TURETILIR. Her sinyal = makine-calisiyor rolesi
                # darbesi (400T o roleyi izler). Son sinyal ESIK sn icindeyse calisiyor;
                # sinyal kesilince durmus. Metal firmware robot_calisiyor=0 gonderir
                # (kullanmaz), o yuzden bu alani metal'de biz uretiyoruz. Vardiya sart:
                # vardiya yokken gelen sinyal hat parazitidir, "calisiyor" demek degildir.
                # Kaynak/montaj: cihazin gercek robot_calisiyor IO sinyali aynen gecer.
                if bolum == 'metal':
                    robot_calisiyor_deger = 0
                    son_sin = kayit.get('son_sinyal')
                    if vardiya_acik and son_sin:
                        try:
                            gecen_sn = (datetime.now() - datetime.strptime(son_sin, '%Y-%m-%d %H:%M:%S')).total_seconds()
                            if 0 <= gecen_sn < METAL_CALISIYOR_ESIK_SN:
                                robot_calisiyor_deger = 1
                        except (ValueError, TypeError):
                            pass
                else:
                    robot_calisiyor_deger = 1 if kayit.get('robot_calisiyor') else 0

                liste.append({
                    'cihaz_id':           cid,
                    'robot_no':           rno,
                    # OPERATÖRÜN GÖRDÜĞÜ ad (2026-08-18): kapama preslerinde firmware
                    # 'Kapama 1' der, sahadaki presin kodu 5'tir → panelde saha kodu
                    # gösterilmezse arıza takibi yanlış prese bakar.
                    'hat_adi':            _hat_ad or rno,
                    'bolum':              bolum,
                    'lokasyon':           dev_lok,
                    # Cok kanalli modulde kart, cihazi DIGER iki presle paylasir:
                    # online/RSSI/firmware ucunde AYNI, sayim ise kanala ozel.
                    'kanal':              kanal,
                    'modul':              beklenen_cihaz.get('modul', ''),
                    'durum':              kayit.get('durum', 'offline'),
                    'ip_adresi':          kayit.get('ip_adresi', ''),
                    'wifi_rssi':          kayit.get('wifi_rssi', 0),
                    'son_heartbeat':      kayit.get('son_heartbeat', ''),
                    'son_heartbeat_dk':   kayit.get('son_heartbeat_dk', 0),
                    'firmware_ver':       kayit.get('firmware_ver', ''),
                    'toplam_sinyal':      kayit.get('toplam_sinyal', 0),
                    'buffer_kuyruk':      kayit.get('buffer_kuyruk', 0),
                    'uptime_sn':          kayit.get('uptime_sn', 0),
                    'free_heap':          kayit.get('free_heap', 0),
                    'bugun_ist1':         ist1_v,
                    'bugun_ist2':         ist2_v,
                    'bugun_toplam':       toplam_v,
                    'robot_calisiyor':    robot_calisiyor_deger,
                    'aktif_vardiya_ts':   vardiya_ts,
                    'aktif_referanslar':  aktif_referanslar,
                    'kayitli':            True,
                })
            else:
                # Hiç bağlanmamış
                liste.append({
                    'cihaz_id':           cid,
                    'robot_no':           rno,
                    'hat_adi':            _hat_ad or rno,
                    'bolum':              bolum,
                    'lokasyon':           dev_lok,
                    'kanal':              kanal,
                    'modul':              beklenen_cihaz.get('modul', ''),
                    'durum':              'beklemede',
                    'aktif_vardiya_ts':   vardiya_ts,
                    'aktif_referanslar':  aktif_referanslar,
                    'kayitli':            False,
                })
        sonuc[bolum] = liste

    return jsonify(sonuc), 200


@app.route('/api/saha_cihazlari/sayac_reset', methods=['POST'])
@panel_gerekli(izin='saha-cihazlari')
def saha_cihazlari_sayac_reset():
    """Manuel sayaç sıfırlama — dashboard kartından çağrılır.

    Body:
      - cihaz_id: 'ABB2-IO', 'MONTAJ-M3', '400T-IO' vb. (zorunlu)
      - istasyon: 0 (tüm), 1..3 (tek kanal) — default 0
      - yapan:    opsiyonel, logging için

    İSTASYON 3 (2026-08-07): kapama presleri üçerli ortak panoda tek modüle bağlı
    (TEL-KAPAMA-N). Kanal başına sıfırlama olmasaydı bir presi sıfırlamak diğer
    ikisinin sayacını da silerdi — bu yüzden aralık 0/1/2'den 0..3'e genişletildi.

    Pilot.db.sayac_reset_noktalari'na INSERT OR REPLACE yapılır.
    /api/saha_cihazlari endpoint'i sayım sırasında bu ts'i de en sona dahil eder.
    """
    import os, sqlite3
    data = request.get_json() or {}
    cihaz_id = (data.get('cihaz_id') or '').strip()
    istasyon = int(data.get('istasyon') or 0)
    yapan    = (data.get('yapan') or '').strip()

    if not cihaz_id:
        return jsonify({'hata': 'cihaz_id zorunlu'}), 400
    if istasyon not in (0, 1, 2, 3):
        return jsonify({'hata': 'istasyon 0..3 olmalı'}), 400

    pilot_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pilot', 'pilot.db')
    if not os.path.exists(pilot_db):
        return jsonify({'hata': 'Pilot DB bulunamadı (pilot_app.py çalışıyor mu?)'}), 503

    try:
        pc = sqlite3.connect(pilot_db)
        try:
            # Tablo henüz yoksa (pilot_app.py yeniden başlatılmamışsa) oluştur
            pc.execute('''
                CREATE TABLE IF NOT EXISTS sayac_reset_noktalari (
                    cihaz_id   TEXT NOT NULL,
                    istasyon   INTEGER NOT NULL DEFAULT 0,
                    reset_ts   TEXT NOT NULL,
                    yapan      TEXT DEFAULT '',
                    PRIMARY KEY (cihaz_id, istasyon)
                )
            ''')
            cur = pc.execute('''
                INSERT OR REPLACE INTO sayac_reset_noktalari
                (cihaz_id, istasyon, reset_ts, yapan)
                VALUES (?, ?, datetime('now','localtime'), ?)
            ''', (cihaz_id, istasyon, yapan))
            pc.commit()
            # Yazılan ts'yi geri oku
            row = pc.execute(
                'SELECT reset_ts FROM sayac_reset_noktalari WHERE cihaz_id=? AND istasyon=?',
                (cihaz_id, istasyon)
            ).fetchone()
            reset_ts = row[0] if row else None
        finally:
            pc.close()
    except Exception as e:
        return jsonify({'hata': f'Reset yazılamadı: {str(e)}'}), 500

    ist_label = 'tüm istasyonlar' if istasyon == 0 else f'İstasyon {istasyon}'
    print(f"[sayac_reset] {cihaz_id} / {ist_label} → reset_ts={reset_ts} (yapan: {yapan or 'belirtilmemiş'})")
    return jsonify({
        'basarili': True,
        'cihaz_id': cihaz_id,
        'istasyon': istasyon,
        'reset_ts': reset_ts,
    }), 200


@app.route('/api/pilot/durus_tanimla', methods=['POST'])
@panel_gerekli(izin='andon-ayarlari')
def pilot_durus_tanimla():
    """Sinyal Analizi'nde 'Kaydedilmemiş' bir boşluğa sebep atar.
    Sinyalden gelen GERÇEK başlangıç saati + süre ile bir duruş kaydı oluşturur
    (baslangic_saati dolu → bir dahaki analizde tam eşleşir). Rapor SUM ile çalışır,
    bu yüzden saatli/ayrı kayıt Andon/OEE toplamlarını bozmaz."""
    data = request.get_json() or {}
    robot_no = (data.get('robot_no') or '').strip()
    tarih    = (data.get('tarih') or '').strip()
    bolum    = (data.get('bolum') or '').strip()
    basla    = (data.get('basla') or '').strip()   # ISO datetime (boşluğun başlangıcı)
    sebep    = (data.get('durus_sebebi') or '').strip()
    tip      = (data.get('durus_tipi') or 'plansiz').strip()
    try:
        sure_dk = int(round(float(data.get('sure_dk') or 0)))
    except (ValueError, TypeError):
        sure_dk = 0
    if not robot_no or not tarih or not sebep or sure_dk <= 0:
        return jsonify({'hata': 'robot_no, tarih, durus_sebebi, sure_dk zorunlu'}), 400
    # Başlangıç saatini HH:MM'e çevir (timed kayıt — gelecekte kesin eşleşme)
    hhmm = ''
    try:
        hhmm = datetime.fromisoformat(basla).strftime('%H:%M') if basla else ''
    except Exception:
        hhmm = ''
    conn = get_db()
    try:
        v = None
        if bolum and _kolon_var(conn, 'vardiyalar', 'bolum'):
            v = conn.execute("SELECT id FROM vardiyalar WHERE robot_no=? AND tarih=? AND bolum=? ORDER BY id DESC LIMIT 1",
                             (robot_no, tarih, bolum)).fetchone()
        if not v:
            v = conn.execute("SELECT id FROM vardiyalar WHERE robot_no=? AND tarih=? ORDER BY id DESC LIMIT 1",
                             (robot_no, tarih)).fetchone()
        if not v:
            return jsonify({'hata': 'Bu cihaz/tarih için vardiya bulunamadı'}), 404
        conn.execute('''INSERT INTO duruslar (vardiya_id, durus_sebebi, aciklama, sure_dk, baslangic_saati, durus_tipi)
                        VALUES (?, ?, ?, ?, ?, ?)''',
                     (v['id'], sebep, (data.get('aciklama') or '').strip(), sure_dk, hhmm, tip))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'basarili': True, 'baslangic_saati': hhmm, 'sure_dk': sure_dk}), 201


@app.route('/api/pilot/sinyal_kalitesi', methods=['GET'])
@panel_gerekli(izin='sinyal-analizi')
def pilot_sinyal_kalitesi():
    """Sinyal Kalitesi / Tani paneli — cihaz başına sayım kararları + kayıp mutabakatı.
    Pilot.db'den okunur: cihaz_kayitlari (tani sayaçları + pulse_ist baseline) + sayac_olaylari.
      - sayildi  : gerçek sayım (firmware kararı)
      - parazit  : elenen gürültü (ham HIGH→LOW kenar) — sistemin ayıkladığı
      - erken    : MIN_PULSE_GAP altında reddedilen
      - kayip    : cihaz NVS sayımı − server'a ulaşan (transit/WiFi kaybı)
      - son_kararlar: son tani olayları (SAYILDI/ERKEN) zaman akışı
    """
    import os, sqlite3
    lokasyon = request.args.get('lokasyon')   # TK1/TK2'ye göre cihaz filtresi (robot_no eşlemesi)
    pilot_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pilot', 'pilot.db')
    cihazlar, son_kararlar = [], []
    if not os.path.exists(pilot_db):
        return jsonify({'cihazlar': [], 'son_kararlar': [], 'hata': 'Pilot DB yok'}), 200
    try:
        pc = sqlite3.connect(pilot_db)
        pc.row_factory = sqlite3.Row
        try:
            rows = pc.execute('''
                SELECT cihaz_id, bolum, robot_no, wifi_rssi,
                       tani_sayildi, tani_parazit, tani_erken,
                       pulse_ist1, pulse_ist2, base_ist1, base_ist2,
                       CAST((julianday('now','localtime')-julianday(son_heartbeat))*1440 AS INTEGER) hb_dk
                FROM cihaz_kayitlari
            ''').fetchall()
            for r in rows:
                cid = r['cihaz_id']
                sc1 = pc.execute("SELECT COUNT(*) FROM sayac_olaylari WHERE cihaz_id=? AND istasyon=1", (cid,)).fetchone()[0]
                sc2 = pc.execute("SELECT COUNT(*) FROM sayac_olaylari WHERE cihaz_id=? AND istasyon=2", (cid,)).fetchone()[0]
                b1 = r['base_ist1'] if r['base_ist1'] is not None else 0
                b2 = r['base_ist2'] if r['base_ist2'] is not None else 0
                kayip = max(0, (r['pulse_ist1'] or 0) + b1 - sc1) + max(0, (r['pulse_ist2'] or 0) + b2 - sc2)
                sayildi = r['tani_sayildi'] or 0
                parazit = r['tani_parazit'] or 0   # TÜM ham HIGH→LOW kenar (gerçek pulse kenarı dahil)
                erken   = r['tani_erken'] or 0
                # Elenen gürültü = ham kenar − sayılan − erken (gerçek pulse'ların kenarı düşülür)
                # Temiz sinyalde ~0 (her kenar sayıldı); 400T gibi gürültülü hatta yüksek.
                elenen_gurultu = max(0, parazit - sayildi - erken)
                cihazlar.append({
                    'cihaz_id': cid, 'bolum': r['bolum'], 'robot_no': r['robot_no'],
                    'wifi_rssi': r['wifi_rssi'], 'online': (r['hb_dk'] if r['hb_dk'] is not None else 999) <= 2,
                    'sayildi': sayildi, 'elenen_gurultu': elenen_gurultu, 'erken': erken,
                    'ham_kenar': parazit, 'kayip': kayip,
                })
            for r in pc.execute('''
                SELECT cihaz_id, robot_no, tip, gap_ms, low_ms, ts
                FROM tani_olaylari ORDER BY id DESC LIMIT 30
            ''').fetchall():
                son_kararlar.append(dict(r))
        finally:
            pc.close()
    except Exception as e:
        return jsonify({'cihazlar': [], 'son_kararlar': [], 'hata': str(e)}), 200
    # Lokasyon filtresi (TK1 = TK1_ROBOT_NOLARI; TK2 = geri kalan)
    if lokasyon == 'TK1':
        cihazlar = [c for c in cihazlar if c.get('robot_no') in TK1_ROBOT_NOLARI]
        son_kararlar = [k for k in son_kararlar if k.get('robot_no') in TK1_ROBOT_NOLARI]
    elif lokasyon == 'TK2':
        cihazlar = [c for c in cihazlar if c.get('robot_no') not in TK1_ROBOT_NOLARI]
        son_kararlar = [k for k in son_kararlar if k.get('robot_no') not in TK1_ROBOT_NOLARI]
    # En çok sinyal işleyen (tani verisi olan) cihazlar üstte
    cihazlar.sort(key=lambda x: (x['sayildi'] + x['ham_kenar']), reverse=True)
    return jsonify({'cihazlar': cihazlar, 'son_kararlar': son_kararlar})


@app.route('/api/pilot/sinyal_analiz', methods=['GET'])
@panel_gerekli(izin='sinyal-analizi')
def pilot_sinyal_analiz():
    """Pilot sayaç pulse'larının zaman analizi.

    Query parametreleri:
      - bolum:    'kaynak' | 'montaj' | 'metal' (zorunlu)
      - robot_no: 'ABB2', 'M3', '400T' vb. (zorunlu)
      - istasyon: '1' | '2' | '' (boş = tüm istasyonlar)
      - tarih:    'YYYY-MM-DD' (default bugün)

    Dönen:
      - olaylar:        her pulse {ts, istasyon, gap_sn}
      - ozet:           toplam_pulse, ort_gap_sn, en_uzun_gap, en_kisa_gap
      - saatlik:        [{hour, pulse_sayisi, ort_gap_sn}]
      - duruslar:       [{basla, biti, sebep, tip, sure_dk}]
      - referans_donemler: o gün üretilen referansların created_at timeline'ı
      - aktif_referanslar: şu an referans_takip'te uretimde olanlar
    """
    import os, sqlite3
    from datetime import datetime, timedelta

    bolum    = request.args.get('bolum', 'kaynak')
    robot_no = request.args.get('robot_no', '').strip()
    istasyon = request.args.get('istasyon', '').strip()
    tarih    = request.args.get('tarih', datetime.now().strftime('%Y-%m-%d'))

    if not robot_no:
        return jsonify({'hata': 'robot_no parametresi zorunlu'}), 400

    # ─── Pilot DB'den pulse'lar ───────────────────────────────────
    pilot_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pilot', 'pilot.db')
    olaylar = []
    if os.path.exists(pilot_db):
        pc = sqlite3.connect(pilot_db)
        pc.row_factory = sqlite3.Row
        try:
            q = '''SELECT ts, istasyon FROM sayac_olaylari
                   WHERE bolum=? AND robot_no=? AND date(ts)=?'''
            params = [bolum, robot_no, tarih]
            if istasyon:
                q += ' AND istasyon=?'
                params.append(int(istasyon))
            q += ' ORDER BY ts'
            rows = pc.execute(q, params).fetchall()

            # Her istasyon için ayrı gap takibi (ist1 ile ist2 birbirine karışmasın)
            son_ts_per_ist = {}
            for r in rows:
                try:
                    ts_dt = datetime.fromisoformat(r['ts'])
                except Exception:
                    continue
                ist = r['istasyon'] or 0
                prev = son_ts_per_ist.get(ist)
                gap_sn = round((ts_dt - prev).total_seconds(), 1) if prev else None
                olaylar.append({
                    'ts':       r['ts'],
                    'istasyon': ist,
                    'gap_sn':   gap_sn,
                })
                son_ts_per_ist[ist] = ts_dt
        except Exception as e:
            print(f'[sinyal_analiz] Pilot DB hata: {e}')
        finally:
            pc.close()

    # ─── Özet metrikleri ──────────────────────────────────────────
    # Duruş eşiği: bu süreyi aşan gap = duruş (gerçek cycle değil). ?durus_esik_sn ile ayarlanır.
    DURUS_ESIK_SN = max(60, int(request.args.get('durus_esik_sn', 300) or 300))
    # Kayıtlı duruşlar (SEBEP eşleştirmesi için). NOT: duruslar tablosu aynı vardiya+sebep'i
    # TOPLAR (Andon toplamı için doğru) → kayıt zamanı timeline'da yanıltıcı. O yüzden
    # duruşları SİNYALDEN (uzun gap) gerçek zamanında türetip sebebi buradan eşleştiriyoruz.
    conn = get_db()
    kayitli_duruslar = []
    try:
        for _r in conn.execute('''
            SELECT d.baslangic_saati, d.sure_dk, d.durus_sebebi, d.durus_tipi, d.aciklama
            FROM duruslar d JOIN vardiyalar v ON d.vardiya_id = v.id
            WHERE v.robot_no=? AND v.tarih=? ORDER BY d.baslangic_saati
        ''', (robot_no, tarih)).fetchall():
            # NOT: kayıtların ~%100'ünde baslangic_saati BOŞ — bu yüzden eşleştirme
            # öncelikle SÜRE ile yapılır (40dk gap ↔ 40dk kayıt). Saat varsa zaman da kullanılır.
            _bs = (_r['baslangic_saati'] or '').strip()
            _sure_sn = (_r['sure_dk'] or 0) * 60
            _bdt = None
            if _bs:
                try:
                    _iso = _bs if ('T' in _bs or len(_bs) > 8) else (f'{tarih}T{_bs}:00' if len(_bs) == 5 else f'{tarih}T{_bs}')
                    _bdt = datetime.fromisoformat(_iso)
                except Exception:
                    _bdt = None
            kayitli_duruslar.append({
                'basla_dt': _bdt,
                'biti_dt':  (_bdt + timedelta(seconds=_sure_sn)) if _bdt else None,
                'sure_sn':  _sure_sn,
                'sebep':    _r['durus_sebebi'] or '',
                'tip':      _r['durus_tipi'] or 'plansiz',
                'aciklama': _r['aciklama'] or '',
                'used':     False,
            })
    except Exception as _e:
        print(f'[sinyal_analiz] kayitli durus hata: {_e}')

    def _durus_gap_mi(prev_dt, cur_dt, gap_sn):
        """Gap bir duruşa mı denk geliyor? (eşik üstü VEYA kayıtlı duruşla örtüşen)"""
        if gap_sn is not None and gap_sn > DURUS_ESIK_SN:
            return True
        if prev_dt is None or cur_dt is None:
            return False
        for kd in kayitli_duruslar:
            if kd['basla_dt'] is not None and prev_dt < kd['biti_dt'] and cur_dt > kd['basla_dt']:
                return True
        return False

    # Cycle istatistikleri SADECE gerçek cycle gap'lerinden (duruş gap'leri hariç) hesaplanır
    cycle_gaps = []
    haric_durus = 0
    for o in olaylar:
        g = o['gap_sn']
        if g is None:
            o['durus'] = False
            continue
        try:
            _cur = datetime.fromisoformat(o['ts'])
            _prev = _cur - timedelta(seconds=g)
        except Exception:
            _cur = _prev = None
        _isdurus = _durus_gap_mi(_prev, _cur, g)
        o['durus'] = _isdurus            # grafik scatter bu bayrakla duruş noktasını atlar
        if _isdurus:
            haric_durus += 1
            continue
        cycle_gaps.append(g)
    gaps = cycle_gaps
    # Medyan
    _medyan = 0
    if gaps:
        _sorted = sorted(gaps)
        _mid = len(_sorted) // 2
        _medyan = round(_sorted[_mid] if len(_sorted) % 2 else (_sorted[_mid-1] + _sorted[_mid]) / 2, 1)
    # Mod (1 sn yuvarlanmış değerlerin en sık görüleni)
    _mod = 0
    if gaps:
        _counts = {}
        for v in gaps:
            r = round(v)
            _counts[r] = _counts.get(r, 0) + 1
        _mod = max(_counts, key=_counts.get)
    ozet = {
        'toplam_pulse':    len(olaylar),
        'ortalama_gap_sn': round(sum(gaps) / len(gaps), 1) if gaps else 0,
        'medyan_gap_sn':   _medyan,
        'mod_gap_sn':      _mod,
        'en_uzun_gap_sn':  round(max(gaps), 1) if gaps else 0,
        'en_kisa_gap_sn':  round(min(gaps), 1) if gaps else 0,
        'haric_durus_gap': haric_durus,   # cycle dışı bırakılan (duruş) gap sayısı
    }

    # ─── Saatlik dağılım ─────────────────────────────────────────
    saatlik_map = {}  # {hour: [gap1, gap2, ...]}
    saatlik_sayim = {}  # {hour: pulse_count}
    for o in olaylar:
        try:
            _cur = datetime.fromisoformat(o['ts'])
        except Exception:
            continue
        h = _cur.hour
        saatlik_sayim[h] = saatlik_sayim.get(h, 0) + 1
        g = o['gap_sn']
        if g is not None and not _durus_gap_mi(_cur - timedelta(seconds=g), _cur, g):
            saatlik_map.setdefault(h, []).append(g)
    saatlik = []
    for h in range(24):
        gaps_h = saatlik_map.get(h, [])
        if h in saatlik_sayim:
            saatlik.append({
                'hour':         h,
                'pulse_sayisi': saatlik_sayim[h],
                'ort_gap_sn':   round(sum(gaps_h) / len(gaps_h), 1) if gaps_h else 0,
            })

    # ─── Duruşlar (ana DB) ──────────────────────────────────────
    # SİNYALDEN türet: birleşik pulse akışında eşik üstü boşluk = bir duruş (GERÇEK zaman).
    # Kayıttaki toplanmış süre/zaman KULLANILMAZ; sebep, kayıtlı duruştan eşleştirilir.
    def _match_durus(prev_dt, cur_dt, gap_sn):
        # 1) Saati DOLU kayıtla zaman örtüşmesi (nadir — kayıtların çoğu saatsiz)
        for kd in kayitli_duruslar:
            if (not kd['used'] and kd['basla_dt'] is not None
                    and prev_dt < kd['biti_dt'] and cur_dt > kd['basla_dt']):
                kd['used'] = True
                return kd
        # 2) SÜRE eşleşmesi (saatsiz kayıtlar): süresi gap'e en yakın, tolerans içinde kayıt
        aday, en_fark = None, None
        tol = max(120, gap_sn * 0.25)   # ±%25 veya en az 2dk
        for kd in kayitli_duruslar:
            if kd['used']:
                continue
            fark = abs(kd['sure_sn'] - gap_sn)
            if fark <= tol and (en_fark is None or fark < en_fark):
                en_fark, aday = fark, kd
        if aday is not None:
            aday['used'] = True
        return aday
    tum_ts = []
    for o in olaylar:
        try:
            tum_ts.append(datetime.fromisoformat(o['ts']))
        except Exception:
            pass
    tum_ts.sort()
    # Kaydedilmemiş (sebebi olmayan) duruş ancak >=10dk ise timeline'a gelsin —
    # kısa boşluklar "Kaydedilmemiş duruş" olarak kalabalık yapmasın. ?kayitsiz_esik_sn ile ayarlanır.
    KAYITSIZ_DURUS_ESIK_SN = max(0, int(request.args.get('kayitsiz_esik_sn', 600) or 600))
    duruslar = []
    for _i in range(1, len(tum_ts)):
        prev_dt, cur_dt = tum_ts[_i-1], tum_ts[_i]
        gap_sn = (cur_dt - prev_dt).total_seconds()
        if gap_sn <= DURUS_ESIK_SN:
            continue
        kd = _match_durus(prev_dt, cur_dt, gap_sn)
        if kd is None and gap_sn < KAYITSIZ_DURUS_ESIK_SN:
            continue  # kısa + kaydedilmemiş → timeline'a ekleme
        duruslar.append({
            'basla':    prev_dt.isoformat(),
            'biti':     cur_dt.isoformat(),
            'sure_dk':  round(gap_sn / 60, 1),
            'sebep':    (kd['sebep'] if kd else '') or 'Kaydedilmemiş',
            'tip':      kd['tip'] if kd else 'plansiz',
            'aciklama': kd['aciklama'] if kd else '',
        })

    # ─── Referans dönemler (üretim kayıtları timeline'ı) ────────
    uretim_rows = conn.execute('''
        SELECT u.referans_kodu, u.ok_adet, u.nok_adet, u.cycle_time_sn, u.created_at,
               r.hedef_cycle_time_sn
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON u.vardiya_id = v.id
        LEFT JOIN referans_listesi r ON r.referans_kodu = u.referans_kodu
                                     AND COALESCE(r.bolum,'kaynak') = COALESCE(v.bolum,'kaynak')
        WHERE v.robot_no=? AND v.tarih=?
        ORDER BY u.created_at
    ''', (robot_no, tarih)).fetchall() if _kolon_var(conn, 'vardiyalar', 'bolum') else conn.execute('''
        SELECT u.referans_kodu, u.ok_adet, u.nok_adet, u.cycle_time_sn, u.created_at,
               r.hedef_cycle_time_sn
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON u.vardiya_id = v.id
        LEFT JOIN referans_listesi r ON r.referans_kodu = u.referans_kodu
        WHERE v.robot_no=? AND v.tarih=?
        ORDER BY u.created_at
    ''', (robot_no, tarih)).fetchall()
    referans_donemler = []
    for r in uretim_rows:
        referans_donemler.append({
            'ts':              r['created_at'],
            'referans_kodu':   r['referans_kodu'],
            'ok_adet':         r['ok_adet'] or 0,
            'nok_adet':        r['nok_adet'] or 0,
            'cycle_time_sn':   r['cycle_time_sn'] or 0,
            'hedef_cycle_sn':  r['hedef_cycle_time_sn'] or 0,
        })

    # ─── Aktif referanslar (şu an üretimde olanlar) ─────────────
    aktif_ref_rows = conn.execute('''
        SELECT referans_kodu, istasyon, hedef_adet
        FROM referans_takip
        WHERE durum='uretimde' AND robot_no=?
    ''', (robot_no,)).fetchall()
    aktif_referanslar = [dict(r) for r in aktif_ref_rows]

    # ─── Referans değişim noktaları (o gün üretime giren referansların başlangıç anları) ───
    # Sinyal Analizi grafiginde dikey cizgi olarak gosterilir
    # Montaj icin robot_no='MONTAJ' veya bos olabilir — fallback ile yakala
    robot_filtre = "(robot_no=? OR robot_no='MONTAJ' OR robot_no='')" if bolum == 'montaj' else "robot_no=?"
    degisim_rows = conn.execute(f'''
        SELECT referans_kodu, uretime_baslama_ts, istasyon
        FROM referans_takip
        WHERE uretime_baslama_ts IS NOT NULL
          AND date(uretime_baslama_ts)=?
          AND {robot_filtre}
        ORDER BY uretime_baslama_ts
    ''', (tarih, robot_no)).fetchall()
    referans_degisim_noktalari = []
    for r in degisim_rows:
        ist = int(r['istasyon'] or 0)
        # Istasyon filtresi varsa eslesmeyenleri at
        if istasyon and ist > 0 and ist != int(istasyon):
            continue
        referans_degisim_noktalari.append({
            'ts':            r['uretime_baslama_ts'],
            'referans_kodu': r['referans_kodu'],
            'istasyon':      ist if ist > 0 else None,
        })

    conn.close()

    return jsonify({
        'cihaz_id':                   f'{robot_no}-IO' if bolum != 'montaj' else f'MONTAJ-{robot_no}',
        'robot_no':                   robot_no,
        'bolum':                      bolum,
        'istasyon':                   int(istasyon) if istasyon else None,
        'tarih':                      tarih,
        'ozet':                       ozet,
        'olaylar':                    olaylar,
        'saatlik':                    saatlik,
        'duruslar':                   duruslar,
        'referans_donemler':          referans_donemler,
        'referans_degisim_noktalari': referans_degisim_noktalari,
        'aktif_referanslar':          aktif_referanslar,
    }), 200


def _kolon_var(conn, tablo, kolon):
    """Migration güvenliği — kolon var mı kontrol."""
    try:
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({tablo})").fetchall()]
        return kolon in cols
    except Exception:
        return False


@app.route('/api/referans/teyit_ozet', methods=['GET'])
def referans_teyit_ozet():
    """Bölüm bazında teyit özeti: toplam / teyitli / teyitsiz / süresiz sayıları.
    ?lokasyon= ile tesise göre filtrelenir (montaj çağrısında TK1/TK2 karışmasın)."""
    bolum = request.args.get('bolum', 'kaynak')
    lokasyon = request.args.get('lokasyon')
    sql = '''
        SELECT
          COUNT(*) as toplam,
          SUM(CASE WHEN COALESCE(sure_teyit,0)=1 THEN 1 ELSE 0 END) as teyitli,
          SUM(CASE WHEN COALESCE(sure_teyit,0)=0 AND COALESCE(hedef_cycle_time_sn,0)>0 THEN 1 ELSE 0 END) as teyitsiz,
          SUM(CASE WHEN COALESCE(hedef_cycle_time_sn,0)=0 THEN 1 ELSE 0 END) as suresiz
        FROM referans_listesi
        WHERE COALESCE(bolum,'kaynak')=?
    '''
    params = [bolum]
    if lokasyon:
        sql += " AND COALESCE(lokasyon,'TK2')=?"
        params.append(lokasyon)
    conn = get_db()
    rows = conn.execute(sql, params).fetchone()
    conn.close()
    d = dict(rows) if rows else {}
    return jsonify({
        'bolum': bolum,
        'lokasyon': lokasyon or '',
        'toplam': d.get('toplam', 0) or 0,
        'teyitli': d.get('teyitli', 0) or 0,
        'teyitsiz': d.get('teyitsiz', 0) or 0,
        'suresiz': d.get('suresiz', 0) or 0,
    }), 200


@app.route('/api/pair_cycle', methods=['GET'])
def api_pair_cycle():
    """Robot kaynak — iki istasyona atanan referansların gerçek paralel çevrim
    süresini hesaplar.

    Query: ?ist1=10.130.X&ist2=10.130.Y (her ikisi de opsiyonel)

    Mantık: max(K1, S2) + max(K2, S1)
      K1=İst1 kaynak süresi, S1=İst1 söktak süresi (referans bazlı)
    """
    from oee import pair_cycle_hesapla
    ist1 = (request.args.get('ist1') or '').strip()
    ist2 = (request.args.get('ist2') or '').strip()
    sonuc = pair_cycle_hesapla(ist1 or None, ist2 or None)
    return jsonify(sonuc), 200


@app.route('/api/referans/sureler', methods=['PATCH'])
def referans_sureler_guncelle():
    """Bir referansın kaynak ve söktak sürelerini günceller.
    Body: {referans_kodu, kaynak_suresi_sn, soktak_suresi_sn}
    Excel'in 'Kaynak Referans' sayfasını da otomatik günceller.
    """
    data = request.get_json() or {}
    kod = (data.get('referans_kodu') or '').strip()
    if not kod:
        return jsonify({'hata': 'referans_kodu zorunlu'}), 400
    try:
        kaynak = float(data.get('kaynak_suresi_sn', 0) or 0)
        soktak = float(data.get('soktak_suresi_sn', 0) or 0)
    except (ValueError, TypeError):
        return jsonify({'hata': 'kaynak/soktak sayısal olmalı'}), 400
    toplam = round(kaynak + soktak, 2)

    conn = get_db()
    try:
        # Bu endpoint kaynak (TK2) içindir → kaynak bölümü + TK2'ye kapat
        # (aynı kodlu TK1 satırına ya da başka bölümün — lazer/pres — satırına dokunma)
        c = conn.cursor()
        c.execute(
            "UPDATE referans_listesi SET kaynak_suresi_sn=?, soktak_suresi_sn=?, hedef_cycle_time_sn=? "
            "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
            "AND COALESCE(bolum,'kaynak')='kaynak' AND COALESCE(lokasyon,'TK2')='TK2'",
            (kaynak, soktak, toplam, kod)
        )
        # Geriye dönük: bu kodla mevcut TK2 KAYNAK üretim kayıtlarındaki cycle_time'ı güncelle
        c.execute(
            "UPDATE uretim_kayitlari SET cycle_time_sn=? "
            "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
            "AND vardiya_id IN (SELECT id FROM vardiyalar WHERE COALESCE(lokasyon,'TK2')='TK2' AND COALESCE(bolum,'kaynak')='kaynak')",
            (toplam, kod)
        )
        conn.commit()
    finally:
        conn.close()

    # Excel'e yaz (Kaynak Referans sayfası — kolon B=kaynak, C=söktak)
    try:
        from import_excel import EXCEL_YOL, BOLUM_SAYFA
        import openpyxl, os
        if os.path.exists(EXCEL_YOL):
            wb = openpyxl.load_workbook(EXCEL_YOL)
            sayfa_adi = BOLUM_SAYFA['kaynak']['ref']
            if sayfa_adi in wb.sheetnames:
                ws = wb[sayfa_adi]
                norm_target = kod.upper().replace(' ', '')
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=1).value
                    if cell and str(cell).upper().replace(' ', '') == norm_target:
                        ws.cell(row=r, column=2, value=kaynak)
                        ws.cell(row=r, column=3, value=soktak)
                        ws.cell(row=r, column=4, value=f'=B{r}+C{r}')
                        break
                wb.save(EXCEL_YOL)
    except Exception as e:
        print(f'[referans/sureler] Excel sync hatası: {e}')

    return jsonify({'basarili': True, 'kaynak': kaynak, 'soktak': soktak, 'toplam': toplam}), 200


@app.route('/api/veri/arsivle_simdi', methods=['POST'])
def veri_arsivle_simdi():
    """Manuel tetik — 18:00'lık otomatik arşivin elle çalıştırılması."""
    try:
        from scheduler import gunluk_arsiv_calistir
        gunluk_arsiv_calistir()
        return jsonify({'basarili': True}), 200
    except Exception as e:
        return jsonify({'hata': str(e), 'basarili': False}), 500


# ─────────────────────────────────────────────────────────────
# EK DURUŞ SEBEPLERİ (Excel dışı)
# ─────────────────────────────────────────────────────────────
# Duruş listesi sunucudaki Excel'den okunur (data/Tk1 Veriler.xlsx → 'Duruş
# Listesi'). Sunucuda Excel KURULU DEĞİL ve dosya git ile taşınmıyor (data/
# gitignore'da) — yeni bir sebep eklemek için dosyayı RDP ile laptopa çekip
# düzenleyip geri kopyalamak gerekiyor. Tek satırlık ekleme için ağır bir zincir.
#
# Buradaki liste Excel'den gelen sebeplerin ÜSTÜNE eklenir. Excel yine ana
# kaynaktır: aynı sebep Excel'e de yazılırsa mükerrer GÖSTERİLMEZ (ad bazlı
# tekilleştirme yapılır), yani ileride Excel'e taşınması sorun çıkarmaz.
#
# TİP SEÇİMİ OEE'Yİ DOĞRUDAN ETKİLER:
#   'planli'  → net plan süresinden DÜŞÜLÜR, kullanılabilirliği düşürmez
#   'plansiz' → kayıp olarak görünür, kullanılabilirliği DÜŞÜRÜR
# Şüphede 'plansiz' seçilir: yanlış 'planli' gerçek bir kaybı GİZLER.
EK_DURUS_SEBEPLERI = {
    # (lokasyon, bolum) — bolum None ise o tesisin TÜM bölümlerinde görünür
    ('TK1', None): [
        # 2026-08-21 kullanıcı isteği. Plansız seçildi: kalıp bakımı planlı
        # yapılıyorsa 'planli' yapılmalı (o zaman kullanılabilirliği düşürmez).
        {'sebep': 'Kalıp Bakımı', 'tip': 'plansiz'},
    ],
    ('TK1', 'tel'): [
        # 2026-08-24 kullanıcı isteği: tel hattında halat taşlama duruşu.
        # 'plansiz' seçildi (modülün kuralı: şüphede plansız — yanlış 'planli'
        # gerçek bir kaybı GİZLER). Planlı bir operasyonsa tipi değiştirilmeli.
        {'sebep': 'Halat Taşlama', 'tip': 'plansiz'},
    ],
}


def _ek_durus_sebepleri(bolum, lokasyon):
    """Excel dışı sebepler — (lokasyon, bolum) ve (lokasyon, None) birleşir."""
    lok = (lokasyon or 'TK2').upper()
    out = []
    for anahtar in ((lok, bolum), (lok, None)):
        out += EK_DURUS_SEBEPLERI.get(anahtar, [])
    return out


@app.route('/api/durus_sebepleri', methods=['GET'])
def durus_sebepleri_api():
    """Bölüme göre duruş sebeplerini Excel'den okuyup döner.

    ?bolum=kaynak → [{'sebep': 'Robot Prog. Çalışması', 'tip': 'planli'}, ...]

    Excel sayfaları: data/uretim_verileri.xlsx
      - Robotik Kaynak Duruş Listesi
      - Montaj Duruş Listesi
      - Metal Enjeksiyon Duruş Listesi

    Cache-Control: no-store → Excel'de yapılan değişiklikler anında yansır.
    """
    bolum = (request.args.get('bolum') or 'kaynak').strip()
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    if bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f"Geçersiz bölüm: {bolum}"}), 400
    sebepler = durus_sebepleri_yukle(bolum, lokasyon)
    # Excel dışı sebepleri ekle — Excel'de zaten varsa TEKRARLAMA (ad bazlı).
    _mevcut = {str(x.get('sebep', '')).strip().upper() for x in sebepler}
    for _ek in _ek_durus_sebepleri(bolum, lokasyon):
        if str(_ek['sebep']).strip().upper() not in _mevcut:
            sebepler.append(dict(_ek))
    # BOŞ LİSTENİN SEBEBİNİ SÖYLE (2026-08-18). Eskiden boş dönerdi ve operatör
    # mobili sessizce 6 maddelik yedek listeye düşerdi: operatörler "duruşlarımız
    # azaldı" diyene kadar kimse fark etmiyor, o arada duruşlar YANLIŞ/GENEL
    # sebeplerle kaydediliyordu. Artık neden boş olduğu API'de görünür.
    uyari = ''
    if not sebepler:
        try:
            import import_excel as _ie
            _yol = _ie.TK1_EXCEL_YOL if (lokasyon or 'TK2').upper() == 'TK1' else _ie.EXCEL_YOL
            if not os.path.exists(_yol):
                uyari = f'Excel dosyası BULUNAMADI: {_yol}'
            else:
                try:
                    import openpyxl as _op, io as _bio
                    # Tanı da BELLEKTEN okur — bozuk dosyayı KİLİTLEMESİN
                    # (bu kod her istekte çalışıyor).
                    with open(_yol, "rb") as _f2:
                        _ham = _f2.read()
                    _wb = _op.load_workbook(_bio.BytesIO(_ham), data_only=True, read_only=True)
                    _sayfa = (_ie.TK1_SAYFA['durus'] if (lokasyon or 'TK2').upper() == 'TK1'
                              else _ie.BOLUM_DURUS_SAYFA.get(bolum, ''))
                    if _sayfa and _sayfa not in _wb.sheetnames:
                        uyari = (f'"{_sayfa}" sayfası dosyada yok. Mevcut sayfalar: '
                                 + ', '.join(_wb.sheetnames[:12]))
                    else:
                        uyari = f'"{_sayfa}" sayfası var ama okunabilir satır yok (liste boş?)'
                    _wb.close()
                except Exception as _e2:
                    uyari = (f'Excel AÇILAMADI ({type(_e2).__name__}: {_e2}) — '
                             f'dosya sunucuda AÇIK/kilitli veya bozuk olabilir')
        except Exception as _e:
            uyari = f'tanı yapılamadı: {_e}'
        print(f'[durus_sebepleri] {bolum}/{lokasyon} BOŞ — {uyari}')
    resp = jsonify({'bolum': bolum, 'lokasyon': lokasyon, 'sebepler': sebepler,
                    'uyari': uyari})
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp, 200


@app.route('/api/export_arsiv', methods=['POST'])
def export_arsiv_endpoint():
    """Verileri masaüstündeki UretimTakipArsiv.xlsx dosyasina yaz."""
    data = request.get_json() or {}
    tarih_bas = data.get('tarih_bas')  # opsiyonel, orn: '2026-02-01'
    tarih_bit = data.get('tarih_bit')  # opsiyonel, orn: '2026-03-10'
    try:
        sonuc = export_arsiv(tarih_bas=tarih_bas, tarih_bit=tarih_bit)
        return jsonify(sonuc), 200
    except Exception as e:
        return jsonify({'basarili': False, 'hata': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# SAHA SAYAÇ TOPLAMI (kullanıcı 2026-08-21)
# ─────────────────────────────────────────────────────────────
# "Dashboard'da makine sayaçlarından aldığımız total veriyi gösteren bir sayaç da
#  koyalım ekranın bir köşesinde."
#
# HAM SİNYAL sayısıdır — üretim adedi DEĞİL. İkisi bilerek ayrı: üretim adedi
# vardiyaya/referansa kapılı, bölenlerden (bukum_operasyon, MAKINE_SINYAL_BOLEN)
# ve paket çarpanından geçer. Buradaki sayaç "saha modülleri bugün kaç kez
# tetiklendi" sorusunu yanıtlar; sistemin canlı olduğunun tek bakışta kanıtıdır.
# Bu yüzden vardiya kapısı da UYGULANMAZ (bkz. project_sayim_vardiya_kapisi):
# vardiya açılmamışken gelen sinyal üretime yazılmaz ama BURADA görünür — zaten
# aranan şey odur.


@app.route('/api/sayac_toplam', methods=['GET'])
def sayac_toplam_api():
    """Saha sayaç modüllerinden BUGÜN gelen toplam sinyal (+ bölüm kırılımı).

    ?lokasyon=TK1 verilirse yalnız o tesisin cihazları sayılır (pilot.db'de
    lokasyon kolonu yok — TK1_ROBOT_NOLARI ile eşlenir, saha_cihazlari ile aynı
    kalıp)."""
    import sqlite3
    lokasyon = (request.args.get('lokasyon') or '').strip().upper()
    tarih = (request.args.get('tarih') or date.today().isoformat()).strip()
    pilot_db = _pilot_db_yolu()
    bos = {'tarih': tarih, 'toplam': 0, 'bugun': 0, 'ilk_gun': None,
           'bolumler': [], 'cihaz_sayisi': 0, 'okunabildi': False}
    if not os.path.exists(pilot_db):
        return jsonify(bos)
    try:
        pc = sqlite3.connect(pilot_db, timeout=5.0)
        pc.row_factory = sqlite3.Row
        try:
            pc.execute('PRAGMA busy_timeout=5000')
            # TOPLAM = BUGÜNE KADARKİ HER ŞEY (kullanıcı 2026-08-25: "sadece
            # günlük olmasın"). Günlük değer ayrıca döner (rozet ipucunda).
            rows = pc.execute(
                "SELECT bolum, robot_no, COUNT(*) AS n, "
                "       SUM(CASE WHEN date(ts)=? THEN 1 ELSE 0 END) AS bugun "
                "FROM sayac_olaylari GROUP BY bolum, robot_no", (tarih,)
            ).fetchall()
            ilk = pc.execute("SELECT MIN(date(ts)) FROM sayac_olaylari").fetchone()
            ilk_gun = ilk[0] if ilk else None
        finally:
            pc.close()
    except Exception as e:
        print(f'[sayac_toplam] pilot.db okunamadı: {e}')
        return jsonify(bos)

    bolum_top, toplam, bugun_top, cihazlar = {}, 0, 0, set()
    for r in rows:
        cihaz = r['robot_no']
        if lokasyon:
            tk1_mi = cihaz in TK1_ROBOT_NOLARI or cihaz in set(HAT_SAYAC_CIHAZI.values())
            if (lokasyon == 'TK1') != bool(tk1_mi):
                continue
        n = int(r['n'] or 0)
        b = r['bolum'] or '—'
        bolum_top[b] = bolum_top.get(b, 0) + n
        toplam += n
        bugun_top += int(r['bugun'] or 0)
        cihazlar.add(cihaz)
    return jsonify({
        'tarih': tarih,
        'toplam': toplam,          # BUGÜNE KADARKİ tüm sinyaller
        'bugun': bugun_top,        # yalnız bugün (ipucunda gösterilir)
        'ilk_gun': ilk_gun,        # "… tarihinden beri" etiketi için
        'cihaz_sayisi': len(cihazlar),
        'bolumler': sorted(
            [{'bolum': b, 'ad': BOLUM_AD.get(b, b), 'sinyal': n} for b, n in bolum_top.items()],
            key=lambda x: -x['sinyal']),
        'okunabildi': True,
    })


# ─────────────────────────────────────────────────────────────
# ÖZET / DASHBOARD VERİSİ
# ─────────────────────────────────────────────────────────────

@app.route('/api/ozet', methods=['GET'])
def ozet():
    """Dashboard için kapsamlı özet verisi."""
    conn = get_db()
    c = conn.cursor()

    bugun = date.today().isoformat()
    tarih_bas = request.args.get('tarih_bas', bugun)
    tarih_bit = request.args.get('tarih_bit', bugun)
    robot = request.args.get('robot')
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    # ÇOKLU BÖLÜM (kullanıcı 2026-08-25): "birden fazla bölüm … seçilebilsin;
    # TK1'de plastik enjeksiyon, montaj ve telin bazı hatlarını tek filtrede
    # görebileyim." 'bolum' virgülle ayrılmış liste kabul eder; TEK değer
    # verildiğinde davranış birebir eskisi gibidir (regresyon yok).
    bolumler = [b.strip() for b in (bolum or '').split(',') if b.strip()]

    # Sorgu bugünü kapsıyorsa: açık vardiyaların auto sayaçlarını ÖNCE tazele —
    # dashboard performans/OEE'si canlı makine adetleriyle hesaplansın
    # (andon rozeti canlıyken performansın bayat kalması sorunu).
    if tarih_bas <= bugun <= tarih_bit:
        # Bölüm başına ayrı çağrı: fonksiyon tek bölüm alıyor ve None geçmek
        # SEÇİLMEYEN bölümleri de senkronlardı (gereksiz yazma + yavaşlık).
        for _b in (bolumler or [None]):
            _acik_auto_vardiyalari_senkronla(conn, bolum=_b, lokasyon=lokasyon or None)

    param_vardiya = [tarih_bas, tarih_bit]
    sart_vardiya = 'tarih BETWEEN ? AND ?'
    if robot:
        sart_vardiya += ' AND robot_no = ?'
        param_vardiya.append(robot)
    if len(bolumler) == 1:
        sart_vardiya += " AND COALESCE(v.bolum, 'kaynak') = ?"
        param_vardiya.append(bolumler[0])
    elif bolumler:
        sart_vardiya += (" AND COALESCE(v.bolum, 'kaynak') IN (%s)"
                         % ','.join('?' * len(bolumler)))
        param_vardiya.extend(bolumler)
    if lokasyon:
        sart_vardiya += " AND COALESCE(v.lokasyon, 'TK2') = ?"
        param_vardiya.append(lokasyon)

    # Vardiya sayıları
    vardiya_sayisi = c.execute(
        f'SELECT COUNT(*) as cnt FROM vardiyalar v WHERE {sart_vardiya}',
        param_vardiya
    ).fetchone()['cnt']

    # Robot bazlı üretim özeti
    robot_uretim = c.execute(f'''
        SELECT v.robot_no,
               SUM(u.ok_adet)  AS toplam_ok,
               SUM(u.nok_adet) AS toplam_nok,
               SUM(u.ok_adet + u.nok_adet) AS toplam_uretim
        FROM vardiyalar v
        LEFT JOIN uretim_kayitlari u ON u.vardiya_id = v.id
        WHERE {sart_vardiya}
        GROUP BY v.robot_no
        ORDER BY v.robot_no
    ''', param_vardiya).fetchall()

    # ── GRUP BAZLI ÜRETİM (2026-08-25) ────────────────────────────────────
    # robot_uretim v.robot_no ile gruplar; hattı üretim kaydında seçilen
    # bölümlerde bu TEK satır demek ('Tel Üretimi' 10 vardiya → tek bar, tek
    # rakam). Kullanıcı: "toplam bir adet yazıyor fakat o adete kesim, kapama,
    # otomatik pres, son montaj dahil, o yüzden işe yarar bir veri olmuyor."
    # Burada kırılım kayit_grubu ile yapılır: tel'de PROSES ADIMI, diğer
    # bölümlerde makine — yani Fabrika Özeti kartındaki barlar adım/makine bazlı.
    _grup_ham = c.execute(f'''
        SELECT v.robot_no,
               COALESCE(v.bolum, 'kaynak') AS bolum,
               COALESCE(u.istasyon, 0) AS istasyon,
               SUM(u.ok_adet)  AS toplam_ok,
               SUM(u.nok_adet) AS toplam_nok,
               COUNT(*)        AS kayit_sayisi
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY v.robot_no, COALESCE(v.bolum, 'kaynak'), COALESCE(u.istasyon, 0)
    ''', param_vardiya).fetchall()
    # ÇOKLU BÖLÜMDE grup adı tek başına yetmez: 'Kapama' telde proses adımı,
    # başka bölümde bir makine adı olabilir. Anahtar (bölüm, grup); satırda
    # bölüm de döner ki panel başlığında yazabilsin.
    _grup_top = {}
    for _r in _grup_ham:
        _b = _r['bolum']
        _g = kayit_grubu(_b, _r['robot_no'], _r['istasyon'])
        _d = _grup_top.setdefault((_b, _g), {'grup': _g, 'bolum': _b,
                                             'toplam_ok': 0, 'toplam_nok': 0,
                                             'kayit_sayisi': 0, 'makineler': []})
        _d['toplam_ok'] += int(_r['toplam_ok'] or 0)
        _d['toplam_nok'] += int(_r['toplam_nok'] or 0)
        _d['kayit_sayisi'] += int(_r['kayit_sayisi'] or 0)
        _m = kayit_hatti(_r['robot_no'], _r['istasyon'])
        if _m and _m not in _d['makineler']:
            _d['makineler'].append(_m)
    grup_uretim = sorted(_grup_top.values(), key=lambda x: -x['toplam_ok'])

    # Referans bazlı üretim
    referans_uretim = c.execute(f'''
        SELECT u.referans_kodu,
               SUM(u.ok_adet)  as toplam_ok,
               SUM(u.nok_adet) as toplam_nok,
               SUM(u.ok_adet + u.nok_adet) as toplam_uretim,
               ROUND(AVG(u.cycle_time_sn), 0) as ort_cycle_time_sn
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY u.referans_kodu
        ORDER BY toplam_uretim DESC
    ''', param_vardiya).fetchall()

    # NOT (2026-08-04): tel bölümünde ayrıca bir "son adım" süzgeci YOK.
    # Her proses adımı kendi referans ekiyle kaydedildiği için ('93.TK.464 KESIM',
    # '93.TK.464 KAPAMA'…) adımlar zaten ayrı satırlarda toplanır — aynı iş birden
    # çok kez sayılmaz. Kapaması/son montajı dışarıda yapılacak ürünler sabit
    # olmadığından referans bazlı son-adım tanımı da kullanılmıyor.

    # ── OPERATÖR × ADIM × REFERANS ÜRETİMİ (2026-08-04) ───────────────────────
    # Kullanıcı isteği (tel üretimi): "1 referans o gün kesim, yarı otomatik ve
    # kapama işlemlerinden geçtiyse ve 200 adet üretildiyse, kesimi yapan
    # operatörün bu referanstan 200 adet KESİM yaptığını görelim."
    # Bu, "üretim yalnız son adımda sayılır" kuralıyla ÇELİŞMEZ: orası ürünün
    # toplam üretimi, burası KİMİN NE İŞ YAPTIĞI. Her adım kendi adediyle görünür.
    # Tel dışındaki bölümlerde de anlamlı (operatör performans kırılımı) → hepsinde
    # döner; frontend yalnız gerekli sayfada gösterir.
    operator_uretim = [dict(r) for r in c.execute(f'''
        SELECT v.operator_adi,
               v.robot_no,
               COALESCE(v.bolum, 'kaynak') AS bolum,
               COALESCE(u.istasyon, 0) AS istasyon,
               u.referans_kodu,
               SUM(u.ok_adet)  as toplam_ok,
               SUM(u.nok_adet) as toplam_nok,
               COUNT(*)        as kayit_sayisi
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY v.operator_adi, v.robot_no, COALESCE(v.bolum, 'kaynak'),
                 COALESCE(u.istasyon, 0), u.referans_kodu
        ORDER BY toplam_ok DESC
    ''', param_vardiya).fetchall()]
    # Hattı ÜRETİM KAYDINDA seçilen bölümlerde (2026-08-18: TK1 montaj + tel,
    # plastik/pres) 'robot_no' sabit hattır — kırılım kaydın istasyonundan gelir.
    # 'hat' alanı gerçek hattı taşır; eski istemciler robot_no'yu okumaya devam eder.
    for _o in operator_uretim:
        _o['hat'] = kayit_hatti(_o.get('robot_no'), _o.get('istasyon'))
    # Tel'de adım karşılığı da verilir ('Kapama 3' → 'Kapama') — frontend adım
    # bazlı gruplayabilsin, hat numarasını ayrıca gösterebilsin.
    if 'tel' in bolumler or not bolumler:
        for _o in operator_uretim:
            if (_o.get('bolum') or '') != 'tel':
                continue
            _o['adim'] = (tel_koddan_adim(_o.get('referans_kodu'))
                          or tel_hat_adimi(_o.get('hat')) or '')

    # Duruş sebep bazlı dağılım
    durus_dagilim = c.execute(f'''
        SELECT d.durus_sebebi,
               d.durus_tipi,
               SUM(d.sure_dk) as toplam_sure,
               COUNT(*) as adet
        FROM duruslar d
        JOIN vardiyalar v ON v.id = d.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY d.durus_sebebi
        ORDER BY toplam_sure DESC
    ''', param_vardiya).fetchall()

    # Planlı / Plansız duruş tip özeti
    durus_tipi_ozet = c.execute(f'''
        SELECT d.durus_tipi,
               SUM(d.sure_dk) as toplam_sure,
               COUNT(*) as adet
        FROM duruslar d
        JOIN vardiyalar v ON v.id = d.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY d.durus_tipi
    ''', param_vardiya).fetchall()

    # Üretim Kayıtları (Detaylı - Düzenlenebilir)
    # istasyon: lazer bölümünde makine (1..6) — frontend 'Lazer N' kırılımı bundan çıkar
    # (lazer'de v.robot_no hep 'Lazer'; makine ayrımı yalnız u.istasyon'da).
    uretim_detay_listesi = c.execute(f'''
        SELECT u.id as uretim_id,
               v.tarih,
               v.robot_no,
               COALESCE(v.bolum, 'kaynak') as bolum,
               v.operator_adi,
               u.referans_kodu,
               u.ok_adet,
               u.nok_adet,
               u.hedef_adet,
               u.cycle_time_sn,
               COALESCE(u.istasyon, 0) as istasyon
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        WHERE {sart_vardiya}
        ORDER BY v.tarih DESC, u.referans_kodu ASC, v.robot_no ASC, v.operator_adi ASC
    ''', param_vardiya).fetchall()
    # PROSES ADIMI (2026-08-21): 'Üretim Detayı' tablosu tel'de adım sütunu
    # gösteriyor (eski 'Operatör Bazlı Üretim' paneli kaldırıldı, sütunu buraya
    # taşındı). Adım hat adından türer — tel_proses tek kaynak, istemci kendi
    # kopyasını tutmaz.
    # 'grup' = Kayıtlar sayfasındaki kırılım başlığı (tel'de proses adımı, diğer
    # bölümlerde makine) — Excel indirmesiyle BİREBİR aynı kural (kayit_grubu).
    # 'makine' = OPERATÖRÜN SEÇTİĞİ ad (2026-08-25). Hattı üretim kaydında
    # seçilen bölümlerde v.robot_no sabit hattır ('TK1 Montaj', 'Tel Üretimi') ve
    # ekranda onu göstermek operatörün gördüğü adla ('MONTAJ - 3', 'Kapama 5')
    # tutmuyordu; filtre de vardiyanın robot_no'suna baktığı için artık
    # kullanılmayan eski hat adları ('LF-LFP', 'Pull', 'Kapama 1') listede
    # kalıyordu. Diğer bölümlerde kayit_hatti zaten robot_no'yu döndürür → fark yok.
    # grup/adim SATIRIN KENDİ bölümünden hesaplanır — çoklu bölüm seçiliyken
    # istek parametresi tek bir bölümü göstermez.
    uretim_detay_listesi = [
        {**dict(r),
         'makine': kayit_hatti(r['robot_no'], r['istasyon']),
         'adim': tel_adim_etiketi(kayit_hatti(r['robot_no'], r['istasyon'])),
         'grup': kayit_grubu(r['bolum'], r['robot_no'], r['istasyon'])}
        for r in uretim_detay_listesi
    ]

    # Robot + Referans bazlı üretim kırılımı
    # İSTASYON DA KIRILIMA GİRER (2026-08-25): hattı üretim kaydında seçilen
    # bölümlerde v.robot_no sabit hattır — yalnız onunla gruplayınca bu tablo
    # 'Tel Üretimi' başlığı altında TEK öbek oluyordu. 'makine' alanı kaydın
    # gerçek makinesini taşır; panel başlığı ondan yazılır.
    # 'grup' de eklenir: panel bu tabloyu üretim listesiyle AYNI süzgeçten
    # geçiriyor (makine + proses adımı) — alanlar aynı adlarda olmalı.
    robot_referans_uretim = [
        {**dict(r),
         'makine': kayit_hatti(r['robot_no'], r['istasyon']),
         'grup': kayit_grubu(r['bolum'], r['robot_no'], r['istasyon'])}
        for r in c.execute(f'''
        SELECT v.robot_no,
               COALESCE(v.bolum, 'kaynak') as bolum,
               COALESCE(u.istasyon, 0) as istasyon,
               u.referans_kodu,
               SUM(u.ok_adet)  as toplam_ok,
               SUM(u.nok_adet) as toplam_nok,
               SUM(u.ok_adet + u.nok_adet) as toplam_uretim,
               ROUND(AVG(u.cycle_time_sn), 0) as ort_cycle_time_sn
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        WHERE {sart_vardiya}
        GROUP BY v.robot_no, COALESCE(v.bolum, 'kaynak'),
                 COALESCE(u.istasyon, 0), u.referans_kodu
        ORDER BY COALESCE(v.bolum, 'kaynak'), v.robot_no,
                 COALESCE(u.istasyon, 0), toplam_uretim DESC
    ''', param_vardiya).fetchall()]

    # Robot + Duruş bazlı kırılım (Güncellendi: Tekil kayıt - Düzenlenebilir)
    robot_durus_kirilim = c.execute(f'''
        SELECT d.id as durus_id,
               v.robot_no,
               d.durus_sebebi,
               d.durus_tipi,
               d.sure_dk as toplam_sure,
               v.operator_adi,
               1 as adet
        FROM duruslar d
        JOIN vardiyalar v ON v.id = d.vardiya_id
        WHERE {sart_vardiya}
        ORDER BY v.robot_no, d.sure_dk DESC
    ''', param_vardiya).fetchall()

    # Vardiya kayıtları (OEE listesi). LİMİT VAR çünkü her satır için hesapla_oee
    # ayrı sorgu kümesi çalıştırıyor — sınırsız aralık paneli dondurur.
    # ?vardiya_limit= ile yükseltilebilir (Kayıtlar sayfası yükseltir): eski sabit
    # 50, geniş tarih aralığında ARANAN VARDİYAYI LİSTEDEN DÜŞÜRÜYORDU ve
    # kullanıcı "sil butonu çıkmıyor" diye bildiriyordu — satırın kendisi yoktu.
    try:
        _v_limit = int(request.args.get('vardiya_limit', 50) or 50)
    except (TypeError, ValueError):
        _v_limit = 50
    _v_limit = max(1, min(1000, _v_limit))
    son_vardiyalar = c.execute(
        f"SELECT v.id, COALESCE(v.bolum,'kaynak') bolum FROM vardiyalar v "
        f'WHERE {sart_vardiya} ORDER BY v.tarih DESC, v.id DESC LIMIT ?',
        param_vardiya + [_v_limit]
    ).fetchall()

    # HANGİ MAKİNEDE ÜRETİLDİ (kullanıcı 2026-08-21) — TK1'de ve pres/plastik/
    # lazer'de vardiyanın robot_no'su SABİT HAT adıdır ('Tel Üretimi',
    # 'Pres Abkant'); gerçek makine her ÜRETİM KAYDINDA (istasyon) durur.
    # Kayıtlar listesi bölüm adını gösterince "hangi makinede üretildi" bilgisi
    # kayboluyordu. Bir vardiyada birden çok makine kullanılabildiği için TEKİL
    # değil LİSTE döner (operatör gün içinde makine değiştirebiliyor).
    vid_listesi = [r['id'] for r in son_vardiyalar]
    makine_map = {}
    if vid_listesi:
        _yer = ','.join('?' * len(vid_listesi))
        for _r in c.execute(
                f"SELECT u.vardiya_id, v.robot_no, u.istasyon, SUM(COALESCE(u.ok_adet,0)) ok "
                f"FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id "
                f"WHERE u.vardiya_id IN ({_yer}) GROUP BY u.vardiya_id, u.istasyon",
                vid_listesi).fetchall():
            ad = kayit_makine_ad(_r['robot_no'], _r['istasyon'] or 0)
            if not ad:
                continue          # sabit hatlı değil → vardiyanın hattı zaten doğru
            makine_map.setdefault(_r['vardiya_id'], []).append((ad, _r['ok'] or 0))
    # En çok üretilen makine başa — listede ilk görünen o olsun
    for _vid, _lst in makine_map.items():
        _lst.sort(key=lambda x: -x[1])

    oee_listesi = []
    for row in son_vardiyalar:
        oee_data = hesapla_oee(row['id'])
        if oee_data:
            _mak = [ad for ad, _ in makine_map.get(row['id'], [])]
            oee_listesi.append({
                'vardiya_id': oee_data['vardiya_id'],
                'tarih': oee_data['tarih'],
                'vardiya': oee_data['vardiya_turu'],
                'robot': oee_data['robot_no'],
                'bolum': row['bolum'],        # çoklu bölüm seçiliyken satır hangi bölüm

                'makineler': _mak,            # boşsa robot_no zaten makinedir
                'operator': oee_data['operator'],
                'baslangic_saati': oee_data.get('baslangic_saati', ''),
                'bitis_saati': oee_data.get('bitis_saati', ''),
                'oee': oee_data['oee'],
                'availability': oee_data['availability'],
                'performance': oee_data['performance'],
                'quality': oee_data['quality'],
                # OEE BİLEŞENLERİ (2026-08-21): Fabrika Özeti'ndeki "Genel OEE"
                # bölüm ortalamalarının ORTALAMASI OLARAK hesaplanamaz — 1
                # vardiyalık bir bölüm 12 vardiyalık bölümle aynı ağırlığı taşır
                # ve rakam anlamını yitirir. İstemci bu bileşenleri TÜM bölümler
                # boyunca toplayıp formülü bir kez uygular (analiz._topla kalıbı).
                'net_plan_dk': oee_data.get('net_plan_dk', 0),
                'calisma_dk': oee_data.get('calisma_suresi_dk', 0),
                # KIRPILMIŞ gönderilir (kullanıcı 2026-08-25): bir vardiya kendi
                # çalışma süresinden fazlasını üretmiş sayılamaz. Ham değer
                # gönderilseydi cycle time'ı cömert bir vardiya (%199) havuzda
                # cycle'ı tanımsız vardiyaları sübvanse eder ve Genel OEE
                # olduğundan ~30 puan yüksek çıkardı. Kural analiz._topla ile AYNI.
                'gercek_uretim_sn': min(oee_data.get('gercek_uretim_sn', 0),
                                        oee_data.get('calisma_suresi_dk', 0) * 60),
                'gercek_uretim_sn_ham': oee_data.get('gercek_uretim_sn', 0),
                'ok': oee_data['toplam_ok'],
                'nok': oee_data['toplam_nok'],
                'hedef': oee_data['toplam_hedef'],
                'planli_durus_dk': oee_data.get('planli_durus_dk', 0),
                'plansiz_durus_dk': oee_data.get('plansiz_durus_dk', 0),
                'durus_dk': oee_data.get('planli_durus_dk', 0) + oee_data.get('plansiz_durus_dk', 0),
            })


    conn.close()

    ort_oee = round(sum(o['oee'] for o in oee_listesi) / len(oee_listesi), 1) if oee_listesi else 0

    return jsonify({
        'tarih_aralik': {'bas': tarih_bas, 'bit': tarih_bit},
        'vardiya_sayisi': vardiya_sayisi,
        'ort_oee': ort_oee,
        'robot_uretim': [dict(r) for r in robot_uretim],
        # Kırılım için robot_uretim yerine BUNU kullanın: hattı üretim kaydında
        # seçilen bölümlerde robot_uretim tek satırdır (bkz. grup_uretim açıklaması).
        'grup_uretim': grup_uretim,
        # Bu bölümde makine/hat ÜRETİM KAYDINDA mı seçiliyor? Panel filtresi ve
        # tablo başlığı buna göre kurulur (etiket boşsa hat vardiyada seçilir).
        'makine_etiketi': kayit_makine_etiketi(
            lokasyon, (bolumler[0] if len(bolumler) == 1 else (bolum or 'kaynak')),
            [r['robot_no'] for r in robot_uretim]),
        # Çoklu bölümde tek etiket anlamsız — her bölümün kendi terimi
        # ('Masa' / 'Makine' / '' = hat vardiyada seçilir).
        'makine_etiketleri': {_b: kayit_makine_etiketi(lokasyon, _b)
                              for _b in (bolumler or [])},
        'secili_bolumler': bolumler,
        'referans_uretim': [dict(r) for r in referans_uretim],
        'operator_uretim': operator_uretim,   # operatör × hat/adım × referans (2026-08-04)
        'uretim_detay_listesi': uretim_detay_listesi,
        'robot_referans_uretim': robot_referans_uretim,
        'durus_dagilim': [dict(r) for r in durus_dagilim],
        'durus_tipi_ozet': [dict(r) for r in durus_tipi_ozet],
        'robot_durus_kirilim': [dict(r) for r in robot_durus_kirilim],
        'vardiyalar': oee_listesi
    })


# ─────────────────────────────────────────────────────────────
# OPERATÖR SÜRE BİRLEŞTİRME (kullanıcı 2026-08-21)
# ─────────────────────────────────────────────────────────────
# Gerekçe ve kurallar analiz.py'de (bkz. _operator_zamani açıklaması). Burası
# Raporlar sayfasının operatör tablosu için aynı hesabı yapar; formül tek yerde
# dursun diye analiz.py'nin yardımcıları kullanılır.


def _operator_sure_birlestir(conn, v_rows):
    """(calisma_dk, planli_dk, plansiz_dk) — robot destekli paralel vardiyalarda
    süre BİR KEZ sayılır. 'Robotla Çalışıyor' işaretsiz vardiyalarda toplama."""
    import analiz as _an
    kayitlar = []
    for x in v_rows:
        try:
            robotlu = int(x['robotla_calisiyor'] or 0) == 1
        except (KeyError, IndexError, TypeError):
            robotlu = False       # eski DB (kolon yok) → eski davranış
        kayitlar.append({
            'tarih': x['tarih'],
            'baslangic': x['baslangic_saati'],
            'toplam_dk': _vardiya_efektif_dk(x),
            'planli_dk': 0, 'plansiz_dk': 0, 'robotlu': robotlu, 'duruslar': [],
            '_id': x['id'],
        })
    kmap = {k['_id']: k for k in kayitlar}
    if kmap:
        yer = ','.join('?' * len(kmap))
        for d in conn.execute(
                f"SELECT vardiya_id, COALESCE(sure_dk,0) AS sure_dk, durus_tipi, "
                f"       baslangic_ts, bitis_ts FROM duruslar WHERE vardiya_id IN ({yer})",
                list(kmap)).fetchall():
            k = kmap.get(d['vardiya_id'])
            if not k:
                continue
            tip = 'planli' if (d['durus_tipi'] or 'plansiz') == 'planli' else 'plansiz'
            k[tip + '_dk'] += d['sure_dk']
            k['duruslar'].append({'tip': tip, 'dk': d['sure_dk'],
                                  'bas_ts': d['baslangic_ts'], 'bit_ts': d['bitis_ts']})
    return _an._operator_zamani(kayitlar)


# ─────────────────────────────────────────────────────────────
# RAPOR API
# ─────────────────────────────────────────────────────────────

@app.route('/api/rapor', methods=['GET'])
def rapor_api():
    """Rapor verisi: operator, referans veya teep."""
    rapor_tipi = request.args.get('rapor_tipi', 'operator')
    tarih_bas = request.args.get('tarih_bas', date.today().isoformat())
    tarih_bit = request.args.get('tarih_bit', date.today().isoformat())
    robot = request.args.get('robot', '')
    operator_filtre = request.args.get('operator', '')
    bolum = request.args.get('bolum', '')
    lokasyon = request.args.get('lokasyon', '')

    conn = get_db()
    c = conn.cursor()

    # Temel vardiya filtresi
    sart = 'v.tarih BETWEEN ? AND ?'
    params = [tarih_bas, tarih_bit]
    if robot:
        sart += ' AND v.robot_no = ?'
        params.append(robot)
    if operator_filtre:
        sart += ' AND v.operator_adi = ?'
        params.append(operator_filtre)
    if bolum:
        sart += " AND COALESCE(v.bolum, 'kaynak') = ?"
        params.append(bolum)
    if lokasyon:
        # TK1/TK2 montaj raporları karışmasın (operatör/referans/TEEP tüm tipler bu sartı paylaşır)
        sart += " AND COALESCE(v.lokasyon, 'TK2') = ?"
        params.append(lokasyon)

    # Rapor bugünü kapsıyorsa: açık vardiyaların auto sayaçlarını ÖNCE tazele —
    # operatör/referans/TEEP raporları canlı makine adetleriyle hesaplansın
    _bugun_r = date.today().isoformat()
    if tarih_bas <= _bugun_r <= tarih_bit:
        _acik_auto_vardiyalari_senkronla(conn, bolum or None, lokasyon or None)

    if rapor_tipi == 'operator':
        # Operator bazli rapor
        rows = c.execute(f'''
            SELECT v.operator_adi,
                   COUNT(DISTINCT v.id) as vardiya_sayisi,
                   SUM(v.toplam_sure_dk) as toplam_calisma_dk,
                   COALESCE(SUM(u.ok_adet), 0) as toplam_ok,
                   COALESCE(SUM(u.nok_adet), 0) as toplam_nok
            FROM vardiyalar v
            LEFT JOIN uretim_kayitlari u ON u.vardiya_id = v.id
            WHERE {sart}
            GROUP BY v.operator_adi
            ORDER BY toplam_ok DESC
        ''', params).fetchall()

        sonuclar = []
        for r in rows:
            op = r['operator_adi']
            # Bu operatorun vardiya id'lerini al
            vp = list(params)
            v_sart = sart + " AND v.operator_adi = ?"
            vp.append(op)
            v_rows = c.execute(
                f'SELECT v.* FROM vardiyalar v WHERE {v_sart}', vp
            ).fetchall()
            vids = [x['id'] for x in v_rows]
            # ── 1 OPERATÖR ÷ 2 MAKİNE (kullanıcı 2026-08-21) ──────────────────
            # Metal enjeksiyonda robot parçayı aldığı için bir operatör iki
            # makineye bakabiliyor ve İKİ vardiya açıyor. Süreleri TOPLANIRSA
            # 07:00-16:00 çalışan kişi 18 saat görünür; adet/saat ve performans
            # yarıya iner. 'Robotla Çalışıyor' işaretli vardiyalarda süre ve duruş
            # BİRLEŞTİRİLİR (union), toplanmaz. İşaretsiz vardiyalar eskisi gibi.
            # Aynı kural analiz.py'de de var (_operator_zamani) — Operatör
            # Performansı sayfası ile bu rapor aynı sayıyı göstermeli.
            efektif_calisma_dk, planli, plansiz = _operator_sure_birlestir(conn, v_rows)

            # OEE ortalamasi
            oee_list = []
            avail_list = []
            perf_list = []
            qual_list = []
            for vid in vids:
                oee_d = hesapla_oee(vid)
                if oee_d:
                    oee_list.append(oee_d['oee'])
                    avail_list.append(oee_d['availability'])
                    perf_list.append(oee_d['performance'])
                    qual_list.append(oee_d['quality'])

            ort_oee = round(sum(oee_list) / len(oee_list), 1) if oee_list else 0
            ort_avail = round(sum(avail_list) / len(avail_list), 1) if avail_list else 0
            ort_perf = round(sum(perf_list) / len(perf_list), 1) if perf_list else 0
            ort_qual = round(sum(qual_list) / len(qual_list), 1) if qual_list else 0

            toplam_uretim = (r['toplam_ok'] or 0) + (r['toplam_nok'] or 0)
            sonuclar.append({
                'operator': op,
                'vardiya_sayisi': r['vardiya_sayisi'],
                'toplam_calisma_dk': efektif_calisma_dk,
                'toplam_ok': r['toplam_ok'] or 0,
                'toplam_nok': r['toplam_nok'] or 0,
                'toplam_uretim': toplam_uretim,
                'planli_durus_dk': planli,
                'plansiz_durus_dk': plansiz,
                'ort_oee': ort_oee,
                'ort_availability': ort_avail,
                'ort_performance': ort_perf,
                'ort_quality': ort_qual,
            })
        conn.close()
        return jsonify({'rapor_tipi': 'operator', 'tarih': {'bas': tarih_bas, 'bit': tarih_bit}, 'veriler': sonuclar})

    elif rapor_tipi == 'referans':
        # Referans bazli rapor
        rows = c.execute(f'''
            SELECT u.referans_kodu,
                   SUM(u.ok_adet) as toplam_ok,
                   SUM(u.nok_adet) as toplam_nok,
                   SUM(u.ok_adet + u.nok_adet) as toplam_uretim,
                   ROUND(AVG(u.cycle_time_sn), 1) as ort_ct,
                   GROUP_CONCAT(DISTINCT v.robot_no) as robotlar,
                   GROUP_CONCAT(DISTINCT v.operator_adi) as operatorler,
                   COUNT(DISTINCT v.id) as vardiya_sayisi
            FROM uretim_kayitlari u
            JOIN vardiyalar v ON v.id = u.vardiya_id
            WHERE {sart}
            GROUP BY u.referans_kodu
            ORDER BY toplam_uretim DESC
        ''', params).fetchall()

        sonuclar = []
        for r in rows:
            toplam = (r['toplam_ok'] or 0) + (r['toplam_nok'] or 0)
            kalite = round((r['toplam_ok'] or 0) / toplam * 100, 1) if toplam > 0 else 0
            # Hedef CT'yi referans listesinden al — rapor lokasyon filtreliyse o tesisin satırı
            # (composite UNIQUE sonrası aynı kod TK1+TK2'de olabilir)
            if lokasyon:
                ref_row = c.execute(
                    "SELECT hedef_cycle_time_sn FROM referans_listesi WHERE referans_kodu = ? AND COALESCE(lokasyon,'TK2') = ?",
                    (r['referans_kodu'], lokasyon)
                ).fetchone()
            else:
                ref_row = c.execute(
                    'SELECT hedef_cycle_time_sn FROM referans_listesi WHERE referans_kodu = ?',
                    (r['referans_kodu'],)
                ).fetchone()
            hedef_ct = ref_row['hedef_cycle_time_sn'] if ref_row else 0

            sonuclar.append({
                'referans_kodu': r['referans_kodu'],
                'toplam_ok': r['toplam_ok'] or 0,
                'toplam_nok': r['toplam_nok'] or 0,
                'toplam_uretim': toplam,
                'kalite_pct': kalite,
                'ort_ct': r['ort_ct'] or 0,
                'hedef_ct': hedef_ct or 0,
                'robotlar': r['robotlar'] or '',
                'operatorler': r['operatorler'] or '',
                'vardiya_sayisi': r['vardiya_sayisi'],
            })
        conn.close()
        return jsonify({'rapor_tipi': 'referans', 'tarih': {'bas': tarih_bas, 'bit': tarih_bit}, 'veriler': sonuclar})

    elif rapor_tipi == 'teep':
        # TEEP Selalesi
        # 1) Takvim suresi
        from datetime import timedelta
        d_bas = datetime.strptime(tarih_bas, '%Y-%m-%d').date()
        d_bit = datetime.strptime(tarih_bit, '%Y-%m-%d').date()
        gun_sayisi = max((d_bit - d_bas).days + 1, 1)
        takvim_dk = gun_sayisi * 24 * 60  # 24h x gun

        # 2) Toplam vardiya suresi — açık vardiyalarda "o ana kadar geçen süre" (canlı)
        v_rows = c.execute(f'SELECT v.* FROM vardiyalar v WHERE {sart}', params).fetchall()
        toplam_vardiya_dk = sum(_vardiya_efektif_dk(v) for v in v_rows)

        # 3) Planli duruslar
        row = c.execute(f'''
            SELECT COALESCE(SUM(d.sure_dk), 0) as toplam
            FROM duruslar d JOIN vardiyalar v ON v.id = d.vardiya_id
            WHERE {sart} AND d.durus_tipi = 'planli'
        ''', params).fetchone()
        planli_durus_dk = row['toplam']

        # 4) Plansiz duruslar
        row = c.execute(f'''
            SELECT COALESCE(SUM(d.sure_dk), 0) as toplam
            FROM duruslar d JOIN vardiyalar v ON v.id = d.vardiya_id
            WHERE {sart} AND d.durus_tipi = 'plansiz'
        ''', params).fetchone()
        plansiz_durus_dk = row['toplam']

        # 5) Net plan suresi
        net_plan_dk = max(0, toplam_vardiya_dk - planli_durus_dk)
        # 6) Fiili calisma
        fiili_calisma_dk = max(0, net_plan_dk - plansiz_durus_dk)

        # 7) Uretim verileri
        uretim_rows = c.execute(f'''
            SELECT u.ok_adet, u.nok_adet, u.cycle_time_sn,
                   r.hedef_cycle_time_sn as guncel_ct
            FROM uretim_kayitlari u
            JOIN vardiyalar v ON v.id = u.vardiya_id
            LEFT JOIN referans_listesi r ON u.referans_kodu = r.referans_kodu
            WHERE {sart}
        ''', params).fetchall()

        toplam_ok = sum(r['ok_adet'] for r in uretim_rows)
        toplam_nok = sum(r['nok_adet'] for r in uretim_rows)
        toplam_uretim = toplam_ok + toplam_nok

        # Ideal uretim suresi (sn -> dk)
        ideal_uretim_sn = 0
        for r in uretim_rows:
            ct = r['guncel_ct'] if (r['guncel_ct'] and r['guncel_ct'] > 0) else (r['cycle_time_sn'] or 0)
            adet = r['ok_adet'] + r['nok_adet']
            ideal_uretim_sn += adet * ct
        ideal_uretim_dk = ideal_uretim_sn / 60

        # Performans kaybi
        performans_kaybi_dk = max(0, fiili_calisma_dk - ideal_uretim_dk)

        # Kalite kaybi
        kalite_kaybi_dk = (toplam_nok / toplam_uretim * ideal_uretim_dk) if toplam_uretim > 0 else 0

        # Net degerli uretim suresi
        net_uretim_dk = max(0, ideal_uretim_dk - kalite_kaybi_dk)

        # TEEP
        teep_pct = round((net_uretim_dk / takvim_dk) * 100, 1) if takvim_dk > 0 else 0
        # OEE (referans)
        oee_pct = round((net_uretim_dk / net_plan_dk) * 100, 1) if net_plan_dk > 0 else 0
        # Loading (kullanim orani)
        loading_pct = round((toplam_vardiya_dk / takvim_dk) * 100, 1) if takvim_dk > 0 else 0
        # Kullanilabilirlik icin dururm sayilari
        plansiz_durus_kalemler = c.execute(f'''
            SELECT d.durus_sebebi, SUM(d.sure_dk) as toplam, COUNT(*) as adet
            FROM duruslar d JOIN vardiyalar v ON v.id = d.vardiya_id
            WHERE {sart} AND d.durus_tipi = 'plansiz'
            GROUP BY d.durus_sebebi ORDER BY toplam DESC
        ''', params).fetchall()
        planli_durus_kalemler = c.execute(f'''
            SELECT d.durus_sebebi, SUM(d.sure_dk) as toplam, COUNT(*) as adet
            FROM duruslar d JOIN vardiyalar v ON v.id = d.vardiya_id
            WHERE {sart} AND d.durus_tipi = 'planli'
            GROUP BY d.durus_sebebi ORDER BY toplam DESC
        ''', params).fetchall()

        conn.close()

        # Uretilmeyen sure (takvimde olup vardiyaya alinmamis)
        uretilmeyen_dk = max(0, takvim_dk - toplam_vardiya_dk)

        return jsonify({
            'rapor_tipi': 'teep',
            'tarih': {'bas': tarih_bas, 'bit': tarih_bit},
            'gun_sayisi': gun_sayisi,
            'selale': {
                'takvim_dk': takvim_dk,
                'uretilmeyen_dk': round(uretilmeyen_dk, 1),
                'toplam_vardiya_dk': round(toplam_vardiya_dk, 1),
                'planli_durus_dk': round(planli_durus_dk, 1),
                'net_plan_dk': round(net_plan_dk, 1),
                'plansiz_durus_dk': round(plansiz_durus_dk, 1),
                'fiili_calisma_dk': round(fiili_calisma_dk, 1),
                'performans_kaybi_dk': round(performans_kaybi_dk, 1),
                'kalite_kaybi_dk': round(kalite_kaybi_dk, 1),
                'net_uretim_dk': round(net_uretim_dk, 1),
            },
            'yuzde': {
                'teep': teep_pct,
                'oee': oee_pct,
                'loading': loading_pct,
            },
            'uretim': {
                'toplam_ok': toplam_ok,
                'toplam_nok': toplam_nok,
                'toplam_uretim': toplam_uretim,
            },
            'plansiz_kalemler': [dict(r) for r in plansiz_durus_kalemler],
            'planli_kalemler': [dict(r) for r in planli_durus_kalemler],
        })

    conn.close()
    return jsonify({'hata': 'Gecersiz rapor_tipi'}), 400


# ─────────────────────────────────────────────────────────────
# FİKSTÜR ADRESLERİ API
# ─────────────────────────────────────────────────────────────

@app.route('/api/fikstur', methods=['GET'])
def get_fikstur_listesi():
    """Tüm fikstür adreslerini getir."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM fikstur_adresleri ORDER BY referans_kodu ASC")
    kayitlar = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(kayitlar)

@app.route('/api/fikstur', methods=['POST'])
def kaydet_fikstur():
    """Yeni fikstür adresi ekle veya mevcut olanı güncelle, ODS'yi de güncelle."""
    data = request.json
    ref = data.get('referans_kodu', '').strip()
    raf = data.get('raf_adresi', '').strip()
    notlar = data.get('notlar', '').strip()

    if not ref:
        return jsonify({'error': 'Referans kodu zorunludur.'}), 400

    conn = get_db()
    c = conn.cursor()
    
    # Eger referans varsa Guncelle, yoksa Ekle
    c.execute("SELECT id FROM fikstur_adresleri WHERE referans_kodu=?", (ref,))
    row = c.fetchone()
    if row:
        c.execute("UPDATE fikstur_adresleri SET raf_adresi=?, notlar=?, created_at=datetime('now', 'localtime') WHERE id=?", 
                  (raf, notlar, row['id']))
    else:
        c.execute("INSERT INTO fikstur_adresleri (referans_kodu, raf_adresi, notlar) VALUES (?,?,?)", 
                  (ref, raf, notlar))
    
    conn.commit()

    # ODS'yi güncelle
    c.execute("SELECT referans_kodu, raf_adresi, notlar FROM fikstur_adresleri ORDER BY referans_kodu ASC")
    tum_kayitlar = [(r['referans_kodu'], r['raf_adresi'], r['notlar']) for r in c.fetchall()]
    conn.close()
    _ods_guncelle(tum_kayitlar)

    return jsonify({'message': 'Fikstür adresi kaydedildi.'})


@app.route('/api/fikstur/ods_import', methods=['POST'])
def fikstur_ods_import():
    """ODS dosyasındaki tüm kayıtları DB'ye aktarır (zaten varsa atlar)."""
    kayitlar = _ods_kayitlari_oku()
    if not kayitlar:
        return jsonify({'error': 'ODS dosyası okunamadı veya boş.'}), 400

    conn = get_db()
    c = conn.cursor()
    eklenen, atlanan = 0, 0
    for ref, raf in kayitlar.items():
        c.execute("SELECT id FROM fikstur_adresleri WHERE referans_kodu=?", (ref,))
        if c.fetchone():
            atlanan += 1
        else:
            c.execute("INSERT INTO fikstur_adresleri (referans_kodu, raf_adresi, notlar) VALUES (?,?,?)",
                      (ref, raf, ''))
            eklenen += 1
    conn.commit()
    conn.close()
    return jsonify({'eklenen': eklenen, 'atlanan': atlanan,
                    'mesaj': f'{eklenen} yeni kayıt eklendi, {atlanan} kayıt zaten mevcuttu.'})


@app.route('/api/fikstur/<int:id>', methods=['DELETE'])
def sil_fikstur(id):
    """Fikstür adresini sil ve ODS'yi güncelle."""
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM fikstur_adresleri WHERE id=?", (id,))
    conn.commit()
    # ODS güncelle
    c.execute("SELECT referans_kodu, raf_adresi, notlar FROM fikstur_adresleri ORDER BY referans_kodu ASC")
    tum_kayitlar = [(r['referans_kodu'], r['raf_adresi'], r['notlar']) for r in c.fetchall()]
    conn.close()
    _ods_guncelle(tum_kayitlar)
    return jsonify({'message': 'Fikstür silindi.'})


# ─────────────────────────────────────────────────────────────
# ROBOT İSTASYON PROGRAMLARI API
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
# REFERANS ROBOT UYUMU API
# ─────────────────────────────────────────────────────────────

@app.route('/api/referans_uyum', methods=['GET'])
def get_referans_uyum():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM referans_robot_uyumu ORDER BY referans_kodu ASC")
    kayitlar = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(kayitlar)

@app.route('/api/referans_uyum', methods=['POST'])
def kaydet_referans_uyum():
    data = request.json
    ref = data.get('referans_kodu', '').strip()
    robot_no = data.get('robot_no', '').strip()
    istasyon = int(data.get('istasyon', 0))
    
    if not ref or not robot_no or not istasyon:
        return jsonify({'error': 'Eksik bilgi.'}), 400
        
    conn = get_db()
    c = conn.cursor()
    # Aynı eşleşme var mı kontrol et
    c.execute("SELECT id FROM referans_robot_uyumu WHERE referans_kodu=? AND robot_no=? AND istasyon=?", (ref, robot_no, istasyon))
    if not c.fetchone():
        c.execute("INSERT INTO referans_robot_uyumu (referans_kodu, robot_no, istasyon) VALUES (?,?,?)", (ref, robot_no, istasyon))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Uyum kaydedildi.'})

@app.route('/api/referans_uyum/<int:id>', methods=['DELETE'])
def sil_referans_uyum(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM referans_robot_uyumu WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Silindi.'})

@app.route('/api/referans_uyum/ods_import', methods=['POST'])
def referans_uyum_ods_import():
    kayitlar = _ods_uyum_oku()
    if not kayitlar:
        return jsonify({'error': 'ODS okunamadı veya eşleşme bulunamadı.'}), 400
        
    conn = get_db()
    c = conn.cursor()
    eklenen, atlanan = 0, 0
    # Eski kayıtları temizleyelim (Opsiyonel: Veya mevcutları koruyup üstüne ekleyebiliriz)
    # Hızlı kurulum için temizlemek faydalı olabilir ancak manuel eklenenleri silmemek adına sadece yeni ekleyelim.
    for ref, r_no, ist in kayitlar:
        c.execute("SELECT id FROM referans_robot_uyumu WHERE referans_kodu=? AND robot_no=? AND istasyon=?", (ref, r_no, ist))
        if c.fetchone():
            atlanan += 1
        else:
            c.execute("INSERT INTO referans_robot_uyumu (referans_kodu, robot_no, istasyon) VALUES (?,?,?)", (ref, r_no, ist))
            eklenen += 1
    conn.commit()
    conn.close()
    return jsonify({'eklenen': eklenen, 'atlanan': atlanan, 'mesaj': f'{eklenen} yeni uyum eklendi, {atlanan} kayıt zaten mevcuttu.'})

@app.route('/api/robot_programlari', methods=['GET'])
def get_robot_programlari():
    """Robot istasyon programlarını getir (tüm kayıtlar, robot+istasyon sırası ile)."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM robot_programlari ORDER BY robot_no ASC, istasyon ASC, guncelleme_tarihi DESC")
    kayitlar = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(kayitlar)

@app.route('/api/robot_programlari', methods=['POST'])
def ekle_robot_programi():
    """Robot istasyon programı ekle (çoklu kayıt - INSERT)."""
    data = request.json
    robot_no = data.get('robot_no')
    istasyon = data.get('istasyon')
    ref = data.get('referans_kodu', '').strip()
    operator = data.get('guncelleyen', '').strip()

    if not robot_no or istasyon is None:
        return jsonify({'error': 'Robot no ve istasyon zorunludur.'}), 400
    if not ref:
        return jsonify({'error': 'Referans kodu zorunludur.'}), 400

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO robot_programlari (robot_no, istasyon, referans_kodu, guncelleyen)
        VALUES (?, ?, ?, ?)
    """, (robot_no, istasyon, ref, operator))
    new_id = c.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Robot programı eklendi.', 'id': new_id})

@app.route('/api/robot_programlari/<int:id>', methods=['DELETE'])
def sil_robot_programi(id):
    """Robot programı kaydını sil."""
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM robot_programlari WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Kayıt silindi.'})



# ─────────────────────────────────────────────────────────────
# ROBOT İŞ ATAMA API
# ─────────────────────────────────────────────────────────────

@app.route('/api/robot_is_atamalari', methods=['GET'])
def is_atamalari_listesi():
    """Tüm aktif iş atamalarını döndür."""
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM robot_is_atamalari WHERE durum="bekliyor" ORDER BY atama_tarihi DESC'
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/robot_is_atamalari', methods=['POST'])
def is_atamasi_ekle():
    """Yeni iş ataması ekle."""
    data = request.get_json() or {}
    robot_no = data.get('robot_no', '').strip()
    referans_kodu = data.get('referans_kodu', '').strip()
    if not robot_no or not referans_kodu:
        return jsonify({'error': 'robot_no ve referans_kodu zorunludur'}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        INSERT INTO robot_is_atamalari (robot_no, istasyon, referans_kodu, aciklama, atayan)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        robot_no,
        int(data.get('istasyon', 0)),
        referans_kodu,
        data.get('aciklama', ''),
        data.get('atayan', '')
    ))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True}), 201

@app.route('/api/robot_is_atamalari/<int:id>', methods=['DELETE'])
def is_atamasi_sil(id):
    """İş atamasını tamamlandı olarak işaretle (sil)."""
    conn = get_db()
    conn.execute('DELETE FROM robot_is_atamalari WHERE id=?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})



# ─────────────────────────────────────────────────────────────
# ANDON ROBOT AYARLARI API
# ─────────────────────────────────────────────────────────────

@app.route('/api/andon_robot_ayarlari', methods=['GET'])
def andon_robot_ayarlari_listesi():
    """?bolum= ve ?lokasyon= ile filtrele; parametresizse hepsi döner.
    NOT: TK1 hatları (Pull/...) ve TK2 montaj (M1-M12) aynı bolum='montaj' altında —
    lokasyon filtresi ikisini ayırır."""
    bolum = request.args.get('bolum', '').strip()
    lokasyon = request.args.get('lokasyon', '').strip()
    sql = "SELECT * FROM andon_robot_ayarlari WHERE 1=1"
    params = []
    if bolum:
        sql += " AND bolum=?"
        params.append(bolum)
    if lokasyon:
        sql += " AND COALESCE(lokasyon,'TK2')=?"
        params.append(lokasyon)
    sql += " ORDER BY bolum, sira"
    conn = get_db()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/andon_robot_ayarlari', methods=['PATCH'])
@panel_gerekli(izin='andon-ayarlari')
def andon_robot_ayarlari_guncelle():
    """Toplu güncelleme: [{robot_no, bolum, goster, sira}, ...]
    Eski payload'larda bolum yoksa 'kaynak' varsayılır (geriye dönük uyumluluk)."""
    data = request.get_json() or []
    arg_lok = request.args.get('lokasyon')
    conn = get_db()
    for item in data:
        bolum = item.get('bolum') or 'kaynak'
        lokasyon = item.get('lokasyon') or arg_lok or 'TK2'
        conn.execute(
            "UPDATE andon_robot_ayarlari SET goster=?, sira=? WHERE robot_no=? AND bolum=? AND COALESCE(lokasyon,'TK2')=?",
            (int(item.get('goster', 1)), int(item.get('sira', 0)), item['robot_no'], bolum, lokasyon)
        )
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})
    
@app.route('/api/ayarlar', methods=['GET'])
def get_ayarlar():
    conn = get_db()
    rows = conn.execute('SELECT anahtar, deger FROM genel_ayarlar').fetchall()
    conn.close()
    return jsonify({r['anahtar']: r['deger'] for r in rows})

@app.route('/api/ayarlar', methods=['POST'])
def save_ayarlar():
    data = request.get_json() or {}
    conn = get_db()
    for k, v in data.items():
        conn.execute('INSERT OR REPLACE INTO genel_ayarlar (anahtar, deger) VALUES (?, ?)', (k, str(v)))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


# ─────────────────────────────────────────────────────────────
# GÜNLÜK ÜRETİM RAPORU E-POSTASI — alıcı yönetimi + manuel gönderim
# SMTP bilgileri mail_config.json'da (gitignore); alıcılar mail_alicilari tablosunda.
# ─────────────────────────────────────────────────────────────
import re as _re
_EMAIL_RE = _re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


@app.route('/api/mail/durum', methods=['GET'])
@panel_gerekli(izin='raporlar')
def mail_durum():
    """Dashboard göstergesi: SMTP config durumu (şifre içermez) + aktif alıcı sayısı."""
    try:
        import mail_raporu
        d = mail_raporu.durum_ozeti()
        conn = get_db()
        d['alici_sayisi'] = conn.execute("SELECT COUNT(*) FROM mail_alicilari WHERE COALESCE(aktif,1)=1").fetchone()[0]
        conn.close()
        return jsonify(d)
    except Exception as e:
        return jsonify({'config_var': False, 'etkin': False, 'hata': str(e)})


@app.route('/api/mail/alicilar', methods=['GET'])
@panel_gerekli(izin='raporlar')
def mail_alicilar_listesi():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, email, ad, COALESCE(aktif,1) as aktif, created_at FROM mail_alicilari ORDER BY email"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/mail/alicilar', methods=['POST'])
@panel_gerekli(izin='raporlar')
def mail_alici_ekle():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    ad = (data.get('ad') or '').strip()
    if not email or not _EMAIL_RE.match(email):
        return jsonify({'hata': 'Geçerli bir e-posta adresi giriniz.'}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO mail_alicilari (email, ad, aktif) VALUES (?, ?, 1)", (email, ad))
        conn.commit()
    except Exception:
        conn.close()
        return jsonify({'hata': 'Bu e-posta zaten ekli.'}), 400
    conn.close()
    return jsonify({'basarili': True}), 201


@app.route('/api/mail/alicilar/<int:aid>', methods=['DELETE'])
@panel_gerekli(izin='raporlar')
def mail_alici_sil(aid):
    conn = get_db()
    conn.execute("DELETE FROM mail_alicilari WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/mail/alicilar/<int:aid>', methods=['PATCH'])
@panel_gerekli(izin='raporlar')
def mail_alici_guncelle(aid):
    """aktif bayrağını aç/kapat (listeden silmeden geçici durdurma)."""
    data = request.get_json() or {}
    aktif = 1 if data.get('aktif') else 0
    conn = get_db()
    conn.execute("UPDATE mail_alicilari SET aktif=? WHERE id=?", (aktif, aid))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


@app.route('/api/mail/gonder_simdi', methods=['POST'])
@panel_gerekli(izin='raporlar')
def mail_gonder_simdi():
    """Günlük raporu ŞİMDİ üretip aktif alıcılara gönderir (17:00'ı beklemeden test/manuel)."""
    data = request.get_json(silent=True) or {}
    tarih = data.get('tarih')  # opsiyonel YYYY-MM-DD; yoksa bugün
    try:
        import mail_raporu
        sonuc = mail_raporu.gunluk_mail_gonder(tarih=tarih)
        return jsonify(sonuc), (200 if sonuc.get('basarili') else 400)
    except Exception as e:
        return jsonify({'basarili': False, 'mesaj': str(e)}), 500


@app.route('/api/mail/test', methods=['POST'])
@panel_gerekli(izin='raporlar')
def mail_test_gonder():
    """Tek adrese test maili — SMTP kurulumunu doğrulamak için."""
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    if not email or not _EMAIL_RE.match(email):
        return jsonify({'basarili': False, 'mesaj': 'Geçerli bir e-posta adresi giriniz.'}), 400
    try:
        import mail_raporu
        sonuc = mail_raporu.test_gonder(email)
        return jsonify(sonuc), (200 if sonuc.get('basarili') else 400)
    except Exception as e:
        return jsonify({'basarili': False, 'mesaj': str(e)}), 500



# ─────────────────────────────────────────────────────────────
# REFERANS TAKİP (LAUNCH / İŞ EMRİ) API
# ─────────────────────────────────────────────────────────────

DURUM_SIRASI = ['launch_alinacak', 'launch_alindi', 'launch_hazir', 'uretimde', 'uretim_tamamlandi']

@app.route('/api/referans_takip', methods=['GET'])
def referans_takip_listesi():
    bolum = request.args.get('bolum')
    lokasyon = request.args.get('lokasyon')
    conn = get_db()
    # Tüm bölümlerde öncelik ASC, NULL olanlar en sonda → oluşturma tarihine göre
    # Kaynak/Söktak süreleri de JOIN ile eklendi (pair_cycle hesabı için)
    # GROUP BY rt.id — FAN-OUT KORUMASI: referans_listesi'nde aynı normalize koda sahip
    # birden çok satır olabilir (örn. '94.LTK.341/10' vs '94. LTK. 341/10' — operatörün
    # boşluklu yazımı auto-create ile ayrı satır açmıştı). JOIN ikisiyle de eşleşince
    # TEK iş emri listede İKİ kez görünüyordu. MAX(): dolu cycle değeri tercih edilir.
    sql = '''
        SELECT rt.*,
               MAX(rl.hedef_cycle_time_sn) as hedef_cycle_time_sn,
               MAX(rl.kaynak_suresi_sn)    as kaynak_suresi_sn,
               MAX(rl.soktak_suresi_sn)    as soktak_suresi_sn
        FROM referans_takip rt
        LEFT JOIN referans_listesi rl ON REPLACE(rt.referans_kodu, ' ', '') = REPLACE(rl.referans_kodu, ' ', '')
            AND COALESCE(rl.lokasyon,'TK2') = COALESCE(rt.lokasyon,'TK2')
        WHERE 1=1'''
    params = []
    if bolum:
        sql += " AND COALESCE(rt.bolum, 'kaynak') = ?"
        params.append(bolum)
    if lokasyon:
        sql += " AND COALESCE(rt.lokasyon, 'TK2') = ?"
        params.append(lokasyon)
    sql += " GROUP BY rt.id ORDER BY (rt.oncelik IS NULL), rt.oncelik ASC, rt.olusturma_tarihi DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

def _oncelik_clamp(c, bolum, yeni_oncelik, eski_oncelik=None, exclude_id=None, lokasyon=None):
    """Yeni öncelik değerini, o bölüm(+lokasyon)daki mevcut öncelikli kayıt sayısına göre clamp eder.

    INSERT (eski_oncelik=None): max = mevcut_sayi + 1 (yeni eklenecek)
    UPDATE (eski var, yeni var): max = mevcut_sayi (toplam değişmiyor)
    UPDATE (eski None, yeni var): max = mevcut_sayi + 1 (öncelik atıyor, +1 olur)

    lokasyon verilirse sayım o tesise kapanır (TK1/TK2 launch sıraları karışmasın).
    None döndürürse clamp gerekmedi (zaten geçerli) veya yeni_oncelik None.
    """
    if yeni_oncelik is None:
        return None
    where_excl = ' AND id != ?' if exclude_id is not None else ''
    excl_args = (exclude_id,) if exclude_id is not None else ()
    lok_sql = " AND COALESCE(lokasyon,'TK2') = ?" if lokasyon else ''
    lok_args = (lokasyon,) if lokasyon else ()
    row = c.execute(f'''
        SELECT COUNT(*) as cnt FROM referans_takip
        WHERE COALESCE(bolum, 'kaynak') = ? AND oncelik IS NOT NULL
        {lok_sql}{where_excl}
    ''', (bolum,) + lok_args + excl_args).fetchone()
    mevcut_sayi = row['cnt'] if row else 0
    # Bu kayıt da öncelikli olacak: max = mevcut_sayi + 1 (kendisi de dahil)
    max_izinli = mevcut_sayi + 1
    if yeni_oncelik > max_izinli:
        return max_izinli
    if yeni_oncelik < 1:
        return 1
    return yeni_oncelik


def _oncelik_kaydir(c, bolum, eski_oncelik, yeni_oncelik, exclude_id=None, lokasyon=None):
    """Bir bölüm kaydının öncelik geçişinde diğer satırları doğru yönde kaydırır.
    Tüm bölümler için çalışır (kaynak/montaj/metal).

    eski_oncelik: kaydın önceki değeri (None = öncesinde öncelik yoktu / yeni kayıt)
    yeni_oncelik: kaydın yeni değeri (None = öncelik kaldırılıyor)
    exclude_id:   bu id'li satıra dokunma (PATCH için kendisini hariç tut)
    lokasyon:     verilirse kaydırma o tesise kapanır (TK1/TK2 launch sıraları karışmasın)
    """
    # Değişiklik yok
    if eski_oncelik == yeni_oncelik:
        return

    where_excl = ' AND id != ?' if exclude_id is not None else ''
    excl_args = (exclude_id,) if exclude_id is not None else ()
    lok_sql = " AND COALESCE(lokasyon,'TK2') = ?" if lokasyon else ''
    lok_args = (lokasyon,) if lokasyon else ()

    # 1) Yeni öncelik atanıyor (önceden yoktu): >= yeni olanları +1 kaydır
    if eski_oncelik is None and yeni_oncelik is not None:
        c.execute(f'''
            UPDATE referans_takip
            SET oncelik = oncelik + 1
            WHERE COALESCE(bolum, 'kaynak') = ?
              AND oncelik IS NOT NULL AND oncelik >= ?
              {lok_sql}{where_excl}
        ''', (bolum, yeni_oncelik) + lok_args + excl_args)

    # 2) Öncelik kaldırılıyor: > eski olanları -1 yukarı çek (boşluğu kapat)
    elif eski_oncelik is not None and yeni_oncelik is None:
        c.execute(f'''
            UPDATE referans_takip
            SET oncelik = oncelik - 1
            WHERE COALESCE(bolum, 'kaynak') = ?
              AND oncelik IS NOT NULL AND oncelik > ?
              {lok_sql}{where_excl}
        ''', (bolum, eski_oncelik) + lok_args + excl_args)

    # 3) Yukarı taşıma (önem artıyor: eski > yeni, ör. 4 -> 2)
    elif yeni_oncelik < eski_oncelik:
        c.execute(f'''
            UPDATE referans_takip
            SET oncelik = oncelik + 1
            WHERE COALESCE(bolum, 'kaynak') = ?
              AND oncelik IS NOT NULL AND oncelik >= ? AND oncelik < ?
              {lok_sql}{where_excl}
        ''', (bolum, yeni_oncelik, eski_oncelik) + lok_args + excl_args)

    # 4) Aşağı taşıma (önem azalıyor: eski < yeni, ör. 1 -> 3)
    else:  # yeni_oncelik > eski_oncelik
        c.execute(f'''
            UPDATE referans_takip
            SET oncelik = oncelik - 1
            WHERE COALESCE(bolum, 'kaynak') = ?
              AND oncelik IS NOT NULL AND oncelik > ? AND oncelik <= ?
              {lok_sql}{where_excl}
        ''', (bolum, eski_oncelik, yeni_oncelik) + lok_args + excl_args)


@app.route('/api/referans_takip', methods=['POST'])
@panel_gerekli(izin='is-yonetimi')
def referans_takip_ekle():
    try:
        data = request.get_json() or {}
        ref = data.get('referans_kodu', '').strip()
        if not ref:
            return jsonify({'error': 'referans_kodu zorunludur'}), 400

        bolum = (data.get('bolum') or 'kaynak').strip()
        if bolum not in GECERLI_BOLUMLER:
            bolum = 'kaynak'
        lokasyon = (data.get('lokasyon') or request.args.get('lokasyon') or 'TK2').strip() or 'TK2'

        # Öncelik (yalnızca montaj için anlamlı). Boş/None: belirtilmemiş.
        oncelik_raw = data.get('oncelik')
        oncelik = None
        if oncelik_raw not in (None, '', 0, '0'):
            try:
                oncelik = int(oncelik_raw)
                if oncelik < 1: oncelik = None
            except (TypeError, ValueError):
                oncelik = None

        conn = get_db()
        c = conn.cursor()
        # Öncelik clamp: kullanıcı çok yüksek bir sayı girdiyse mevcut listedeki
        # ref sayısına göre maksimum izinliye düşür (örn. 3 ref var, 5 girilmişse → 4)
        oncelik = _oncelik_clamp(c, bolum, oncelik, lokasyon=lokasyon)
        # Yeni kayıt: eski_oncelik=None, yeni_oncelik girildiyse mevcut >=N olanları +1
        _oncelik_kaydir(c, bolum, None, oncelik, lokasyon=lokasyon)
        durum_init = data.get('durum', 'launch_alinacak')
        # Eger referans dogrudan 'uretimde' durumda ekleniyorsa, pilot sayac sifirlama
        # zamanini su an olarak isaretle (yeni referans secimi = sayac sifir)
        uretime_baslama_init = "datetime('now','localtime')" if durum_init == 'uretimde' else None
        if uretime_baslama_init:
            c.execute(f'''
                INSERT INTO referans_takip (referans_kodu, hedef_adet, aciklama, durum, olusturan, robot_no, istasyon, bolum, oncelik, lokasyon, uretime_baslama_ts)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, {uretime_baslama_init})
            ''', (ref, int(data.get('hedef_adet', 0)), data.get('aciklama', ''),
                  durum_init, data.get('olusturan', ''),
                  data.get('robot_no', ''), int(data.get('istasyon', 0)), bolum, oncelik, lokasyon))
        else:
            c.execute('''
                INSERT INTO referans_takip (referans_kodu, hedef_adet, aciklama, durum, olusturan, robot_no, istasyon, bolum, oncelik, lokasyon)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (ref, int(data.get('hedef_adet', 0)), data.get('aciklama', ''),
                  durum_init, data.get('olusturan', ''),
                  data.get('robot_no', ''), int(data.get('istasyon', 0)), bolum, oncelik, lokasyon))
        conn.commit()

        # PUSH bildirim — bölümün sorumlu operatörlerinin telefonuna (arka planda).
        # LOKASYON: BILDIRIM_ALICILARI TK2 (ana fabrika) sorumluları — TK1 launch'ı
        # TK2 montaj sorumlusuna GITMESIN (ayrı fabrika). TK1 için alıcı tanımlanınca
        # burada (bolum, lokasyon) anahtarlı yapıya geçilir.
        alicilar = (BILDIRIM_ALICILARI.get(bolum) or []) if lokasyon == 'TK2' else []
        if alicilar:
            parcalar = [ref]
            if data.get('robot_no'):
                parcalar.append(str(data.get('robot_no')))
            if data.get('istasyon'):
                parcalar.append(f"İst.{data.get('istasyon')}")
            if data.get('hedef_adet'):
                parcalar.append(f"{data.get('hedef_adet')} adet")
            if data.get('olusturan'):
                parcalar.append(f"Ekleyen: {data.get('olusturan')}")
            bolum_ad = BOLUM_AD.get(bolum, bolum)
            _push_gonder_async(alicilar, f'🚀 Yeni Launch — {bolum_ad}', ' · '.join(parcalar))

        return jsonify({'basarili': True}), 201
    except Exception as e:
        print(f"HATA (referans_takip_ekle): {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/referans_takip/<int:id>', methods=['PATCH'])
@panel_gerekli(izin='is-yonetimi')
def referans_takip_guncelle(id):
    data = request.get_json() or {}
    conn = get_db()
    c = conn.cursor()

    # Öncelik değişiyorsa shift mantığı çalıştır (tüm bölümler için)
    if 'oncelik' in data:
        mevcut = c.execute("SELECT bolum, oncelik, COALESCE(lokasyon,'TK2') as lokasyon FROM referans_takip WHERE id=?", (id,)).fetchone()
        if mevcut:
            bolum_kayit = (mevcut['bolum'] or 'kaynak')
            lok_kayit = mevcut['lokasyon'] or 'TK2'
            eski_onc = mevcut['oncelik']  # None olabilir
            yeni_raw = data.get('oncelik')
            yeni_onc = None
            if yeni_raw not in (None, '', 0, '0'):
                try:
                    yeni_onc = int(yeni_raw)
                    if yeni_onc < 1: yeni_onc = None
                except (TypeError, ValueError):
                    yeni_onc = None
            # Clamp: girilen sayı listedeki ref sayısını aşıyorsa düşür (o tesise kapalı)
            yeni_onc = _oncelik_clamp(c, bolum_kayit, yeni_onc, eski_oncelik=eski_onc, exclude_id=id, lokasyon=lok_kayit)
            _oncelik_kaydir(c, bolum_kayit, eski_onc, yeni_onc, exclude_id=id, lokasyon=lok_kayit)
            data['oncelik'] = yeni_onc  # normalize edildi

    fields, vals = [], []
    if 'durum' in data:
        fields.append('durum=?'); vals.append(data['durum'])
        # Durum 'uretimde'ye geciliyorsa pilot sayac sifirlama referansi olarak
        # uretime_baslama_ts'yi simdiye set et. (Duraklatip tekrar baslatirsa
        # sayac yeniden sifirlanir — su anki uretim hizini gosterir.)
        if data['durum'] == 'uretimde':
            fields.append("uretime_baslama_ts=datetime('now','localtime')")
    if 'aciklama' in data:
        fields.append('aciklama=?'); vals.append(data['aciklama'])
    if 'depo_notu' in data:
        fields.append('depo_notu=?'); vals.append((data.get('depo_notu') or '').strip())
    if 'hedef_adet' in data:
        try:
            fields.append('hedef_adet=?'); vals.append(int(data.get('hedef_adet') or 0))
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'error': f"hedef_adet sayısal olmalı: {data.get('hedef_adet')!r}"}), 400
    if 'referans_kodu' in data:
        fields.append('referans_kodu=?'); vals.append(data['referans_kodu'])
    if 'robot_no' in data:
        fields.append('robot_no=?'); vals.append(data['robot_no'])
    if 'istasyon' in data:
        try:
            fields.append('istasyon=?'); vals.append(int(data.get('istasyon') or 0))
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'error': f"istasyon sayısal olmalı: {data.get('istasyon')!r}"}), 400
    if 'oncelik' in data:
        fields.append('oncelik=?'); vals.append(data['oncelik'])  # None olabilir
    if not fields:
        conn.close()
        return jsonify({'error': 'Güncellenecek alan yok'}), 400
    fields.append("guncelleme_tarihi=datetime('now','localtime')")
    vals.append(id)
    c.execute(f"UPDATE referans_takip SET {','.join(fields)} WHERE id=?", vals)
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})

@app.route('/api/robot/<robot_no>/atamalar', methods=['GET'])
def robot_atama_listesi(robot_no):
    """Belirli bir robota atanmış aktif işleri getir."""
    conn = get_db()
    rows = conn.execute('''
        SELECT * FROM referans_takip 
        WHERE robot_no = ? AND durum != 'uretim_tamamlandi'
        ORDER BY guncelleme_tarihi DESC
    ''', (robot_no,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/referans_takip/<int:id>', methods=['DELETE'])
@panel_gerekli(izin='is-yonetimi')
def referans_takip_sil(id):
    conn = get_db()
    conn.execute('DELETE FROM referans_takip WHERE id=?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'basarili': True})


# ─────────────────────────────────────────────────────────────

# Bellekte tutulan kayan duyuru mesajı
_andon_mesaj = {'metin': '', 'yazar': ''}

@app.route('/api/andon', methods=['GET'])
def andon_veri():
    """Andon TV için bugünün tüm üretim verilerini tek sorguda döndürür. ?bolum= ile filtrelenebilir."""
    bugun = date.today().isoformat()
    bolum = request.args.get('bolum', '')
    lokasyon = request.args.get('lokasyon', '')
    conn = get_db()
    c = conn.cursor()

    # Bölüm + lokasyon filtresi
    bolum_sart = ''
    bolum_params = [bugun]
    if bolum:
        bolum_sart += ' AND v.bolum = ?'
        bolum_params.append(bolum)
    if lokasyon:
        bolum_sart += " AND COALESCE(v.lokasyon,'TK2') = ?"
        bolum_params.append(lokasyon)

    # Bugünkü tüm vardiyalar
    vardiyalar = c.execute(
        f'SELECT * FROM vardiyalar v WHERE v.tarih = ?{bolum_sart} ORDER BY v.baslangic_saati',
        bolum_params
    ).fetchall()

    # Otomatik molalar (çay/yemek) — saati gelenleri bu vardiyalara işle (idempotent, canlı)
    for _v in vardiyalar:
        _otomatik_mola_uygula(c, _v)
    conn.commit()

    # SAYAÇ SENKRONU — açık vardiyaların TÜM auto kayıtları (ESP32 + test cihazı).
    # Andon 30sn'de bir çağrılır: canlı makine adetleri üretim/performans hesabına
    # da anlık yansır (eskiden yalnız operatör mobili açıkken tazeleniyordu →
    # pilot rozeti canlı ama adet/performans bayat görünüyordu).
    _acik_auto_vardiyalari_senkronla(conn, bolum=bolum or None, lokasyon=lokasyon or None)

    # Vardiya id'leri
    vardiya_ids = [v['id'] for v in vardiyalar]
    if not vardiya_ids:
        vardiya_ids_placeholder = '(-1)'
    else:
        vardiya_ids_placeholder = '(' + ','.join(str(v) for v in vardiya_ids) + ')'

    # Bugünkü tüm üretim kayıtları
    uretim_rows = c.execute(f'''
        SELECT u.* FROM uretim_kayitlari u
        WHERE u.vardiya_id IN {vardiya_ids_placeholder}
    ''').fetchall()

    # Bugünkü duruşlar
    durus_rows = c.execute(f'''
        SELECT d.* FROM duruslar d
        WHERE d.vardiya_id IN {vardiya_ids_placeholder}
        ORDER BY d.sure_dk DESC
    ''').fetchall()

    # Toplamlar
    toplam_ok  = sum(r['ok_adet']  for r in uretim_rows)
    toplam_nok = sum(r['nok_adet'] for r in uretim_rows)
    toplam_hedef = sum(r['hedef_adet'] for r in uretim_rows)
    toplam_uretim = toplam_ok + toplam_nok
    kalite_pct = round((toplam_ok / toplam_uretim * 100), 1) if toplam_uretim > 0 else 0

    # Toplam planlı süre ve duruş
    # Açık vardiyalarda "o ana kadar geçen süre" (canlı), kapalılarda gerçek süre
    toplam_planli_dk = sum(_vardiya_efektif_dk(v) for v in vardiyalar)
    toplam_durus_dk  = sum(d['sure_dk'] or 0 for d in durus_rows)
    planli_durus_dk  = sum(d['sure_dk'] or 0 for d in durus_rows if d['durus_tipi'] == 'planli')
    plansiz_durus_dk = toplam_durus_dk - planli_durus_dk

    # Basit OEE (birleşik)
    if toplam_planli_dk > 0:
        kullanilabilirlik = max(0, (toplam_planli_dk - plansiz_durus_dk) / toplam_planli_dk * 100)
    else:
        kullanilabilirlik = 0

    # Performans (Adet bazlı basitleştirilmiş)
    performans = (toplam_ok / toplam_hedef * 100) if toplam_hedef > 0 else 0
    if performans > 100: performans = 100 # %100'ü geçmesin görsel amaçlı
    
    # OEE = K * P * Q
    oee = (kullanilabilirlik / 100) * (performans / 100) * (kalite_pct / 100) * 100
    
    # Plansız duruşlar (uyarı bandı için)
    plansiz_duruslar = [
        {'sebep': d['durus_sebebi'], 'sure_dk': d['sure_dk']}
        for d in durus_rows if d['durus_tipi'] == 'plansiz'
    ]

    # ── İş atamaları (aktif) — bölüme göre filtrelenir ──
    if bolum:
        atama_rows = c.execute('''
            SELECT id, istasyon, referans_kodu, aciklama, olusturan as atayan, robot_no
            FROM referans_takip
            WHERE robot_no != '' AND robot_no IS NOT NULL
              AND durum != 'uretim_tamamlandi'
              AND COALESCE(bolum, 'kaynak') = ?
            ORDER BY guncelleme_tarihi DESC
        ''', (bolum,)).fetchall()
    else:
        atama_rows = c.execute('''
            SELECT id, istasyon, referans_kodu, aciklama, olusturan as atayan, robot_no
            FROM referans_takip
            WHERE robot_no != '' AND robot_no IS NOT NULL
              AND durum != 'uretim_tamamlandi'
            ORDER BY guncelleme_tarihi DESC
        ''').fetchall()

    # Robot bazlı duruşları çek (bölüm filtresine uygun)
    robot_durus_rows = c.execute(f'''
        SELECT v.robot_no, COALESCE(SUM(d.sure_dk), 0) as toplam_durus_dk, COUNT(d.id) as durus_adet
        FROM vardiyalar v
        LEFT JOIN duruslar d ON d.vardiya_id = v.id
        WHERE v.id IN {vardiya_ids_placeholder}
        GROUP BY v.robot_no
    ''').fetchall()

    robot_durus_detay = c.execute(f'''
        SELECT v.robot_no, d.durus_sebebi, d.durus_tipi, d.sure_dk
        FROM duruslar d
        JOIN vardiyalar v ON d.vardiya_id = v.id
        WHERE v.id IN {vardiya_ids_placeholder}
        ORDER BY v.robot_no, d.sure_dk DESC
    ''').fetchall()

    robot_durus_map = {}
    for r in robot_durus_rows:
        robot_durus_map[r['robot_no']] = {'toplam_durus_dk': r['toplam_durus_dk'], 'durus_adet': r['durus_adet'], 'detay': []}
    for d in robot_durus_detay:
        rn = d['robot_no']
        if rn not in robot_durus_map: robot_durus_map[rn] = {'toplam_durus_dk': 0, 'durus_adet': 0, 'detay': []}
        robot_durus_map[rn]['detay'].append({'sebep': d['durus_sebebi'], 'sure_dk': d['sure_dk'], 'tip': d['durus_tipi']})

    # Vardiya bazlı duruş haritası (her vardiya için ayrı duruş istatistiği)
    # — montaj/metal'de aynı hat/makinede birden fazla vardiya açıkken duruşları
    # operatörler arasında karıştırmamak için.
    vardiya_durus_rows = c.execute(f'''
        SELECT v.id as vid, COALESCE(SUM(d.sure_dk), 0) as toplam_durus_dk, COUNT(d.id) as durus_adet
        FROM vardiyalar v
        LEFT JOIN duruslar d ON d.vardiya_id = v.id
        WHERE v.id IN {vardiya_ids_placeholder}
        GROUP BY v.id
    ''').fetchall()
    vardiya_durus_detay = c.execute(f'''
        SELECT d.vardiya_id as vid, d.durus_sebebi, d.durus_tipi, d.sure_dk
        FROM duruslar d
        WHERE d.vardiya_id IN {vardiya_ids_placeholder}
        ORDER BY d.vardiya_id, d.sure_dk DESC
    ''').fetchall()
    vardiya_durus_map = {}
    for r in vardiya_durus_rows:
        vardiya_durus_map[r['vid']] = {'toplam_durus_dk': r['toplam_durus_dk'], 'durus_adet': r['durus_adet'], 'detay': []}
    for d in vardiya_durus_detay:
        vid = d['vid']
        if vid not in vardiya_durus_map:
            vardiya_durus_map[vid] = {'toplam_durus_dk': 0, 'durus_adet': 0, 'detay': []}
        vardiya_durus_map[vid]['detay'].append({'sebep': d['durus_sebebi'], 'sure_dk': d['sure_dk'], 'tip': d['durus_tipi']})

    # ── aktif_vardiyalar: her açık vardiya = ayrı kart ──
    # Bu, andon ekranında aynı hat/makineye birden fazla operatör açtığında
    # her operatörün kendi kartını ve kendi üretim/duruş kayıtlarını görmesini sağlar.
    aktif_vardiyalar = []
    for v in vardiyalar:
        if v['durum'] == 'kapali':   # 'acik' VEYA 'aktif' (yeniden acilan) goster; sadece 'kapali' gizle
            continue
        durus_info = vardiya_durus_map.get(v['id'], {'toplam_durus_dk': 0, 'durus_adet': 0, 'detay': []})
        # robotla_calisiyor — sütun yoksa eski DB; güvenli erişim
        try:
            rc = 1 if v['robotla_calisiyor'] else 0
        except (KeyError, IndexError):
            rc = 0
        item = {
            'vardiya_id': v['id'],
            'robot_no': v['robot_no'],
            'operator': v['operator_adi'],
            'vardiya': v['vardiya_turu'],
            'baslangic': v['baslangic_saati'],
            'bitis': v['bitis_saati'],
            'istasyon_1': [], 'istasyon_2': [], 'diger': [], 'atamalar': [],
            'durus_dk': durus_info['toplam_durus_dk'],
            'durus_adet': durus_info['durus_adet'],
            'durus_detay': durus_info['detay'],
            'robotla_calisiyor': rc,
            'pair_cycle': None,  # kaynak için iki istasyondaki aktif referansların gerçek çevrimi
        }
        for u in uretim_rows:
            if u['vardiya_id'] == v['id']:
                # test_cihaz_id: eski DB'de kolon olmayabilir — güvenli erişim
                try:
                    _tc = u['test_cihaz_id']
                except (KeyError, IndexError):
                    _tc = None
                row = {
                    'ref': u['referans_kodu'],
                    'launch': u['launch_adet'] or 0,
                    'tamamlandi': 1 if u['tamamlandi'] else 0,
                    'teyit': 0,        # aşağıda ref_durum_map'ten doldurulacak
                    'suresiz': 0,
                    'ct': u['cycle_time_sn'] or 0,   # metal andon CT gösterimi (referans_listesi hedefi ile override edilir)
                    'ok': u['ok_adet'] or 0,
                    # Test cihazı sayaç kaynağı — andon ref satırında canlı adet rozeti göstermek için
                    'test_cihaz': 1 if _tc else 0,
                }
                ist = u['istasyon'] or 0
                if ist == 1:   item['istasyon_1'].append(row)
                elif ist == 2: item['istasyon_2'].append(row)
                else:          item['diger'].append(row)
        aktif_vardiyalar.append(item)

    # Referans durumu (teyit / süresiz) — andonda işaretlemek için
    # Tek SQL ile tüm bölüm için map oluştur, sonra her satıra ekle
    ref_durum_map = {}
    if bolum == 'kaynak':
        for r in c.execute(
            "SELECT referans_kodu, COALESCE(sure_teyit,0) as teyit, COALESCE(hedef_cycle_time_sn,0) as ct "
            "FROM referans_listesi WHERE COALESCE(bolum,'kaynak')='kaynak'"
        ).fetchall():
            norm = str(r['referans_kodu'] or '').upper().replace(' ', '')
            ref_durum_map[norm] = {
                'teyit': int(r['teyit'] or 0),
                'suresiz': 1 if (r['ct'] or 0) <= 0 else 0,
            }
        # Her aktif_vardiya'nın her ref satırına teyit/süresiz bayraklarını yaz
        for it in aktif_vardiyalar:
            for koleksiyon in ('istasyon_1', 'istasyon_2', 'diger'):
                for row in it[koleksiyon]:
                    norm = str(row.get('ref') or '').upper().replace(' ', '')
                    durum = ref_durum_map.get(norm)
                    if durum:
                        row['teyit'] = durum['teyit']
                        row['suresiz'] = durum['suresiz']
    elif bolum == 'metal':
        # Metal andon: her referans satırına güncel hedef cycle time'ı yaz.
        # referans_listesi hedefi varsa o önceliklidir (panelden güncellenince andon yansır);
        # yoksa row['ct'] zaten kayıttaki cycle_time_sn (fallback) olarak kalır.
        ref_ct_map = {}
        for r in c.execute(
            "SELECT referans_kodu, COALESCE(hedef_cycle_time_sn,0) as ct "
            "FROM referans_listesi WHERE COALESCE(bolum,'kaynak')='metal'"
        ).fetchall():
            norm = str(r['referans_kodu'] or '').upper().replace(' ', '')
            ref_ct_map[norm] = r['ct']
        for it in aktif_vardiyalar:
            for koleksiyon in ('istasyon_1', 'istasyon_2', 'diger'):
                for row in it[koleksiyon]:
                    norm = str(row.get('ref') or '').upper().replace(' ', '')
                    hedef = ref_ct_map.get(norm)
                    if hedef and hedef > 0:
                        row['ct'] = hedef

    # Atamaları her uygun shift kartına ekle (aynı robot_no'lu kartlara)
    # Montaj'da "atama" kavramı kullanılmıyor — operatöre sıradaki iş olarak
    # bireysel atama yansımaz (öncelik listesi yalnızca yönetici görünümünde anlamlı).
    if bolum != 'montaj':
        for a in atama_rows:
            rn = a['robot_no']
            atama_item = {'id': a['id'], 'istasyon': a['istasyon'], 'referans_kodu': a['referans_kodu'], 'aciklama': a['aciklama'], 'atayan': a['atayan']}
            for it in aktif_vardiyalar:
                if it['robot_no'] == rn:
                    it['atamalar'].append(atama_item)

    # Pair cycle hesabı (sadece kaynak bölümü için).
    # Her vardiyada İst.1 ve İst.2'deki en son üretim referansını al, çevrim hesapla.
    if bolum == 'kaynak':
        from oee import pair_cycle_hesapla
        for it in aktif_vardiyalar:
            ist1_kod = it['istasyon_1'][-1]['ref'] if it['istasyon_1'] else None
            ist2_kod = it['istasyon_2'][-1]['ref'] if it['istasyon_2'] else None
            if ist1_kod or ist2_kod:
                try:
                    it['pair_cycle'] = pair_cycle_hesapla(ist1_kod, ist2_kod)
                except Exception as e:
                    print(f'[andon pair_cycle] hata: {e}')
                    it['pair_cycle'] = None

    # Eski 'robotlar' alanı (legacy andon.html için backward compat — robot bazında ilk vardiya)
    robotlar = {}
    for v in vardiyalar:
        if v['durum'] == 'kapali':   # 'acik' VEYA 'aktif' (yeniden acilan) goster; sadece 'kapali' gizle
            continue
        r = v['robot_no']
        durus_info = robot_durus_map.get(r, {'toplam_durus_dk': 0, 'durus_adet': 0, 'detay': []})
        if r not in robotlar:
            robotlar[r] = {
                'robot_no': r, 'operator': v['operator_adi'], 'vardiya': v['vardiya_turu'],
                'baslangic': v['baslangic_saati'], 'bitis': v['bitis_saati'],
                'istasyon_1': [], 'istasyon_2': [], 'diger': [], 'atamalar': [],
                'durus_dk': durus_info['toplam_durus_dk'], 'durus_adet': durus_info['durus_adet'],
                'durus_detay': durus_info.get('detay', [])
            }
        for u in uretim_rows:
            if u['vardiya_id'] == v['id']:
                row = {'ref': u['referans_kodu'], 'launch': u['launch_adet'] or 0, 'tamamlandi': 1 if u['tamamlandi'] else 0}
                ist = u['istasyon'] or 0
                if ist == 1: robotlar[r]['istasyon_1'].append(row)
                elif ist == 2: robotlar[r]['istasyon_2'].append(row)
                else: robotlar[r]['diger'].append(row)

    for a in atama_rows:
        rn = a['robot_no']
        if rn not in robotlar:
            robotlar[rn] = {'robot_no':rn, 'operator':'', 'vardiya':'', 'baslangic':'', 'bitis':'', 'istasyon_1':[], 'istasyon_2':[], 'diger':[], 'atamalar':[]}
        robotlar[rn]['atamalar'].append({'id': a['id'], 'istasyon': a['istasyon'], 'referans_kodu': a['referans_kodu'], 'aciklama': a['aciklama'], 'atayan': a['atayan']})

    # Referans takip listesi — bölüm VE lokasyona göre filtrelenir
    # (TK1 andonunda TK2 montaj iş emirleri görünmesin / tersi)
    # GROUP BY rt.id — fan-out koruması (normalize-duplike referans satırları JOIN'de
    # tek iş emrini çoğaltmasın; bkz. /api/referans_takip'teki aynı koruma)
    rt_sql = '''
        SELECT rt.*, MAX(rl.hedef_cycle_time_sn) as hedef_cycle_time_sn
        FROM referans_takip rt
        LEFT JOIN referans_listesi rl ON REPLACE(rt.referans_kodu, ' ', '') = REPLACE(rl.referans_kodu, ' ', '')
            AND COALESCE(rl.lokasyon,'TK2') = COALESCE(rt.lokasyon,'TK2')
        WHERE 1=1
    '''
    rt_params = []
    if bolum:
        rt_sql += " AND COALESCE(rt.bolum, 'kaynak') = ?"
        rt_params.append(bolum)
    if lokasyon:
        rt_sql += " AND COALESCE(rt.lokasyon, 'TK2') = ?"
        rt_params.append(lokasyon)
    rt_sql += " GROUP BY rt.id ORDER BY rt.olusturma_tarihi DESC"
    rt_rows = c.execute(rt_sql, rt_params).fetchall()
    referans_takip_list = [dict(row) for row in rt_rows]

    # Tum ayarlari yukle
    ayarlar_rows = c.execute("SELECT anahtar, deger FROM genel_ayarlar").fetchall()
    
    response_data = {
        'tarih': bugun,
        'toplam_ok': toplam_ok,
        'toplam_nok': toplam_nok,
        'toplam_hedef': toplam_hedef,
        'kalite_pct': kalite_pct,
        'kullanilabilirlik_pct': round(kullanilabilirlik, 1),
        'performans_pct': round(performans, 1),
        'oee_pct': round(oee, 1),
        'toplam_durus_dk': toplam_durus_dk,
        'plansiz_durus_dk': plansiz_durus_dk,
        'vardiya_sayisi': len(vardiyalar),
        'robotlar': list(robotlar.values()),
        'aktif_vardiyalar': aktif_vardiyalar,
        'plansiz_duruslar': plansiz_duruslar,
        'referans_takip': referans_takip_list,
        'duyuru': _andon_mesaj
    }
    
    # Ayarlari response_data'ya entegre et
    for row in ayarlar_rows:
        response_data[row['anahtar']] = row['deger']
        
    return jsonify(response_data)




@app.route('/api/andon/mesaj', methods=['POST'])
def andon_mesaj_guncelle():
    """Andon TV kayan duyuru mesajını güncelle."""
    data = request.get_json() or {}
    _andon_mesaj['metin'] = data.get('metin', '')
    _andon_mesaj['yazar'] = data.get('yazar', '')
    return jsonify({'basarili': True})


# ─────────────────────────────────────────────────────────────
# GÜNLÜK ÜRETİM RAPORU (Excel Görünümü) API
# ─────────────────────────────────────────────────────────────

@app.route('/api/rapor/gunluk_detay', methods=['GET'])
def gunluk_rapor_detay():
    """Belirli bir tarih, vardiya ve bolum icin uretim detaylarini getirir.
    Sira (key) bolume gore degisir:
      - kaynak: 1..9, M
      - montaj: o gun calisan distinct hat'lar (HAT 1, HAT 2 ...)
      - metal:  300T, 400T, 550T, Şerit Testere
    """
    tarih = request.args.get('tarih')
    vardiya_turu = request.args.get('vardiya')
    bolum = (request.args.get('bolum') or 'kaynak').strip()
    if bolum not in GECERLI_BOLUMLER:
        bolum = 'kaynak'
    # Lokasyon verilmezse TK2 varsay: TK1 vardiyaları da bolum='montaj' olduğundan
    # filtresiz sorgu TK1 operatörlerini TK2 montaj raporuna karıştırıyordu.
    # (TK1 raporu isteyen istemciler — mobil/rapor sayfası — lokasyon=TK1 gönderir.)
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip() or 'TK2'

    # ALT GRUP FILTRESI (kullanici 2026-08-21): "tel uretimi icin gunluk uretim
    # paylasilirken alt grup ozelinde de rapor paylasilabilsin — tam otomatik
    # makine icin ayri uretim raporu uretilebilsin."
    # Deger PROSES ADIMI'dir ('Kapama', 'Tam Otomatik'...), hat adi DEGIL: bir
    # adimin birden cok makinesi var (Kapama'da 12 pres) ve kullanici genelde
    # adimin tamamini istiyor. Tek makine istenirse ?hat= ile daraltilir.
    adim_filtre = (request.args.get('adim') or '').strip()
    hat_filtre = (request.args.get('hat') or '').strip()

    if not tarih or not vardiya_turu:
        return jsonify({'hata': 'tarih ve vardiya parametreleri zorunludur'}), 400

    try:
        conn = get_db()

        # Bugünün detayı isteniyorsa auto sayaçları tazele (canlı adetlerle raporla)
        if tarih == date.today().isoformat():
            _acik_auto_vardiyalari_senkronla(conn, bolum or None, lokasyon or None)

        # LAZER: makine vardiyada değil ÜRETİM KAYDINDA seçilir (istasyon 1..6) —
        # rapor makine bazında kırılır, gruplamaya istasyon eklenir.
        lazer_modu = (bolum == 'lazer')
        # HATTI ÜRETİM KAYDINDA SEÇİLEN BÖLÜMLER (2026-08-18: TK1 montaj + tel,
        # ayrıca plastik/pres): vardiya.robot_no sabit hattır, GERÇEK hat kaydın
        # istasyonundadır → gruplama istasyona göre yapılmalı, yoksa raporun
        # tamamı tek satıra ('TK1 Montaj') çöker.
        kayitta_hat = sabit_hat_adi(lokasyon, bolum)
        ist_modu = lazer_modu or bool(kayitta_hat)
        ist_kolon = ", COALESCE(u.istasyon, 0) as istasyon" if ist_modu else ""
        sql = f"""
            SELECT
                v.robot_no,
                v.operator_adi,
                u.referans_kodu,
                SUM(u.ok_adet)    as ok_toplam,
                SUM(u.nok_adet)   as nok_toplam,
                SUM(u.tamir_adet) as tamir_toplam{ist_kolon}
            FROM vardiyalar v
            LEFT JOIN uretim_kayitlari u ON v.id = u.vardiya_id
            WHERE v.tarih = ?
              AND UPPER(REPLACE(REPLACE(v.vardiya_turu,'ü','u'),'Ü','U')) =
                  UPPER(REPLACE(REPLACE(?,'ü','u'),'Ü','U'))
              AND COALESCE(v.bolum, 'kaynak') = ?
        """
        params = [tarih, vardiya_turu, bolum]
        if lokasyon:
            sql += " AND COALESCE(v.lokasyon, 'TK2') = ?"
            params.append(lokasyon)
        if ist_modu:
            sql += " GROUP BY v.robot_no, v.operator_adi, u.referans_kodu, COALESCE(u.istasyon, 0) ORDER BY istasyon, v.operator_adi"
        else:
            sql += " GROUP BY v.robot_no, v.operator_adi, u.referans_kodu ORDER BY v.robot_no, v.operator_adi"
        rows = conn.execute(sql, params).fetchall()
        conn.close()

        # ── ALT GRUP SÜZGECİ (2026-08-21) ───────────────────────────────────
        # Adım, kaydın istasyonundan türüyor (istasyon → makine → tel adımı);
        # SQL'de çözülemez çünkü eşleme tel_proses'te. Satırlar kurulduktan
        # sonra Python'da süzülür — hem 'adim' (Kapama, Tam Otomatik…) hem tek
        # makine ('hat') destekli. ist_modu YOKSA (kaynak/metal gibi hattı
        # vardiyada seçilen bölümler) süzgeç ANLAMSIZDIR, uygulanmaz.
        # SEÇENEK LİSTESİ SÜZMEDEN ÖNCE hesaplanır — süzülmüş satırlardan
        # üretilseydi seçici tek seçeneğe çöker ve kullanıcı geri dönemezdi.
        _tum_adimlar = []
        if ist_modu and bolum == 'tel':
            for r in rows:
                _a = tel_hat_adimi(kayit_hatti(r['robot_no'], r['istasyon']))
                if _a and _a not in _tum_adimlar:
                    _tum_adimlar.append(_a)
            _tum_adimlar = ([a for a in TEL_ADIMLARI if a in _tum_adimlar]
                            + [a for a in _tum_adimlar if a not in TEL_ADIMLARI])

        if (adim_filtre or hat_filtre) and ist_modu:
            _suzulmus = []
            for r in rows:
                _h = kayit_hatti(r['robot_no'], r['istasyon'])
                if hat_filtre and _h != hat_filtre:
                    continue
                if adim_filtre and tel_hat_adimi(_h) != adim_filtre:
                    continue
                _suzulmus.append(r)
            rows = _suzulmus

        # Bolume gore robot/hat/makine sirasi
        if bolum == 'kaynak':
            robot_listesi = ['1','2','3','4','5','6','7','8','9','M']
        elif bolum == 'metal':
            robot_listesi = ['300T', '400T', '550T', 'Şerit Testere']
        elif lazer_modu:
            robot_listesi = [f'Lazer {i}' for i in range(1, 7)]
        elif kayitta_hat:
            # Yalnız O GÜN çalışılmış hatlar (22 tel hattının tamamını boş boş
            # listelemek raporu okunmaz yapar); sıra kanonik hat sırasıdır.
            _kanon = HAT_MAKINELERI.get(kayitta_hat, [])
            _gorulen = []
            for r in rows:
                _h = kayit_hatti(r['robot_no'], r['istasyon'])
                if _h and _h not in _gorulen:
                    _gorulen.append(_h)
            robot_listesi = ([h for h in _kanon if h in _gorulen]
                             + [h for h in _gorulen if h not in _kanon])
        else:  # montaj / isleme / pres
            # O gun calisan distinct hat'lari sirayla
            seen = []
            for r in rows:
                rn = (r['robot_no'] or '').strip()
                if rn and rn not in seen:
                    seen.append(rn)
            robot_listesi = seen

        rapor_data = {r: [] for r in robot_listesi}

        def normalize_robot_kaynak(r_no):
            r = str(r_no or '').strip().upper()
            for prefix in ('ABB-', 'ABB', 'ROBOT-', 'ROBOT'):
                if r.startswith(prefix):
                    tail = r[len(prefix):].strip()
                    if tail in robot_listesi:
                        return tail
            if r == 'M':
                return 'M'
            return r

        for row in rows:
            if bolum == 'kaynak':
                target_r = normalize_robot_kaynak(row['robot_no'])
            elif lazer_modu:
                # Kayıt hangi makinedeyse o gruba (istasyon 1..6 → 'Lazer N');
                # makinesiz eski/boş kayıtlar genel 'Lazer' grubuna düşer.
                ist = int(row['istasyon'] or 0)
                target_r = f'Lazer {ist}' if ist > 0 else 'Lazer'
            elif kayitta_hat:
                # Kaydın hattı (istasyon) — eski kayıtlarda vardiyanın hattına düşer
                target_r = kayit_hatti(row['robot_no'], row['istasyon'])
            else:
                target_r = (row['robot_no'] or '').strip()
            if not target_r:
                continue

            if not row['referans_kodu'] and (row['ok_toplam'] or 0) == 0:
                continue

            entry = {
                'operator':   row['operator_adi'] or '',
                'parca_kodu': row['referans_kodu'] or '',
                'adet':       row['ok_toplam']    or 0,
                'tamir':      row['tamir_toplam'] or 0,
                'hurda':      row['nok_toplam']   or 0,
            }

            if target_r not in rapor_data:
                rapor_data[target_r] = []
            rapor_data[target_r].append(entry)

        # ── TEL: ÜRÜN BAZLI ADIM KIRILIMI (kullanıcı 2026-08-18) ───────────────
        # Sorun: aynı 100 parçanın kapaması + son montajı ayrı satırlar olarak
        # yazılınca rapordaki TOPLAM 200 diyordu — adım adedi YAPILAN İŞTİR, ürün
        # adedi değil. Çözüm: satır = ÜRÜN (ek atılmış base kod), kolon = ADIM.
        # '93.TK.464 | Kapama 100 | Son Montaj 100' okunduğunda aynı 100 parçanın
        # iki adımı olduğu gözle görülür; kimse satırı toplamaz. Adım toplamları
        # ayrı verilir, TEK BİR genel toplam ÜRETİLMEZ.
        adim_kirilimi = None
        if bolum == 'tel':
            _satir = {}       # base kod -> {adim: adet}
            _op = {}          # base kod -> {adim: set(operatör)}
            _adim_toplam = {}
            for row in rows:
                kod = row['referans_kodu'] or ''
                adet = int(row['ok_toplam'] or 0)
                if not kod:
                    continue
                adim = tel_koddan_adim(kod) or tel_hat_adimi(
                    kayit_hatti(row['robot_no'], row['istasyon'])) or '—'
                base = tel_ek_ayikla(kod)
                _satir.setdefault(base, {}).setdefault(adim, 0)
                _satir[base][adim] += adet
                _op.setdefault(base, {}).setdefault(adim, set()).add(row['operator_adi'] or '')
                _adim_toplam[adim] = _adim_toplam.get(adim, 0) + adet
            adim_kirilimi = {
                'adimlar': [a for a in TEL_ADIMLARI if a in _adim_toplam]
                           + [a for a in _adim_toplam if a not in TEL_ADIMLARI],
                'satirlar': [
                    {'referans': b, 'adet': _satir[b],
                     'operatorler': {a: ', '.join(sorted(o)) for a, o in _op[b].items()}}
                    for b in sorted(_satir)
                ],
                'adim_toplam': _adim_toplam,
            }

        return jsonify({
            'tarih':   tarih,
            'vardiya': vardiya_turu,
            'bolum':   bolum,
            'siralama': robot_listesi,
            'adim_kirilimi': adim_kirilimi,
            # Alt grup seçicisi için: o gün GERÇEKTEN çalışılmış adımlar
            # (22 hattın tamamını listelemek seçiciyi kullanışsız yapar).
            'adim_secenekleri': _tum_adimlar,
            'adim_filtre': adim_filtre or hat_filtre or '',
            'data':    rapor_data,
        })
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({'hata': f'Rapor hatasi: {str(e)}'}), 500


# ═════════════════════════════════════════════════════════════════════════════
#  OPERATÖR PERFORMANSI + OTOMATİK ANALİZ (kullanıcı 2026-08-21)
# ═════════════════════════════════════════════════════════════════════════════
# Hesap analiz.py'de: app.py, mail_raporu.py ve dashboard aynı bulguları oradan
# alır. Buradaki uç noktalar sadece parametre doğrulayıp JSON'a çevirir.

def _analiz_araligi(varsayilan_gun=7):
    """?bas=&bit= — verilmezse son N gün. Ters aralık gelirse düzeltilir."""
    bit = (request.args.get('bit') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    bas = (request.args.get('bas') or '').strip()
    if not bas:
        bas = (datetime.strptime(bit, '%Y-%m-%d')
               - timedelta(days=varsayilan_gun - 1)).strftime('%Y-%m-%d')
    if bas > bit:
        bas, bit = bit, bas
    return bas, bit


@app.route('/api/rapor/operator_performans', methods=['GET'])
@panel_gerekli(izin='operator-performans')
def operator_performans_api():
    """Operatör bazlı performans + OEE tablosu.

    'hat_farki' sütunu ADİL KARŞILAŞTIRMA için: ham OEE makineye çok bağlıdır
    (yavaş hatta çalışan hep sonda çıkar), bu sütun kişinin kendi çalıştığı
    hatların dönem ortalamasına göre kaç puan yukarıda/aşağıda olduğunu verir.
    """
    import analiz as _an
    bas, bit = _analiz_araligi()
    lokasyon = (request.args.get('lokasyon') or '').strip() or None
    bolum = (request.args.get('bolum') or '').strip() or None
    if bolum and bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f'Geçersiz bölüm: {bolum}'}), 400
    conn = get_db()
    vardiyalar = _an.vardiya_metrikleri(conn, bas, bit, lokasyon, bolum)
    resp = jsonify({
        'donem': {'bas': bas, 'bit': bit, 'lokasyon': lokasyon or 'HEPSİ',
                  'bolum': bolum or 'HEPSİ',
                  # Bölüm seçilmediyse cycle süresi tanımsız bölümler hariç
                  # (bkz. analiz.OEE_DISI_BOLUMLER) — ekranda yazılır.
                  'haric_bolumler': ([] if bolum else list(_an.OEE_DISI_BOLUMLER)),
                  'gun': len({v['tarih'] for v in vardiyalar})},
        'ozet': _an._topla(vardiyalar),
        'operatorler': _an.operator_performans(vardiyalar),
        'hatlar': _an.hat_performans(vardiyalar),
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/analiz', methods=['GET'])
@panel_gerekli(izin='analiz')
def analiz_api():
    """Dönem analizi: yerel bulgular (+ ?yorum=1 ise Claude yönetici özeti).

    Yorum katmanı ai_config.json'a bağlıdır ve HATASI ANALİZİ DURDURMAZ —
    'yorum_hata' alanında sebep döner, bulgular her hâlükârda gelir.
    """
    import analiz as _an
    bas, bit = _analiz_araligi()
    lokasyon = (request.args.get('lokasyon') or '').strip() or None
    bolum = (request.args.get('bolum') or '').strip() or None
    if bolum and bolum not in GECERLI_BOLUMLER:
        return jsonify({'hata': f'Geçersiz bölüm: {bolum}'}), 400
    # Aralık üst sınırı: analiz her vardiyayı tek tek çözer; çok uzun aralık
    # hem yavaşlar hem de bulguları anlamsızca genelleştirir.
    if (datetime.strptime(bit, '%Y-%m-%d') - datetime.strptime(bas, '%Y-%m-%d')).days > 120:
        return jsonify({'hata': 'En fazla 120 günlük aralık analiz edilebilir'}), 400
    yorum = request.args.get('yorum') in ('1', 'true', 'evet')
    sonuc = _an.analiz_yap(bas, bit, lokasyon, bolum, yorum=yorum, conn=get_db())
    resp = jsonify(sonuc)
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/analiz/gun_ozet', methods=['GET'])
@panel_gerekli(izin='ozet')
def analiz_gun_ozet_api():
    """Fabrika özeti sayfasının alt bloğu: ÖNCEKİ GÜNÜN özeti + öne çıkanlar.
    Yorum katmanı burada ÇALIŞMAZ (her sayfa açılışında çağrılır — ücret olurdu).
    """
    import analiz as _an
    gun = (request.args.get('gun') or '').strip() or \
        (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    lokasyon = (request.args.get('lokasyon') or '').strip() or None
    sonuc = _an.gunluk_ozet(gun, lokasyon, conn=get_db())
    sonuc['gun'] = gun
    resp = jsonify(sonuc)
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/analiz/ai_durum', methods=['GET'])
@panel_gerekli(izin='analiz')
def analiz_ai_durum():
    """AI yorum katmanı kullanılabilir mi? (panelde butonun durumunu belirler)
    ANAHTAR DÖNDÜRÜLMEZ — yalnız var/yok bilgisi."""
    import analiz as _an
    cfg, cfg_hata = _an.ai_config(hata_ile=True)
    try:
        import anthropic  # noqa: F401
        paket = True
    except ImportError:
        paket = False
    # TANI ALANLARI (2026-08-21): "pip install yaptım ama hâlâ kurulu değil diyor"
    # vakası. Sebep genelde pip'in KULLANICI klasörüne kurması ve servisin BAŞKA
    # bir hesapla çalışması — o zaman paket ortada durur ama servis göremez.
    # Yorumcunun hangi python'u ve hangi hesabı kullandığını göstermek tek adımda
    # ayırt ettirir.
    import getpass, sys as _sys   # modül düzeyinde import YOK (dosya kalıbı)
    # SAĞLAYICIYA GÖRE HAZIRLIK (2026-08-21, "kredi yüklemek istemiyorum"):
    #   anthropic → paket + anahtar gerekir
    #   gemini    → yalnız anahtar (requests hep kurulu, paket şartı YOK)
    #   ollama    → ne paket ne anahtar; yerel servise erişilebilmeli
    saglayici = (cfg.get('saglayici') or 'anthropic').strip().lower()
    anahtar_var = bool((cfg.get('api_anahtari')
                        or os.environ.get('ANTHROPIC_API_KEY' if saglayici == 'anthropic'
                                          else 'GEMINI_API_KEY') or '').strip())
    ollama_erisim = None
    if saglayici == 'ollama':
        try:
            import requests as _rq
            _r = _rq.get((cfg.get('ollama_url') or 'http://localhost:11434').rstrip('/')
                         + '/api/version', timeout=2)
            ollama_erisim = (_r.status_code == 200)
        except Exception:
            ollama_erisim = False
    hazir = bool(cfg.get('etkin')) and (
        (saglayici == 'anthropic' and paket and anahtar_var)
        or (saglayici == 'gemini' and anahtar_var)
        or (saglayici == 'ollama' and bool(ollama_erisim)))
    etkin_model = {'anthropic': cfg.get('model'),
                   'gemini': cfg.get('gemini_model'),
                   'ollama': cfg.get('ollama_model')}.get(saglayici, cfg.get('model'))
    return jsonify({
        'config_var': os.path.exists(_an.AI_CONFIG),
        'config_hata': cfg_hata,          # JSON bozuksa/BOM'luysa sebebi
        'config_alanlar': cfg.get('_alanlar') or [],   # dosyada BULUNAN anahtar adları
        'python_yolu': _sys.executable,
        'servis_kullanici': (getpass.getuser() if hasattr(getpass, 'getuser') else ''),
        'saglayici': saglayici,
        'ollama_url': cfg.get('ollama_url') if saglayici == 'ollama' else None,
        'ollama_erisim': ollama_erisim,
        'etkin': bool(cfg.get('etkin')),
        'paket_kurulu': paket,
        'anahtar_var': anahtar_var,
        'model': etkin_model,
        'hazir': hazir,
    })


@app.route('/api/rapor/vardiya_listesi', methods=['GET'])
def rapor_vardiya_listesi():
    """Belirtilen tarihteki ve bolumdeki mevcut vardiya turlerini dondurur."""
    tarih = request.args.get('tarih')
    bolum = request.args.get('bolum')
    # Lokasyon verilmezse TK2 varsay (gunluk_rapor_detay ile aynı gerekçe:
    # TK1 vardiya türleri TK2 rapor dropdown'una karışmasın).
    lokasyon = (request.args.get('lokasyon') or 'TK2').strip() or 'TK2'
    if not tarih:
        return jsonify({'hata': 'tarih zorunludur'}), 400
    try:
        conn = get_db()
        sql = "SELECT DISTINCT vardiya_turu FROM vardiyalar WHERE tarih = ?"
        params = [tarih]
        if bolum:
            sql += " AND COALESCE(bolum, 'kaynak') = ?"
            params.append(bolum)
        if lokasyon:
            sql += " AND COALESCE(lokasyon, 'TK2') = ?"
            params.append(lokasyon)
        sql += " ORDER BY vardiya_turu"
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        return jsonify([r['vardiya_turu'] for r in rows])
    except Exception as e:
        return jsonify({'hata': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# AS400 (ERP) TEYIT — günün üretimi ↔ açık launch eşlemesi
# Kaynak: as400/launch_esle.py → canlı TKC0301F.XPRO90 (pyodbc; şifre Windows
# Kimlik Bilgileri Yöneticisi'nde). Salt-okunur; ERP'ye hiçbir şey yazmaz.
# ─────────────────────────────────────────────────────────────
@app.route('/api/as400/teyit_listesi', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_listesi():
    """Sabah teyit KUYRUĞU. İki mod:
      ?gunler=3[&bugun=1]  → son 3 ÜRETİM gününün listesi (takvim günü değil —
                             hafta sonu/tatil/duruş pencereyi kaydırmasın);
                             bugun=1 ise bugünü de ekler (varsayılan mod)
      ?tarih=YYYY-MM-DD    → tek gün (geçmiş inceleme)
    AS400 okumaları tek sefer yapılır (esle_coklu). Yanıt her gün için
    kategorili liste + son robot gönderimleri (as400_teyit_log) içerir."""
    tarih = (request.args.get('tarih') or '').strip()
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _le
    except Exception as e:
        return jsonify({'hata': f'AS400 modülü yüklenemedi: {e}'}), 502
    if tarih:
        tarihler = [tarih]
    else:
        try:
            gunler = max(1, min(7, int(request.args.get('gunler') or 3)))
        except ValueError:
            gunler = 3
        tarihler = []
        if (request.args.get('bugun') or '') in ('1', 'true'):
            tarihler.append(date.today().isoformat())
        # SON N ÜRETİM GÜNÜ (2026-08-17) — takvim günü DEĞİL. Üretimsiz günler
        # (hafta sonu / tatil / duruş) pencereyi yiyordu: her pazartesi pencere
        # pratikte tek güne iniyor, 07.08'deki bir haftalık aradan sonra da o günün
        # teyit edilmemiş kodları "Yenile" ile HİÇ görünmüyordu (yalnız elle tarih
        # filtresiyle). Üretim yapılan günleri seçince ara pencereyi kaydırmaz.
        # DB okunamazsa eski takvim davranışına düşülür (kuyruk hiç boş kalmasın).
        tarihler += (_le.son_uretim_gunleri(gunler) or
                     [(date.today() - timedelta(days=i)).isoformat() for i in range(1, gunler + 1)])
    try:
        coklu = _le.esle_coklu(tarihler)
    except Exception as e:
        # AS400 kapalı/şifre yok/ağ hatası → paneli kırmadan anlaşılır mesaj
        return jsonify({'hata': f'AS400 sorgusu başarısız: {e}'}), 502
    conn = get_db()
    son_gonderimler = [dict(r) for r in conn.execute(
        "SELECT id, created_at, uretim_tarihi, yil, launch_no, referans, adet, sonuc, mesaj, olusturan "
        "FROM as400_teyit_log ORDER BY id DESC LIMIT 15").fetchall()]

    # COP verilmiş hurdalar (2026-07-24): "♻ Hurda → COP" bölümü zaten COP'lanmış
    # satırları TEKRAR göstermesin (sayfa yenilenince mükerrer görünüyordu). Anahtar =
    # "uretim_tarihi|article" (article = COP log'daki launch_no). Dedup ile aynı mantık.
    cop_verildi = {}
    cop_hatali = {}   # basarisiz COP denemeleri (ERP'de girilmis OLABILIR)
    try:
        for _cr in conn.execute(
            "SELECT uretim_tarihi, launch_no, SUM(adet) tp FROM as400_teyit_log "
            "WHERE yil='CO' AND sonuc='ok' GROUP BY uretim_tarihi, launch_no").fetchall():
            cop_verildi[_cop_key(_cr['uretim_tarihi'], _cr['launch_no'])] = _cr['tp']
        # BASARISIZ COP DENEMELERI (2026-08-17). Panel "COP verildi"yi YALNIZ kendi
        # log'undan (sonuc='ok') turetiyor; ERP'ye bakmiyor. Hareket sorgusu da COP'u
        # kapsamiyor (MGCACD IN ('RPR','CFI')) — COP'u oraya eklemek YANLIS olur,
        # hurda hareketi uretim teyidi sayilip mukerrer frenini bozardi.
        # Sonuc: robot COP'u ERP'ye yazdigi halde dogrulama tutmazsa satir sessizce
        # "hic COP verilmemis" gibi gorunuyordu (10.130.6206b vakasi: ERP'de kayit
        # 269526 duruyordu, panel bekliyor diyordu). Artik basarisiz deneme de
        # doner ve satirda UYARI olarak gosterilir; tekrar gondermek ERP'de mevcut
        # hareketi bulup kaydi 'ok' ile ESITLER (mukerrer COP YAZMAZ).
        for _cr in conn.execute(
            "SELECT uretim_tarihi, launch_no, MAX(created_at) son, COUNT(*) n FROM as400_teyit_log "
            "WHERE yil='CO' AND sonuc<>'ok' GROUP BY uretim_tarihi, launch_no").fetchall():
            _k = _cop_key(_cr['uretim_tarihi'], _cr['launch_no'])
            if _k not in cop_verildi:
                cop_hatali[_k] = {'son': _cr['son'], 'deneme': _cr['n']}
    except Exception as _e:
        print(f'[teyit_listesi] cop_verildi atlandı: {_e}')

    # İŞARETLER (2026-07-20): günlük 'kontrol edildi' + kalıcı 'gerek_yok'.
    # Backend kanonik eşleştirir → her satıra 'isaret'/'kalici_haric' enjekte edilir
    # (frontend'in normalize etmesine gerek kalmaz). Tablo yoksa (eski DB) sessiz geç.
    try:
        gun_isaret = {}
        for t in tarihler:
            gun_isaret[t] = {r['referans']: {'durum': r['durum'], 'aciklama': r['aciklama']}
                             for r in conn.execute(
                                 "SELECT referans, durum, aciklama FROM as400_teyit_isaret "
                                 "WHERE kapsam='gun' AND uretim_tarihi=?", (t,)).fetchall()}
        kalici_isaret = {r['referans']: r['aciklama'] for r in conn.execute(
            "SELECT referans, aciklama FROM as400_teyit_isaret WHERE kapsam='kalici' AND durum='gerek_yok'").fetchall()}
        for t in tarihler:
            for satirlar in coklu[t].values():
                for r in satirlar:
                    kn = _le.kanonik(r.get('referans', ''))
                    if kn in gun_isaret[t]:
                        r['isaret'] = gun_isaret[t][kn]
                    if kn in kalici_isaret:
                        r['kalici_haric'] = True
    except Exception as e:
        print(f'[teyit_listesi] işaret enjeksiyonu atlandı: {e}')

    return jsonify({
        'gunler': [{'tarih': t,
                    'ozet': {k: len(v) for k, v in coklu[t].items()},
                    'liste': coklu[t]} for t in tarihler],
        'son_gonderimler': son_gonderimler,
        'cop_verildi': cop_verildi,
        'cop_hatali': cop_hatali,
    })


@app.route('/api/as400/teyit_isaret', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_isaret():
    """Teyit ekranında bir satırı işaretle / işareti kaldır.
    Body: {kapsam:'gun'|'kalici', durum?:'kontrol'|'gerek_yok', referans, bolum?,
           uretim_tarihi?, aciklama?, kaldir?}
    kapsam='kalici' → referans HER GÜN teyit dışı (HARIC). kapsam='gun' → o güne özel.
    durum='teyit_ver' (2026-07-27, 4536B vakası): satır Dikkat/varyant/ara-op'a düşse
    bile 'kod DOĞRU, teyit ver' demektir → CFI kuyruğuna zorlanır (robot + oto koşu).
    Referans launch_esle.kanonik ile normalize saklanır (eşleşme tutarlı)."""
    data = request.get_json(silent=True) or {}
    kapsam = (data.get('kapsam') or 'gun').strip()
    durum = (data.get('durum') or '').strip()
    referans = (data.get('referans') or '').strip()
    bolum = (data.get('bolum') or '').strip()
    u_tarih = (data.get('uretim_tarihi') or '').strip()
    aciklama = (data.get('aciklama') or '').strip()
    kaldir = bool(data.get('kaldir'))
    if kapsam not in ('gun', 'kalici') or not referans:
        return jsonify({'hata': 'kapsam ve referans zorunlu'}), 400
    if kapsam == 'kalici':
        u_tarih = ''
        durum = 'gerek_yok'
    else:
        if not u_tarih:
            return jsonify({'hata': 'günlük işaret için uretim_tarihi zorunlu'}), 400
        if durum not in ('kontrol', 'gerek_yok', 'teyit_ver'):
            durum = 'kontrol'
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _le
        ref_norm = _le.kanonik(referans)
    except Exception:
        ref_norm = referans.upper().replace(' ', '')
    conn = get_db()
    kullanici = g.panel_ku['kullanici_adi']
    try:
        if kaldir:
            conn.execute(
                "DELETE FROM as400_teyit_isaret WHERE kapsam=? AND uretim_tarihi=? AND bolum=? AND referans=?",
                (kapsam, u_tarih, bolum, ref_norm))
        else:
            conn.execute(
                "INSERT INTO as400_teyit_isaret (kapsam, uretim_tarihi, bolum, referans, durum, aciklama, olusturan) "
                "VALUES (?,?,?,?,?,?,?) "
                "ON CONFLICT(kapsam, uretim_tarihi, bolum, referans) "
                "DO UPDATE SET durum=excluded.durum, aciklama=excluded.aciklama, olusturan=excluded.olusturan",
                (kapsam, u_tarih, bolum, ref_norm, durum, aciklama, kullanici))
        conn.commit()
        return jsonify({'basarili': True, 'kapsam': kapsam, 'durum': durum, 'referans': ref_norm, 'kaldirildi': kaldir})
    except Exception as e:
        conn.rollback()
        return jsonify({'hata': str(e)}), 400


# ── AS400 teyit GÖNDERİMİ (yazma) ──
# Session B'yi tek seferde tek istek sürebilir (fiziksel tek ekran) → global kilit.
import threading as _threading
_AS400_KILIT = _threading.Lock()
# Nazik durdurma bayrağı: kullanıcı "Durdur"a basınca set edilir. Döngü HER satır
# başında kontrol eder → set ise kalan satırlar 'atlandi'. Çalışan cscript ASLA
# kill edilmez (record-lock riski); mevcut launch biter, SONRAKİLERE geçilmez.
_AS400_DURDUR = _threading.Event()

# ── TEYIT-AGENT KÖPRÜSÜ (2026-07-24) ──
# Server'da app NSSM SERVİSİ olarak Session 0'da koşar; PCOMM (pcsws.exe) ise RDP
# kullanıcı oturumundadır. Session 0'dan başlatılan cscript PCOMM oturumlarını GÖREMEZ
# (Windows oturum izolasyonu) → robot, RDP oturumunda koşan as400/teyit_agent.py'ye
# HTTP ile devredilir (127.0.0.1:5010). Agent yoksa (laptop/interaktif) yerel subprocess
# — config dosyası YOK, tespit otomatik (localhost probe; agent kapalıysa anında reddedilir).
_TEYIT_AGENT_URL = 'http://127.0.0.1:5010'


def _teyit_agent_var():
    try:
        import requests
        r = requests.get(_TEYIT_AGENT_URL + '/durum', timeout=1.5)
        return r.status_code == 200 and (r.json() or {}).get('agent') == 'cofle-teyit'
    except Exception:
        return False


def _oturum_0_mi():
    """App bir Windows SERVİSİ olarak mı koşuyor (Session 0)? Sunucuda cofle-app
    NSSM servisi Session 0'dadır; PCOMM pencereleri promanage'ın RDP oturumunda.
    Oturum izolasyonu yüzünden servisin başlattığı cscript PCOMM'u GÖREMEZ —
    bu yüzden teyit-agent var (bkz. as400/SERVER_AS400_KURULUM.md).
    Belirlenemezse False (laptop davranışı) döner."""
    try:
        import ctypes
        sid = ctypes.c_ulong()
        if ctypes.windll.kernel32.ProcessIdToSessionId(os.getpid(), ctypes.byref(sid)):
            return sid.value == 0
    except Exception:
        pass
    return False


def _robot_log_kuyrugu(saniye=240, satir=4):
    """Robotun KENDİ log dosyasının son satırları (2026-08-17).

    teyit_gir.js/cfi_gir.js daha PCOMM'a dokunmadan ÖNCE
    as400/teyit_loglari/<zaman>_<launch>.txt dosyasını açar ve her adımı oraya da
    yazar. Robot stdout üretmeden öldüğünde tek kanıt bu dosyadır; app buraya hiç
    bakmıyordu. Dosya YOKSA script daha log satırına gelmeden düşmüş demektir
    (bunu da söylüyoruz — 'sessiz' hatanın en ayırt edici işareti)."""
    try:
        import time
        d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400', 'teyit_loglari')
        if not os.path.isdir(d):
            return ''
        simdi = time.time()
        yeni = [(os.path.getmtime(os.path.join(d, f)), f) for f in os.listdir(d)
                if f.lower().endswith('.txt')]
        yeni = [(m, f) for m, f in yeni if (simdi - m) <= saniye]
        if not yeni:
            # SEBEP İDDİA ETME: dosya yokluğu tek başına cscript/WSH sorunu demek
            # değil (ör. ECL37110'da script log satırına gelemeden de düşebilir).
            return ('  · Robot log dosyası oluşmadı (as400\\teyit_loglari) — script log '
                    'satırına gelmeden düşmüş.')
        m, f = max(yeni)
        with open(os.path.join(d, f), 'r', encoding='cp1254', errors='replace') as fh:
            satirlar = [l.strip() for l in fh if l.strip()]
        if not satirlar:
            return (f'  · Robot logu ({f}) BOŞ — script log dosyasını açtı ama tek satır '
                    f'yazamadan düştü (PCOMM.autECLSession / Session B bağlanamadı).')
        son = ' | '.join(satirlar[-satir:])[:400]
        return f'  · Robot logu ({f}) son satırlar: {son}'
    except Exception as e:
        return f'  · Robot logu okunamadı: {e}'


def _robot_sessiz_mesaj(yol, rc, hata_cikti):
    """Robot HİÇ stdout üretmedi → sebebi anlaşılır yaz (2026-08-17).

    NEDEN: cscript'in stderr'i yakalanıyordu ama HİÇ KULLANILMIYORDU; stdout boş
    olunca satır "Robot iptal:" diye BOMBOŞ bir mesajla loglanıyordu. 14 satırın
    14'ü aynı boş mesajla düşüp operatöre tek bir ipucu vermiyordu.
    En sık sebep: teyit_gir.js'in 2. satırındaki
        new ActiveXObject("PCOMM.autECLSession") / SetConnectionByName("B")
    PCOMM kapalı/oturum açılmamışsa FIRLATIR → JScript hatası stderr'e gider,
    stdout boş kalır (SONUC= satırı hiç basılmaz)."""
    st = ' '.join((hata_cikti or '').split())[:300]
    # ECL37110 = "emulasyon arayuzu kullanilamiyor" (2026-08-17 sahada gorulen hata,
    # PCOMM Italyanca: 'Interfaccia di emulazione non disponibile'). COM nesnesi
    # OLUSUYOR (yani PCOMM KURULU) ama o Windows oturumunda ACIK EMULATOR YOK.
    # En sik sebebi: PCOMM ile teyit-agent FARKLI RDP oturumlarinda (ya da PCOMM
    # hic acik degil). cscript bu hatada bile 0 ile cikar — rc'ye guvenme.
    if 'ECL37110' in st or 'emulazione' in st.lower() or 'emulation interface' in st.lower():
        return (f'PCOMM emülasyon arayüzü YOK (ECL37110) — PCOMM kurulu ama robotun '
                f'çalıştığı Windows oturumunda AÇIK bir emülatör oturumu bulunamadı. '
                f'Genelde PCOMM ile teyit-agent FARKLI RDP oturumlarındadır: sunucuda '
                f'`query session` ile promanage\'ın KAÇ oturumu olduğuna bakın, fazlasını '
                f'logoff edin; A+B pencerelerini agent\'ın koştuğu oturumda açıp sign-on '
                f'yapın (as400\\Robot_Tani.bat bunu yerinde gösterir).' + _robot_log_kuyrugu())
    pcomm_kokusu = (not st) or any(k in st for k in (
        'PCOMM', 'autECL', 'ActiveX', '80040154', '800401F3', 'SetConnectionByName',
        'Automation', 'sunucu'))
    # NOT: mesajda cp1254 DIŞI karakter kullanma (ok/emoji). Bu metin sunucu
    # konsoluna da düşebiliyor; '➜' orada UnicodeEncodeError veriyordu.
    ipucu = (' >> PCOMM Session B AÇIK ve OTURUM AÇILMIŞ olmalı. Sunucuya RDP ile girip '
             'A+B pencerelerini açın, sign-on yapın, teyit-agent penceresinin ayakta '
             'olduğunu görün, sonra RDP\'yi LOGOFF değil DISCONNECT edin '
             '(bkz. as400/SERVER_AS400_KURULUM.md · madde 7).') if pcomm_kokusu else ''
    # "AS400'e hiçbir şey yazılmadı" DEME: burada bunu bilemeyiz (çıkış kodu 0 ise
    # script kendini başarılı sanmış olabilir). Yazılıp yazılmadığını çağıran,
    # BPROF0 teyitli değerini önce/sonra okuyarak söyler.
    return (f'Robot hiç çıktı üretmedi ({yol}, çıkış kodu {rc}).'
            + (f' Robot hatası: {st}' if st else '')
            + ipucu + _robot_log_kuyrugu())


def _as400_robot_calistir(script_dosya, args, timeout_sn):
    """Robot cscript'ini çalıştır (agent varsa onda, yoksa yerel).
    Döner: (cikti:str, hata:str|None) — hata doluysa satır 'hata' olarak loglanmalı.
    KRİTİK: agent'a istek GİTTİKTEN sonra bağlantı koparsa yerel TEKRAR DENENMEZ
    (robot agent'ta hâlâ koşuyor olabilir → çift giriş riski)."""
    args = [str(a) for a in args]
    if _teyit_agent_var():
        try:
            import requests
            r = requests.post(_TEYIT_AGENT_URL + '/calistir',
                              json={'script': script_dosya, 'args': args, 'timeout': timeout_sn},
                              timeout=(3, timeout_sn + 15))
            if r.status_code == 200:
                j = r.json() or {}
                cikti = j.get('cikti') or ''
                if not cikti.strip():
                    # Agent cscript'i çalıştırdı ama stdout BOŞ → script başlarken
                    # fırlamış (çoğunlukla PCOMM/Session B yok). rc + stderr ile anlat.
                    return '', _robot_sessiz_mesaj('teyit-agent', j.get('rc'), j.get('hata_cikti'))
                return cikti, None
            if r.status_code == 504:
                return '', f'Robot zaman aşımı ({timeout_sn} sn) — sıradakine geçildi'
            return '', f'Teyit-agent hatası (HTTP {r.status_code}): {(r.json() or {}).get("hata", "")}'
        except Exception as e:
            return '', f'Teyit-agent bağlantısı koptu ({e}) — robot durumu belirsiz, Session B\'yi kontrol edin'
    # Agent YOK. Servis olarak koşuyorsak yerel cscript PCOMM'u göremez — denemek
    # yerine sebebi söyle (eskiden denenip boş çıktıyla "Robot iptal:" yazılıyordu).
    if _oturum_0_mi():
        return '', ('Teyit-agent çalışmıyor ve uygulama Windows SERVİSİ olarak (Session 0) '
                    'koşuyor — servisin başlattığı robot PCOMM pencerelerini GÖREMEZ, teyit '
                    'verilemez. Sunucuya RDP ile girip PCOMM A+B\'yi açın, sign-on yapın ve '
                    'as400\\Teyit_Agent_Baslat.bat\'ı çalıştırın (bkz. SERVER_AS400_KURULUM.md).')
    import subprocess
    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400', script_dosya)
    try:
        pr = subprocess.run([r'C:\Windows\SysWOW64\cscript.exe', '//nologo', script] + args,
                            capture_output=True, timeout=timeout_sn)
        cikti = (pr.stdout or b'').decode('cp1254', errors='replace')
        if not cikti.strip():
            return '', _robot_sessiz_mesaj('yerel cscript', pr.returncode,
                                           (pr.stderr or b'').decode('cp1254', errors='replace'))
        return cikti, None
    except subprocess.TimeoutExpired:
        # Timeout → cscript KILL edilir → yarım transaction + record-lock riski;
        # sonraki satıra devam edilir (sonraki robot G0'da record-lock'u kurtarır).
        return '', f'Robot zaman aşımı ({timeout_sn} sn) — sıradakine geçildi'
    except FileNotFoundError:
        return '', (r'cscript.exe bulunamadı (C:\Windows\SysWOW64\cscript.exe) — '
                    '32-bit cscript şart (PCOMM COM 32-bit).')


# ── SESSION B ÖN KONTROLÜ (kullanıcı 2026-08-25) ────────────────────────────
# "AS400 bazen uzun süre işlem yapılmadığında düşebiliyor. Az önce teyit vermek
#  istedim bundan dolayı veremedim."
# Oturum düştüğünde BAĞLANTI AÇIK KALABİLİYOR ('Fine del lavoro' ekranı), yani
# robot bağlanır ama menüye giremez ve satır satır hata verir. Gözcü bunu 5
# dakikada bir düzeltiyor — kullanıcı o 5 dakikayı beklemesin diye TOPLU GÖNDERİM
# BAŞLARKEN bir kez kontrol edilir. oturum_ac.js sağlıklıysa 'ZATEN' deyip hemen
# döner; ölü oturumda Disconnetti + Connetti + sign-on yapar.
#
# ⚠ ASLA GÖNDERİMİ ENGELLEMEZ: agent kapalıysa / şifre kasada yoksa bu kontrol
# başarısız olur, ama satırlar yine denenir ve kendi hata mesajlarını üretir.
# Engelleseydi, çalışan bir kurulumda tek bir yan koşul tüm teyidi durdururdu.
def _as400_oturum_hazirla():
    """Session B'yi doğrula/kurtar. Döner: (durum, mesaj) — durum: OK|ZATEN|None."""
    try:
        import requests
        r = requests.get(_TEYIT_AGENT_URL + '/durum', timeout=2)
        kullanici = ((r.json() or {}).get('gozcu') or {}).get('kullanici') or ''
    except Exception as e:
        return None, f'agent durumu okunamadı ({e}) — oturum ön kontrolü atlandı'
    if not kullanici:
        return None, 'oturum_config.json içinde kullanıcı tanımlı değil — ön kontrol atlandı'
    cikti, hata = _as400_robot_calistir('oturum_ac.js', [kullanici], 180)
    if hata:
        return None, f'oturum kontrolü çalıştırılamadı: {hata}'
    if 'SONUC=OK' in cikti:
        return 'OK', 'Session B düşmüştü — yeniden bağlanıp giriş yapıldı.'
    if 'SONUC=ZATEN' in cikti:
        return 'ZATEN', 'Session B sağlıklı.'
    son = ' | '.join(l.strip() for l in cikti.splitlines() if l.strip())[-300:]
    return None, f'oturum kurtarılamadı: {son}'


def _as400_launch_durum(yil, no):
    """BPROF0'dan launch'ın (teyitli Q0QTRI, durum Q0AVAN) — bağımsız doğrulama.
    Döner: (teyitli:float|None, durum:int|None)."""
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import keyring, pyodbc
    import as400_config as _cfg
    pw = _cfg.sifre_al()
    cn = pyodbc.connect(_cfg.baglanti_dizesi(pw), timeout=15, autocommit=True)
    try:
        r = cn.cursor().execute(
            "SELECT Q0QTRI, Q0AVAN FROM tkc0301F.BPROF0 WHERE Q0RED2=? AND Q0RENU=?",
            (int(yil), int(no))).fetchone()
        if not r:
            return None, None
        try:
            durum = int(str(r[1]).strip())
        except (TypeError, ValueError):
            durum = None
        return float(r[0]), durum
    finally:
        cn.close()


def _article_gecersiz_mi(article):
    """ERP article master kontrolü — robot ekrana yazmadan ÖNCE (2026-07-30).
    Döner: None (kod geçerli / doğrulanamadı) | reddetme mesajı (str).

    Master OKUNAMAZSA None döner: bilinmiyorken meşru teyidi bloklamak yanlış
    yön. Bu bir "bozuk kodu erken yakala" katmanı, tek fren değil."""
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _la
        var = _la.article_tanimli([article])
        if var is None:          # sorgu/bağlantı hatası — karar verme
            return None
        if _la.kanonik(article) in var:
            return None
        return (f'ERP\'de böyle bir article YOK: "{article}" — operatör kodu yanlış '
                f'yazmış olabilir (örn. kod alanına operasyon bilgisi karışmış). '
                f'Robota gönderilmedi; üretim kaydındaki referans kodunu düzeltin. '
                f'Kodun DOĞRU olduğundan eminseniz "⚠ yine de gönder" ile bu kontrolü '
                f'atlayabilirsiniz (kod gerçekten tanımsızsa robot ekrana yazamaz, '
                f'yalnız bu satır hata verir — kuyruk durmaz).')
    except Exception as e:
        print(f'[article] kontrol yapılamadı ({article}): {e}')
    return None


def _kapasite_reddi(conn, referans, uretim_tarihi, adet):
    """SON SAVUNMA (2026-07-30) — robot çalıştırmadan hemen önce, DB'den BAĞIMSIZ
    kapasite kontrolü. Döner: None (sorun yok) | reddetme mesajı (str).

    NEDEN AYRI HESAP: frontend listesi bayat olabilir, kuyruk elle POST edilebilir,
    /api/as400/teyit_gonder doğrudan çağrılabilir. Hayali stok girişini ENGELLEYEN
    asıl yer burasıdır; UI'daki Dikkat kovası yalnızca erken uyarıdır.

    Cycle time tanımsız (0/NULL) → None döner, kontrol YOK (kullanıcı kararı).
    'teyit_ver' işareti kullanıcının açık onayı → kontrol atlanır.
    Hesap hata verirse None döner: belirsizken meşru teyidi bloklamak yanlış yön
    (asıl fren zaten operatör kaydının kendisi, bu ek bir emniyet katmanı)."""
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _lk
        kanon = _lk.kanonik(referans)
        # Kullanıcı bu satırı açıkça onayladıysa geç
        onay = conn.execute(
            "SELECT 1 FROM as400_teyit_isaret WHERE kapsam='gun' AND uretim_tarihi=? "
            "AND referans=? AND durum='teyit_ver' LIMIT 1", (uretim_tarihi, kanon)).fetchone()
        if onay:
            return None
        # O gün o referansı üreten vardiyalar + cycle time (bölüm/lokasyon eşleşmeli)
        # ct SKALAR alt-sorgu (LEFT JOIN DEĞİL): normalize aynı koda düşen birden
        # fazla referans satırı varsa JOIN satır çoğaltır (bkz. launch_esle
        # gun_uretimi'ndeki 94.LTK.340 vakası). Burada adet toplanmadığı için
        # zararı sınırlıydı ama aynı tuzağı bırakmıyoruz.
        rows = conn.execute("""
            SELECT (SELECT MAX(rl.hedef_cycle_time_sn) FROM referans_listesi rl
                     WHERE UPPER(REPLACE(rl.referans_kodu,' ','')) = UPPER(REPLACE(u.referans_kodu,' ',''))
                       AND COALESCE(rl.bolum,'kaynak') = COALESCE(v.bolum,'kaynak')
                       AND COALESCE(rl.lokasyon,'TK2') = COALESCE(v.lokasyon,'TK2')) ct,
                   v.id vid,
                   COALESCE(v.toplam_sure_dk,0) sure_dk,
                   COALESCE(v.baslangic_saati,'') b, COALESCE(v.bitis_saati,'') e
            FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id
            WHERE v.tarih = ?
              AND UPPER(REPLACE(u.referans_kodu,' ','')) = UPPER(REPLACE(?,' ',''))
        """, (uretim_tarihi, referans)).fetchall()
        if not rows:
            return None
        ct = 0.0
        sureler = {}
        for r in rows:
            try:
                c = float(r['ct'] or 0)
            except (TypeError, ValueError):
                c = 0.0
            if c > 0 and ct <= 0:
                ct = c
            sureler[r['vid']] = _lk._vardiya_suresi_dk(r['sure_dk'], r['b'], r['e'])
        if ct <= 0:
            return None                       # referans tanımsız → kontrol yok
        sure_dk = sum(v for v in sureler.values() if v and v > 0) or _lk.KAPASITE_VARSAYILAN_DK
        teorik = (sure_dk * 60.0) / ct
        azami = teorik * _lk.KAPASITE_TOLERANS
        if adet > azami:
            gerek = (adet * ct) / 3600.0
            return (f'KAPASİTE AŞIMI — {adet} adet × {ct:g} sn = {gerek:.1f} saat gerekir, '
                    f'vardiya {sure_dk / 60.0:.1f} saat (makul üst sınır ~{int(azami)} adet). '
                    f'Operatör kaydı hatalı olabilir; doğruysa panelden "➕ teyit ver" ile onaylayın.')
    except Exception as e:
        print(f'[kapasite] kontrol yapılamadı ({referans} {uretim_tarihi}): {e}')
    return None


def _teyit_gonder_calistir(conn, satirlar, kullanici, varsayilan_tarih='', zorla=False):
    """Launch teyit satırlarını robotla işler — endpoint VE 17:10 oto koşusu ortak
    çekirdeği. ÇAĞIRAN _AS400_KILIT'i tutuyor olmalı. Döner: sonuclar listesi.
    Her satır BAĞIMSIZ denenir — bir launch hata/iptal verirse SONRAKİ launch'lara
    DEVAM edilir (kullanıcı 2026-07-20: hata alınca ilk sayfaya çıkıp sıradaki
    launch'tan devam et). Robot her çağrıda G0'da B'yi temizler + record-lock kurtarır."""
    sonuclar = []
    # Session B ön kontrolü (bkz. _as400_oturum_hazirla) — ölü oturumda
    # launch teyitleri de satır satır hata veriyordu.
    _od, _om = _as400_oturum_hazirla()
    if _od == 'OK':
        print(f'[TEYIT] {_om}')
    elif _od is None:
        print(f'[TEYIT] oturum ön kontrolü: {_om}')

    for idx, s in enumerate(satirlar):
        # NAZIK DURDURMA: kullanıcı "Durdur"a bastıysa kalan satırlara GEÇME (çalışan
        # cscript öldürülMEZ; mevcut satır zaten bittiğinde bu kontrole geliriz).
        if _AS400_DURDUR.is_set():
            for kalan in satirlar[idx:]:
                sonuclar.append({
                    'yil': str(kalan.get('yil') or '').strip(),
                    'no': str(kalan.get('no') or '').strip(),
                    'article': str(kalan.get('article') or '').strip(),
                    'referans': str(kalan.get('referans') or '').strip(),
                    'adet': kalan.get('adet'), 'sonuc': 'atlandi',
                    'mesaj': 'Kullanıcı durdurdu'})
            break
        yil = str(s.get('yil') or '').strip()
        no = str(s.get('no') or '').strip()
        article = str(s.get('article') or '').strip()
        referans = str(s.get('referans') or '').strip()
        # Üretim tarihi SATIR bazlı (kuyruk görünümü birden çok günü birlikte gönderir)
        u_tarih = (str(s.get('uretim_tarihi') or '').strip() or varsayilan_tarih)
        try:
            adet = int(s.get('adet') or 0)
        except (TypeError, ValueError):
            adet = 0
        kayit = {'yil': yil, 'no': no, 'article': article, 'referans': referans,
                 'adet': adet, 'uretim_tarihi': u_tarih}

        if not (yil.isdigit() and len(yil) == 2 and no.isdigit() and 0 < adet <= 99999 and article and u_tarih):
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': 'Geçersiz satır parametresi'})
            continue
        # Mükerrer koruması 1: kendi gönderim log'umuz — ADET DAHİL (2026-08-06).
        # ESKİDEN gün+launch yeterliydi: o güne bir kez teyit verildiyse İKİNCİSİ
        # asla gönderilemiyordu. Gerçek olay (94.LTK.890): aynı gün AYNI referansı
        # İKİ OPERATÖR üretti, yalnız birinin adedi teyit edildi; kalan adet için
        # gönderim "zaten gönderilmiş" diye sessizce atlanıyor ve teyit hiç
        # tamamlanamıyordu. Adet karşılaştırması kısmi teyide izin verir, gerçek
        # mükerrerliği (aynı satırı iki kez göndermek) yine engeller.
        # Aynı gün aynı adet MEŞRU şekilde iki kez gerekiyorsa (iki operatör eşit
        # adet üretmişse) 'zorla' ile gönderilir — arayüzde "yine de gönder" butonu.
        var = conn.execute(
            "SELECT id, adet FROM as400_teyit_log WHERE uretim_tarihi=? AND yil=? AND launch_no=? "
            "AND sonuc='ok' AND CAST(adet AS INTEGER)=?",
            (u_tarih, yil, no, int(adet))).fetchone()
        if var and not zorla:
            sonuclar.append({**kayit, 'sonuc': 'atlandi',
                             'mesaj': f'Bu launch\'a bu gün için {int(adet)} adet ZATEN gönderilmiş '
                                      f'(log #{var["id"]}). Farklı adet göndermek serbest; aynı adedi '
                                      f'tekrar göndermek gerekiyorsa "yine de gönder" seçeneğini kullanın.'})
            continue
        # Mükerrer koruması 1b — REFERANS+GÜN (2026-07-30, 94.LTK.487/10 olayı):
        # yukarıdaki kontrol launch numarasına bağlı; ilk launch kapanıp YENİ launch
        # açılınca (26-150830 → 26-153549) anahtar değişiyor ve aynı üretim günü
        # İKİNCİ KEZ teyit alabiliyordu. Aynı gün + aynı referans için başarılı bir
        # launch teyidi varsa dur. Meşru bölünmüş teyit için 'zorla' ile aşılabilir.
        # CFI GÖNDERİMLERİ DE SAYILIR (2026-08-21). Eskiden bu sorgu
        # `yil NOT IN ('CF','CO')` ile CFI kayıtlarını DIŞLIYORDU; CFI tarafındaki
        # eş kontrol de yalnız yil='CF' satırlarına bakıyordu. Yani iki gönderim
        # yolu birbirini HİÇ görmüyordu ve aynı üretim bir kez CFI'dan, bir kez
        # launch'tan teyit edilebiliyordu — 94.LTK.215'te ERP'de tam bu şekil var
        # (2×RPR + 1×CFI, hepsi aynı gün). Artık 'CO' (COP/hurda — bağımsız bir
        # depo hareketi) dışında her başarılı gönderim engel sayılır.
        if not zorla:
            var2 = conn.execute(
                "SELECT id, yil, launch_no, adet FROM as400_teyit_log "
                "WHERE uretim_tarihi=? AND sonuc='ok' AND yil != 'CO' "
                "AND UPPER(REPLACE(referans,' ',''))=UPPER(REPLACE(?,' ','')) LIMIT 1",
                (u_tarih, referans)).fetchone()
            if var2:
                # ── ADEDE DUYARLI FREN (kullanıcı 2026-08-25) ────────────────
                # Eski hâli ADEDE HİÇ BAKMIYORDU: o güne bir kez teyit verildiyse
                # ikinci parça ASLA gönderilemiyor, meşru bölünmüş teyit için her
                # seferinde "yine de gönder" gerekiyordu — ve o düğme freni TAMAMEN
                # kaldırdığı için gerçek mükerrerlik de aynı düğmeden geçiyordu.
                # Yeni ölçüt: GÜNÜN ÜRETİMİNİ AŞIYOR MUYUZ?
                #   gönderilen + gönderilecek <= o gün üretilen  → MEŞRU, izin ver
                #   aşıyorsa                                     → dur (asıl risk bu)
                # Gerçek olay (10.300.2756W, 2026-08-21): gün 98 üretti (16+10+41+31),
                # 41'i gönderilmişti; panel 98 daha öneriyordu → 139 olurdu.
                _gonderilen = _gun_ref_gonderilen(conn, u_tarih, referans)
                _uretilen = _gun_ref_uretim(u_tarih, referans)
                _nereden = ('CFI olarak' if var2['yil'] == 'CF'
                            else f'launch {var2["launch_no"]} ile')
                if _uretilen is None:
                    # Üretim okunamadı → ESKİ katı davranış (emniyetli taraf)
                    sonuclar.append({**kayit, 'sonuc': 'atlandi',
                                     'mesaj': f'Bu üretim günü ({u_tarih}) için bu referansa zaten teyit '
                                              f'verilmiş: {_nereden}, {var2["adet"]} adet (log #{var2["id"]}). '
                                              f'Günün üretimi okunamadığı için bölünmüş teyit kontrolü '
                                              f'yapılamadı — emin olduğunuz durumda "zorla" ile gönderin.'})
                    continue
                if _gonderilen + adet > _uretilen:
                    _kalan = max(0, int(round(_uretilen - _gonderilen)))
                    sonuclar.append({**kayit, 'sonuc': 'atlandi',
                                     'mesaj': f'ÜRETİMİ AŞIYOR — {u_tarih} günü bu referanstan '
                                              f'{int(_uretilen)} adet üretildi, {int(_gonderilen)} adedi '
                                              f'zaten teyit edildi ({_nereden}, log #{var2["id"]}). '
                                              f'{adet} daha gönderilirse ERP\'ye toplam '
                                              f'{int(_gonderilen + adet)} yazılır. '
                                              + (f'Gönderilebilecek kalan: {_kalan} adet.'
                                                 if _kalan else 'Bu günün teyidi TAMAMLANMIŞ.')})
                    continue
                # Aşmıyor → MEŞRU bölünmüş teyit, engelleme yok (loga düş)
                print(f'[TEYIT] bölünmüş teyit: {referans} {u_tarih} — '
                      f'gönderilen {int(_gonderilen)} + {adet} <= üretim {int(_uretilen)}')
        # Mükerrer koruması 2 (ASIL): ERP'nin GERÇEK teyit hareketleri —
        # operatörün ELLE girdiği teyitler de burada görünür. Frontend
        # atlansa/eski liste gönderilse bile burada yakalanır.
        if not zorla:
            try:
                import launch_esle as _le2
                hrk = _le2.teyit_hareketleri([article]).get(_le2.kanonik(article), [])
                zt, ilgili = _le2._zaten_teyitli(hrk, u_tarih, adet,
                                                 _le2.ref_uretim_gecmisi(referans, u_tarih))
            except Exception:
                zt, ilgili = None, []
            if zt == 'kesin':
                det = ', '.join(f"{h['tarih']}: {h['adet']:g}" for h in ilgili)
                sonuclar.append({**kayit, 'sonuc': 'atlandi',
                                 'mesaj': f'ERP\'de bu üretim için zaten teyit var ({det}) — mükerrer olurdu'})
                continue
            # Mükerrer koruması 3: KAPASİTE (2026-07-30) — hayali stok girişini engeller
            kap = _kapasite_reddi(conn, referans, u_tarih, adet)
            if kap:
                sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': kap})
                continue

        bayrak = 'S' if str(s.get('bayrak') or 'A').upper() == 'S' else 'A'
        try:
            once, once_durum = _as400_launch_durum(yil, no)
        except Exception as e:
            sonuclar.append({**kayit, 'bayrak': bayrak, 'sonuc': 'hata', 'mesaj': f'Ön okuma hatası: {e}'})
            continue

        cikti, robot_hata = _as400_robot_calistir('teyit_gir.js', [yil, no, article, adet, bayrak], 150)
        if robot_hata:
            # ROBOT HATA VERDİ AMA YAZMIŞ OLABİLİR (2026-08-17). Eskiden burada
            # sonra-okuması YAPILMIYORDU: "robot iptal" denip geçiliyor, ekranda
            # gerçekten teyit girilmişse KİMSE görmüyordu. Çıktısız/hatalı koşuda
            # tek güvenilir kanıt BPROF0'ın teyitli değeridir — okuyup söyle.
            try:
                sonra_h, _ = _as400_launch_durum(yil, no)
            except Exception:
                sonra_h = None
            if once is not None and sonra_h is not None and abs(sonra_h - once) > 0.001:
                robot_hata += (f'  ⚠ DİKKAT: robot hata verdi AMA launch teyitlisi DEĞİŞTİ '
                               f'({once:g} → {sonra_h:g}) — ERP\'ye yazılmış olabilir, TEKRAR '
                               f'GÖNDERMEDEN ÖNCE Session B\'yi ve launch\'ı kontrol edin.')
                # Log'a da 'hata' olarak düşsün ki mükerrer freni bunu görebilsin
                conn.execute(
                    "INSERT INTO as400_teyit_log (uretim_tarihi, yil, launch_no, referans, article, adet, bayrak, sonuc, mesaj, teyitli_once, teyitli_sonra, olusturan) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (u_tarih, yil, no, referans, article, adet, bayrak, 'hata', robot_hata, once, sonra_h, kullanici))
                conn.commit()
            elif once is not None and sonra_h is not None:
                robot_hata += f'  ✔ Kontrol edildi: launch teyitlisi DEĞİŞMEDİ ({once:g}) — ERP\'ye yazılmadı.'
            sonuclar.append({**kayit, 'bayrak': bayrak, 'sonuc': 'hata', 'mesaj': robot_hata,
                             'teyitli_once': once, 'teyitli_sonra': sonra_h})
            continue

        robot_ok = 'SONUC=OK' in cikti
        try:
            sonra, sonra_durum = _as400_launch_durum(yil, no)
        except Exception:
            sonra, sonra_durum = None, None
        # Doğrulama: teyitli tam +adet artmalı; S ise launch KAPANMIŞ (durum 70) olmalı.
        teyit_ok = (once is not None and sonra is not None and abs(sonra - once - adet) < 0.001)
        kapanma_ok = (bayrak == 'A') or (sonra_durum == 70)
        dogrulandi = robot_ok and teyit_ok and kapanma_ok
        son_satirlar = ' | '.join(l for l in cikti.splitlines() if l.strip())[-500:]
        sonuc = 'ok' if dogrulandi else 'hata'
        if dogrulandi:
            mesaj = f'Doğrulandı: teyitli {once:g} → {sonra:g}' + (
                f' · launch KAPANDI (durum {sonra_durum})' if bayrak == 'S' else '')
            mesaj += _is_emri_dus_router(conn, referans, adet)
            # COP (hurda) BURADA GİRİLMEZ (kullanıcı 2026-07-23: robot Rientro↔07>01
            # mekik dokuyordu) — hurda ayrı /cop_gonder fazında EN SONA girilir.
        elif robot_ok and teyit_ok and not kapanma_ok:
            mesaj = f'Teyit işlendi ({once:g}→{sonra:g}) ama S ile KAPANMADI (durum {sonra_durum}) — kontrol edin'
        elif robot_ok:
            mesaj = f'Robot OK ama doğrulama tutmadı (önce {once} sonra {sonra}, durum {sonra_durum})'
        else:
            mesaj = (f'Robot iptal: {son_satirlar}' if son_satirlar
                     else 'Robot iptal — çıktı yok (robot ekrana hiç ulaşamadı)')
        conn.execute(
            "INSERT INTO as400_teyit_log (uretim_tarihi, yil, launch_no, referans, article, adet, bayrak, sonuc, mesaj, teyitli_once, teyitli_sonra, olusturan) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (u_tarih, yil, no, referans, article, adet, bayrak, sonuc, mesaj, once, sonra, kullanici))
        conn.commit()
        sonuclar.append({**kayit, 'bayrak': bayrak, 'sonuc': sonuc, 'mesaj': mesaj,
                         'teyitli_once': once, 'teyitli_sonra': sonra})
        # Hata olsa bile SONRAKİ satıra devam (durdurma YOK).

    return sonuclar


@app.route('/api/as400/teyit_gonder', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_gonder():
    """Seçili satırları AS400'e teyit olarak girer (Session B ekran robotu).
    Body: {uretim_tarihi, satirlar: [{yil, no, article, referans, adet, bayrak}], zorla}
    Çekirdek _teyit_gonder_calistir'da (17:10 oto koşusuyla ORTAK). Aynı üretim
    günü + launch için ikinci gönderim reddedilir (zorla=true hariç). Her satır
    sonrası BPROF0.Q0QTRI ile bağımsız doğrulama."""
    data = request.get_json(silent=True) or {}
    varsayilan_tarih = (data.get('uretim_tarihi') or '').strip()
    satirlar = data.get('satirlar') or []
    zorla = bool(data.get('zorla'))
    if not satirlar:
        return jsonify({'hata': 'satirlar zorunlu'}), 400
    if len(satirlar) > 60:
        return jsonify({'hata': 'Tek seferde en fazla 60 satır'}), 400
    if not _AS400_KILIT.acquire(blocking=False):
        return jsonify({'hata': 'Devam eden bir AS400 gönderimi var — bitmesini bekleyin'}), 409
    try:
        _AS400_DURDUR.clear()   # yeni gönderim → önceki "durdur" bayrağını sıfırla
        sonuclar = _teyit_gonder_calistir(get_db(), satirlar, g.panel_ku['kullanici_adi'],
                                          varsayilan_tarih, zorla)
        ozet = {k: sum(1 for r in sonuclar if r['sonuc'] == k) for k in ('ok', 'hata', 'atlandi')}
        return jsonify({'ozet': ozet, 'sonuclar': sonuclar, 'durduruldu': _AS400_DURDUR.is_set()})
    finally:
        _AS400_DURDUR.clear()
        _AS400_KILIT.release()


@app.route('/api/as400/teyit_durum', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_durum():
    """Şu an bir AS400 gönderimi çalışıyor mu? (Gönder butonundan sonra Durdur'u
    doğru göstermek için sayfa açılışında sorulur.)"""
    return jsonify({'calisiyor': _AS400_KILIT.locked(), 'durduruluyor': _AS400_DURDUR.is_set()})


@app.route('/api/as400/teyit_durdur', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_durdur():
    """Devam eden gönderimi NAZİKÇE durdurur: bayrağı set eder; döngü işlenmekte
    olan launch'ı bitirince kalan satırlara GEÇMEZ. Çalışan cscript ASLA öldürülmez
    (record-lock riski). Gönderim yoksa 'bos' döner."""
    if not _AS400_KILIT.locked():
        return jsonify({'durum': 'bos', 'mesaj': 'Şu an devam eden gönderim yok'})
    _AS400_DURDUR.set()
    return jsonify({'durum': 'durduruluyor',
                    'mesaj': 'Durdurma istendi — işlenmekte olan launch bitince duracak'})


def _is_emri_dus(conn, referans, adet):
    """Başarılı teyit sonrası İŞ YÖNETİMİ kit düşümü (kullanıcı 2026-07-22):
    kaynak/montaj iş emrinden (referans_takip) teyit edilen adet düşülür;
    kalan <= 3 ise (tam / fazla / 3 eksik üretim) iş emri LİSTEDEN SİLİNİR —
    'üretildi ama listeden silinmesi unutuldu' problemi biter.
    Eşleşme gevşek (boşluk+nokta+büyük/küçük duyarsız). Döner: ek mesaj | ''."""
    try:
        adet = int(adet or 0)
    except (TypeError, ValueError):
        return ''
    if not referans or adet <= 0:
        return ''
    try:
        satir = conn.execute(
            "SELECT id, referans_kodu, hedef_adet, bolum FROM referans_takip "
            "WHERE COALESCE(bolum,'') IN ('kaynak','montaj') "
            "  AND COALESCE(lokasyon,'TK2')='TK2' "
            "  AND COALESCE(hedef_adet,0) > 0 "
            "  AND UPPER(REPLACE(REPLACE(referans_kodu,' ',''),'.','')) = "
            "      UPPER(REPLACE(REPLACE(?,' ',''),'.','')) "
            "ORDER BY id LIMIT 1", (referans,)).fetchone()
        if not satir:
            return ''
        kalan = int(satir['hedef_adet']) - adet
        if kalan <= 3:
            conn.execute("DELETE FROM referans_takip WHERE id=?", (satir['id'],))
            conn.commit()
            return f" · iş emri listeden SİLİNDİ ({satir['bolum']}, hedef {satir['hedef_adet']}, kalan {kalan})"
        conn.execute(
            "UPDATE referans_takip SET hedef_adet=?, guncelleme_tarihi=datetime('now','localtime') WHERE id=?",
            (kalan, satir['id']))
        conn.commit()
        return f" · iş emrinden düşüldü ({satir['bolum']}: kalan {kalan})"
    except Exception as e:
        return f" · iş emri düşümü yapılamadı: {e}"


# ── LAPTOP→SERVER İŞ EMRİ DÜŞÜM KÖPRÜSÜ (2026-07-24) ──
# Teyit LAPTOPTA koşuyor (PCOMM) ama iş takip verisi SERVER'da (canlı DB). Bu yüzden
# laptop'un yerel düşümü kullanıcının gördüğü listeye ulaşmıyordu. Köprü: LAPTOP'ta
# server_sync.json ({"url":"http://<server>:5000"}) VARSA teyit sonrası düşüm SERVER'a
# HTTP ile yollanır; server yoksa/ulaşılamazsa yerel düşüme düşülür. SERVER tarafında EK
# CONFIG/TOKEN GEREKMEZ — endpoint yalnız İÇ AĞ'dan (LAN) kabul eder, dış (cloudflared) red.
def _server_sync_url():
    """Laptop client modu: server_sync.json içindeki url (yoksa None = yerel mod).
    Bu dosya makineye özel (gitignore) — SADECE laptopta bulunur, server'da YOK."""
    try:
        p = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'server_sync.json')
        if not os.path.exists(p):
            return None
        with open(p, encoding='utf-8') as f:
            return (json.load(f).get('url') or None)
    except Exception as e:
        print(f'[server_sync] config okunamadı: {e}')
        return None


def _is_ic_ag_istegi():
    """dus_sync makine-arası İÇ uç noktası: yalnız doğrudan LAN'dan (192.168.x) kabul;
    Cloudflare tüneli (CF-Ray/CF-Connecting-IP header'lı DIŞ istek) reddedilir.
    (App ProxyFix kullanmıyor → remote_addr gerçek TCP peer, spoof edilemez.)"""
    if request.headers.get('CF-Ray') or request.headers.get('CF-Connecting-IP'):
        return False
    return (request.remote_addr or '').startswith('192.168.')


def _is_emri_sunucuya_bildir(url, referans, adet):
    """Laptop: başarılı teyit sonrası server'ın iş takip düşümünü HTTP ile tetikler.
    Best-effort. Döner: (ok:bool, mesaj:str)."""
    try:
        import requests
        r = requests.post(url.rstrip('/') + '/api/is_emri/dus_sync',
                          json={'referans': referans, 'adet': adet}, timeout=8)
        if r.status_code == 200:
            return True, ((r.json() or {}).get('mesaj', '') or '')
        return False, f" · ⚠ iş emri server düşümü başarısız (HTTP {r.status_code})"
    except Exception as e:
        return False, f" · ⚠ iş emri server'a ULAŞILAMADI ({e})"


def _is_emri_dus_router(conn, referans, adet):
    """server_sync url varsa (laptop) düşümü SERVER'da yap (authoritative); server
    ulaşılamazsa YEREL düşüme düş + uyarı (interim kayıp olmasın). url yoksa yerel."""
    url = _server_sync_url()
    if not url:
        return _is_emri_dus(conn, referans, adet)          # server/standalone
    ok, mesaj = _is_emri_sunucuya_bildir(url, referans, adet)
    if ok:
        return mesaj                                       # server authoritative (boş = eşleşme yok)
    return _is_emri_dus(conn, referans, adet) + mesaj + ' — elle kontrol'


@app.route('/api/is_emri/dus_sync', methods=['POST'])
def api_is_emri_dus_sync():
    """Makine-arası iş emri düşümü (laptop→server). Yalnız İÇ AĞ'dan (LAN) kabul edilir
    (dış/cloudflared reddedilir); panel oturumu/token GEREKMEZ. Server KENDİ
    referans_takip'inde _is_emri_dus çalıştırır."""
    if not _is_ic_ag_istegi():
        return jsonify({'hata': 'yalnız iç ağ (LAN) erişimi'}), 403
    data = request.get_json(silent=True) or {}
    referans = (data.get('referans') or '').strip()
    try:
        adet = int(data.get('adet') or 0)
    except (TypeError, ValueError):
        adet = 0
    conn = get_db()
    mesaj = _is_emri_dus(conn, referans, adet)
    return jsonify({'mesaj': mesaj, 'dustu': bool(mesaj)})


def _as400_cfi_bugun(article, causal='CFI'):
    """BUGÜN girilen CFI/COP hareket adetleri (bağımsız doğrulama kanalı).
    Döner: [adet, ...] — BMMAF0 MGCACD=causal, hareket tarihi = bugün."""
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import keyring, pyodbc
    import as400_config as _cfg
    bugun = date.today()
    pw = _cfg.sifre_al()
    cn = pyodbc.connect(_cfg.baglanti_dizesi(pw), timeout=20, autocommit=True)
    try:
        # HARF DUYARLILIGI (2026-08-17): MGARCD=? karsilastirmasi harf duyarli.
        # '10.130.6206b' COP'u ERP'ye BASARIYLA girildi (ekranda 10.130.6206B
        # olarak duruyor) ama bu sorgu kucuk harfli kodu arayip bulamadi →
        # dogrulama 'False' dondu → satir 'hata' loglandi → panel COP'u hala
        # BEKLIYOR gosterdi. Kodun iki bicimini de sor (indeks korunur).
        rows = cn.cursor().execute(
            "SELECT MGQTA FROM tkc0301F.BMMAF0 WHERE MGCACD=? AND MGARCD IN (?, ?) "
            "AND MGDSSO=? AND MGDAAO=? AND MGDMMO=? AND MGDGGO=?",
            (causal, article, str(article).upper(),
             bugun.year // 100, bugun.year % 100, bugun.month, bugun.day)).fetchall()
        return [float(r[0] or 0) for r in rows]
    finally:
        cn.close()


# TK1 plastikte teyit AYNI DEPOYA gitmez: her urunun Warehouse cd ve
# Counterpart War.cd'si farkli (kullanici 2026-08-25). Kodlar referans kartinda
# tanimlanir (referans_listesi.depo_kodu / karsi_depo_kodu).
# BOSSA TEYIT VERILMEZ (kullanici karari): yanlis depoya stok yazmak, gec
# teyitten cok daha pahalidir ve fark edilmesi zordur.
CFI_VARSAYILAN_DEPO = ('01D', '01D')      # TK2 davranisi — cfi_gir.js varsayilani
CFI_DEPO_ZORUNLU_KAPSAM = {('TK1', 'plastik')}


def _cfi_depo_kodlari(conn, referans, tesis, bolum):
    """(warehouse, counterpart, hata_mesaji). Hata varsa satir GONDERILMEZ."""
    kapsam = (str(tesis or '').upper(), str(bolum or ''))
    if kapsam not in CFI_DEPO_ZORUNLU_KAPSAM:
        return CFI_VARSAYILAN_DEPO[0], CFI_VARSAYILAN_DEPO[1], None
    r = conn.execute(
        "SELECT COALESCE(depo_kodu,'') d, COALESCE(karsi_depo_kodu,'') k "
        "FROM referans_listesi "
        "WHERE UPPER(REPLACE(referans_kodu,' ',''))=UPPER(REPLACE(?,' ','')) "
        "  AND COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')=? LIMIT 1",
        (referans, kapsam[1], kapsam[0])).fetchone()
    d = (r['d'] or '').strip() if r else ''
    k = (r['k'] or '').strip() if r else ''
    if not d:
        return None, None, ('Depo kodu tanımlı değil — Referanslar sayfasında bu referansın '
                            '"Warehouse cd" alanını doldurun (yanlış depoya stok yazılmasın diye '
                            'teyit gönderilmedi).')
    if not k:
        return None, None, ('Counterpart War.cd tanımlı değil — Referanslar sayfasında doldurun '
                            '(teyit gönderilmedi).')
    return d, k, None


def _gun_ref_gonderilen(conn, uretim_tarihi, referans):
    """Bu ÜRETİM GÜNÜ + referans için BİZİM başarıyla gönderdiğimiz toplam adet.
    COP (yil='CO') sayılmaz — hurda ayrı bir depo hareketidir, üretim teyidi değil."""
    r = conn.execute(
        "SELECT COALESCE(SUM(CAST(adet AS REAL)),0) t FROM as400_teyit_log "
        "WHERE uretim_tarihi=? AND sonuc='ok' AND yil != 'CO' "
        "AND UPPER(REPLACE(referans,' ',''))=UPPER(REPLACE(?,' ',''))",
        (uretim_tarihi, referans)).fetchone()
    return float(r['t'] or 0) if r else 0.0


def _gun_ref_uretim(uretim_tarihi, referans, _onbellek={}):
    """O günün o referans için ÜRETİLEN toplamı (rework hariç).

    launch_esle.gun_uretimi ile AYNI kuralları kullanır (rework ayıklama, tesis/
    bölüm kapsamı) — burada ayrı bir SQL yazmak iki kuralın zamanla ayrışması
    demekti. Gün başına bir kez okunur, sonuç önbelleğe alınır."""
    if uretim_tarihi not in _onbellek:
        try:
            import sys as _sys
            _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
            if _d not in _sys.path:
                _sys.path.insert(0, _d)
            import launch_esle as _le
            harita = {}
            for u in _le.gun_uretimi(uretim_tarihi):
                harita[_le.kanonik(u.get('referans'))] = (
                    harita.get(_le.kanonik(u.get('referans')), 0) + (u.get('adet') or 0))
            _onbellek[uretim_tarihi] = harita
        except Exception as e:
            print(f'[_gun_ref_uretim] okunamadi ({uretim_tarihi}): {e}')
            _onbellek[uretim_tarihi] = None
    harita = _onbellek.get(uretim_tarihi)
    if harita is None:
        return None            # okunamadı → çağıran ESKİ (katı) davranışa düşer
    try:
        import launch_esle as _le
        return harita.get(_le.kanonik(referans), 0)
    except Exception:
        return None


def _cfi_gonder_calistir(conn, satirlar, kullanici, zorla=False):
    """CFI depo girişi satırlarını robotla işler — endpoint VE 17:10 oto koşusu
    ortak çekirdeği. ÇAĞIRAN _AS400_KILIT'i tutuyor olmalı. Döner: sonuclar.
    Kullanıcı akışı (2026-07-20): Warehouse=01D, Causal=CFI, Counterpart=01D,
    kod + adet + Enter. Mükerrer koruması: (1) as400_teyit_log (yil='CF'),
    (2) BMMAF0 RPR+CFI hareketleri (_zaten_teyitli). Doğrulama: bugünkü CFI
    hareketlerinde adet birebir aranır."""
    sonuclar = []
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import launch_esle as _le
    # Session B ön kontrolü — ölü oturumda her satır tek tek hata verirdi
    _od, _om = _as400_oturum_hazirla()
    if _od == 'OK':
        print(f'[CFI] {_om}')
    elif _od is None:
        print(f'[CFI] oturum ön kontrolü: {_om}')
    for idx, s in enumerate(satirlar):
        if _AS400_DURDUR.is_set():
            for kalan in satirlar[idx:]:
                sonuclar.append({'referans': str(kalan.get('referans') or '').strip(),
                                 'article': str(kalan.get('article') or '').strip(),
                                 'adet': kalan.get('adet'), 'sonuc': 'atlandi',
                                 'mesaj': 'Kullanıcı durdurdu'})
            break
        referans = str(s.get('referans') or '').strip()
        article = (str(s.get('article') or '').strip() or referans)
        u_tarih = str(s.get('uretim_tarihi') or '').strip()
        try:
            adet = int(s.get('adet') or 0)
        except (TypeError, ValueError):
            adet = 0
        kayit = {'referans': referans, 'article': article, 'adet': adet, 'uretim_tarihi': u_tarih}
        if not (article and 0 < adet <= 99999 and u_tarih):
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': 'Geçersiz satır parametresi'})
            continue
        # Mükerrer 1: kendi log'umuz (CFI kayıtları yil='CF', launch_no=article).
        # ADET DAHİL (2026-08-06) — bkz. launch tarafındaki aynı düzeltme: aynı gün
        # aynı referansı iki operatör üretince kalan adet gönderilemiyordu.
        var = conn.execute(
            "SELECT id, adet FROM as400_teyit_log WHERE uretim_tarihi=? AND yil='CF' AND launch_no=? "
            "AND sonuc='ok' AND CAST(adet AS INTEGER)=?",
            (u_tarih, article, int(adet))).fetchone()
        if var and not zorla:
            sonuclar.append({**kayit, 'sonuc': 'atlandi',
                             'mesaj': f'Bu koda bu gün için {int(adet)} adet CFI ZATEN gönderilmiş '
                                      f'(log #{var["id"]}). Farklı adet göndermek serbest.'})
            continue
        # Mükerrer 1b — LAUNCH GÖNDERİMLERİ DE SAYILIR (2026-08-21, 94.LTK.215).
        # Yukarıdaki kontrol yalnız yil='CF' satırlarına bakar; launch tarafındaki
        # eş kontrol de CFI'yı dışlıyordu → aynı üretim bir kez launch'tan, bir kez
        # CFI'dan teyit edilebiliyordu. İki yol artık birbirinin log'unu görüyor.
        if not zorla:
            var1b = conn.execute(
                "SELECT id, yil, launch_no, adet FROM as400_teyit_log "
                "WHERE uretim_tarihi=? AND sonuc='ok' AND yil NOT IN ('CF','CO') "
                "AND UPPER(REPLACE(referans,' ',''))=UPPER(REPLACE(?,' ','')) LIMIT 1",
                (u_tarih, referans)).fetchone()
            if var1b:
                sonuclar.append({**kayit, 'sonuc': 'atlandi',
                                 'mesaj': f'Bu üretim günü ({u_tarih}) için bu referansa LAUNCH '
                                          f'teyidi zaten verilmiş: {var1b["yil"]}-{var1b["launch_no"]}, '
                                          f'{var1b["adet"]} adet (log #{var1b["id"]}). Aynı üretim hem '
                                          f'launch hem CFI ile teyit edilmemeli — gerçekten gerekiyorsa '
                                          f'"zorla" ile gönderin.'})
                continue
        # Mükerrer 2: ERP hareketleri (RPR+CFI) — HEM article HEM üretim referansı
        # (2026-07-21 4082W dersi: launch teyidi gerçek article'a, CFI adayı çıplak
        # koda bakıyordu → aynı üretim iki kez gitti). Aynı-gün + geçmiş-açıklama kurallı.
        try:
            hrk_map = _le.teyit_hareketleri([article, referans])
            hrk = list(hrk_map.get(_le.kanonik(article), []))
            if _le.kanonik(referans) != _le.kanonik(article):
                hrk += hrk_map.get(_le.kanonik(referans), [])
            zt, ilgili = _le._zaten_teyitli(hrk, u_tarih, adet,
                                            _le.ref_uretim_gecmisi(referans, u_tarih))
        except Exception:
            zt, ilgili = None, []
        if zt == 'kesin' and not zorla:
            h0 = ilgili[0] if ilgili else {}
            sonuclar.append({**kayit, 'sonuc': 'atlandi',
                             'mesaj': f"ERP'de karşılığı var: {h0.get('tarih','')} {h0.get('adet',0):g} adet ({h0.get('tur','')})"})
            continue
        # Mükerrer 3: KAPASİTE (2026-07-30) — CFI de depoya stok yazar, aynı fren
        if not zorla:
            kap = _kapasite_reddi(conn, referans, u_tarih, adet)
            if kap:
                sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': kap})
                continue
        # KOD DOĞRULAMA (2026-07-30): article ERP master'ında yoksa robota GÖNDERME.
        # '10.300.1992A.OP.1' vakası: operatör kod alanına operasyon bilgisi yazmış;
        # robot ekrana yazamayınca alan taştı ve AYNI KOŞUDAKİ DİĞER satırlar da
        # gönderilemedi. Tek bozuk kod tüm kuyruğu durduruyordu.
        # ZORLA = kullanıcının açık onayı → master kontrolü ATLANIR (2026-08-17).
        # Gerekçe: bu kontrol ERP master'ını okuyamadığında/kod harf farkıyla
        # bulunamadığında MEŞRU teyidi de reddediyordu ('10.130.6202b' vakası).
        # Kod artık ekranı KİLİTLEMİYOR (cfi_gir iptalinde ana menüye dönüyor),
        # yani zorlamanın bedeli tek satırın hatası — kuyruk durmuyor.
        gecersiz = None if zorla else _article_gecersiz_mi(article)
        if gecersiz:
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': gecersiz})
            continue
        # DEPO KODLARI — TK1 plastikte referanstan gelir, eksikse SATIR ATLANIR.
        # 'zorla' bunu ATLAMAZ: eksik depo kodu bir kod-eşleşme şüphesi değil,
        # bilginin HİÇ olmamasıdır; zorlamak yanlış depoya yazmak demektir.
        _wh, _cp, _depo_hata = _cfi_depo_kodlari(
            conn, referans, s.get('tesis'), s.get('bolum'))
        if _depo_hata:
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': _depo_hata})
            continue
        _robot_arg = [article, adet, 'CFI', f'WH={_wh}', f'CP={_cp}']
        cikti, robot_hata = _as400_robot_calistir('cfi_gir.js', _robot_arg, 120)
        if robot_hata:
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': robot_hata})
            continue
        robot_ok = 'SONUC=OK' in cikti
        # Bağımsız doğrulama: bugünkü CFI hareketlerinde adet birebir var mı
        try:
            dogru = any(abs(q - adet) < 0.001 for q in _as400_cfi_bugun(article))
        except Exception:
            dogru = None   # okunamadı — yalnız robot çıktısına güven
        dogrulandi = robot_ok and (dogru is not False)
        son_satirlar = ' | '.join(l for l in cikti.splitlines() if l.strip())[-400:]
        sonuc = 'ok' if dogrulandi else 'hata'
        if dogrulandi:
            mesaj = f'CFI girildi: {article} → {adet} adet' + ('' if dogru else ' (hareket sorgusu doğrulanamadı, robot OK)')
            mesaj += _is_emri_dus_router(conn, referans, adet)
            # COP (hurda) ayrı fazda EN SONA girilir (bkz. /api/as400/cop_gonder)
        elif robot_ok:
            mesaj = f'Robot OK ama bugünkü CFI hareketlerinde {adet} bulunamadı — elle kontrol edin'
        else:
            mesaj = (f'Robot iptal: {son_satirlar}' if son_satirlar
                     else 'Robot iptal — çıktı yok (robot ekrana hiç ulaşamadı)')
        conn.execute(
            "INSERT INTO as400_teyit_log (uretim_tarihi, yil, launch_no, referans, article, adet, bayrak, sonuc, mesaj, olusturan) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (u_tarih, 'CF', article, referans, article, adet, 'CFI', sonuc, mesaj, kullanici))
        conn.commit()
        sonuclar.append({**kayit, 'sonuc': sonuc, 'mesaj': mesaj})
    return sonuclar


@app.route('/api/as400/cfi_gonder', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_cfi_gonder():
    """Launch'sız üretim için CFI depo girişi (Session B robotu, 07>01>F1 ekranı).
    Body: {satirlar: [{referans, article?, adet, uretim_tarihi}], zorla}
    Çekirdek _cfi_gonder_calistir'da (17:10 oto koşusuyla ORTAK)."""
    data = request.get_json(silent=True) or {}
    satirlar = data.get('satirlar') or []
    zorla = bool(data.get('zorla'))
    if not satirlar:
        return jsonify({'hata': 'satirlar zorunlu'}), 400
    if len(satirlar) > 60:
        return jsonify({'hata': 'Tek seferde en fazla 60 satır'}), 400
    if not _AS400_KILIT.acquire(blocking=False):
        return jsonify({'hata': 'Devam eden bir AS400 gönderimi var — bitmesini bekleyin'}), 409
    try:
        _AS400_DURDUR.clear()
        sonuclar = _cfi_gonder_calistir(get_db(), satirlar, g.panel_ku['kullanici_adi'], zorla)
        ozet = {k: sum(1 for r in sonuclar if r['sonuc'] == k) for k in ('ok', 'hata', 'atlandi')}
        return jsonify({'ozet': ozet, 'sonuclar': sonuclar, 'durduruldu': _AS400_DURDUR.is_set()})
    finally:
        _AS400_DURDUR.clear()
        _AS400_KILIT.release()


# ── AS400 PLANLAMA (kaynak MRP yardımcısı, kullanıcı 2026-07-23 akışı öğretti) ──
# Manuel akış: 07>10>01 stok → sipariş yoksa üst kodlar (implosion) → üst siparişleri
# → üretilecekse alt parça stok kontrolü → launch. Dosyalar (keşif 2026-07-23,
# çapalarla doğrulandı): STOK=TKC0300F.BSMAF0 (partita nedeniyle DAİMA GROUP BY);
# SİPARİŞ=TKC0301F.BORDF1 (açık: OCDOSO<>'C' AND OCQTOR>OCQTCO; PR/BFABF0 KULLANMA —
# operasyon süreli MRP önerileri); BOM=TKC0301F.BSPEF2 (Y2ARTI üst, Y2RICD alt,
# Y2IVQT birim miktar — Y2QTA hep 0, kullanma). CHAR(21) alanlar boşluk dolgulu →
# eşitlik parametreyle (DB2 blank-pad), TRIM yalnız SELECT'te.

def _plan_stoklar(cn, articles):
    """{artikel: {'depolar': {depo: {...}}, 'kullanilabilir': 01D+CF2+MK2+MT2 stok toplamı}}"""
    if not articles:
        return {}
    yer = ','.join('?' * len(articles))
    sonuc = {}
    for r in cn.cursor().execute(
            f"SELECT TRIM(S0ARTI), TRIM(S0MGCD), SUM(S0GIAD), SUM(S0IMPP)+SUM(S0IMPC), "
            f"SUM(S0ORDF)+SUM(S0ORDP) FROM TKC0300F.BSMAF0 "
            f"WHERE S0SOCI='01' AND S0ARTI IN ({yer}) GROUP BY TRIM(S0ARTI), TRIM(S0MGCD)",
            list(articles)):
        art, depo = r[0], r[1]
        stok, engaged, ordered = float(r[2] or 0), float(r[3] or 0), float(r[4] or 0)
        if not (stok or engaged or ordered):
            continue
        a = sonuc.setdefault(art, {'depolar': {}, 'kullanilabilir': 0.0, 'toplam_stok': 0.0})
        a['depolar'][depo] = {'stok': stok, 'engaged': engaged, 'ordered': ordered,
                              'avail': stok - engaged + ordered}
        a['toplam_stok'] += stok
        if depo in ('01D', 'CF2', 'MK2', 'MT2'):
            a['kullanilabilir'] += stok
    return sonuc


def _plan_siparisler(cn, articles):
    """{artikel: {'acik_toplam': N, 'satirlar': [{siparis, musteri, adet, acik, teslim}]}}"""
    if not articles:
        return {}
    yer = ','.join('?' * len(articles))
    sonuc = {}
    for r in cn.cursor().execute(
            f"SELECT TRIM(d.OCARCD), d.OCRED2, d.OCRENU, TRIM(a.ABRGSC), "
            f"d.OCQTOR, d.OCQTCO, d.OCCCD2, d.OCCCD3, d.OCCCD4 "
            f"FROM TKC0301F.BORDF1 d LEFT JOIN TKC0301F.BANAF0 a ON a.ABCFCD=d.OCCFCD "
            f"WHERE d.OCARCD IN ({yer}) AND d.OCDOSO<>'C' AND d.OCQTOR>d.OCQTCO "
            f"ORDER BY d.OCCCD2, d.OCCCD3, d.OCCCD4",
            list(articles)):
        art = r[0]
        acik = float(r[4] or 0) - float(r[5] or 0)
        s = sonuc.setdefault(art, {'acik_toplam': 0.0, 'satirlar': []})
        s['acik_toplam'] += acik
        if len(s['satirlar']) < 10:
            # EBCDIC(1026)→ODBC çevirisinde Türkçe harfler bozuk gelir: "=Ü £=Ö $=İ §=Ş
            musteri = str(r[3] or '').strip().translate(str.maketrans('"£$§', 'ÜÖİŞ'))
            s['satirlar'].append({
                'siparis': f'{int(r[1])}-{int(r[2])}',
                'musteri': musteri,
                'acik': acik,
                'teslim': f'20{int(r[6] or 0):02d}-{int(r[7] or 0):02d}-{int(r[8] or 0):02d}',
            })
    return sonuc


def _cop_gonder_calistir(conn, satirlar, kullanici, zorla=False):
    """HURDA (COP) girişi satırlarını robotla işler — endpoint VE 17:10 oto koşusu
    ortak çekirdeği. ÇAĞIRAN _AS400_KILIT'i tutuyor olmalı. Döner: sonuclar.
    COP'lar EN SON faz (kullanıcı 2026-07-23: launch/CFI fazlarına ARAYA GİRMESİN;
    robot Rientro↔07>01 mekik dokuyordu). Tüm hurdalar ardışık girilir: B,
    07>01>F1 ekranında kalır (alanlar kalıcı — hızlı)."""
    sonuclar = []
    for idx, s in enumerate(satirlar):
        if _AS400_DURDUR.is_set():
            for kalan in satirlar[idx:]:
                sonuclar.append({'referans': str(kalan.get('referans') or '').strip(),
                                 'article': str(kalan.get('article') or '').strip(),
                                 'adet': kalan.get('adet'), 'bayrak': 'COP',
                                 'sonuc': 'atlandi', 'mesaj': 'Kullanıcı durdurdu'})
            break
        referans = str(s.get('referans') or '').strip()
        article = (str(s.get('article') or '').strip() or referans)
        u_tarih = str(s.get('uretim_tarihi') or '').strip()
        try:
            adet = int(s.get('adet') or 0)
        except (TypeError, ValueError):
            adet = 0
        kayit = {'referans': referans, 'article': article, 'adet': adet,
                 'uretim_tarihi': u_tarih, 'bayrak': 'COP'}
        if not (article and 0 < adet <= 99999 and u_tarih):
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': 'Geçersiz satır parametresi'})
            continue
        # Mükerrer: aynı üretim günü + article + AYNI ADET için COP girildiyse atla.
        # ADET DAHİL (2026-08-06) — launch/CFI ile aynı gerekçe: aynı gün iki
        # operatör hurda çıkarırsa ikincisinin adedi de girilebilmeli.
        # KARSILASTIRMA NORMALIZE (2026-08-17): eskiden launch_no=? HARF/BOSLUK
        # DUYARLI esitlikti. Panel ise anahtari normalize ederek (bosluk sil+upper)
        # kuruyor. Ikisi ayrismasi KISIR DONGU uretiyordu: gonderim "zaten girilmis"
        # diye atliyor, panel ayni satiri "bekliyor" gostermeye devam ediyordu ve
        # satir hicbir zaman listeden dusmuyordu (kullanici 10.130.6206b vakasi).
        # Artik iki taraf AYNI normalizasyonu kullanir.
        _norm = "REPLACE(REPLACE(REPLACE(UPPER(launch_no),' ',''),CHAR(9),''),CHAR(10),'')"
        var = conn.execute(
            f"SELECT id, uretim_tarihi, launch_no, adet, created_at FROM as400_teyit_log "
            f"WHERE uretim_tarihi=? AND yil='CO' AND sonuc='ok' AND CAST(adet AS INTEGER)=? "
            f"AND {_norm} = ?",
            (u_tarih, int(adet), re.sub(r'\s+', '', str(article or '')).upper())).fetchone()
        if var:
            # Mesaj log satirini KIMLIKLENDIRIR: panel yine bekliyor gosteriyorsa
            # anahtarin nerede ayristigi buradan gorulur (sessiz dongu bitti).
            sonuclar.append({**kayit, 'sonuc': 'atlandi',
                             'mesaj': (f'Bu üretim günü için hurda COP zaten girilmiş '
                                       f'(log #{var["id"]}: {var["uretim_tarihi"]} · '
                                       f'kod "{var["launch_no"]}" · {var["adet"]} adet · '
                                       f'{var["created_at"]}). Satır listeden düşmüyorsa '
                                       f'bu kaydı bildir.')})
            continue
        # Mükerrer koruması 2 — ERP'NİN KENDİSİ (2026-07-30, 10.300.4180A olayı):
        # yukarıdaki kontrol yalnızca KENDİ log'umuza bakar. Robot COP'u ERP'ye
        # yazdığı hâlde 'SONUC=OK' dönmezse (ekran akışı beklenmedik bitti,
        # bağlantı koptu) log'a 'hata' düşer; satır COP bekleyenlerde kalır ve
        # tekrar gönderilirse ERP'ye İKİNCİ kez COP girilir. 4180A'da ERP'de COP
        # 18 vardı ama panel hâlâ bekliyor gösteriyordu. ERP'de bugün bu adet
        # zaten varsa gönderme; log'a 'ok' yaz ki panel de düzelsin.
        if not zorla:
            try:
                _mevcut = _as400_cfi_bugun(article, causal='COP')
            except Exception:
                _mevcut = None          # ERP okunamadı → engelleme (güvenli yön)
            if _mevcut and any(abs(q - adet) < 0.001 for q in _mevcut):
                _msg = (f'ERP\'de bugün {article} için {adet:g} adetlik COP hareketi ZATEN var '
                        f'(robot daha önce yazmış ama log\'a "ok" düşmemiş). Mükerrer COP '
                        f'girilmedi; kayıt eşitlendi.')
                conn.execute(
                    "INSERT INTO as400_teyit_log (uretim_tarihi, yil, launch_no, referans, article, "
                    "adet, bayrak, sonuc, mesaj, olusturan) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (u_tarih, 'CO', article, referans, article, adet, 'COP', 'ok', _msg, kullanici))
                conn.commit()
                sonuclar.append({**kayit, 'sonuc': 'atlandi', 'mesaj': _msg})
                continue
        # COP da aynı ekranı (07>01>F1) kullanır — bozuk kod burada da kuyruğu kilitler.
        # zorla ile atlanabilir (bkz. CFI tarafındaki aynı gerekçe).
        gecersiz = None if zorla else _article_gecersiz_mi(article)
        if gecersiz:
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': gecersiz})
            continue
        cikti, robot_hata = _as400_robot_calistir('cfi_gir.js', [article, adet, 'COP'], 120)
        if robot_hata:
            sonuclar.append({**kayit, 'sonuc': 'hata', 'mesaj': robot_hata})
            continue
        robot_ok = 'SONUC=OK' in cikti
        try:
            dogru = any(abs(q - adet) < 0.001 for q in _as400_cfi_bugun(article, causal='COP'))
        except Exception:
            dogru = None
        dogrulandi = robot_ok and (dogru is not False)
        sonuc = 'ok' if dogrulandi else 'hata'
        if dogrulandi:
            mesaj = f'♻ Hurda COP girildi: {article} → {adet} adet'
        elif robot_ok:
            mesaj = f'Robot OK ama bugünkü COP hareketlerinde {adet} bulunamadı — elle kontrol edin'
        else:
            son_satirlar = ' | '.join(l for l in cikti.splitlines() if l.strip())[-300:]
            mesaj = (f'Robot iptal: {son_satirlar}' if son_satirlar
                     else 'Robot iptal — çıktı yok (robot ekrana hiç ulaşamadı)')
        conn.execute(
            "INSERT INTO as400_teyit_log (uretim_tarihi, yil, launch_no, referans, article, adet, bayrak, sonuc, mesaj, olusturan) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (u_tarih, 'CO', article, referans, article, adet, 'COP', sonuc, mesaj, kullanici))
        conn.commit()
        sonuclar.append({**kayit, 'sonuc': sonuc, 'mesaj': mesaj})
    return sonuclar


@app.route('/api/as400/cop_gonder', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_cop_gonder():
    """HURDA (COP) girişleri — EN SON faz. Çekirdek _cop_gonder_calistir'da
    (17:10 oto koşusuyla ORTAK). Body: {satirlar: [{referans, article, adet(hurda),
    uretim_tarihi}]}"""
    data = request.get_json(silent=True) or {}
    satirlar = data.get('satirlar') or []
    if not satirlar:
        return jsonify({'hata': 'satirlar zorunlu'}), 400
    if len(satirlar) > 60:
        return jsonify({'hata': 'Tek seferde en fazla 60 satır'}), 400
    if not _AS400_KILIT.acquire(blocking=False):
        return jsonify({'hata': 'Devam eden bir AS400 gönderimi var — bitmesini bekleyin'}), 409
    try:
        _AS400_DURDUR.clear()
        sonuclar = _cop_gonder_calistir(get_db(), satirlar, g.panel_ku['kullanici_adi'],
                                        zorla=bool(data.get('zorla')))
        ozet = {k: sum(1 for r in sonuclar if r['sonuc'] == k) for k in ('ok', 'hata', 'atlandi')}
        return jsonify({'ozet': ozet, 'sonuclar': sonuclar, 'durduruldu': _AS400_DURDUR.is_set()})
    finally:
        _AS400_DURDUR.clear()
        _AS400_KILIT.release()


@app.route('/api/as400/teyit_gecmisi', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_teyit_gecmisi():
    """Geçmişe dönük teyit kaydı (2026-07-30, kullanıcı isteği: "hangi gün ne
    teyit verildiğini görebileceğimiz bir filtre bölümü olsun").

    Sayfadaki listeler yalnızca SON birkaç günü ve robotun SON 15 gönderimini
    gösteriyordu; "geçen hafta şu koda teyit verdik mi" sorusunun cevabı yoktu.
    Kaynak as400_teyit_log — yani BİZİM gönderdiklerimiz (ERP'nin tamamı değil).

    Parametreler: bas, bit (YYYY-MM-DD) · alan=uretim|gonderim (hangi tarihe göre
    filtrelenecek) · tip=hepsi|launch|cfi|cop · sonuc=hepsi|ok|hata|atlandi ·
    ref (kod parçası, opsiyonel)."""
    bas = (request.args.get('bas') or '').strip()
    bit = (request.args.get('bit') or '').strip()
    if not (bas and bit):
        bit = date.today().isoformat()
        bas = (date.today() - timedelta(days=7)).isoformat()
    alan = 'gonderim' if (request.args.get('alan') == 'gonderim') else 'uretim'
    tip = (request.args.get('tip') or 'hepsi').lower()
    sonuc_f = (request.args.get('sonuc') or 'hepsi').lower()
    ref = (request.args.get('ref') or '').strip()

    # gonderim tarihi created_at'in gun kismi; uretim tarihi zaten YYYY-MM-DD
    tarih_ifade = "DATE(created_at)" if alan == 'gonderim' else "uretim_tarihi"
    kosul = [f"{tarih_ifade} BETWEEN ? AND ?"]
    par = [bas, bit]
    if tip == 'launch':
        kosul.append("yil NOT IN ('CF','CO')")
    elif tip == 'cfi':
        kosul.append("yil='CF'")
    elif tip == 'cop':
        kosul.append("yil='CO'")
    if sonuc_f in ('ok', 'hata', 'atlandi'):
        kosul.append("sonuc=?"); par.append(sonuc_f)
    if ref:
        kosul.append("UPPER(REPLACE(referans,' ','')) LIKE ?")
        par.append('%' + ref.upper().replace(' ', '') + '%')

    conn = get_db()
    satirlar = [dict(r) for r in conn.execute(
        "SELECT id, created_at, uretim_tarihi, yil, launch_no, referans, article, adet, "
        "bayrak, sonuc, mesaj, olusturan FROM as400_teyit_log "
        f"WHERE {' AND '.join(kosul)} ORDER BY id DESC LIMIT 500", par).fetchall()]
    for s in satirlar:
        s['tip'] = 'COP' if s['yil'] == 'CO' else ('CFI' if s['yil'] == 'CF' else 'LAUNCH')

    ozet = {'satir': len(satirlar), 'ok': 0, 'hata': 0, 'atlandi': 0,
            'LAUNCH': 0, 'CFI': 0, 'COP': 0, 'adet_ok': 0.0}
    gunler = {}
    for s in satirlar:
        ozet[s['sonuc']] = ozet.get(s['sonuc'], 0) + 1
        ozet[s['tip']] += 1
        g = (s['created_at'] or '')[:10] if alan == 'gonderim' else s['uretim_tarihi']
        d = gunler.setdefault(g, {'tarih': g, 'ok': 0, 'hata': 0, 'atlandi': 0, 'adet': 0.0})
        d[s['sonuc']] = d.get(s['sonuc'], 0) + 1
        if s['sonuc'] == 'ok':
            ozet['adet_ok'] += float(s['adet'] or 0)
            d['adet'] += float(s['adet'] or 0)
    return jsonify({'satirlar': satirlar, 'ozet': ozet,
                    'gunler': sorted(gunler.values(), key=lambda x: x['tarih'] or '', reverse=True),
                    'filtre': {'bas': bas, 'bit': bit, 'alan': alan, 'tip': tip,
                               'sonuc': sonuc_f, 'ref': ref},
                    'limit_asildi': len(satirlar) >= 500})


# ══════════════════════════════════════════════════════════════════════
# KAYNAK PLANI (2026-07-31) — haftalık plan + ERP'den üretilebilirlik takibi
# Kullanıcı akışı: planlama haftalık .xlsb çıkarır → buraya alınır → ERP'den
# 1. seviye ağaç + 01D/CF2 stoku ile üretilebilir adet hesaplanır → hafta
# boyunca "Yenile" ile tazelenir; ÖNCEKİ ölçüm saklandığı için "stoğa bir şey
# geldi mi" farkı doğrudan görünür. Notlar yenilemede KORUNUR.
# ══════════════════════════════════════════════════════════════════════
def _kp_modul():
    """kaynak_plan_kontrol modülü (plan okuma + ERP sorguları orada)."""
    import sys as _sys
    _kok = os.path.dirname(os.path.abspath(__file__))
    if _kok not in _sys.path:
        _sys.path.insert(0, _kok)
    import kaynak_plan_kontrol as _kp
    return _kp


@app.route('/api/kaynak_plan', methods=['GET'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_liste():
    """Plan satırları + son ölçüm + önceki ölçüme göre değişim."""
    conn = get_db()
    try:
        satirlar = [dict(r) for r in conn.execute(
            "SELECT * FROM kaynak_plan WHERE aktif=1 ORDER BY sira").fetchall()]
    except Exception as e:
        return jsonify({'hata': f'Tablo yok ya da okunamadı: {e}'}), 500
    for s in satirlar:
        onceki, simdi = s.get('onceki_uretilebilir'), s.get('uretilebilir')
        s['degisim'] = (simdi - onceki) if (onceki is not None and simdi is not None) else None
        s['agac_farki'] = [x for x in (s.get('agac_farki') or '').split('|') if x]
        # "Bu ürünü kaynatmam gerekiyor mu?" — panelde ayrı sütun (2026-07-31)
        s['kaynatilmali'] = (s.get('gereken') or 0) > 0
    ozet = {'toplam': len(satirlar)}
    for s in satirlar:
        ozet[s['karar'] or 'YOK'] = ozet.get(s['karar'] or 'YOK', 0) + 1
    ozet['kaynatilmali'] = sum(1 for s in satirlar if (s.get('gereken') or 0) > 0)
    ozet['gecikmis_adet'] = sum(1 for s in satirlar if (s.get('gecikmis') or 0) > 0)
    ozet['artan'] = sum(1 for s in satirlar if (s['degisim'] or 0) > 0)
    ozet['azalan'] = sum(1 for s in satirlar if (s['degisim'] or 0) < 0)
    son = max((s.get('olculdu') or '' for s in satirlar), default='')
    dosya = next((s.get('plan_dosya') for s in satirlar if s.get('plan_dosya')), '')
    return jsonify({'satirlar': satirlar, 'ozet': ozet, 'son_olcum': son,
                    'plan_dosya': dosya,
                    'plan_yuklendi': next((s.get('plan_yuklendi') for s in satirlar
                                           if s.get('plan_yuklendi')), '')})


@app.route('/api/kaynak_plan/parcalar', methods=['GET'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_parcalar():
    """Tek kodun alt parça kırılımı (satır genişletildiğinde)."""
    kod = (request.args.get('kod') or '').strip()
    if not kod:
        return jsonify({'hata': 'kod zorunlu'}), 400
    conn = get_db()
    p = [dict(r) for r in conn.execute(
        "SELECT * FROM kaynak_plan_parca WHERE kaynak_kod=? ORDER BY alt_kod", (kod,)).fetchall()]
    for x in p:
        o = x.get('onceki_stok')
        x['degisim'] = (x['stok_sayilan'] - o) if o is not None else None
    return jsonify({'parcalar': p})


@app.route('/api/kaynak_plan/<int:pid>/not', methods=['PATCH'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_not(pid):
    """Referans notu — hafta boyunca takip için. Yenilemede KORUNUR."""
    data = request.get_json(silent=True) or {}
    metin = (data.get('aciklama') or '').strip()[:500]
    conn = get_db()
    cur = conn.execute(
        "UPDATE kaynak_plan SET aciklama=?, not_guncelleyen=?, "
        "not_guncellendi=datetime('now','localtime') WHERE id=?",
        (metin, g.panel_ku['kullanici_adi'], pid))
    conn.commit()
    if not cur.rowcount:
        return jsonify({'hata': 'Satır bulunamadı'}), 404
    return jsonify({'ok': True, 'aciklama': metin})


def _kaynak_plan_olc(conn, kodlar=None):
    """ERP'den ağaç + stok çekip ölçümü günceller. Notlara DOKUNMAZ.
    Önceki ölçüm onceki_* alanlarına taşınır → değişim görünür."""
    kp = _kp_modul()
    satirlar = [dict(r) for r in conn.execute(
        "SELECT * FROM kaynak_plan WHERE aktif=1" +
        (" AND kaynak_kod IN (%s)" % ','.join('?' * len(kodlar)) if kodlar else ""),
        kodlar or []).fetchall()]
    if not satirlar:
        return {'olculen': 0}
    for s in satirlar:
        s['plan_parcalar'] = []          # ağaç farkı yalnız yüklemede hesaplanır
    cn = kp.erp_baglan()
    try:
        agac = kp.urun_agaci(cn, [s['kaynak_kod'] for s in satirlar])
        alt = sorted({a for lst in agac.values() for a, _, _ in lst})
        stok = kp.stoklar(cn, alt)
        # Referansın KENDİ stoğu da ERP'den (kullanıcı 2026-07-31) — plan
        # dosyasındaki stok sütunu dosyanın çekildiği günün fotoğrafı.
        ref_stok = kp.stoklar(cn, [s['kaynak_kod'] for s in satirlar])
        # İhtiyaç da ERP'den tazelenir (OPR'ler gün içinde değişir)
        opr = kp.opr_ihtiyaclari(cn, [s['kaynak_kod'] for s in satirlar])
    finally:
        try:
            cn.close()
        except Exception:
            pass
    for s in satirlar:
        d = opr.get(s['kaynak_kod'])
        if d is not None:
            s['iht_6h'] = d['ihtiyac']
            s['en_eski_opr'] = d['en_eski'] or ''
            s['opr_sayisi'] = d['opr_sayisi']
            s['gecikmis'] = d['gecikmis']
        else:
            s['iht_6h'] = s.get('iht_6h') or 0
            s['en_eski_opr'] = s.get('en_eski_opr') or ''
            s['opr_sayisi'] = s.get('opr_sayisi') or 0
            s['gecikmis'] = s.get('gecikmis') or 0
    kp.hesapla(satirlar, agac, stok, ref_stok)
    # G GI = 01D + MDT (ERP ekranındaki stok) — GEREKEN bunun üzerinden
    for s in satirlar:
        d = ref_stok.get(s['kaynak_kod'], {})
        s['stok_ggi'] = d.get('01D', 0) + d.get('MDT', 0)
        s['toplam_stok'] = s['stok_ggi']
        s['gereken'] = max(0.0, (s.get('iht_6h') or 0) - s['stok_ggi'])
        s['kaynatilmali'] = s['gereken'] > 0
        u = s.get('uretilebilir')
        if not s['kaynatilmali']:
            s['karar'] = 'GEREK YOK'
        elif u is None:
            s['karar'] = 'ELLE BAK'
        elif u >= s['gereken']:
            s['karar'] = 'TALIMAT VER'
        elif u > 0:
            s['karar'] = 'KISMI'
        else:
            s['karar'] = 'MALZEME YOK'
    simdi = datetime.now().strftime('%Y-%m-%d %H:%M')
    for s in satirlar:
        eski = conn.execute("SELECT uretilebilir, olculdu FROM kaynak_plan WHERE kaynak_kod=?",
                            (s['kaynak_kod'],)).fetchone()
        # İlk ölçümde "önceki" yok; sonrakilerde bir öncekini taşı.
        onc_u = eski['uretilebilir'] if eski else None
        onc_t = eski['olculdu'] if eski else ''
        conn.execute(
            "UPDATE kaynak_plan SET uretilebilir=?, kisitlayan=?, kisit_stok=?, parca_sayisi=?, "
            "durum=?, karar=?, eksi_var=?, olculdu=?, onceki_uretilebilir=?, onceki_olculdu=?, "
            "toplam_stok=?, gereken=?, ref_depolar=?, iht_6h=?, en_eski_opr=?, "
            "opr_sayisi=?, gecikmis=?, stok_ggi=? "
            "WHERE kaynak_kod=?",
            (s.get('uretilebilir'), s.get('kisitlayan', ''),
             next((p['stok_sayilan'] for p in s.get('parcalar', [])
                   if p['kod'] == s.get('kisitlayan')), 0),
             len(s.get('parcalar', [])), s.get('durum', ''), s.get('karar', ''),
             1 if s.get('eksi_var') else 0, simdi, onc_u, onc_t,
             s.get('toplam_stok') or 0, s.get('gereken') or 0,
             ' '.join(f'{d}:{v:g}' for d, v in (s.get('ref_depolar') or {}).items()),
             s.get('iht_6h') or 0, s.get('en_eski_opr') or '', s.get('opr_sayisi') or 0,
             s.get('gecikmis') or 0, s.get('stok_ggi') or 0,
             s['kaynak_kod']))
        # parça kırılımı: önceki stoğu koru, satırları tazele
        onceki_stoklar = {r['alt_kod']: r['stok_sayilan'] for r in conn.execute(
            "SELECT alt_kod, stok_sayilan FROM kaynak_plan_parca WHERE kaynak_kod=?",
            (s['kaynak_kod'],)).fetchall()}
        conn.execute("DELETE FROM kaynak_plan_parca WHERE kaynak_kod=?", (s['kaynak_kod'],))
        for p in s.get('parcalar', []):
            conn.execute(
                "INSERT INTO kaynak_plan_parca (kaynak_kod, alt_kod, birim, um, stok_01d, stok_cf2, "
                "stok_sayilan, kapasite, eksi_bakiye, diger_depolar, onceki_stok, olculdu) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (s['kaynak_kod'], p['kod'], p['birim'], p['um'],
                 p['depolar'].get('01D', 0), p['depolar'].get('CF2', 0), p['stok_sayilan'],
                 p['kapasite'], 1 if p['eksi_bakiye'] else 0,
                 ' '.join(f'{d}:{v:g}' for d, v in p['depolar'].items()
                          if d not in kp.SAYILAN_DEPOLAR),
                 onceki_stoklar.get(p['kod']), simdi))
    conn.commit()
    return {'olculen': len(satirlar), 'olculdu': simdi}


@app.route('/api/kaynak_plan/yukle', methods=['POST'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_yukle():
    """Haftalık plan dosyasını içe alır. Body: {yol} (sunucudan okunur) ya da
    multipart 'dosya'. Notlar KORUNUR — aynı kod tekrar gelirse not silinmez.
    Planda olmayan eski kodlar aktif=0 yapılır (geçmiş notlar kaybolmasın)."""
    kp = _kp_modul()
    gecici = None
    try:
        if 'dosya' in request.files:
            f = request.files['dosya']
            if not (f.filename or '').lower().endswith(('.xlsb', '.xlsx')):
                return jsonify({'hata': 'Yalnız .xlsb / .xlsx dosyası'}), 400
            import tempfile
            # Uzantı KORUNMALI: okuyucu buna bakarak seçiliyor (.xlsb→pyxlsb,
            # .xlsx→openpyxl). Sabit .xlsb yazmak yüklenen .xlsx'i yanlış
            # okuyucuya gönderiyordu (2026-07-31).
            _uz = os.path.splitext(f.filename or '')[1].lower() or '.xlsb'
            gecici = os.path.join(tempfile.gettempdir(), f'kaynak_plan_{os.getpid()}{_uz}')
            f.save(gecici)
            yol, gosterim = gecici, f.filename
        else:
            data = request.get_json(silent=True) or {}
            yol = (data.get('yol') or kp.VARSAYILAN_PLAN).strip()
            gosterim = os.path.basename(yol)
            if not os.path.exists(yol):
                return jsonify({'hata': f'Dosya bulunamadı: {yol} — sunucu bu yolu '
                                        f'göremiyorsa dosyayı yükleyin'}), 404
        try:
            satirlar = kp.plan_oku(yol)
        except Exception as e:
            return jsonify({'hata': f'Plan dosyası okunamadı: {e}'}), 400
        if not satirlar:
            return jsonify({'hata': "Dosyada 'Kaynak kodu' sütunu bulunamadı"}), 400

        conn = get_db()
        simdi = datetime.now().strftime('%Y-%m-%d %H:%M')
        gelen = {s['kaynak_kod'] for s in satirlar}
        conn.execute("UPDATE kaynak_plan SET aktif=0")
        for s in satirlar:
            conn.execute(
                "INSERT INTO kaynak_plan (kaynak_kod, sira, urun, acik_launch, gereken, "
                "kontrol6, gun_stok, bakiye, toplam_stok, plan_stok, iht_2h, iht_6h, "
                "plan_dosya, plan_yuklendi, agac_farki, aktif) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'',1) "
                "ON CONFLICT(kaynak_kod) DO UPDATE SET sira=excluded.sira, urun=excluded.urun, "
                "acik_launch=excluded.acik_launch, "
                "gereken=excluded.gereken, kontrol6=excluded.kontrol6, gun_stok=excluded.gun_stok, "
                "bakiye=excluded.bakiye, toplam_stok=excluded.toplam_stok, plan_stok=excluded.plan_stok, "
                "iht_2h=excluded.iht_2h, iht_6h=excluded.iht_6h, plan_dosya=excluded.plan_dosya, "
                "plan_yuklendi=excluded.plan_yuklendi, aktif=1",
                (s['kaynak_kod'], s['sira'], s['urun'], s['acik_launch'],
                 s['gereken'], (None if s.get('kontrol6') is None else int(s['kontrol6'])),
                 s['gun_stok'], s['bakiye'], s['toplam_stok'], s['toplam_stok'],
                 s['iht_2h'], s['iht_6h'],
                 gosterim, simdi))
            # planlamanın elle yazdığı parçalar — ölçümde ağaç farkı için saklanır
            conn.execute("UPDATE kaynak_plan SET agac_farki=? WHERE kaynak_kod=?",
                         ('|'.join(s.get('plan_parcalar') or []), s['kaynak_kod']))
        conn.commit()
        pasif = conn.execute("SELECT COUNT(*) c FROM kaynak_plan WHERE aktif=0").fetchone()['c']
        return jsonify({'ok': True, 'satir': len(satirlar), 'dosya': gosterim,
                        'pasiflenen': pasif,
                        'mesaj': f'{len(satirlar)} kaynak kodu alındı. '
                                 f'Şimdi "Stokları Yenile" ile ERP hesabını çalıştırın.'})
    finally:
        if gecici and os.path.exists(gecici):
            try:
                os.remove(gecici)
            except Exception:
                pass


@app.route('/api/kaynak_plan/erpden_kur', methods=['POST'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_erpden_kur():
    """Listeyi TAMAMEN ERP'den kurar — plan dosyasına gerek yok (2026-07-31).

    Kullanıcı: "plan için bu exceli planlama biriminden hiç beklemeyelim."
    Kod kümesi : sistemdeki kaynak referansları (referans_listesi bolum='kaynak')
                 + daha önce listeye girmiş kodlar (Excel'den gelmiş olabilir)
    İhtiyaç    : XPRO90 açık OPR'lerin ufuk içindeki KALAN toplamı
    Öncelik    : en eski OPR teslim tarihi (aciliyet sırası)
    Stok       : G GI = 01D + MDT
    Notlar KORUNUR; ihtiyacı kalmayan kodlar aktif=0 olur (not kaybolmasın).
    Body: {ufuk_gun: 42}"""
    data = request.get_json(silent=True) or {}
    try:
        ufuk = max(1, min(365, int(data.get('ufuk_gun') or 42)))
    except (TypeError, ValueError):
        ufuk = 42
    kp = _kp_modul()
    conn = get_db()
    kodlar = {r['referans_kodu'].strip() for r in conn.execute(
        "SELECT referans_kodu FROM referans_listesi "
        "WHERE COALESCE(bolum,'kaynak')='kaynak' AND COALESCE(referans_kodu,'')<>''").fetchall()}
    kodlar |= {r['kaynak_kod'] for r in conn.execute(
        "SELECT kaynak_kod FROM kaynak_plan").fetchall()}
    kodlar = sorted(k for k in kodlar if k)
    if not kodlar:
        return jsonify({'hata': 'Sistemde kaynak referansı yok — önce referans listesini kurun'}), 400
    try:
        cn = kp.erp_baglan()
    except Exception as e:
        return jsonify({'hata': f'AS400 bağlantısı kurulamadı: {e}'}), 502
    try:
        opr = kp.opr_ihtiyaclari(cn, kodlar, ufuk)
    finally:
        try:
            cn.close()
        except Exception:
            pass
    simdi = datetime.now().strftime('%Y-%m-%d %H:%M')
    # Aciliyet sırası: en eski OPR tarihi önce; tarihi olmayan sona
    sirali = sorted(opr.items(), key=lambda kv: (kv[1]['en_eski'] or '9999-12-31', kv[0]))
    conn.execute("UPDATE kaynak_plan SET aktif=0")
    for i, (kod, d) in enumerate(sirali, start=1):
        conn.execute(
            "INSERT INTO kaynak_plan (kaynak_kod, sira, urun, iht_6h, acik_launch, gereken, "
            "en_eski_opr, opr_sayisi, gecikmis, plan_dosya, plan_yuklendi, agac_farki, aktif) "
            "VALUES (?,?,'',?,0,?,?,?,?,'AS400 (OPR)',?,'',1) "
            "ON CONFLICT(kaynak_kod) DO UPDATE SET sira=excluded.sira, "
            "iht_6h=excluded.iht_6h, gereken=excluded.gereken, "
            "en_eski_opr=excluded.en_eski_opr, opr_sayisi=excluded.opr_sayisi, "
            "gecikmis=excluded.gecikmis, plan_dosya=excluded.plan_dosya, "
            "plan_yuklendi=excluded.plan_yuklendi, aktif=1",
            (kod, i, d['ihtiyac'], d['ihtiyac'], d['en_eski'] or '',
             d['opr_sayisi'], d['gecikmis'], simdi))
    conn.commit()
    return jsonify({'ok': True, 'satir': len(sirali), 'ufuk_gun': ufuk,
                    'taranan_kod': len(kodlar),
                    'mesaj': f'{len(sirali)} referansın açık üretim emri var '
                             f'({ufuk} günlük ufuk). Sıralama en eski OPR tarihine göre. '
                             f'Şimdi "Stokları Yenile" ile malzeme kontrolünü çalıştırın.'})


@app.route('/api/kaynak_plan/oprler', methods=['GET'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_oprler():
    """Tek kodun açık üretim emirleri (satır genişletildiğinde gösterilir)."""
    kod = (request.args.get('kod') or '').strip()
    if not kod:
        return jsonify({'hata': 'kod zorunlu'}), 400
    kp = _kp_modul()
    try:
        cn = kp.erp_baglan()
    except Exception as e:
        return jsonify({'hata': f'AS400 bağlantısı kurulamadı: {e}'}), 502
    try:
        d = kp.opr_ihtiyaclari(cn, [kod]).get(kod) or {'satirlar': [], 'ihtiyac': 0}
    finally:
        try:
            cn.close()
        except Exception:
            pass
    return jsonify({'kod': kod, **d})


@app.route('/api/kaynak_plan/yenile', methods=['POST'])
@panel_gerekli(izin='kaynak-plan')
def kaynak_plan_yenile():
    """ERP'den ağaç + stokları tazeler. Notlar ve plan satırları korunur."""
    data = request.get_json(silent=True) or {}
    kodlar = data.get('kodlar') or None
    try:
        sonuc = _kaynak_plan_olc(get_db(), kodlar)
    except Exception as e:
        return jsonify({'hata': f'ERP ölçümü başarısız: {e}'}), 502
    return jsonify({'ok': True, **sonuc})


@app.route('/api/as400/planlama', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_planlama():
    """Tek referans için planlama görünümü: depo stokları + doğrudan açık
    siparişler + üst kodlar (3 seviye implosion; her biri stok+sipariş ile) +
    alt parçalar (1 seviye explosion; 01D/CF2/MK2/MT2 stoklarıyla)."""
    ref = (request.args.get('ref') or '').strip()
    if not ref or len(ref) > 21:
        return jsonify({'hata': 'ref zorunlu (en fazla 21 karakter)'}), 400
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import keyring, pyodbc
    import launch_esle as _le
    import as400_config as _cfg
    ref = _le.bosluk_nokta(ref)
    # ASCII dışı karakter (Türkçe İ/Ü/Ö...) AS400 sorgusunu CWBNL0107 ile patlatır
    import re as _re2
    if not _re2.match(r'^[A-Za-z0-9./\-]{3,21}$', ref):
        return jsonify({'hata': f"Geçersiz referans: '{ref}' — yalnız harf/rakam/nokta/tire (Türkçe karakter olmadan)"}), 400
    try:
        pw = _cfg.sifre_al()
        cn = pyodbc.connect(_cfg.baglanti_dizesi(pw), timeout=30, autocommit=True)
    except Exception as e:
        return jsonify({'hata': f'AS400 bağlantısı kurulamadı: {e}'}), 502
    try:
        # ÜST KODLAR — 3 seviyeye kadar implosion (iteratif; döngü/patlama korumalı)
        ustler = []          # [{kod, seviye, birim}]
        gorulen = {ref}
        seviye_kume = [ref]
        for seviye in (1, 2, 3):
            if not seviye_kume or len(gorulen) > 40:
                break
            yer = ','.join('?' * len(seviye_kume))
            yeni = []
            for r in cn.cursor().execute(
                    f"SELECT DISTINCT TRIM(Y2ARTI), TRIM(Y2RICD), Y2IVQT "
                    f"FROM TKC0301F.BSPEF2 WHERE Y2RICD IN ({yer})", seviye_kume):
                ust = r[0]
                if ust and ust not in gorulen:
                    gorulen.add(ust)
                    ustler.append({'kod': ust, 'seviye': seviye,
                                   'alt': r[1], 'birim': float(r[2] or 0)})
                    yeni.append(ust)
            seviye_kume = yeni
        # ALT PARÇA AĞACI — ÇOK SEVİYELİ explosion (kullanıcı 2026-07-23 kademeli
        # kontrol: bu kademenin stoğu yetmiyorsa bir alt kademenin parçalarıyla
        # üretilebilir mi? — derinlik 4, toplam ≤ 80 düğüm, döngü korumalı).
        agac_kodlari = set()

        def _patlat(kod, derinlik, yol):
            if derinlik <= 0 or len(agac_kodlari) > 80 or kod in yol:
                return []
            cocuklar = []
            for r in cn.cursor().execute(
                    "SELECT TRIM(Y2RICD), Y2IVQT, Y2IVUM FROM TKC0301F.BSPEF2 "
                    "WHERE Y2ARTI = ? ORDER BY Y2RIPR", (kod,)):
                alt_kod = str(r[0] or '').strip()
                if not alt_kod:
                    continue
                agac_kodlari.add(alt_kod)
                cocuklar.append({
                    'kod': alt_kod, 'birim': float(r[1] or 0), 'um': str(r[2] or '').strip(),
                    'cocuklar': _patlat(alt_kod, derinlik - 1, yol | {kod}),
                })
            return cocuklar

        agac = _patlat(ref, 4, set())
        # STOK + SİPARİŞLER (ref + üstler + ağaçtaki tüm kodlar tek seferde)
        tum_kodlar = [ref] + [u['kod'] for u in ustler] + sorted(agac_kodlari)
        stoklar = _plan_stoklar(cn, tum_kodlar)
        siparisler = _plan_siparisler(cn, [ref] + [u['kod'] for u in ustler])
        for u in ustler:
            u['stok'] = stoklar.get(u['kod'], {'depolar': {}, 'kullanilabilir': 0, 'toplam_stok': 0})
            u['siparis'] = siparisler.get(u['kod'], {'acik_toplam': 0, 'satirlar': []})

        def _stok_bagla(dugumler):
            for d in dugumler:
                s = stoklar.get(d['kod'], {'depolar': {}, 'kullanilabilir': 0, 'toplam_stok': 0})
                d['stok'] = s
                # MDT = FASONDA (kullanıcı 2026-07-23): üretilmiş, dış işlemde; dönüşte
                # ÜST koda girer → arz sayılır ama ayrı gösterilir.
                d['fasonda'] = float((s['depolar'].get('MDT') or {}).get('stok', 0) or 0)
                _stok_bagla(d['cocuklar'])
        _stok_bagla(agac)
        kok_stok = stoklar.get(ref, {'depolar': {}, 'kullanilabilir': 0, 'toplam_stok': 0})
        return jsonify({
            'ref': ref,
            'stok': kok_stok,
            'fasonda': float((kok_stok['depolar'].get('MDT') or {}).get('stok', 0) or 0),
            'dogrudan_siparis': siparisler.get(ref, {'acik_toplam': 0, 'satirlar': []}),
            'ustler': ustler,
            'agac': agac,
        })
    except Exception as e:
        return jsonify({'hata': f'AS400 planlama sorgusu başarısız: {e}'}), 502
    finally:
        cn.close()


def _acik_transferler_sorgula(gunler=60):
    """BMMAF0'daki iptal edilmemiş TA 02→01 transferlerini döndürür (liste).
    GET endpoint'i VE 16:45 oto iptal koşusu ortak sorgusu. Hata → raise."""
    bas = date.today() - timedelta(days=gunler)
    esik = bas.year * 10000 + bas.month * 100 + bas.day   # YYYYMMDD karşılaştırma
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import keyring, pyodbc
    import as400_config as _cfg
    pw = _cfg.sifre_al()
    cn = pyodbc.connect(_cfg.baglanti_dizesi(pw), timeout=30, autocommit=True)
    try:
        rows = cn.cursor().execute(
            "SELECT MGDSSO, MGDAAO, MGDMMO, MGDGGO, MGARCD, MGQTA, MGUTVA, MGANRE, MGNURE, MGPRRE "
            "FROM tkc0301F.BMMAF0 "
            "WHERE MGCACD='TA' AND MGMGCD='02' AND MGMCCD='01' "
            "  AND (MGDSSO*1000000 + MGDAAO*10000 + MGDMMO*100 + MGDGGO) >= ? "
            "ORDER BY (MGDSSO*1000000 + MGDAAO*10000 + MGDMMO*100 + MGDGGO) DESC, MGNURE DESC, MGPRRE",
            (esik,)).fetchall()
        return [{
            'tarih': f'{int(r[0])}{int(r[1]):02d}-{int(r[2]):02d}-{int(r[3]):02d}',
            'article': str(r[4]).strip(),
            'adet': float(r[5] or 0),
            'kullanici': str(r[6]).strip(),
            'kayit': f'{int(r[7])}-{int(r[8])}',
            'satir': int(r[9]),
        } for r in rows]
    finally:
        cn.close()


@app.route('/api/as400/acik_transferler', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_acik_transferler():
    """TK1 'hayali stok' transferleri (kullanıcı 2026-07-22 öğretti): alt parça
    yolda iken launch açabilmek için 07>01>F1'den TA causal ile 02→01 stok
    taşınır (02 eksiye düşer); parça gelince Shift+F4 ile İPTAL edilmeli —
    iptal kaydı fiziksel SİLER. Dolayısıyla BMMAF0'da hâlâ duran her TA 02→01
    hareketi = iptal edilmemiş transfer. Tarih bazlı liste döner."""
    try:
        gunler = max(7, min(365, int(request.args.get('gunler') or 60)))
    except (TypeError, ValueError):
        gunler = 60
    try:
        transferler = _acik_transferler_sorgula(gunler)
        return jsonify({'gunler': gunler, 'transferler': transferler})
    except Exception as e:
        return jsonify({'hata': f'AS400 sorgusu başarısız: {e}'}), 502


def _transfer_kayit_duruyor_mu(kayit, satir):
    """İptal DOĞRULAMASI: BMMAF0'da (MGANRE-MGNURE, MGPRRE) hâlâ var mı?
    İptal (Shift+F4) kaydı FİZİKSEL siler → yoksa iptal kesinleşmiştir.
    Döner: True (duruyor) | False (silinmiş) | None (okunamadı)."""
    try:
        yil, no = str(kayit).split('-', 1)
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import keyring, pyodbc
        import as400_config as _cfg
        pw = _cfg.sifre_al()
        cn = pyodbc.connect(_cfg.baglanti_dizesi(pw), timeout=15, autocommit=True)
        try:
            r = cn.cursor().execute(
                "SELECT COUNT(*) FROM tkc0301F.BMMAF0 "
                "WHERE MGCACD='TA' AND MGANRE=? AND MGNURE=? AND MGPRRE=?",
                (int(yil), int(no), int(satir))).fetchone()
            return bool(r and int(r[0]) > 0)
        finally:
            cn.close()
    except Exception:
        return None


def _transfer_iptal_calistir(conn, satirlar, kullanici, dryrun=False):
    """Transfer iptal satırlarını robotla işler (transfer_iptal.js) — manuel
    endpoint VE 16:45 oto koşusu ortak çekirdeği. ÇAĞIRAN _AS400_KILIT'i tutmalı.
    satirlar: [{kayit:'26-245458', satir:200, article, adet, tarih?}]
    Doğrulama: robot SONUC=OK dese bile BMMAF0'da kayıt aranır — silinmişse 'ok'.
    Her sonuç as400_transfer_log'a yazılır. Döner: sonuclar."""
    sonuclar = []
    for idx, s in enumerate(satirlar):
        if _AS400_DURDUR.is_set():
            for kalan in satirlar[idx:]:
                sonuclar.append({'kayit': str(kalan.get('kayit') or ''), 'satir': kalan.get('satir'),
                                 'article': str(kalan.get('article') or ''), 'adet': kalan.get('adet'),
                                 'sonuc': 'atlandi', 'mesaj': 'Kullanıcı durdurdu'})
            break
        kayit = str(s.get('kayit') or '').strip()
        article = str(s.get('article') or '').strip()
        tarih = str(s.get('tarih') or '').strip()
        try:
            satir = int(s.get('satir') or 0)
        except (TypeError, ValueError):
            satir = 0
        try:
            adet = float(s.get('adet') or 0)
        except (TypeError, ValueError):
            adet = 0
        kaydet = {'kayit': kayit, 'satir': satir, 'article': article, 'adet': adet, 'tarih': tarih}
        m = re.match(r'^(\d{2})-(\d{1,7})$', kayit)
        if not (m and satir > 0 and article and adet > 0):
            sonuclar.append({**kaydet, 'sonuc': 'hata', 'mesaj': 'Geçersiz satır parametresi'})
            continue
        # Mükerrer/bayat koruması: kayıt AS400'de hâlâ duruyor mu? (Başka biri
        # elle iptal etmiş olabilir; robot boşuna ekrana girmesin.)
        duruyor = _transfer_kayit_duruyor_mu(kayit, satir)
        if duruyor is False:
            sonuclar.append({**kaydet, 'sonuc': 'atlandi', 'mesaj': 'Kayıt AS400\'de zaten yok (iptal edilmiş)'})
            _transfer_logla(conn, kaydet, 'atlandi', 'Zaten iptal edilmiş', kullanici)
            continue
        adet_arg = ('%g' % adet)
        robot_args = [m.group(1), m.group(2), satir, article, adet_arg] + (['DRYRUN'] if dryrun else [])
        cikti, robot_hata = _as400_robot_calistir('transfer_iptal.js', robot_args, 90)
        if robot_hata:
            sonuclar.append({**kaydet, 'sonuc': 'hata', 'mesaj': robot_hata})
            _transfer_logla(conn, kaydet, 'hata', robot_hata, kullanici)
            continue
        robot_ok = ('SONUC=OK' in cikti) or (dryrun and 'SONUC=DRYRUN-OK' in cikti)
        son_satirlar = ' | '.join(l for l in cikti.splitlines() if l.strip())[-400:]
        if dryrun:
            sonuc = 'ok' if robot_ok else 'hata'
            mesaj = ('DRYRUN doğrulandı (Shift+F4 basılmadı)' if robot_ok
                     else f'DRYRUN iptal: {son_satirlar}')
        else:
            # ASIL doğrulama: kayıt BMMAF0'dan silindi mi?
            silindi = _transfer_kayit_duruyor_mu(kayit, satir)
            if robot_ok and silindi is False:
                sonuc, mesaj = 'ok', f'İptal doğrulandı: {kayit} satır {satir} ({article}, {adet:g}) BMMAF0\'dan silindi'
            elif robot_ok and silindi is None:
                sonuc, mesaj = 'ok', 'Robot OK; BMMAF0 doğrulaması okunamadı — sabah kontrol edin'
            elif robot_ok:
                sonuc, mesaj = 'hata', 'Robot OK dedi ama kayıt BMMAF0\'da HÂLÂ DURUYOR — elle kontrol edin'
            else:
                sonuc, mesaj = 'hata', (f'Robot iptal: {son_satirlar}' if son_satirlar
                                        else 'Robot iptal — çıktı yok (robot ekrana hiç ulaşamadı)')
        _transfer_logla(conn, kaydet, sonuc, mesaj, kullanici)
        sonuclar.append({**kaydet, 'sonuc': sonuc, 'mesaj': mesaj})
    return sonuclar


def _transfer_logla(conn, k, sonuc, mesaj, kullanici):
    try:
        conn.execute(
            "INSERT INTO as400_transfer_log (kayit, satir, article, adet, hareket_tarihi, sonuc, mesaj, olusturan) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (k['kayit'], k['satir'], k['article'], k['adet'], k.get('tarih') or '', sonuc, mesaj, kullanici))
        conn.commit()
    except Exception as e:
        print(f'[transfer_iptal] log yazılamadı: {e}')


@app.route('/api/as400/transfer_iptal', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_transfer_iptal():
    """Seçili açık transferleri ROBOTLA iptal eder (07>01, satır başına 2 + Enter,
    Shift+F4). Body: {satirlar: [{kayit, satir, article, adet, tarih?}], dryrun?}
    dryrun=true → robot ekrana girer, satırı doğrular ama Shift+F4'e BASMAZ."""
    data = request.get_json(silent=True) or {}
    satirlar = data.get('satirlar') or []
    dryrun = bool(data.get('dryrun'))
    if not satirlar:
        return jsonify({'hata': 'satirlar zorunlu'}), 400
    if len(satirlar) > 80:
        return jsonify({'hata': 'Tek seferde en fazla 80 satır'}), 400
    if not _AS400_KILIT.acquire(blocking=False):
        return jsonify({'hata': 'Devam eden bir AS400 gönderimi var — bitmesini bekleyin'}), 409
    try:
        _AS400_DURDUR.clear()
        sonuclar = _transfer_iptal_calistir(get_db(), satirlar, g.panel_ku['kullanici_adi'], dryrun)
        ozet = {k: sum(1 for r in sonuclar if r['sonuc'] == k) for k in ('ok', 'hata', 'atlandi')}
        return jsonify({'ozet': ozet, 'sonuclar': sonuclar, 'durduruldu': _AS400_DURDUR.is_set()})
    finally:
        _AS400_DURDUR.clear()
        _AS400_KILIT.release()


# ─────────────────────────────────────────────────────────────
# OTOMATİK AKŞAM KOŞULARI (kullanıcı 2026-07-27):
#   16:45 — açık TA 02→01 transferlerinin TÜMÜ robotla iptal edilir (07>01,
#           satır başına 2 + Enter, Shift+F4). Kullanıcı kararı: hepsi.
#   17:10 — günün TÜM teyit kuyruğu: launch teyitleri (A / tamamlayana S) +
#           CFI girişleri + hurda COP. Bugün dahil (gün 17:00'da bitmiş sayılır).
# Emin olunamayan satırlar GÖNDERİLMEZ → as400_oto_kosu 'sabah_kontrol'
# listesine yazılır; dashboard teyit sayfasındaki "Sabah Kontrol" paneli gösterir.
# MAKİNE GUARD'I: yalnız teyit-agent'lı makinede koşar (server). Laptop'ta app
# açık kalsa bile çifte koşu OLMAZ (agent 127.0.0.1:5010 yalnız server'da).
# ─────────────────────────────────────────────────────────────
_OTO_CONFIG_YOL = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400', 'oto_config.json')
_OTO_VARSAYILAN = {
    'transfer_iptal': {'etkin': True, 'saat': '16:45', 'gunler': 60, 'dryrun': False},
    'oto_teyit':      {'etkin': True, 'saat': '17:10', 'gunler': 3},
    # ERKEN TEYİT (2026-08-18, kullanıcı isteği): operatör bir referansı
    # "tamamlandı" işaretleyince gecikme_dk sonra o referansın teyidi gün
    # İÇİNDE gönderilir; gün sonunu beklemez.
    # VARSAYILAN KAPALI — bu, ERP'ye gün içinde YAZAN yeni bir otomasyon.
    # Sahada bir kez izlenmeden kendiliğinden başlamamalı (oto_config.json'da
    # etkin:true ile açılır, restart gerekmez — her turda taze okunur).
    'erken_teyit':    {'etkin': False, 'gecikme_dk': 5, 'kontrol_dk': 2},
}


# oto_config.json son okuma hatasi (bos = sorun yok). Sabah Kontrol paneli okur.
_OTO_CONFIG_HATA = {'mesaj': '', 'zaman': ''}


def _oto_config():
    """oto_config.json + varsayılanlar. utf-8-sig: BOM toleransı (mail_config dersi)."""
    cfg = json.loads(json.dumps(_OTO_VARSAYILAN))   # derin kopya
    try:
        with open(_OTO_CONFIG_YOL, encoding='utf-8-sig') as f:
            dosya = json.load(f) or {}
        for k, v in dosya.items():
            if isinstance(v, dict) and k in cfg:
                cfg[k].update(v)
    except FileNotFoundError:
        pass
    except Exception as e:
        # BOZUK CONFIG SESSIZ KALMASIN (2026-08-18): tek bir eksik virgül bile
        # TÜM dosyayı düşürür ve tüm ayarlar varsayılana döner — kullanıcı
        # "erken teyit'i açtım" sanırken hiçbir şey değişmemiş olur (yaşandı).
        # Hata modülde tutulur; Sabah Kontrol paneli bunu ekranda gösterir.
        _OTO_CONFIG_HATA['mesaj'] = str(e)
        _OTO_CONFIG_HATA['zaman'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'[OTO] oto_config.json okunamadı ({e}) — varsayılanlar kullanılıyor')
        return cfg
    _OTO_CONFIG_HATA['mesaj'] = ''      # basarili okuma -> hata temizlenir
    return cfg


def _oto_config_yaz(cfg):
    with open(_OTO_CONFIG_YOL, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _oto_telafi_gerek_mi(kosu_turu):
    """scheduler'a verilecek 'bugün kaçırıldı mı?' yordamını üretir (closure).

    True döner ⇔ as400_oto_kosu'da BUGÜNE ait o türden HİÇ kayıt yok. Yani iş ne
    çalıştı ne de atlandı → uygulama planlı saatte ayakta değildi.
    Atlama da kayıt bıraktığı için (_oto_atlandi_kaydet), "ayar kapalıydı" durumu
    telafi TETİKLEMEZ — kapalı iş zaten tekrar atlanacaktı, boşuna koşmasın."""
    def _kontrol():
        try:
            _c = db_connect()
            try:
                r = _c.execute(
                    "SELECT 1 FROM as400_oto_kosu WHERE tur=? "
                    "AND date(created_at) = date('now','localtime') LIMIT 1",
                    (kosu_turu,)
                ).fetchone()
                return r is None
            finally:
                _c.close()
        except Exception as e:
            # Okuyamadıysak telafi ETME — belirsizken robot koşturmak yanlış yön
            print(f'[SCHED] telafi kontrolü okunamadı ({kosu_turu}): {e}')
            return False
    return _kontrol


def _oto_atlandi_kaydet(tur, ozet_metin, neden):
    """Erken ÇIKIŞ (atlama) durumlarını da as400_oto_kosu'ya yazar — böylece Sabah
    Kontrol paneli 'hiç çalışmadı mı yoksa atladı mı' ayırt edilebilir (2026-07-28
    dersi: agent kapalı/bölüm kapalı olunca job sessizce dönüyordu → panelde iz yok,
    'dün akşam neden verilmedi' cevaplanamıyordu). Kendi bağlantısını açar/kapatır."""
    try:
        _c = db_connect()
        try:
            _oto_kosu_kaydet(_c, tur, [], [{'tip': 'atlandi', 'neden': neden}], ozet_metin)
        finally:
            _c.close()
    except Exception as e:
        print(f'[OTO-{tur.upper()}] atlanma kaydı yazılamadı: {e}')


def _oto_kosu_kaydet(conn, tur, sonuclar, sabah_kontrol, ozet_metin):
    """Koşu raporunu as400_oto_kosu'ya yazar (Sabah Kontrol paneli buradan okur)."""
    try:
        say = {k: sum(1 for r in sonuclar if r.get('sonuc') == k) for k in ('ok', 'hata', 'atlandi')}
        conn.execute(
            "INSERT INTO as400_oto_kosu (tur, ozet, detay, ok, hata, atlandi, kontrol) VALUES (?,?,?,?,?,?,?)",
            (tur, ozet_metin, json.dumps({'sonuclar': sonuclar, 'sabah_kontrol': sabah_kontrol},
                                         ensure_ascii=False, default=str),
             say['ok'], say['hata'], say['atlandi'], len(sabah_kontrol)))
        # Eski koşuları buda (son 120 kalsın)
        conn.execute("DELETE FROM as400_oto_kosu WHERE id NOT IN "
                     "(SELECT id FROM as400_oto_kosu ORDER BY id DESC LIMIT 120)")
        conn.commit()
        print(f'[OTO-{tur.upper()}] {ozet_metin}')
    except Exception as e:
        print(f'[OTO-{tur.upper()}] koşu raporu yazılamadı: {e}')


def _kilit_bekleyerek_al(maks_sn, etiket):
    """_AS400_KILIT'i sabırla alır (16:45 koşusu uzarsa 17:10 bekleyebilsin).
    Döner: True (alındı) | False (süre doldu)."""
    import time as _t
    son = _t.time() + maks_sn
    while _t.time() < son:
        if _AS400_KILIT.acquire(blocking=False):
            return True
        print(f'[{etiket}] AS400 kilidi meşgul — 15 sn sonra tekrar denenecek')
        _t.sleep(15)
    return False


def oto_transfer_iptal_job():
    """16:45 — açık TA 02→01 transferlerinin TÜMÜNÜ robotla iptal eder."""
    cfg = _oto_config()['transfer_iptal']
    if not cfg.get('etkin'):
        print('[OTO-TRANSFER] kapalı (oto_config) — atlandı')
        _oto_atlandi_kaydet('transfer', 'ATLANDI: oto transfer iptali KAPALI (oto_config)',
                            'Ayar kapalı — Sabah Kontrol panelinden aç.')
        return
    if not _teyit_agent_var():
        print('[OTO-TRANSFER] teyit-agent yok — bu makine robot koşusu için yapılandırılmamış, atlandı')
        _oto_atlandi_kaydet('transfer', 'ATLANDI: teyit-agent kapalı (PCOMM/Session B yok)',
                            'teyit-agent 127.0.0.1:5010 yanıt vermedi — RDP oturumunda PCOMM+agent açık olmalı.')
        return
    conn = db_connect()
    try:
        try:
            transferler = _acik_transferler_sorgula(int(cfg.get('gunler') or 60))
        except Exception as e:
            _oto_kosu_kaydet(conn, 'transfer', [], [{'neden': f'AS400 sorgusu başarısız: {e}'}],
                             f'HATA: açık transferler okunamadı ({e})')
            return
        if not transferler:
            _oto_kosu_kaydet(conn, 'transfer', [], [], 'Açık transfer yok — iptal edilecek hareket bulunmadı')
            return
        # Aynı kayıt numarasının satırları ardışık işlensin (robot her çağrıda
        # kayıt no'yu yeniden girer; sıralı gidince ekran akışı kısa kalır).
        satirlar = sorted(transferler, key=lambda t: (t['kayit'], t['satir']))
        if not _kilit_bekleyerek_al(20 * 60, 'OTO-TRANSFER'):
            _oto_kosu_kaydet(conn, 'transfer', [], [{'neden': 'AS400 kilidi 20 dk boşalmadı — koşu yapılamadı'}],
                             'HATA: kilit alınamadı')
            return
        try:
            _AS400_DURDUR.clear()
            dryrun = bool(cfg.get('dryrun'))
            sonuclar = _transfer_iptal_calistir(conn, satirlar, 'oto-16:45', dryrun=dryrun)
        finally:
            _AS400_DURDUR.clear()
            _AS400_KILIT.release()
        sabah = [{'tip': 'transfer', 'neden': s.get('mesaj', ''), **{k: s.get(k) for k in ('kayit', 'satir', 'article', 'adet', 'tarih')}}
                 for s in sonuclar if s.get('sonuc') != 'ok']
        ok_n = sum(1 for s in sonuclar if s.get('sonuc') == 'ok')
        _oto_kosu_kaydet(conn, 'transfer', sonuclar, sabah,
                         f'{len(satirlar)} açık transfer: {ok_n} iptal edildi'
                         + (f', {len(sabah)} sabah kontrole ayrıldı' if sabah else '')
                         + (' [DRYRUN]' if dryrun else ''))
    finally:
        conn.close()


def _oto_kuyruk_olustur(conn, tarihler):
    """17:10 koşusu için kuyrukları kurar — dashboard as4Render bucket mantığının
    Python karşılığı. Döner: (launchQ, cfiQ, copQ, sabah_kontrol).
    KURALLAR (UI ile aynı): zaten_teyitli 'kesin' → dokunma; 'olasi'/zombi/varyant →
    sabah kontrol; işaret 'kontrol' → elle ilgilenildi, dokunma; işaret 'teyit_ver' →
    kod DOĞRU, CFI kuyruğuna zorla (4536B vakası). Bugün DAHİL (17:00'da gün bitti)."""
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import launch_esle as _le
    coklu = _le.esle_coklu(tarihler)
    gun_isaret = {}
    for t in tarihler:
        gun_isaret[t] = {r['referans']: r['durum'] for r in conn.execute(
            "SELECT referans, durum FROM as400_teyit_isaret WHERE kapsam='gun' AND uretim_tarihi=?",
            (t,)).fetchall()}
    cop_verildi = {f"{r['uretim_tarihi']}|{r['launch_no']}" for r in conn.execute(
        "SELECT uretim_tarihi, launch_no FROM as400_teyit_log WHERE yil='CO' AND sonuc='ok'").fetchall()}

    launchQ, cfiQ, copQ, sabah = [], [], [], []

    def isr(t, r):
        return gun_isaret.get(t, {}).get(_le.kanonik(r.get('referans', '')), '')

    def sabaha(tip, t, r, neden, launch=''):
        sabah.append({'tip': tip, 'tarih': t, 'bolum': r.get('bolum', ''),
                      'referans': r.get('referans', ''), 'adet': r.get('adet', 0),
                      'launch': launch, 'neden': neden})

    def cfiye(t, r, article=''):
        # SADECE HURDA satırı (ok_adet=0, nok_adet>0) — 2026-07-31: bu satırlar
        # artık listeden düşmüyor, ama teyit edilecek adetleri YOK. Kuyruğa
        # girerlerse gönderim 'Geçersiz satır parametresi' hatası üretir; hurdaları
        # zaten aşağıdaki COP fazında toplanır. ACIK kolunda aynı kapı zaten var.
        _adet = int(round(float(r.get('adet') or 0)))
        if _adet <= 0:
            return
        cfiQ.append({'referans': r['referans'], 'article': (article or r['referans']),
                     'adet': _adet, 'uretim_tarihi': t,
                     'bolum': r.get('bolum', '')})

    def kapasite_asti(t, r):
        """Kapasite aşımı → oto koşuda HİÇBİR kuyruğa girmez, sabah kontrole düşer.
        UI ile aynı kural (2026-07-30): cycle time tanımlı + vardiya süresinde
        fiziken üretilemeyecek adet. '➕ teyit ver' işareti kullanıcı onayıdır."""
        k = r.get('kapasite_asim')
        if not k or isr(t, r) == 'teyit_ver':
            return False
        sabaha('kapasite', t, r,
               f"ADET ŞÜPHELİ: {k.get('adet')} adet × {k.get('ct')} sn = "
               f"{k.get('gerekli_saat')} saat gerekir, vardiya "
               f"{round((k.get('sure_dk') or 0) / 60.0, 1)} saat — makul üst sınır "
               f"~{k.get('azami')} adet. Operatör kaydını düzeltin ya da ➕ teyit ver ile onaylayın.",
               (( r.get('launchlar') or [{}])[0] or {}).get('launch', ''))
        return True

    for t in tarihler:
        L = coklu.get(t) or {}
        for r in L.get('ACIK', []):
            ls = r.get('launchlar') or []
            if any(l.get('zaten_teyitli') == 'kesin' for l in ls):
                continue
            # SATIR düzeyi teyit: teyit KENDİ koduna CFI ile verilmiş olabilir —
            # launch'ın article'ında görünmez (2026-07-30, 6175B vakası).
            if r.get('zaten_teyitli') == 'kesin':
                continue
            if kapasite_asti(t, r):
                continue
            im = isr(t, r)
            if im == 'kontrol':
                continue
            temiz = [l for l in ls if not l.get('zombi')]
            if not temiz:
                sabaha('zombi', t, r, 'Launch kalanı ≤ 0 (S unutulmuş olabilir) — elle S ile kapatın',
                       (ls[0] or {}).get('launch', '') if ls else '')
                continue
            # 'olasi' şüphesi kullanıcı ➕ teyit ver işaretiyle EZİLEBİLİR (karar insanın)
            # Launch article'ı VEYA satırın kendi kodu 'olasi' ise elle kontrol.
            if ((any(l.get('zaten_teyitli') == 'olasi' for l in ls)
                 or r.get('zaten_teyitli') == 'olasi') and im != 'teyit_ver'):
                sabaha('olasi', t, r, 'Kısmi/uyuşmayan teyit görünüyor — elle kontrol edin',
                       temiz[0].get('launch', ''))
                continue
            l0 = temiz[0]
            adet = int(round(float(r.get('adet') or 0)))
            if adet <= 0:
                continue
            bayrak = 'S' if adet >= float(l0.get('kalan') or 0) - 1 else 'A'
            launchQ.append({'yil': l0['red2'], 'no': l0['renu'], 'article': l0['article'],
                            'referans': r['referans'], 'adet': adet, 'uretim_tarihi': t,
                            'bayrak': bayrak})
        for r in L.get('SUPHELI', []):
            ls = r.get('launchlar') or []
            if any(l.get('zaten_teyitli') == 'kesin' for l in ls):
                continue
            # SATIR düzeyi teyit: varyantta teyit KENDİ koduna CFI ile verilmiş
            # olabilir (2026-07-30, 6175B) — launch article'ında görünmez.
            if r.get('zaten_teyitli') == 'kesin':
                continue
            if kapasite_asti(t, r):
                continue
            im = isr(t, r)
            if im == 'kontrol':
                continue
            if r.get('zaten_teyitli') == 'olasi' and im != 'teyit_ver':
                sabaha('olasi', t, r, 'Kısmi/uyuşmayan teyit görünüyor — elle kontrol edin',
                       (ls[0] or {}).get('launch', '') if ls else '')
                continue
            if im == 'teyit_ver':
                cfiye(t, r)   # kullanıcı 'kod doğru' dedi → kendi koduna CFI
                continue
            sabaha('varyant', t, r, 'Yazım/varyant farkı — panelden 🔁 düzelt ya da ➕ teyit ver',
                   (ls[0] or {}).get('launch', '') if ls else '')
        for kat in ('OPR10', 'YOK'):
            for r in L.get(kat, []):
                if r.get('zaten_teyitli') == 'kesin':
                    continue
                if kapasite_asti(t, r):
                    continue
                im = isr(t, r)
                if im == 'kontrol':
                    continue
                if r.get('zaten_teyitli') == 'olasi' and im != 'teyit_ver':
                    sabaha('olasi', t, r, 'Kısmi/uyuşmayan teyit görünüyor — elle kontrol edin')
                    continue
                art = ''
                if kat == 'OPR10':
                    ls = r.get('launchlar') or []
                    art = (ls[0] or {}).get('article', '') if ls else ''
                cfiye(t, r, art)
        for r in L.get('KAPALI', []):
            im = isr(t, r)
            if r.get('zaten_teyitli') == 'kesin' or im == 'kontrol':
                continue
            if kapasite_asti(t, r):
                continue
            if im == 'teyit_ver':
                cfiye(t, r)
            # değilse: TK1 akışı (durum 45/50) — bilgi, sabah listesini şişirme
        for r in L.get('ARAOP', []):
            if isr(t, r) == 'teyit_ver':
                cfiye(t, r)
        # ── Hurda → COP (HARIC ve ARAOP hariç tüm kovalar) ──
        for kat, rows in L.items():
            if kat in ('HARIC', 'ARAOP'):
                continue
            for r in rows:
                try:
                    h = int(round(float(r.get('hurda') or 0)))
                except (TypeError, ValueError):
                    h = 0
                if h <= 0:
                    continue
                ls = r.get('launchlar') or []
                if kat == 'SUPHELI':
                    sabaha('cop_supheli', t, r, f'Hurda {h} adet: varyant eşleşme — COP kodunu doğrulayıp elle gönderin')
                    continue
                art = ((ls[0] or {}).get('article') or r['referans']) if ls else r['referans']
                if f'{t}|{art}' in cop_verildi:
                    continue
                copQ.append({'referans': r['referans'], 'article': art, 'adet': h,
                             'uretim_tarihi': t, 'bolum': r.get('bolum', '')})
    return launchQ, cfiQ, copQ, sabah


def _erken_teyit_adaylari(conn, tarih, gecikme_dk):
    """ERKEN TEYİT adayı referanslar (2026-08-18).

    Aday olmanın şartları:
      1) O günün üretimi, teyit kapsamındaki bölüm + TK2 (launch_esle ile aynı kapsam)
      2) Referansın O GÜNKÜ TÜM kayıtları 'tamamlandı' işaretli.
         NEDEN "tümü": kuyruk motoru (gun_uretimi) satırları REFERANS BAZINDA
         toplar. Referansın açık (devam eden) bir kaydı varken teyit gönderirsek
         o an ki kısmi adet gider, kalanı gün sonunda FARK olarak takılır.
         Aynı referansa sonradan DÖNÜLÜRSE yeni kayıt açılır ve o kayıt
         tamamlanana kadar aday listesinden düşer — kendiliğinden doğru.
      3) İşaretten bu yana gecikme_dk geçmiş (operatör adedi düzeltebilsin /
         işareti geri alabilsin diye bilerek bekleniyor)
      4) ok_adet > 0 (yalnız hurdalı satır teyit almaz)
      5) O gün + o referans için as400_teyit_log'da BAŞARILI kayıt yok
         (COP hariç — yil='CO' hurda hareketidir, üretim teyidi değil)

    Döner: {kanonik_referans: {'referans','toplam','son_ts'}}
    """
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _le
    except Exception as e:
        print(f'[ERKEN] launch_esle yüklenemedi: {e}')
        return {}
    bolumler = _le.LAUNCH_BOLUMLERI
    try:
        rows = conn.execute(f"""
            SELECT u.referans_kodu ref,
                   SUM(COALESCE(u.ok_adet,0)) toplam,
                   MAX(u.tamamlandi_ts) son_ts,
                   SUM(CASE WHEN COALESCE(u.tamamlandi,0)=1 THEN 0 ELSE 1 END) acik
            FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id
            WHERE v.tarih = ?
              AND COALESCE(v.lokasyon,'TK2') = ?
              AND COALESCE(v.bolum,'kaynak') IN ({','.join('?'*len(bolumler))})
              AND u.referans_kodu IS NOT NULL AND u.referans_kodu != ''
            GROUP BY u.referans_kodu
            HAVING acik = 0 AND toplam > 0 AND son_ts IS NOT NULL
               AND son_ts <= datetime('now','localtime', ?)
        """, (tarih, _le.LOKASYON) + tuple(bolumler) + (f'-{int(gecikme_dk)} minutes',)).fetchall()
    except Exception as e:
        print(f'[ERKEN] aday sorgusu hatası: {e}')
        return {}
    # Zaten teyit edilmişleri düş (bosuna robot koşturma)
    teyitli = set()
    try:
        for r in conn.execute(
            "SELECT referans FROM as400_teyit_log "
            "WHERE uretim_tarihi=? AND sonuc='ok' AND yil <> 'CO'", (tarih,)).fetchall():
            teyitli.add(_le.kanonik(r['referans']))
    except Exception:
        pass
    adaylar = {}
    for r in rows:
        kn = _le.kanonik(r['ref'])
        if kn in teyitli:
            continue
        adaylar[kn] = {'referans': r['ref'], 'toplam': r['toplam'], 'son_ts': r['son_ts']}
    return adaylar


def _mutabakat_farklari(conn, tarih):
    """GÜN SONU MUTABAKATI (2026-08-18): o gün teyit EDİLEN adet ile üretim
    kaydındaki GÜNCEL adet tutuyor mu?

    Gerekçe (kullanıcı): erken teyit gün içinde gidiyor, ama operatör gün sonunda
    adedi değiştirebiliyor. Değişikliği kimse fark etmezse ERP ile üretim kaydı
    kalıcı olarak ayrışır.

    Döner: [{'referans','uretim','teyit','fark'}] — yalnız fark != 0 olanlar.
      fark > 0  → EKSİK teyit (gün sonu adet arttı) → farkı göndermek GEREKİR
      fark < 0  → FAZLA teyit (adet düşürüldü) → ERP'de fazla stok var; geri almak
                  NEGATİF hareket demektir, otomatik YAPILMAZ, yalnız raporlanır.
    COP (yil='CO') hurda hareketidir, üretim teyidi sayılmaz — hesaba katılmaz.
    """
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import launch_esle as _le
    bolumler = _le.LAUNCH_BOLUMLERI
    uretim = {}
    for r in conn.execute(f"""
            SELECT u.referans_kodu ref, SUM(COALESCE(u.ok_adet,0)) toplam
            FROM uretim_kayitlari u JOIN vardiyalar v ON v.id = u.vardiya_id
            WHERE v.tarih = ? AND COALESCE(v.lokasyon,'TK2') = ?
              AND COALESCE(v.bolum,'kaynak') IN ({','.join('?'*len(bolumler))})
              AND u.referans_kodu IS NOT NULL AND u.referans_kodu != ''
            GROUP BY u.referans_kodu""",
            (tarih, _le.LOKASYON) + tuple(bolumler)).fetchall():
        kn = _le.kanonik(r['ref'])
        g = uretim.setdefault(kn, {'referans': r['ref'], 'uretim': 0, 'teyit': 0})
        g['uretim'] += int(r['toplam'] or 0)
    for r in conn.execute(
            "SELECT referans, SUM(CAST(adet AS INTEGER)) tp FROM as400_teyit_log "
            "WHERE uretim_tarihi=? AND sonuc='ok' AND yil <> 'CO' "
            "GROUP BY referans", (tarih,)).fetchall():
        kn = _le.kanonik(r['referans'])
        g = uretim.setdefault(kn, {'referans': r['referans'], 'uretim': 0, 'teyit': 0})
        g['teyit'] += int(r['tp'] or 0)
    out = []
    for kn, g in uretim.items():
        fark = g['uretim'] - g['teyit']
        if fark != 0:
            out.append({**g, 'kanonik': kn, 'fark': fark})
    out.sort(key=lambda x: -abs(x['fark']))
    return out


def _mutabakat_uygula(conn, tarih, kullanici='oto-mutabakat'):
    """Farkları işle: ARTAN farkı gönder, AZALANI yalnız raporla.

    Artan fark için satır, o günün AS400 eşlemesinden kurulur (açık launch varsa
    launch teyidi, yoksa CFI) ama ADET = FARK'tır — tam adet DEĞİL. Tam adedi
    tekrar göndermek mükerrer olurdu; zaten mükerrer frenleri de reddederdi.
    Döner: (gonderilen_sonuclar, rapor_satirlari)
    """
    farklar = _mutabakat_farklari(conn, tarih)
    if not farklar:
        return [], []
    import sys as _sys
    _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
    if _d not in _sys.path:
        _sys.path.insert(0, _d)
    import launch_esle as _le
    try:
        gun = _le.esle_coklu([tarih])[tarih]
    except Exception as e:
        return [], [{**f, 'karar': f'AS400 eşlemesi okunamadı: {e}'} for f in farklar]

    # kanonik referans -> (kategori, satir)
    harita = {}
    for kat, satirlar in gun.items():
        for r in satirlar:
            harita.setdefault(_le.kanonik(r.get('referans') or ''), (kat, r))

    launchQ, cfiQ, rapor = [], [], []
    for f in farklar:
        if f['fark'] < 0:
            rapor.append({**f, 'karar': "FAZLA TEYİT — ERP'de fazla stok var, "
                                        "geri alma NEGATİF hareket gerektirir, elle düzeltin"})
            continue
        kat_satir = harita.get(f['kanonik'])
        if not kat_satir:
            rapor.append({**f, 'karar': 'Eksik teyit ama AS400 eşlemesi bulunamadı — elle bakın'})
            continue
        kat, r = kat_satir
        launchlar = [l for l in (r.get('launchlar') or []) if l.get('durum') == 40]
        if launchlar:
            l = launchlar[0]
            launchQ.append({'uretim_tarihi': tarih, 'yil': l['red2'], 'no': l['renu'],
                            'article': l['article'], 'referans': r['referans'],
                            'adet': f['fark'], 'bayrak': 'A'})
            rapor.append({**f, 'karar': f'EKSİK TEYİT → launch {l["launch"]} ile {f["fark"]} adet gönderildi'})
        else:
            cfiQ.append({'uretim_tarihi': tarih, 'referans': r['referans'],
                         'article': r['referans'], 'adet': f['fark']})
            rapor.append({**f, 'karar': f'EKSİK TEYİT → CFI ile {f["fark"]} adet gönderildi'})
    sonuclar = []
    if launchQ:
        sonuclar += [{**x, 'faz': 'mutabakat-launch'} for x in
                     _teyit_gonder_calistir(conn, launchQ, kullanici)]
    if cfiQ:
        sonuclar += [{**x, 'faz': 'mutabakat-cfi'} for x in
                     _cfi_gonder_calistir(conn, cfiQ, kullanici)]
    return sonuclar, rapor


def erken_teyit_job():
    """Gün İÇİ teyit: 'tamamlandı' işaretli referansları gecikme sonrası gönderir.

    Kuyruk motoru YENİDEN YAZILMAZ — _oto_kuyruk_olustur (17:10 koşusuyla ORTAK)
    bugünün kuyruğunu kurar, biz yalnız ADAY referanslara süzeriz. Üçüncü bir
    kova mantığı kopyası çıkarmak drift riskidir (bkz. _oto_kuyruk_olustur notu).
    """
    cfg = _oto_config().get('erken_teyit') or {}
    if not cfg.get('etkin'):
        return
    if not _teyit_agent_var():
        return          # bu makine robot koşusu için yapılandırılmamış — sessiz geç
    bugun = date.today().isoformat()
    conn = db_connect()
    try:
        adaylar = _erken_teyit_adaylari(conn, bugun, cfg.get('gecikme_dk', 5))
        if not adaylar:
            return
        # Kilit BEKLEMEDEN alınır: meşguls (elle gönderim / 17:10) bir sonraki tura kalsın.
        if not _AS400_KILIT.acquire(blocking=False):
            print(f'[ERKEN] AS400 meşgul — {len(adaylar)} aday sonraki tura ertelendi')
            return
        try:
            _AS400_DURDUR.clear()
            try:
                launchQ, cfiQ, copQ, _sabah = _oto_kuyruk_olustur(conn, [bugun])
            except Exception as e:
                print(f'[ERKEN] kuyruk oluşturulamadı: {e}')
                return
            import launch_esle as _le

            def _suz(kuyruk):
                return [r for r in kuyruk
                        if _le.kanonik(r.get('referans') or '') in adaylar]
            lq, cq = _suz(launchQ), _suz(cfiQ)
            # COP (hurda) erken fazda GÖNDERİLMEZ: hurda gün sonunda tek seferde
            # girilir (17:10 koşusunun COP fazı) — gün içinde ekranlar arası mekik
            # dokumamak için alınmış karar (2026-07-23) burada da geçerli.
            if not lq and not cq:
                return
            print(f'[ERKEN] {len(lq)} launch + {len(cq)} CFI gönderiliyor '
                  f'(adaylar: {", ".join(a["referans"] for a in adaylar.values())})')
            sonuclar = []
            if lq:
                sonuclar += [{**r, 'faz': 'launch'} for r in
                             _teyit_gonder_calistir(conn, lq, 'oto-erken')]
            if cq:
                sonuclar += [{**r, 'faz': 'cfi'} for r in
                             _cfi_gonder_calistir(conn, cq, 'oto-erken')]
            ok = sum(1 for r in sonuclar if r.get('sonuc') == 'ok')
            print(f'[ERKEN] sonuç: {ok} ok / {len(sonuclar)} satır')
            try:
                _oto_kosu_kaydet(conn, 'erken', sonuclar, [],
                                 f'Erken teyit: {ok}/{len(sonuclar)} ok')
            except Exception:
                pass
        finally:
            _AS400_DURDUR.clear()
            _AS400_KILIT.release()
    finally:
        try:
            conn.close()
        except Exception:
            pass


def oto_teyit_job():
    """17:10 — günün tüm teyit kuyruğu: launch (A/S) → CFI → COP."""
    cfg = _oto_config()['oto_teyit']
    if not cfg.get('etkin'):
        print('[OTO-TEYIT] kapalı (oto_config) — atlandı')
        _oto_atlandi_kaydet('teyit', 'ATLANDI: oto teyit KAPALI (oto_config)',
                            'Ayar kapalı — Sabah Kontrol panelinden aç.')
        return
    if not _teyit_agent_var():
        print('[OTO-TEYIT] teyit-agent yok — bu makine robot koşusu için yapılandırılmamış, atlandı')
        _oto_atlandi_kaydet('teyit', 'ATLANDI: teyit-agent kapalı (PCOMM/Session B yok)',
                            'teyit-agent 127.0.0.1:5010 yanıt vermedi — RDP oturumunda PCOMM+agent açık olmalı.')
        return
    gunler = max(1, min(7, int(cfg.get('gunler') or 3)))
    tarihler = [date.today().isoformat()] + [(date.today() - timedelta(days=i)).isoformat()
                                             for i in range(1, gunler + 1)]
    conn = db_connect()
    try:
        if not _kilit_bekleyerek_al(45 * 60, 'OTO-TEYIT'):
            _oto_kosu_kaydet(conn, 'teyit', [], [{'neden': 'AS400 kilidi 45 dk boşalmadı — koşu yapılamadı'}],
                             'HATA: kilit alınamadı')
            return
        try:
            _AS400_DURDUR.clear()
            try:
                launchQ, cfiQ, copQ, sabah = _oto_kuyruk_olustur(conn, tarihler)
            except Exception as e:
                _oto_kosu_kaydet(conn, 'teyit', [], [{'neden': f'Kuyruk oluşturulamadı: {e}'}],
                                 f'HATA: AS400 eşleme başarısız ({e})')
                return
            sonuclar = []
            mutabakat_rapor = []
            for faz, kuyruk in (('launch', launchQ), ('cfi', cfiQ), ('cop', copQ)):
                if _AS400_DURDUR.is_set():
                    break
                if not kuyruk:
                    continue
                if faz == 'launch':
                    s = _teyit_gonder_calistir(conn, kuyruk, 'oto-17:10')
                elif faz == 'cfi':
                    s = _cfi_gonder_calistir(conn, kuyruk, 'oto-17:10')
                else:
                    s = _cop_gonder_calistir(conn, kuyruk, 'oto-17:10')
                sonuclar += [{**r, 'faz': faz} for r in s]
            # GÜN SONU MUTABAKATI (2026-08-18) — normal kuyruktan SONRA.
            # Erken teyit gün içinde kısmi adet göndermiş olabilir; operatör gün
            # sonunda adedi değiştirmiş olabilir. Burada teyit edilen toplam ile
            # üretim kaydı karşılaştırılır: ARTAN fark gönderilir, AZALAN yalnız
            # raporlanır (negatif hareket otomatik yapılmaz).
            if not _AS400_DURDUR.is_set():
                try:
                    _ms, mutabakat_rapor = _mutabakat_uygula(conn, date.today().isoformat())
                    sonuclar += _ms
                    for _r in mutabakat_rapor:
                        print(f"[MUTABAKAT] {_r['referans']}: üretim {_r['uretim']} / "
                              f"teyit {_r['teyit']} / fark {_r['fark']:+d} — {_r['karar']}")
                except Exception as _e:
                    print(f'[MUTABAKAT] hata: {_e}')
        finally:
            _AS400_DURDUR.clear()
            _AS400_KILIT.release()
        # Hatalı gönderimler de sabah kontrolüne düşsün
        for r in sonuclar:
            if r.get('sonuc') == 'hata':
                sabah.append({'tip': 'gonderim_hata', 'tarih': r.get('uretim_tarihi', ''),
                              'bolum': r.get('bolum', ''), 'referans': r.get('referans', ''),
                              'adet': r.get('adet', 0), 'launch': r.get('no', '') or r.get('article', ''),
                              'neden': f"{r.get('faz', '')}: {r.get('mesaj', '')}"})
        ok_n = sum(1 for r in sonuclar if r.get('sonuc') == 'ok')
        atl_n = sum(1 for r in sonuclar if r.get('sonuc') == 'atlandi')
        _oto_kosu_kaydet(conn, 'teyit', sonuclar, sabah,
                         f'{len(launchQ)} launch + {len(cfiQ)} CFI + {len(copQ)} COP kuyruğa alındı: '
                         f'{ok_n} ok, {atl_n} atlandı'
                         + (f', {len(sabah)} satır sabah kontrole ayrıldı' if sabah else ''))
    finally:
        conn.close()


@app.route('/api/as400/oto_kosu', methods=['GET'])
@panel_gerekli(izin='as400-teyit')
def as400_oto_kosu():
    """Sabah Kontrol paneli: son otomatik koşular (16:45 transfer + 17:10 teyit)."""
    try:
        limit = max(1, min(30, int(request.args.get('limit') or 6)))
    except (TypeError, ValueError):
        limit = 6
    conn = get_db()
    _oto_config()          # taze oku — bozuksa _OTO_CONFIG_HATA dolar
    kosular = []
    try:
        for r in conn.execute(
                "SELECT id, tur, ozet, detay, ok, hata, atlandi, kontrol, created_at "
                "FROM as400_oto_kosu ORDER BY id DESC LIMIT ?", (limit,)).fetchall():
            d = dict(r)
            try:
                d['detay'] = json.loads(d.get('detay') or '{}')
            except Exception:
                d['detay'] = {}
            kosular.append(d)
    except Exception as e:
        return jsonify({'hata': str(e)}), 500

    # ── ÇÖZÜLMÜŞ SATIRLARI İŞARETLE (2026-07-30, kullanıcı: "kontrol kısmında
    # düzeltilen şeyleri göstermeyelim") ─────────────────────────────────────
    # Sabah kontrol listesi koşu ANINDA donuyor: o gün emin olunamayan satırlar
    # JSON'a yazılıyor ve bir daha güncellenmiyordu. Sonradan teyit verilse veya
    # kullanıcı "gerek yok"/"kontrol edildi" işaretlese bile satır listede
    # kalıyor, panel çözülmüş işleri bekliyormuş gibi gösteriyordu.
    # Burada her satır CANLI durumla karşılaştırılır; çözülmüşler cozuldu=True
    # alır ve panelde gizlenir (sayısı ayrıca bildirilir — sessiz kesme yok).
    try:
        import sys as _sys
        _d = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'as400')
        if _d not in _sys.path:
            _sys.path.insert(0, _d)
        import launch_esle as _le

        # (üretim günü, kanonik kod) → o güne o koda BAŞARILI gönderim yapılmış
        basarili = set()
        for r in conn.execute(
                "SELECT uretim_tarihi, referans, article FROM as400_teyit_log WHERE sonuc='ok'").fetchall():
            for kod in (r['referans'], r['article']):
                if kod:
                    basarili.add((r['uretim_tarihi'] or '', _le.kanonik(kod)))
        # kullanıcı işaretleri: kalıcı 'gerek_yok' + o güne özel her işaret
        kalici = _le.kalici_haric_set()
        gunluk = set()
        for r in conn.execute(
                "SELECT uretim_tarihi, referans FROM as400_teyit_isaret WHERE kapsam='gun'").fetchall():
            gunluk.add((r['uretim_tarihi'] or '', _le.kanonik(r['referans'] or '')))

        # ERP KANALI (2026-07-30, 94.ltk.704 olayı): robot hareketi ERP'ye YAZDIĞI
        # hâlde ekran doğrulaması tutmazsa log'a 'hata' düşer ve satır sonsuza
        # kadar "kontrol bekliyor" görünür. 704'te ERP'de 07-29 CFI 24 vardı ama
        # panel hata gösteriyordu. Kendi log'umuz tek kanal olamaz — ERP'ye de bak.
        # Sorgu yalnız listede GERÇEKTEN açık kalan kodlar için yapılır (pahalı).
        acik_adaylar = set()
        for k in kosular:
            for s in ((k.get('detay') or {}).get('sabah_kontrol') or []):
                kod = (s.get('referans') or s.get('article') or '').strip()
                kn = _le.kanonik(kod)
                if kod and kn and kn not in kalici and (s.get('tarih') or '', kn) not in basarili \
                        and (s.get('tarih') or '', kn) not in gunluk:
                    # ERP sorgusu TAM eşleşme; operatör kodu küçük harfle yazmış
                    # olabilir (94.ltk.704 ⇄ ERP'de 94.LTK.704). Her iki yazımı da
                    # sor — sonuç zaten kanonik anahtarla toplanıyor.
                    acik_adaylar.add(kod)
                    acik_adaylar.add(kod.upper())
        erp_hrk = {}
        if acik_adaylar:
            try:
                erp_hrk = _le.teyit_hareketleri(sorted(acik_adaylar)) or {}
            except Exception as _e:
                print(f'[oto_kosu] ERP kanali atlandi: {_e}')

        for k in kosular:
            sabah = (k.get('detay') or {}).get('sabah_kontrol') or []
            acik = 0
            for s in sabah:
                kn = _le.kanonik(s.get('referans') or s.get('article') or '')
                t = s.get('tarih') or ''
                if not kn:
                    acik += 1
                    continue
                if kn in kalici:
                    s['cozuldu'], s['cozum'] = True, 'kalıcı olarak "teyit gerekmez" işaretli'
                elif (t, kn) in basarili:
                    s['cozuldu'], s['cozum'] = True, 'bu üretim günü için teyit sonradan verilmiş'
                elif (t, kn) in gunluk:
                    s['cozuldu'], s['cozum'] = True, 'panelden elle işaretlenmiş'
                else:
                    # ERP'de üretim gününde VEYA sonrasında bu adetle hareket var mı?
                    _h = [x for x in (erp_hrk.get(kn) or [])
                          if (not t or x.get('tarih', '') >= t)
                          and abs(float(x.get('adet') or 0) - float(s.get('adet') or 0)) < 0.001]
                    if _h:
                        s['cozuldu'] = True
                        s['cozum'] = (f"ERP'de kayıtlı: {_h[0]['tarih']} · {_h[0].get('tur', '')} "
                                      f"{_h[0]['adet']:g} adet (robot yazdı, ekran doğrulaması tutmamış)")
                    else:
                        acik += 1
            k['acik_kontrol'] = acik
            k['cozulen_kontrol'] = len(sabah) - acik
    except Exception as e:
        # Tespit edilemezse hiçbir satır gizlenmez (güvenli yön: eksik gösterme)
        print(f'[oto_kosu] cozuldu tespiti atlandi: {e}')

    # config_hata: oto_config.json bozuksa panel bunu GOSTERIR. Aksi halde tek bir
    # eksik virgul tum ayarlari varsayilana dusurur ve kullanici farketmez.
    return jsonify({'kosular': kosular, 'config': _oto_config(),
                    'config_hata': dict(_OTO_CONFIG_HATA),
                    'agent': _teyit_agent_var()})


@app.route('/api/as400/oto_config', methods=['POST'])
@panel_gerekli(izin='as400-teyit')
def as400_oto_config_degistir():
    """Oto koşu aç/kapat (+ transfer dryrun). Body: {tur, etkin?, dryrun?}.
    Saat değişikliği DOSYADAN yapılır ve restart ister (thread saati startta okur)."""
    data = request.get_json(silent=True) or {}
    tur = (data.get('tur') or '').strip()
    if tur not in _OTO_VARSAYILAN:
        return jsonify({'hata': f'geçersiz tur: {tur}'}), 400
    cfg = _oto_config()
    if 'etkin' in data:
        cfg[tur]['etkin'] = bool(data.get('etkin'))
    if 'dryrun' in data and tur == 'transfer_iptal':
        cfg[tur]['dryrun'] = bool(data.get('dryrun'))
    try:
        _oto_config_yaz(cfg)
    except Exception as e:
        return jsonify({'hata': f'config yazılamadı: {e}'}), 500
    return jsonify({'basarili': True, 'config': cfg})


# ─────────────────────────────────────────────────────────────

# BAŞLATMA
# ─────────────────────────────────────────────────────────────

# Modul yüklenirken DB migration'larını çalıştır (debug auto-reload sonrası da
# eksik kolonların eklendiğinden emin olmak için).
init_db()


if __name__ == '__main__':
    # Windows konsolu (cp1254) Unicode sembolleri (→ • ✓ emoji) encode EDEMEZ →
    # print() UnicodeEncodeError fırlatır; bu bir yanıt yolunda olursa (örn. mail
    # başarı logu) BAŞARI, HATA gibi görünür. errors='replace' → print ASLA çökmesin.
    try:
        import sys as _sysio
        _sysio.stdout.reconfigure(errors='replace')
        _sysio.stderr.reconfigure(errors='replace')
    except Exception:
        pass

    # Flask debug modundayken (reloader) çift çalışmayı önlemek için kontrol
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        init_db()

    # Scheduler — sadece bir kez başlat (reloader child process'inde başlat,
    # debug modda da main process'te çift olmasın)
    if not os.environ.get('FLASK_DEBUG') or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        try:
            from scheduler import start_scheduler
            # AS400 oto koşuları (2026-07-27): saatler oto_config.json'dan.
            # Jobs kendi içinde etkin/agent kontrolü yapar (kapatmak restart istemez;
            # saat DEĞİŞİKLİĞİ restart ister — thread saati burada okur).
            _ek = []
            _ocfg = _oto_config()
            # kosu_turu: as400_oto_kosu.tur — telafi kontrolü bu kayda bakar
            for _tur, _fn, _et, _vs, _kt in (
                    ('transfer_iptal', oto_transfer_iptal_job, 'AS400 Transfer İptali (02→01)', (16, 45), 'transfer'),
                    ('oto_teyit', oto_teyit_job, 'AS400 Oto Teyit Kuyruğu', (17, 10), 'teyit')):
                try:
                    _h, _m = (int(x) for x in str(_ocfg[_tur].get('saat') or '').split(':'))
                except (TypeError, ValueError):
                    _h, _m = _vs
                # TELAFİ: uygulama planlı saatte ayakta değilse (restart/çökme) o gün
                # sessizce kaybolurdu. Açılışta "bugün hiç koşmadıysa bir kez çalıştır".
                # Güvenli: her iki iş de kendi içinde mükerrer korumalı (as400_teyit_log
                # 'ok' kaydı olan launch tekrar gönderilmez), ayrıca kapalıysa zaten atlar.
                _ek.append((_h, _m, _fn, _et, _oto_telafi_gerek_mi(_kt)))
            # ERKEN TEYİT: günde bir kez değil, gün içinde PERİYODİK.
            # İş kendi içinde etkin/agent/kilit kontrolü yapar; kapalıyken bedeli
            # birkaç saniyelik bir sorgudur. Aralık config'ten (varsayılan 2 dk).
            try:
                _ekd = max(1, int((_ocfg.get('erken_teyit') or {}).get('kontrol_dk') or 2))
            except (TypeError, ValueError):
                _ekd = 2
            start_scheduler(ek_gorevler=_ek,
                            periyodik_gorevler=[(_ekd, erken_teyit_job, 'AS400 Erken Teyit')])
        except Exception as _e:
            print(f'[SCHED] başlatılamadı: {_e}')

    print("\n" + "="*55)
    print("  COFLE MANAGE - URETIM TAKIP SISTEMI CALISIYOR")
    print("="*55)
    print("  Operator Formu : https://coflemanage.online")
    print("  Yonetici Panel : https://coflemanage.online/dashboard")
    print("  Andon Ekrani   : https://coflemanage.online/andon")
    print("="*55)
    print("  Sistem artik bu adresten yayinlanmaktadir.")
    print("  Otomatik arsiv : Her gun 18:00 (data/arsiv/)")
    print("="*55 + "\n")
    # GUVENLIK: internete acik sistemde debug=True = uzaktan kod calistirma (RCE).
    # Varsayilan KAPALI; yerel gelistirmede ortam degiskeni ile ac: COFLE_DEBUG=1
    # threaded=True ŞART: teyit gönderimi bir robot çağrısında 15-150sn bloklar;
    # tek-thread'de bu sürede TÜM app donar (Durdur yoklaması, operatör mobil, andon
    # cevap alamaz — "sistem durdu" belirtisi). Çok-thread'de web responsive kalır,
    # robot yine _AS400_KILIT ile tek tek çalışır. (get_db flask.g request-scoped →
    # her thread kendi bağlantısını açar, SQLite thread-güvenli; yazımlar busy-timeout.)
    app.run(host='0.0.0.0', port=5000, threaded=True, debug=(os.environ.get('COFLE_DEBUG') == '1'))
