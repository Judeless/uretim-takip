import openpyxl
import os

excel_path = r'C:\Users\selcu\OneDrive\Masaüstü\Kaynakhane.xlsx'
if not os.path.exists(excel_path):
    print("Excel not found")
    exit()

wb = openpyxl.load_workbook(excel_path, data_only=True)
search_terms = ['6416', '2062']

for s_name in wb.sheetnames:
    ws = wb[s_name]
    print(f"Checking sheet: {s_name}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] and any(term in str(row[0]) for term in search_terms):
            print(f"Row {i+1}: {row}")
