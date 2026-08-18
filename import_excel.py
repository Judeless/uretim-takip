# -*- coding: utf-8 -*-
"""
data/uretim_verileri.xlsx dosyasından referansları ve operatörleri
uretim.db veritabanına aktarır.

Excel yapısı (bölüm başına referans + operatör sayfası):
  - Kaynak Referans      | Kaynak Operator
  - Montaj Referans      | Montaj Operator
  - Metal Referans       | Metal Operator
  - İşleme Referans      | İşleme Operatör
  - Lazer Referans       | Lazer Operatör
  - Pres Abkant Referans | Pres Abkant Operatör

Her referans sayfası: 1. sütun kod, 2. sütun cycle time (sn).
Her operator sayfası: 2. sütun operatör adı (1. sütun No).
"""
import openpyxl
import sqlite3
import io
import os

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_YOL   = os.path.join(PROJECT_DIR, 'data', 'uretim_verileri.xlsx')
DB_PATH     = os.path.join(PROJECT_DIR, 'uretim.db')

BOLUM_SAYFA = {
    'kaynak': {'ref': 'Kaynak Referans', 'op': 'Kaynak Operator'},
    'montaj': {'ref': 'Montaj Referans', 'op': 'Montaj Operator'},
    'metal':  {'ref': 'Metal Referans',  'op': 'Metal Operator'},
    # Yeni TK2 bölümleri (2026-07): iş takibi/sayaç/andon YOK — sadece vardiya + üretim girişi.
    # Sayfa adları Excel'dekiyle BAYT-BAYT aynı olmalı (yenilerde Türkçe 'Operatör', eskilerde ASCII 'Operator').
    'isleme': {'ref': 'İşleme Referans',      'op': 'İşleme Operatör'},
    'lazer':  {'ref': 'Lazer Referans',       'op': 'Lazer Operatör'},
    'pres':   {'ref': 'Pres Abkant Referans', 'op': 'Pres Abkant Operatör'},
}

# Bölüm bazlı duruş sebepleri sayfaları.
# Yeni bölümlerin sayfası Excel'de henüz yok → durus_sebepleri_yukle sayfa bulunamazsa
# 'Montaj Duruş Listesi'ne düşer (genel sebepler). Kendi sayfası eklenirse otomatik kullanılır.
BOLUM_DURUS_SAYFA = {
    'kaynak': 'Robotik Kaynak Duruş Listesi',
    'montaj': 'Montaj Duruş Listesi',
    'metal':  'Metal Enjeksiyon Duruş Listesi',
    'isleme': 'İşleme Duruş Listesi',
    'lazer':  'Lazer Duruş Listesi',
    'pres':   'Pres Abkant Duruş Listesi',
}

# Ek sayfalar (kaynak bölümüne özel — diğer bölümler için gerekmez)
ROBOT_PROGRAM_SAYFA = 'Robot Program Listesi'
FIKSTUR_RAF_SAYFA   = 'Fikstür Raf Listesi'

# ── TK1 (yan tesis) — ayrı Excel, montaj mantığı, lokasyon='TK1' ──
TK1_EXCEL_YOL = os.path.join(PROJECT_DIR, 'data', 'Tk1 Veriler.xlsx')
TK1_SAYFA = {
    'ref':   'Refernaslar',     # tek kolon 'Ürün Kodu' (sayfa adı Excel'de bu typo ile)
    'op':    'Operatörler',     # BAŞLIK SATIRI YOK — row 0 = ilk operatör
    'durus': 'Duruş Listesi',   # No | Duruş Listesi | Planlı/Plansız
    # ── Plastik enjeksiyon (2026-08-04) ──
    # TK1'de montajdan AYRI bölüm (bolum='plastik'): makineleri 320T / 407T /
    # Yapistirma / Sizdirmazlik Test. Yapıştırma AYRI BÖLÜM DEĞİL, bu bölümün bir
    # makinesi — kodları (sonu 'G' ile biten) aynı referans sayfasında durur.
    # Sayfa adları Excel'dekiyle BİREBİR — ikisinde de yazım hatası var
    # ('Enjeksyion'), düzeltmeyin: Excel'de düzeltilirse burası da güncellenmeli.
    'plastik_ref': 'Plastik Enjeksiyon Referanslar',
    'plastik_op':  'Plastik Enjeksyion Operatörler',
    # ── Tel üretimi (2026-08-04) ──
    # TK1'de ayrı bölüm (bolum='tel'). Referans sayfası ÇOK KOLONLU: 1. kolon ürün
    # kodu, sonraki kolonlar PROSES ADIMLARI (başlık satırından okunur, işaretli
    # olanlar referansın adımları olur). Bkz. TEL_ADIM_BASLIK.
    'tel_ref': 'Tel Referanslar',
    'tel_op':  'Tel Operatörler',
}

# Tel referans sayfasının adım kolonları. Excel başlığı (küçük harfe indirgenmiş)
# → sistemdeki adım adı. Başlık yazımı esnek olsun diye birkaç varyant tanınır;
# tanınmayan kolon sessizce YOK SAYILIR (yeni kolon eklenirse buraya da eklenmeli).
TEL_ADIM_BASLIK = {
    'halat kesme': 'Halat Kesme', 'kesim': 'Halat Kesme', 'halat kesim': 'Halat Kesme',
    'yarı otomatik': 'Yarı Otomatik', 'yari otomatik': 'Yarı Otomatik',
    'yarı otomat': 'Yarı Otomatik', 'yari otomat': 'Yarı Otomatik',
    'tam otomatik': 'Tam Otomatik', 'tam otomat': 'Tam Otomatik',
    'kapama': 'Kapama',
    'son montaj': 'Son Montaj', 'montaj': 'Son Montaj',
}
# Hücre "bu adım var" mı diyor? X / x / 1 / EVET / VAR / ✓ kabul edilir.
_TEL_ISARET = {'x', '1', 'evet', 'var', '✓', '✔', 'e', 'yes', 'true'}


# Üretim sırası — tel_adimlar bu sıraya göre yazılır (app.TEL_ADIMLARI ile aynı).
# Sıralı tutmak "son adım" hesabını kolaylaştırır ve panelde okunaklı gösterir.
TEL_ADIM_SIRASI = ('Halat Kesme', 'Yarı Otomatik', 'Tam Otomatik', 'Kapama', 'Son Montaj')


def _tel_isaretli_mi(hucre):
    if hucre is None:
        return False
    s = str(hucre).strip().lower()
    if not s:
        return False
    if s in _TEL_ISARET:
        return True
    # Sayısal 1 (openpyxl int/float döndürebilir)
    try:
        return float(s) == 1.0
    except ValueError:
        return False

# Başlık satırı tespiti (2026-08-04): TK1 sayfalarının bazısında başlık VAR
# ('Refernaslar' → 'Ürün Kodu'), bazısında YOK ('Operatörler' → row 0 = ilk kişi).
# Yeni plastik sayfalarında hangisinin olacağı garanti değil; sabit "ilk satırı atla"
# kuralı başlıksız sayfada İLK KAYDI YUTAR. Bu yüzden içeriğe bakıyoruz.
_TK1_BASLIKLAR = {
    'ürün kodu', 'urun kodu', 'ürün kod', 'parça kodu', 'parca kodu', 'parça kod',
    'kod', 'kodu', 'referans', 'referans kodu', 'referanslar',
    'ad', 'adi', 'adı', 'isim', 'ad soyad', 'operatör', 'operator',
    'operatörler', 'operatorler', 'no',
}


def _tk1_baslik_mi(deger):
    """Hücre bir başlık mı, veri mi? (TK1 sayfalarında başlık satırı tutarsız)"""
    return str(deger or '').strip().lower() in _TK1_BASLIKLAR


# ── TK1 OPERATÖR DAĞILIMI (kullanıcı 2026-08-04) ────────────────────────────
# Excel'deki tek 'Operatörler' sayfası TK1'in TÜM operatörlerini tutuyor; bölüm
# ayrımı Excel'de yok, kural burada:
#   montaj  → yalnız aşağıdaki iki kişi
#   tel     → kalan HERKES
#   plastik → kendi sayfasından gelir (Mustafa Kaya, Musa Kolip)
# Ad karşılaştırması Türkçe karakter ve boşluk farklarına DAYANIKLI (_ad_normal):
# Excel'de 'BİRCAN KILIÇ' / 'Bircan Kilic' / çift boşluk hepsi eşleşir.
TK1_MONTAJ_OPERATORLERI = ('BİRCAN KILIÇ', 'OSMAN İMAT')


def _ad_normal(ad):
    """Ad eşleştirme anahtarı: Türkçe karakterler sadeleşir, boşluklar tekilleşir."""
    s = str(ad or '').strip().upper()
    for a, b in (('İ', 'I'), ('Ş', 'S'), ('Ğ', 'G'), ('Ü', 'U'), ('Ö', 'O'), ('Ç', 'C')):
        s = s.replace(a, b)
    return ' '.join(s.split())


_TK1_MONTAJ_NORMAL = {_ad_normal(a) for a in TK1_MONTAJ_OPERATORLERI}


def tk1_operator_bolumu(ad):
    """TK1 ana operatör sayfasındaki bir isim hangi bölüme ait? 'montaj' | 'tel'"""
    return 'montaj' if _ad_normal(ad) in _TK1_MONTAJ_NORMAL else 'tel'


def durus_sebepleri_yukle(bolum, lokasyon='TK2'):
    """Bölüme (TK2) veya lokasyona (TK1) ait duruş sebeplerini Excel'den okur.

    Sayfa formatı: No | Duruş Listesi | Planlı/Plansız
    Döner: [{'sebep': str, 'tip': 'planli'|'plansiz'}, ...]

    TK1 → data/Tk1 Veriler.xlsx 'Duruş Listesi' (bolum'dan bağımsız, tek liste).
    TK2 (default) → data/uretim_verileri.xlsx, bolum-spesifik sayfa.
    Excel/sayfa yoksa boş liste.
    """
    if (lokasyon or 'TK2').upper() == 'TK1':
        excel_yol = TK1_EXCEL_YOL
        sayfa_adi = TK1_SAYFA['durus']
    else:
        if bolum not in BOLUM_DURUS_SAYFA:
            return []
        excel_yol = EXCEL_YOL
        sayfa_adi = BOLUM_DURUS_SAYFA[bolum]
    if not os.path.exists(excel_yol):
        return []

    try:
        # DOSYAYI BELLEGE OKU, openpyxl"e BytesIO ver (2026-08-18).
        # NEDEN: bozuk bir xlsx"te openpyxl uye okurken BadZipFile firlatiyor ve
        # arsivi ACIK birakiyor -> her istekte bir tutamak sizip dosya KILITLENIYOR.
        # Sahada yasandi: bozuk dosya ne tasinabildi ne degistirilebildi
        # ("baska bir islem tarafindan kullaniliyor"). BytesIO ile OS tutamagi
        # "with" bitince kapanir; bozuk dosya bile kendini kilitlemez.
        with open(excel_yol, "rb") as _fh:
            _veri = _fh.read()
        wb = openpyxl.load_workbook(io.BytesIO(_veri), data_only=True)
        if sayfa_adi not in wb.sheetnames:
            # Bölümün kendi duruş sayfası yoksa genel listeye düş (yeni bölümler:
            # işleme/lazer/pres — Excel'e kendi sayfaları eklenince otomatik geçilir).
            sayfa_adi = BOLUM_DURUS_SAYFA.get('montaj', '')
            if sayfa_adi not in wb.sheetnames:
                return []
        ws = wb[sayfa_adi]
        sonuc = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # Başlık
            if not row or row[1] is None:
                continue
            sebep = str(row[1]).strip()
            if not sebep:
                continue
            # Türkçe ı/i karakter farkı: 'Plansız' lower → 'plansız' (dotless ı),
            # ama 'siz' substring'i (regular i) bulunmaz. Bu yüzden açık string match.
            tip_raw = str(row[2] or '').strip().lower()
            if 'plansız' in tip_raw or 'plansiz' in tip_raw:
                tip = 'plansiz'
            elif 'planlı' in tip_raw or 'planli' in tip_raw:
                tip = 'planli'
            else:
                tip = 'plansiz'  # boş/bilinmeyen → güvenli yan: plansız
            sonuc.append({'sebep': sebep, 'tip': tip})
        return sonuc
    except Exception as e:
        print(f"[durus_sebepleri] Hata: {e}")
        return []


def _bolum_import(conn, wb, bolum):
    """Tek bölümün referans + operatör verisini import eder."""
    sayfalar = BOLUM_SAYFA[bolum]
    c = conn.cursor()

    # ── Referans sayfası ──
    if sayfalar['ref'] not in wb.sheetnames:
        return {
            'referanslar_eklenen': 0,
            'referanslar_guncellenen': 0,
            'referanslar_silinen': 0,
            'operatorler_eklenen': 0,
            'hata': f"'{sayfalar['ref']}' sayfası bulunamadı"
        }
    ref_sayfa = wb[sayfalar['ref']]
    print(f"  [{bolum.upper()}] Referans sayfası: '{ref_sayfa.title}'")

    ref_sayisi = 0
    ref_guncellenen = 0
    tel_kodu_atlanan = 0        # TK2'ye yazılmayan 93.* (TK1 tel) kodları
    excel_kodlari_norm = set()

    # Kaynak bölümü için Excel artık 4 kolon: Kod | Kaynak Süresi | Söktak Süresi | Toplam Cycle
    # Pres Abkant 3 kolon: Kod | Açıklama | Süre (operatör mobilde açıklamayı görür)
    # Montaj/Metal/diğerleri eski 2 kolon: Kod | Cycle Time (tek değer)
    kaynak_modu = (bolum == 'kaynak')
    pres_modu = (bolum == 'pres')

    for i, row in enumerate(ref_sayfa.iter_rows(values_only=True)):
        if i == 0:
            continue  # Başlık
        if not row or row[0] is None:
            continue
        kod = str(row[0]).strip()
        if not kod or len(kod) < 2:
            continue

        # Süreleri parse et
        kaynak_sn = 0.0
        soktak_sn = 0.0
        cycle = 0.0
        aciklama = ''
        bukum_op = 1   # yalnız pres modunda D kolonundan okunur; diğerlerinde 1 = bölme yok
        if kaynak_modu:
            # B = Kaynak Süresi, C = Söktak Süresi, D = Toplam (formül)
            try:
                kaynak_sn = float(row[1]) if (len(row) > 1 and row[1] is not None) else 0.0
            except (ValueError, TypeError):
                kaynak_sn = 0.0
            try:
                soktak_sn = float(row[2]) if (len(row) > 2 and row[2] is not None) else 0.0
            except (ValueError, TypeError):
                soktak_sn = 0.0
            cycle = round(kaynak_sn + soktak_sn, 2)
        elif pres_modu:
            # B = Açıklama (metin), C = Süre, D = Büküm Op. (2026-07-29)
            aciklama = str(row[1]).strip() if (len(row) > 1 and row[1] is not None) else ''
            try:
                cycle = float(row[2]) if (len(row) > 2 and row[2] is not None) else 0.0
            except (ValueError, TypeError):
                cycle = 0.0
            # Büküm operasyon sayısı: bir parçaya sırayla uygulanan büküm adedi.
            # Sayaç bölücüsü olarak kullanılır (3 op → 3. sinyal = 1 parça).
            # Boş/bozuk = 1 (bölme yok). D kolonu eski dosyalarda hiç yok → len kontrolü şart.
            try:
                bukum_op = int(row[3]) if (len(row) > 3 and row[3] is not None) else 1
            except (ValueError, TypeError):
                bukum_op = 1
            bukum_op = max(1, min(99, bukum_op))
        else:
            # Montaj/Metal/diğerleri — tek değer
            try:
                cycle = float(row[1]) if (len(row) > 1 and row[1] is not None) else 0.0
            except (ValueError, TypeError):
                cycle = 0.0

        # ── TK1 TEL KODU TK2'YE YAZILMAZ (kullanıcı 2026-08-06) ──────────────
        # Kural: "93. ile başlayan bütün referanslar tel referansıdır" → tel
        # üretimi TK1'de yapılır, bu kodların TK2 listesinde işi yok.
        # OLAY: TK2 Excel'inin 'Montaj Referans' sayfasında 1460 adet 93.* kod
        # duruyordu; hepsi süresiz olduğu için TK2 montajın "Süre Tanımı Bekleyen
        # Referanslar" panelini şişiriyor (2384 satırın 1459'u bunlardı) ve gerçek
        # eksikler arasında kayboluyordu.
        # Koda ATLAMA yeterli: bu kodlar excel_kodlari_norm'a girmediği için
        # aşağıdaki mirror-sync onları DB'den de temizler — Excel'i elle
        # düzeltmeye gerek kalmaz, sonraki içe aktarımlarda geri gelmezler.
        # Ölçüm (2026-08-06): TK2'de 93.* kodların HİÇBİRİNİN süresi tanımlı
        # değildi, yani silinen bir bilgi yok.
        if kod.strip().startswith('93.'):
            tel_kodu_atlanan += 1
            continue

        excel_kodlari_norm.add(kod.upper().replace(' ', ''))

        # LOKASYON GÜVENLİĞİ: _bolum_import YALNIZCA TK2 Excel'ini (EXCEL_YOL) işler.
        # Tüm referans_listesi okuma/yazma işlemleri lokasyon='TK2' ile kapatılır ki
        # TK2 import'u TK1 (yan tesis) kayıtlarını eşleştirip ezmesin/silmesin.
        # BÖLÜM KAPSAMASI: UNIQUE(referans_kodu, bolum, lokasyon) sonrası aynı kod birden
        # fazla bölümde farklı süreyle yaşayabilir (16 kod metal+işleme'de ortak) —
        # eşleştirme ve cycle geri-yayılımı bu bölümün satırı/vardiyalarıyla sınırlı,
        # yoksa bölümler birbirinin kaydını çalar/cycle'ını ezer.
        # TAM-EŞ TERCİHİ: DB'de aynı koda normalize olan birden fazla yazım varyantı
        # olabilir (eski auto-create kalıntısı: '94.LTK.10' + '94.ltk.10'). Excel'deki
        # yazımla birebir eşleşen satır varsa ONU güncelle — yoksa varyantı Excel
        # yazımına çevirirken tam-eş satırla UNIQUE çakışması oluşur.
        mevcut = c.execute(
            "SELECT id, referans_kodu FROM referans_listesi "
            "WHERE UPPER(REPLACE(referans_kodu, ' ', '')) = UPPER(REPLACE(?, ' ', '')) "
            "AND COALESCE(bolum, 'kaynak') = ? AND COALESCE(lokasyon, 'TK2') = 'TK2' "
            "ORDER BY (referans_kodu = ?) DESC, id LIMIT 1",
            (kod, bolum, kod)
        ).fetchone()

        if mevcut:
            if pres_modu:
                # Pres'te açıklama VE büküm operasyon sayısı Excel'den yönetilir —
                # import her seferinde günceller (operatörün mobilden girdiği değer de
                # zaten Excel'e yazılıyor, iki yön aynı hücreye bakar).
                c.execute(
                    'UPDATE referans_listesi SET hedef_cycle_time_sn = ?, kaynak_suresi_sn = ?, soktak_suresi_sn = ?, referans_kodu = ?, aciklama = ?, bukum_operasyon = ? WHERE id = ?',
                    (cycle, kaynak_sn, soktak_sn, kod, aciklama, bukum_op, mevcut[0])
                )
            else:
                # Diğer bölümlerde aciklama'ya DOKUNMA (dashboard'dan girilmiş olabilir)
                c.execute(
                    'UPDATE referans_listesi SET hedef_cycle_time_sn = ?, kaynak_suresi_sn = ?, soktak_suresi_sn = ?, referans_kodu = ? WHERE id = ?',
                    (cycle, kaynak_sn, soktak_sn, kod, mevcut[0])
                )
            if cycle > 0:
                c.execute(
                    "UPDATE uretim_kayitlari SET cycle_time_sn = ? "
                    "WHERE UPPER(REPLACE(referans_kodu, ' ', '')) = UPPER(REPLACE(?, ' ', '')) "
                    "AND vardiya_id IN (SELECT id FROM vardiyalar WHERE COALESCE(lokasyon, 'TK2') = 'TK2' AND COALESCE(bolum, 'kaynak') = ?)",
                    (cycle, kod, bolum)
                )
            ref_guncellenen += 1
        else:
            c.execute(
                "INSERT INTO referans_listesi (referans_kodu, hedef_cycle_time_sn, kaynak_suresi_sn, soktak_suresi_sn, aciklama, bolum, lokasyon, bukum_operasyon) VALUES (?, ?, ?, ?, ?, ?, 'TK2', ?)",
                (kod, cycle, kaynak_sn, soktak_sn, aciklama, bolum, bukum_op)
            )
            if cycle > 0:
                c.execute(
                    "UPDATE uretim_kayitlari SET cycle_time_sn = ? "
                    "WHERE UPPER(REPLACE(referans_kodu, ' ', '')) = UPPER(REPLACE(?, ' ', '')) "
                    "AND vardiya_id IN (SELECT id FROM vardiyalar WHERE COALESCE(lokasyon, 'TK2') = 'TK2' AND COALESCE(bolum, 'kaynak') = ?)",
                    (cycle, kod, bolum)
                )
            ref_sayisi += 1

    print(f"  Referanslar: {ref_sayisi} eklendi, {ref_guncellenen} güncellendi"
          + (f", {tel_kodu_atlanan} adet '93.*' TK1 tel kodu atlandı" if tel_kodu_atlanan else ""))

    # ── MIRROR SYNC: Excel'de olmayan referansları bu bölümden temizle ──
    ref_silinen = 0
    if excel_kodlari_norm:
        bolum_refs = c.execute(
            "SELECT id, referans_kodu FROM referans_listesi "
            "WHERE COALESCE(bolum, 'kaynak') = ? AND COALESCE(lokasyon, 'TK2') = 'TK2'",
            (bolum,)
        ).fetchall()
        for ref_row in bolum_refs:
            ref_norm = str(ref_row[1] or '').upper().replace(' ', '')
            if ref_norm and ref_norm not in excel_kodlari_norm:
                c.execute('DELETE FROM referans_listesi WHERE id = ?', (ref_row[0],))
                ref_silinen += 1
        if ref_silinen:
            print(f"  Referanslar: {ref_silinen} adet (Excel'de olmayan) silindi")
    else:
        print("  UYARI: Excel'den hiçbir referans okunamadı, silme atlandı")

    # ── Operatör sayfası ──
    op_sayisi = 0
    if sayfalar['op'] in wb.sheetnames:
        op_sayfa = wb[sayfalar['op']]
        print(f"  [{bolum.upper()}] Operatör sayfası: '{op_sayfa.title}'")
        for i, row in enumerate(op_sayfa.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row or len(row) < 2 or row[1] is None:
                continue
            ad = str(row[1]).strip()
            if not ad:
                continue
            try:
                # LOKASYON: TK2 import'u yalnız TK2 operatörlerine bakar/yazar —
                # TK1'de aynı ad varsa TK2 operatörü sessizce atlanmasın (ayrı fabrika).
                # ÇOKLU BÖLÜM: aynı kişi birden fazla bölümde çalışabilir → bölüm başına
                # satır (UNIQUE(ad, bolum, lokasyon)). Kişi başka bölümde zaten varsa
                # yeni bölüm satırı ONUN PIN'iyle açılır (bir kişi = tek PIN).
                mevcut_op = c.execute(
                    "SELECT id FROM operatorler WHERE UPPER(ad) = UPPER(?) AND bolum = ? AND COALESCE(lokasyon,'TK2')='TK2'",
                    (ad, bolum)
                ).fetchone()
                if not mevcut_op:
                    ayni_isim = c.execute(
                        "SELECT id, pin FROM operatorler WHERE UPPER(ad) = UPPER(?) AND COALESCE(lokasyon,'TK2')='TK2'",
                        (ad,)
                    ).fetchone()
                    # conn row_factory'siz (tuple) — pin = index 1
                    pin = (ayni_isim[1] or '0000') if ayni_isim else '0000'
                    c.execute("INSERT INTO operatorler (ad, bolum, pin, lokasyon) VALUES (?, ?, ?, 'TK2')", (ad, bolum, pin))
                    op_sayisi += 1
            except Exception as e:
                print(f"  Operatör eklenemedi ({ad}): {e}")
        print(f"  Operatörler: {op_sayisi} eklendi")

    return {
        'referanslar_eklenen': ref_sayisi,
        'referanslar_guncellenen': ref_guncellenen,
        'referanslar_silinen': ref_silinen,
        'operatorler_eklenen': op_sayisi
    }


def import_tk1(conn=None):
    """TK1 (yan tesis) referans + operatör verisini import eder — lokasyon='TK1', bolum='montaj'.

    _bolum_import'tan AYRI (mirror-sync YOK → TK2 verisini silmez). referans_kodu global
    UNIQUE olduğu için INSERT OR IGNORE (TK2 ile çakışan ~47 kod atlanır; kod yine elle
    girilebilir). TK1 'Operatörler' sayfasında BAŞLIK SATIRI YOK → row 0 dahil. PIN default
    '0000' (panelden değiştirilir). Idempotent — tekrar çalıştırılabilir.
    """
    kapat = False
    if conn is None:
        conn = sqlite3.connect(DB_PATH, timeout=20.0)
        kapat = True
    c = conn.cursor()
    if not os.path.exists(TK1_EXCEL_YOL):
        return {'basarili': False, 'hata': f"'{TK1_EXCEL_YOL}' bulunamadı"}
    wb = openpyxl.load_workbook(TK1_EXCEL_YOL, data_only=True)

    def _sayfa_aktar(sayfa_adi, bolum, tur):
        """Tek kolonlu TK1 sayfasını aktarır. tur: 'ref' | 'op'. Döner: eklenen satır.

        Başlık satırı SABİT İNDEKSLE DEĞİL İÇERİKLE atlanır (bkz. _tk1_baslik_mi):
        montaj referans sayfasında başlık var, operatör sayfasında yok; plastik
        sayfalarında hangisinin olacağı belli değil ve sabit kural ya ilk kaydı
        yutar ya da başlığı referans/operatör olarak veritabanına yazar."""
        if sayfa_adi not in wb.sheetnames:
            return 0
        eklenen = 0
        ilk_veri = True
        for row in wb[sayfa_adi].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            deger = str(row[0]).strip()
            if not deger:
                continue
            # Başlık YALNIZ ilk dolu satırda aranır. Kelime listesini her satıra
            # uygulamak MEŞRU kodları elerdi: TK1 listesinde 'TEST', 'ÖZEL ÜRÜN',
            # 'HALAT TAHRIBAT' gibi rakamsız ama GERÇEK referanslar var (aynı
            # sebeple "kod rakam içermeli" kuralı da kullanılamaz).
            if ilk_veri:
                ilk_veri = False
                if _tk1_baslik_mi(deger):
                    continue
            if tur == 'ref':
                if len(deger) < 2:
                    continue
                # 93.* KODLARI DOĞRUDAN TELE YAZILIR (kullanıcı 2026-08-04).
                # Ana sayfa montaj + tel kodlarını birlikte tutuyor. Önce montaja
                # yazıp sonra UPDATE ile taşımak İKİ KAYIT üretiyordu (tel'e taşınan
                # eski satır + ana sayfadan tekrar eklenen montaj satırı) ve ikinci
                # içe aktarımda UNIQUE(kod,bolum,lokasyon) hatası veriyordu.
                _ref_bolum = ('tel' if (bolum == 'montaj' and deger.startswith('93.'))
                              else bolum)
                cur = c.execute(
                    "INSERT OR IGNORE INTO referans_listesi "
                    "(referans_kodu, hedef_cycle_time_sn, bolum, lokasyon) VALUES (?, 0, ?, 'TK1')",
                    (deger, _ref_bolum))
            else:
                # Ana operatör sayfasında bölüm İSME göre belirlenir (montaj = 2 kişi,
                # kalan herkes tel). Plastik/tel kendi sayfalarından gelirse bolum sabit.
                _op_bolum = tk1_operator_bolumu(deger) if bolum == 'montaj' else bolum
                cur = c.execute(
                    "INSERT OR IGNORE INTO operatorler (ad, pin, bolum, lokasyon) "
                    "VALUES (?, '0000', ?, 'TK1')",
                    (deger, _op_bolum))
            eklenen += cur.rowcount
        return eklenen

    # ── ÖNCE MEVCUT VERİYİ DÜZELT, SONRA EKLE ────────────────────────────────
    # Sıra kritik: yeni kayıtlar isme/koda göre doğru bölümle ekleniyor. Eğer önce
    # eklersek, aynı kişi/kod hem eski 'montaj' satırı hem yeni 'tel' satırı olarak
    # bulunur ve taşıma UNIQUE(ad,bolum,lokasyon) kısıtına takılır.
    #
    # 1) Kod kuralı — 93.* → tel. İDEMPOTENT ve GERİYE DÖNÜK. Tersi YAPILMAZ
    #    (tel sayfasından gelen 93.* olmayan kod tel kalır — orada bilinçli tanım var).
    # Satır satır: aynı kod tel'de zaten varsa (eski veriden kalma çift kayıt)
    # toplu UPDATE tüm işlemi patlatırdı — o satır silinir, tel kaydı korunur.
    tel_tasinan = 0
    for _rr in c.execute(
            "SELECT id, referans_kodu FROM referans_listesi "
            "WHERE COALESCE(lokasyon,'TK2')='TK1' AND COALESCE(bolum,'')='montaj' "
            "AND TRIM(referans_kodu) LIKE '93.%'").fetchall():
        try:
            c.execute("UPDATE referans_listesi SET bolum='tel' WHERE id=?", (_rr[0],))
        except sqlite3.IntegrityError:
            c.execute("DELETE FROM referans_listesi WHERE id=?", (_rr[0],))
        tel_tasinan += 1

    # 2) Operatör dağılımı — montaj listesinde OLMAYAN herkes tele taşınır.
    #    'Admin' (her bölümde görünen özel kullanıcı) dışarıda bırakılır.
    #    Çakışma olursa (kişi zaten tel'de kayıtlı) eski montaj satırı silinir ama
    #    PIN'i korunur — operatör kendi PIN'iyle girmeye devam etsin.
    op_tel_tasinan = 0
    for _r in c.execute(
            "SELECT id, ad, COALESCE(pin,'0000') FROM operatorler "
            "WHERE COALESCE(lokasyon,'TK2')='TK1' AND COALESCE(bolum,'')='montaj' "
            "AND ad != 'Admin'").fetchall():
        if tk1_operator_bolumu(_r[1]) != 'tel':
            continue
        try:
            c.execute("UPDATE operatorler SET bolum='tel' WHERE id=?", (_r[0],))
        except sqlite3.IntegrityError:
            # Aynı kişi tel'de zaten var → PIN'i oraya taşı, montaj satırını sil
            c.execute("UPDATE operatorler SET pin=? WHERE ad=? AND bolum='tel' "
                      "AND COALESCE(lokasyon,'TK2')='TK1' AND COALESCE(pin,'0000')='0000'",
                      (_r[2], _r[1]))
            c.execute("DELETE FROM operatorler WHERE id=?", (_r[0],))
        op_tel_tasinan += 1

    # ── Montaj (mevcut akış) ──
    # TEL KURALI (kullanıcı 2026-08-04): TK1'de "93." ile başlayan HER referans
    # tel referansıdır. Ana referans sayfası montaj + tel kodlarını birlikte
    # tutuyor; bölüm koda bakılarak ayrılır (ayrı sayfa tutmaya gerek yok).
    ref_eklenen = _sayfa_aktar(TK1_SAYFA['ref'], 'montaj', 'ref')
    op_eklenen  = _sayfa_aktar(TK1_SAYFA['op'],  'montaj', 'op')

    # ── Plastik enjeksiyon (2026-08-04) — TK1'de AYRI bölüm ──
    # Operatör ve referansları montajdan bağımsız: plastik operatörü montaj
    # referanslarını görmemeli, mobilde bölüm seçilince kendi listesi gelmeli.
    p_ref_eklenen = _sayfa_aktar(TK1_SAYFA['plastik_ref'], 'plastik', 'ref')
    p_op_eklenen  = _sayfa_aktar(TK1_SAYFA['plastik_op'],  'plastik', 'op')

    # ── Tel üretimi (2026-08-04) — OPSİYONEL sayfa ──
    # Tel referansları ARTIK BU SAYFADAN GELMİYOR: TK1'de '93.' ile başlayan her
    # kod tel referansıdır (yukarıdaki kural) ve operatörler ana sayfadan isimle
    # ayrışır. Bu blok yalnız GERİYE UYUM için duruyor — sayfa varsa okunur,
    # yoksa hiçbir şey olmaz (normal durum budur).
    # tel_adimlar (proses tanımı) artık KULLANILMIYOR: kapaması/son montajı
    # dışarıda yapılacak ürünler sabit olmadığı için referans bazlı akış tanımı
    # terk edildi; her adım kendi referans ekiyle kaydediliyor (bkz. tel_proses.py).
    t_ref_eklenen = t_adim_guncel = 0
    if TK1_SAYFA['tel_ref'] in wb.sheetnames:
        ws = wb[TK1_SAYFA['tel_ref']]
        sutun_adim = {}          # kolon indeksi → adım adı
        baslik_okundu = False
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None or not str(row[0]).strip():
                continue
            ilk = str(row[0]).strip()
            if not baslik_okundu:
                baslik_okundu = True
                if _tk1_baslik_mi(ilk):
                    # Başlık satırı: adım kolonlarını buradan öğren
                    for i, h in enumerate(row[1:], start=1):
                        ad = TEL_ADIM_BASLIK.get(str(h or '').strip().lower())
                        if ad:
                            sutun_adim[i] = ad
                    continue      # başlık satırı veri değil
            kod = ilk
            if len(kod) < 2:
                continue
            # İşaretli adımları ÜRETİM SIRASINDA topla (kolon sırası değil —
            # Excel'de kolonlar karışık dizilse bile sıra doğru kalsın).
            secili = [sutun_adim[i] for i in sorted(sutun_adim)
                      if i < len(row) and _tel_isaretli_mi(row[i])]
            sirali = [a for a in TEL_ADIM_SIRASI if a in secili]
            adim_metni = ','.join(sirali)
            cur = c.execute(
                "INSERT OR IGNORE INTO referans_listesi "
                "(referans_kodu, hedef_cycle_time_sn, bolum, lokasyon, tel_adimlar) "
                "VALUES (?, 0, 'tel', 'TK1', ?)", (kod, adim_metni))
            t_ref_eklenen += cur.rowcount
            if adim_metni:
                cur2 = c.execute(
                    "UPDATE referans_listesi SET tel_adimlar=? "
                    "WHERE referans_kodu=? AND COALESCE(bolum,'')='tel' "
                    "AND COALESCE(lokasyon,'TK2')='TK1' AND COALESCE(tel_adimlar,'')!=?",
                    (adim_metni, kod, adim_metni))
                t_adim_guncel += cur2.rowcount
    t_op_eklenen = _sayfa_aktar(TK1_SAYFA['tel_op'], 'tel', 'op')

    conn.commit()
    if kapat:
        conn.close()
    print(f"[TK1] montaj: {ref_eklenen} referans + {op_eklenen} operatör | "
          f"plastik: {p_ref_eklenen} referans + {p_op_eklenen} operatör | "
          f"tel: {t_ref_eklenen} referans + {t_op_eklenen} operatör "
          f"({t_adim_guncel} proses tanımı güncellendi, {tel_tasinan} kod '93.*' kuralıyla, "
          f"{op_tel_tasinan} operatör tele taşındı) (lokasyon=TK1)")
    ref_eklenen += p_ref_eklenen + t_ref_eklenen
    op_eklenen  += p_op_eklenen + t_op_eklenen
    # NOT: referans_kodu GLOBAL UNIQUE → TK2 ile çakışan TK1 kodları INSERT OR IGNORE ile
    # atlanır (atlanan = bu kodlar TK2'de var). Tam ayrışma için UNIQUE(referans_kodu,lokasyon)
    # migration gerekir (bkz. lokasyon denetimi). 'referanslar_guncellenen' UI mesajı için 0.
    return {'basarili': True, 'referanslar_eklenen': ref_eklenen,
            'referanslar_guncellenen': 0, 'referanslar_silinen': 0,
            'operatorler_eklenen': op_eklenen}


def import_data(bolum=None):
    """Excel'den verileri import eder.

    bolum=None  → tüm bölümler (BOLUM_SAYFA anahtarları)
    bolum='kaynak' / 'montaj' / 'metal' / 'isleme' / 'lazer' / 'pres' → sadece o bölüm
    """
    if bolum and bolum not in BOLUM_SAYFA:
        return {'basarili': False, 'hata': f"Geçersiz bölüm: {bolum}"}

    if not os.path.exists(EXCEL_YOL):
        return {'basarili': False, 'hata': f'Excel dosyası bulunamadı: {EXCEL_YOL}'}

    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS operatorler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ad TEXT UNIQUE NOT NULL,
            bolum TEXT DEFAULT 'kaynak'
        )
    ''')
    try:
        c.execute("ALTER TABLE operatorler ADD COLUMN bolum TEXT DEFAULT 'kaynak'")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE referans_listesi ADD COLUMN bolum TEXT DEFAULT 'kaynak'")
    except Exception:
        pass

    wb = openpyxl.load_workbook(EXCEL_YOL, data_only=True)

    sonuclar = {}
    toplam = {'referanslar_eklenen': 0, 'referanslar_guncellenen': 0,
              'referanslar_silinen': 0, 'operatorler_eklenen': 0}

    islenen = [bolum] if bolum else list(BOLUM_SAYFA.keys())

    for b in islenen:
        print(f"\n{'='*50}")
        print(f"  {b.upper()} import başlıyor...")
        print(f"{'='*50}")
        sonuc = _bolum_import(conn, wb, b)
        sonuclar[b] = sonuc
        for k in toplam:
            toplam[k] += sonuc.get(k, 0)

    conn.commit()
    conn.close()

    return {
        'basarili': True,
        **toplam,
        'detay': sonuclar
    }


def _program_listesi_import(conn, wb):
    """Robot Program Listesi sayfası → robot_programlari tablosu.
    Sayfa formatı (matrix):
       Satır 0: ROBOT | RAFNO | ABB-1 | ABB-1 | ABB-2 | ABB-2 | ... (her robot 2 kez)
       Satır 1: İSTASYON | <boş> | İST-1 | İST-2 | İST-1 | İST-2 | ...
       Satır 2+: <referans_kodu> | <raf_no> | √ | <boş> | √ | √ | ...
    Bu matrix'i düzleştirip her √ işareti için bir satır INSERT eder.
    """
    if wb is None or ROBOT_PROGRAM_SAYFA not in wb.sheetnames:
        return {'eklenen': 0, 'silinen': 0, 'hata': 'sayfa yok'}

    ws = wb[ROBOT_PROGRAM_SAYFA]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {'eklenen': 0, 'silinen': 0, 'hata': 'yetersiz satır'}

    # Satır 0: robot adları (MERGED — birden çok kolonu kapsar)
    # Satır 1: istasyon. Kolon 0=referans, 1=raf
    robot_satiri = rows[0]
    istasyon_satiri = rows[1]
    # Merged cell mantığı: None olan kolonda son görülen robot devam eder
    robot_son = ''
    kolon_eslesme = []  # [(col_idx, robot_no, istasyon), ...]
    for j in range(2, max(len(robot_satiri), len(istasyon_satiri))):
        # Robot adı — yeni değer varsa al, yoksa son görüleni kullan (merged)
        r_raw = robot_satiri[j] if j < len(robot_satiri) and robot_satiri[j] is not None else None
        if r_raw is not None:
            robot_son = str(r_raw).strip()
        if not robot_son:
            continue
        # İstasyon
        i_raw = istasyon_satiri[j] if j < len(istasyon_satiri) and istasyon_satiri[j] is not None else None
        if i_raw is None:
            continue
        i_str = str(i_raw).strip()
        robot_no = robot_son.replace('-', '').replace(' ', '')  # ABB-1 → ABB1
        # İstasyon: "İST-1" → 1
        ist = 0
        if '1' in i_str: ist = 1
        elif '2' in i_str: ist = 2
        elif '3' in i_str: ist = 3
        if ist > 0:
            kolon_eslesme.append((j, robot_no, ist))

    c = conn.cursor()
    # Mevcut programları temizle (Excel master)
    c.execute('DELETE FROM robot_programlari')

    eklenen = 0
    for r in rows[2:]:
        if not r or r[0] is None: continue
        ref = str(r[0]).strip()
        if not ref or len(ref) < 2: continue
        for col_idx, robot_no, ist in kolon_eslesme:
            if col_idx >= len(r): continue
            val = str(r[col_idx] or '').strip()
            if val and val != '':  # √ veya başka bir işaret varsa
                c.execute(
                    'INSERT INTO robot_programlari (robot_no, istasyon, referans_kodu, guncelleyen) VALUES (?, ?, ?, ?)',
                    (robot_no, ist, ref, 'Excel İçe Aktar')
                )
                eklenen += 1

    print(f"  Robot Program: {eklenen} satır eklendi")
    return {'eklenen': eklenen}


def _fikstur_raf_import(conn, wb):
    """Fikstür Raf Listesi sayfası → fikstur_raf tablosu.
    Sayfa formatı: 3 raf yan yana (A, B, C). Her raf 2 kolon: kod | raf_no.
    Aralarda boş kolon olabilir.
    """
    if wb is None or FIKSTUR_RAF_SAYFA not in wb.sheetnames:
        return {'eklenen': 0, 'hata': 'sayfa yok'}

    ws = wb[FIKSTUR_RAF_SAYFA]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {'eklenen': 0, 'hata': 'yetersiz satır'}

    c = conn.cursor()
    c.execute('DELETE FROM fikstur_raf')

    eklenen = 0
    # Her satırdaki tüm (kod, raf_no) çiftlerini topla
    for r in rows[1:]:  # Başlık satırını atla
        if not r: continue
        # Her 3 kolonda bir grup: (kod_col, raf_col, boş)
        i = 0
        while i < len(r):
            kod = str(r[i] or '').strip() if i < len(r) else ''
            raf = str(r[i+1] or '').strip() if i+1 < len(r) else ''
            if kod and raf:
                c.execute(
                    'INSERT INTO fikstur_raf (referans_kodu, raf_no) VALUES (?, ?)',
                    (kod, raf)
                )
                eklenen += 1
            i += 3  # Sonraki grup

    print(f"  Fikstür Raf: {eklenen} satır eklendi")
    return {'eklenen': eklenen}


def kaynak_ek_import():
    """Robot Program Listesi + Fikstür Raf sayfalarını içe alır — kaynak alanının ek
    sayfaları. Eski 'Toplu Veri Yönetimi' panelinden taşındı: dashboard'da kaynak bölümü
    için 'Excel'den Aktar' artık bunları da kapsar (tek buton, tek akış)."""
    if not os.path.exists(EXCEL_YOL):
        return {'program_eklenen': 0, 'fikstur_eklenen': 0}
    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    try:
        wb = openpyxl.load_workbook(EXCEL_YOL, data_only=True)
        sonuc = {}
        try:
            p = _program_listesi_import(conn, wb)
            sonuc['program_eklenen'] = p.get('eklenen', 0)
        except Exception as e:
            print(f"  Robot Program HATA: {e}")
            sonuc['program_eklenen'] = 0
        try:
            f = _fikstur_raf_import(conn, wb)
            sonuc['fikstur_eklenen'] = f.get('eklenen', 0)
        except Exception as e:
            print(f"  Fikstür HATA: {e}")
            sonuc['fikstur_eklenen'] = 0
        conn.commit()
        return sonuc
    finally:
        conn.close()


def import_tum(yedek_al=False):
    """Excel'deki TÜM sayfaları okuyup DB'yi günceller:
       - Tüm bölüm referansları (cycle time)
       - Tüm bölüm operatörleri
       - Robot programları (kaynak için matrix)
       - Fikstür raf listesi
       - Duruş sebepleri (validasyon — read on-demand)
    """
    if not os.path.exists(EXCEL_YOL):
        return {'basarili': False, 'hata': f'Excel bulunamadı: {EXCEL_YOL}'}

    # Önce normal referans+operator (mevcut)
    sonuc = import_data()

    # Sonra ek sayfalar
    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    try:
        wb = openpyxl.load_workbook(EXCEL_YOL, data_only=True)

        # Robot Program
        try:
            prog_sonuc = _program_listesi_import(conn, wb)
            sonuc['program_eklenen'] = prog_sonuc.get('eklenen', 0)
        except Exception as e:
            print(f"  Robot Program HATA: {e}")
            sonuc['program_eklenen'] = 0

        # Fikstür Raf
        try:
            fik_sonuc = _fikstur_raf_import(conn, wb)
            sonuc['fikstur_eklenen'] = fik_sonuc.get('eklenen', 0)
        except Exception as e:
            print(f"  Fikstür HATA: {e}")
            sonuc['fikstur_eklenen'] = 0

        conn.commit()
    finally:
        conn.close()

    # Duruş sebepleri validasyonu — her bölüm için sayıyı raporla
    # (gerçek DB import yok, read-on-demand; ama kullanıcı sayıyı görsün)
    durus_ozet = {}
    for b in BOLUM_DURUS_SAYFA.keys():
        durus_ozet[b] = len(durus_sebepleri_yukle(b))
    sonuc['durus_sebepleri'] = durus_ozet

    return sonuc


def export_referans_cycle_times(bolum=None, lokasyon='TK2'):
    """DB'deki cycle_time'ları Excel'in <Bolum> Referans sayfa(lar)ına yazar (SADECE TK2).
    Diğer veriler (operatör, duruş, program, fikstür) korunur.

    lokasyon='TK1' ise ATLA: TK1 referansları cycle time kullanmaz ve ayrı dosyadadır
    (data/Tk1 Veriler.xlsx, tek kolon) — TK1 verisi TK2 Excel'ine sızmamalı.
    """
    if (lokasyon or 'TK2').upper() == 'TK1':
        return {'basarili': True, 'atlandi': 'TK1 (cycle time export yok)', 'yazilan': 0}
    if not os.path.exists(EXCEL_YOL):
        return {'basarili': False, 'hata': f'Excel bulunamadı: {EXCEL_YOL}'}

    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    conn.row_factory = sqlite3.Row
    bolum_listesi = [bolum] if bolum else list(BOLUM_SAYFA.keys())
    toplam_yazilan = 0
    wb = openpyxl.load_workbook(EXCEL_YOL)

    for b in bolum_listesi:
        sayfa_adi = BOLUM_SAYFA[b]['ref']
        if sayfa_adi not in wb.sheetnames:
            continue
        ws = wb[sayfa_adi]
        kaynak_modu = (b == 'kaynak')
        # Mevcut Excel satırlarını oku, kod → (satır, sayfadaki ham yazım) map'i
        kod_satir = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1: continue  # Başlık
            if not row or row[0] is None: continue
            kod = str(row[0]).strip()
            kod_satir[kod.upper().replace(' ', '')] = (i, kod)

        # DB'deki bu bölüme ait referansları çek — SADECE TK2 (TK1 ayrı dosyada/cycle'sız).
        # SÜRESİZLER (ct=0/NULL) HARİÇ: Excel resmî/tanımlı referans listesidir; operatör
        # üretim girişinden auto-create ile doğan "Süre Tanımı Bekleyen" kodlar Excel'e
        # YAZILMAZ (yeni satır olarak eklenmez, Excel'de zaten varsa değeri de ezilmez).
        # Süre tanımlanınca (ct>0) bir sonraki sync'te Excel'e girer.
        db_rows = conn.execute(
            "SELECT referans_kodu, hedef_cycle_time_sn, kaynak_suresi_sn, soktak_suresi_sn, "
            "COALESCE(sure_teyit,0) as sure_teyit, COALESCE(aciklama,'') as aciklama, "
            "COALESCE(bukum_operasyon,1) as bukum_operasyon FROM referans_listesi "
            "WHERE COALESCE(bolum,'kaynak')=? AND COALESCE(lokasyon,'TK2')='TK2' "
            "AND COALESCE(hedef_cycle_time_sn,0) > 0 ORDER BY referans_kodu",
            (b,)
        ).fetchall()

        # Kaynak sayfası 4+1 kolonlu: Kod | Kaynak(B) | Söktak(C) | Toplam(D formül) | Süre Teyit(E)
        # (ESKİ BUG: her bölümde B'ye toplam cycle yazılıyordu — kaynakta B=Kaynak Süresi
        #  kolonunu eziyordu. Montaj/metal 2 kolonlu: Kod | Cycle(B) — o davranış doğru.)
        if kaynak_modu and (ws.cell(row=1, column=5).value or '') == '':
            ws.cell(row=1, column=5, value='Süre Teyit')
        # Pres sayfası 3 kolonluydu — D başlığı yoksa aç (2026-07-29 büküm operasyonu)
        if b == 'pres' and (ws.cell(row=1, column=4).value or '') == '':
            ws.cell(row=1, column=4, value='Büküm Op.')

        yazilan = 0
        # Aynı koda normalize olan birden fazla DB satırı (eski yazım varyantları:
        # '94.LTK.10' + '94.ltk.10') aynı Excel satırına düşer — sayfadaki yazımla
        # birebir eşleşen satır ÖNCELİKLİDİR; bayat varyantın taze değeri ezmesine izin verme.
        yazilan_norm = {}
        for r in db_rows:
            kod = (r['referans_kodu'] or '').strip()
            norm = kod.upper().replace(' ', '')
            if norm in kod_satir:
                ri, sayfa_kod = kod_satir[norm]
                tam_es = (kod == sayfa_kod)
            else:
                ri = ws.max_row + 1
                ws.cell(row=ri, column=1, value=kod)
                kod_satir[norm] = (ri, kod)
                tam_es = True
            if yazilan_norm.get(norm) and not tam_es:
                continue  # bu koda tam-eş (veya ilk) yazım zaten yazıldı — varyantla ezme
            if kaynak_modu:
                ws.cell(row=ri, column=2, value=r['kaynak_suresi_sn'] or 0)
                ws.cell(row=ri, column=3, value=r['soktak_suresi_sn'] or 0)
                ws.cell(row=ri, column=4, value=f'=B{ri}+C{ri}')
                ws.cell(row=ri, column=5, value='EVET' if r['sure_teyit'] else '')
            elif b == 'pres':
                # Pres 4 kolonlu: B=Açıklama, C=Süre, D=Büküm Op.
                # (süreyi B'ye yazmak açıklamayı ezerdi)
                ws.cell(row=ri, column=2, value=r['aciklama'] or '')
                ws.cell(row=ri, column=3, value=r['hedef_cycle_time_sn'] or 0)
                ws.cell(row=ri, column=4, value=int(r['bukum_operasyon'] or 1))
            else:
                ws.cell(row=ri, column=2, value=r['hedef_cycle_time_sn'] or 0)
            yazilan_norm[norm] = True
            yazilan += 1
        toplam_yazilan += yazilan
        print(f"  {b}: {yazilan} satır Excel'e yazıldı")

    conn.close()
    try:
        wb.save(EXCEL_YOL)
    except PermissionError:
        # Dosya Excel'de/OneDrive'da AÇIK — kullanıcı sessiz kayıp yaşamasın, net mesaj dön
        return {'basarili': False,
                'hata': 'Excel dosyası şu an açık (uretim_verileri.xlsx) — kapatıp tekrar deneyin.'}
    return {'basarili': True, 'yazilan': toplam_yazilan, 'dosya': EXCEL_YOL}


if __name__ == '__main__':
    res = import_tum()
    print("\n✅ Import tamamlandı:", res)
