"""
export_excel.py
Tüm üretim verilerini Excel dosyasına aktarır.
Dosya yolu: C:\\Users\\selcu\\OneDrive\\Masaüstü\\UretimTakipArsiv.xlsx
"""
import os
from datetime import datetime
from database import get_db
from oee import hesapla_oee

MASAUSTU = r'C:\Users\selcu\OneDrive\Masaüstü'
DOSYA_ADI = 'UretimTakipArsiv.xlsx'
DOSYA_YOLU = os.path.join(MASAUSTU, DOSYA_ADI)


def _stil(ws, renk_hex, kalin=True, hizalama='left', sarilabilir=False):
    """Hücre stilini yardımcı olarak döndürür (import edilen stil nesneleri)."""
    from openpyxl.styles import PatternFill, Font, Alignment
    return {
        'fill': PatternFill(fill_type='solid', fgColor=renk_hex),
        'font': Font(bold=kalin, color='FFFFFF' if kalin else '1e293b'),
        'alignment': Alignment(horizontal=hizalama, wrap_text=sarilabilir),
    }


def _baslik_satiri(ws, sutunlar, renk='1E3A5F'):
    """İlk satırı başlık olarak formatla."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    for col_idx, baslik in enumerate(sutunlar, start=1):
        hucre = ws.cell(row=1, column=col_idx, value=baslik)
        hucre.fill = PatternFill(fill_type='solid', fgColor=renk)
        hucre.font = Font(bold=True, color='FFFFFF', size=10)
        hucre.alignment = Alignment(horizontal='center', vertical='center')
        ince = Side(style='thin', color='FFFFFF')
        hucre.border = Border(left=ince, right=ince, bottom=ince)


def _zebra(ws, row_num, sutun_sayisi):
    """Zebra rengi uygula."""
    from openpyxl.styles import PatternFill
    if row_num % 2 == 0:
        for col in range(1, sutun_sayisi + 1):
            ws.cell(row=row_num, column=col).fill = PatternFill(fill_type='solid', fgColor='F8FAFC')


def _kolonlari_genislet(ws):
    """İçeriğe göre kolon genişliği ayarla."""
    for col in ws.columns:
        max_w = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_w = max(max_w, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_w + 4, 10), 40)


def export_arsiv(tarih_bas=None, tarih_bit=None):
    """
    Verileri UretimTakipArsiv.xlsx dosyasına yaz.
    tarih_bas / tarih_bit: 'YYYY-MM-DD' formatında opsiyonel filtre.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return {'basarili': False, 'hata': 'openpyxl kurulu değil. pip install openpyxl'}

    conn = get_db()
    c = conn.cursor()

    # Tarih filtresi
    tarih_sart = ''
    params_list = []
    if tarih_bas and tarih_bit:
        tarih_sart = 'WHERE tarih BETWEEN ? AND ?'
        params_list = [tarih_bas, tarih_bit]
    elif tarih_bas:
        tarih_sart = 'WHERE tarih >= ?'
        params_list = [tarih_bas]

    wb = Workbook()

    # ─── SAYFA 1: VARDİYALAR ───────────────────────────────────
    ws_v = wb.active
    ws_v.title = 'Vardiyalar'
    ws_v.row_dimensions[1].height = 22

    basliklar_v = [
        'ID', 'Tarih', 'Lokasyon', 'Vardiya Türü', 'Robot No', 'Operatör',
        'Başlangıç', 'Bitiş', 'Toplam Süre (dk)',
        'Planlı Duruş (dk)', 'Plansız Duruş (dk)',
        'Net Plan (dk)', 'Çalışma (dk)',
        'OEE %', 'Availability %', 'Performance %', 'Quality %',
        'OK Adet', 'NOK Adet', 'Durum'
    ]
    _baslik_satiri(ws_v, basliklar_v)

    vardiyalar = c.execute(
        f'SELECT * FROM vardiyalar {tarih_sart} ORDER BY tarih DESC, id DESC',
        params_list
    ).fetchall()

    for row_num, v in enumerate(vardiyalar, start=2):
        oee_data = hesapla_oee(v['id']) or {}
        satir = [
            v['id'], v['tarih'], dict(v).get('lokasyon') or 'TK2',
            v['vardiya_turu'], v['robot_no'], v['operator_adi'],
            v['baslangic_saati'], v['bitis_saati'], v['toplam_sure_dk'],
            oee_data.get('planli_durus_dk', ''),
            oee_data.get('plansiz_durus_dk', ''),
            oee_data.get('net_plan_dk', ''),
            oee_data.get('calisma_suresi_dk', ''),
            oee_data.get('oee', ''),
            oee_data.get('availability', ''),
            oee_data.get('performance', ''),
            oee_data.get('quality', ''),
            oee_data.get('toplam_ok', ''),
            oee_data.get('toplam_nok', ''),
            v['durum'],
        ]
        for col_idx, deger in enumerate(satir, start=1):
            ws_v.cell(row=row_num, column=col_idx, value=deger)
        _zebra(ws_v, row_num, len(basliklar_v))
    _kolonlari_genislet(ws_v)

    # ─── SAYFA 2: ÜRETİM KAYITLARI ────────────────────────────
    ws_u = wb.create_sheet('Uretim Kayitlari')
    ws_u.row_dimensions[1].height = 22

    basliklar_u = [
        'ID', 'Vardiya ID', 'Tarih', 'Lokasyon', 'Robot', 'Operatör',
        'Referans Kodu', 'OK Adet', 'NOK Adet',
        'Toplam', 'Hedef Adet', 'Cycle Time (sn)',
        'İş Süresi (dk)', 'Kalite %'
    ]
    _baslik_satiri(ws_u, basliklar_u, renk='16A34A')

    uretim = c.execute('''
        SELECT u.*, v.tarih, v.robot_no, v.operator_adi, COALESCE(v.lokasyon,'TK2') as v_lokasyon
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        ''' + ('WHERE v.' + tarih_sart[6:] if tarih_sart else '') + '''
        ORDER BY v.tarih DESC, u.id
    ''', params_list).fetchall()

    for row_num, u in enumerate(uretim, start=2):
        toplam = (u['ok_adet'] or 0) + (u['nok_adet'] or 0)
        ct = u['cycle_time_sn'] or 0
        is_dk = round(toplam * ct / 60, 1) if ct > 0 else ''
        kalite = round(u['ok_adet'] / toplam * 100, 1) if toplam > 0 else ''
        satir = [
            u['id'], u['vardiya_id'], u['tarih'], u['v_lokasyon'], u['robot_no'], u['operator_adi'],
            u['referans_kodu'], u['ok_adet'], u['nok_adet'],
            toplam, u['hedef_adet'], ct,
            is_dk, kalite
        ]
        for col_idx, deger in enumerate(satir, start=1):
            ws_u.cell(row=row_num, column=col_idx, value=deger)
        _zebra(ws_u, row_num, len(basliklar_u))
    _kolonlari_genislet(ws_u)

    # ─── SAYFA 3: DURUŞLAR ─────────────────────────────────────
    ws_d = wb.create_sheet('Duruslar')
    ws_d.row_dimensions[1].height = 22

    basliklar_d = [
        'ID', 'Vardiya ID', 'Tarih', 'Lokasyon', 'Robot', 'Operatör',
        'Duruş Sebebi', 'Duruş Tipi', 'Süre (dk)',
        'Başlangıç Saati', 'Açıklama'
    ]
    _baslik_satiri(ws_d, basliklar_d, renk='DC2626')

    duruslar = c.execute('''
        SELECT d.*, v.tarih, v.robot_no, v.operator_adi, COALESCE(v.lokasyon,'TK2') as v_lokasyon
        FROM duruslar d
        JOIN vardiyalar v ON v.id = d.vardiya_id
        ''' + ('WHERE v.' + tarih_sart[6:] if tarih_sart else '') + '''
        ORDER BY v.tarih DESC, d.id
    ''', params_list).fetchall()

    for row_num, d in enumerate(duruslar, start=2):
        dr = dict(d)  # sqlite3.Row -> dict
        satir = [
            dr['id'], dr['vardiya_id'], dr['tarih'], dr['v_lokasyon'], dr['robot_no'], dr['operator_adi'],
            dr['durus_sebebi'],
            'Planlı' if dr.get('durus_tipi') == 'planli' else 'Plansız',
            dr['sure_dk'], dr['baslangic_saati'], dr['aciklama']
        ]
        for col_idx, deger in enumerate(satir, start=1):
            ws_d.cell(row=row_num, column=col_idx, value=deger)
        _zebra(ws_d, row_num, len(basliklar_d))
    _kolonlari_genislet(ws_d)

    # ─── SAYFA 4: REFERANS BAZLI ÖZET ─────────────────────────
    ws_ref = wb.create_sheet('Referans Ozet')
    ws_ref.row_dimensions[1].height = 22

    basliklar_ref = [
        'Referans Kodu', 'Lokasyon', 'Toplam Üretim', 'OK Adet', 'NOK Adet',
        'Kalite %', 'Ort. Cycle Time (sn)'
    ]
    _baslik_satiri(ws_ref, basliklar_ref, renk='7C3AED')

    # Lokasyon kırılımı: TK1/TK2 aynı kodlu üretimleri tek satırda TOPLAMASIN (ayrı fabrikalar)
    ref_ozet = c.execute('''
        SELECT u.referans_kodu,
               COALESCE(v.lokasyon,'TK2') as v_lokasyon,
               SUM(u.ok_adet + u.nok_adet) as toplam,
               SUM(u.ok_adet) as ok,
               SUM(u.nok_adet) as nok,
               ROUND(AVG(u.cycle_time_sn), 0) as ort_ct
        FROM uretim_kayitlari u
        JOIN vardiyalar v ON v.id = u.vardiya_id
        ''' + ('WHERE v.' + tarih_sart[6:] if tarih_sart else '') + '''
        GROUP BY u.referans_kodu, COALESCE(v.lokasyon,'TK2')
        ORDER BY toplam DESC
    ''', params_list).fetchall()

    for row_num, r in enumerate(ref_ozet, start=2):
        toplam = r['toplam'] or 0
        ok = r['ok'] or 0
        kalite = round(ok / toplam * 100, 1) if toplam > 0 else ''
        satir = [r['referans_kodu'], r['v_lokasyon'], toplam, ok, r['nok'] or 0, kalite, r['ort_ct'] or '']
        for col_idx, deger in enumerate(satir, start=1):
            ws_ref.cell(row=row_num, column=col_idx, value=deger)
        _zebra(ws_ref, row_num, len(basliklar_ref))
    _kolonlari_genislet(ws_ref)

    conn.close()

    # ─── KAYDET ────────────────────────────────────────────────
    os.makedirs(MASAUSTU, exist_ok=True)
    wb.save(DOSYA_YOLU)

    zaman = datetime.now().strftime('%d.%m.%Y %H:%M')
    return {
        'basarili': True,
        'dosya': DOSYA_YOLU,
        'zaman': zaman,
        'vardiya_sayisi': len(vardiyalar),
        'uretim_kayit': len(uretim),
        'durus_kayit': len(duruslar),
    }
