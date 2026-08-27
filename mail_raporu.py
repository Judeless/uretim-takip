# -*- coding: utf-8 -*-
"""
mail_raporu.py — Günlük üretim raporunu Excel olarak üretir ve e-posta ile gönderir.

Akış (scheduler her gün gönderim saatinde çağırır):
  1. O günün tüm üretim kayıtlarını çeker (tarih, tesis, bölüm, operatör, referans, adet)
  2. Bir .xlsx dosyası oluşturur (data/gunluk_rapor/YYYY-MM-DD_UretimRaporu.xlsx)
  3. mail_alicilari tablosundaki aktif alıcılara SMTP ile e-posta atar (Excel ekli)

SMTP bilgileri mail_config.json'da tutulur (gitignore — git'e girmez).
Config yoksa/etkin değilse özellik sessizce devre dışıdır (cofle_test paterni).
"""
import os
import json
import smtplib
import sqlite3
import ssl
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formataddr

# Tel üretiminde adım, referans kodunun EKİNDEN çözülür ('93.TK.464 KAPAMA').
# app.py import EDİLEMEZ (Flask'ı ayağa kaldırır) — ortak mantık tel_proses'te.
from tel_proses import tel_koddan_adim, tel_ek_ayikla


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH     = os.path.join(PROJECT_DIR, 'uretim.db')
CONFIG_YOL  = os.path.join(PROJECT_DIR, 'mail_config.json')
RAPOR_DIR   = os.path.join(PROJECT_DIR, 'data', 'gunluk_rapor')
RAPOR_GUN_LIMIT = 60  # data/gunluk_rapor'da en fazla bu kadar günlük dosya tutulur

# Bölüm kod → görünen ad (app.py BOLUM_AD ile aynı — rapor okunur olsun)
BOLUM_AD = {
    'kaynak': 'Robot Kaynak',
    'montaj': 'Montaj',
    'metal':  'Metal Enjeksiyon',
    'isleme': 'İşleme',
    'lazer':  'Lazer Kesim',
    'pres':   'Pres Abkant',
    'plastik': 'Plastik Enjeksiyon',
    'tel':    'Tel Üretimi',
}


# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
def config_yukle():
    """mail_config.json'u okur. Yoksa/bozuksa None döner.
    encoding='utf-8-sig': Notepad (özellikle Server 2019) UTF-8 dosyaya BOM ekler;
    utf-8-sig BOM'u otomatik yutar (BOM'suz dosyayı da sorunsuz okur)."""
    if not os.path.exists(CONFIG_YOL):
        return None
    try:
        with open(CONFIG_YOL, encoding='utf-8-sig') as f:
            return json.load(f)
    except Exception as e:
        print(f'[MAIL] Config okunamadı: {e}')
        return None


def config_hatasi():
    """Dosya VAR ama JSON'u bozuksa hata metni, sorun yoksa ''.

    2026-07-30: config elle duzenlenirken bir satirin sonunda virgul unutuldu;
    dosya okunamaz oldu ve panel "SMTP tanimli degil" dedi. Mesaj yaniltiyordu
    (dosya oradaydi) ve o gun rapor maili hic gitmeyecekti. Gercek sebebi ve
    hatanin satirini panele tasiyoruz."""
    if not os.path.exists(CONFIG_YOL):
        return ''
    try:
        with open(CONFIG_YOL, encoding='utf-8-sig') as f:
            json.load(f)
        return ''
    except Exception as e:
        return str(e)


def yontem_al(c=None):
    """Gönderim yöntemi: 'outlook' (açık Outlook masaüstü — şifre gerekmez) | 'smtp'."""
    c = c or config_yukle() or {}
    return str(c.get('yontem', 'smtp')).lower()


def host_uygun(c=None):
    """config'te 'sadece_host' varsa, sadece o makinede gönderim yapılır.

    2026-07-29: sistem hem laptopta hem canlı sunucuda (ProManage) ayakta olduğu
    için alıcılara İKİ mail gitti. Laptopun veritabanı bayat olduğundan onunki
    "üretim bulunmamaktadır" diyordu — yani yanlış makinenin gönderdiği mail
    doğrusuyla çelişiyor. Kopya kurulumda config de kopyalandığı için 'etkin'
    bayrağı tek başına yetmiyor; makine kimliği gerekiyor.
    Alan yoksa davranış değişmez (geriye uyumlu)."""
    c = c if c is not None else (config_yukle() or {})
    beklenen = (c.get('sadece_host') or '').strip()
    if not beklenen:
        return True
    try:
        import socket
        return socket.gethostname().strip().upper() == beklenen.upper()
    except Exception:
        return True          # host okunamadı → engelleme (güvenli yön: mevcut davranış)


def etkin():
    """Mail entegrasyonu kullanıma hazır mı? (config var + etkin + doğru makine
    + yönteme göre zorunlu alanlar)"""
    c = config_yukle()
    if not c or not c.get('etkin'):
        return False
    if not host_uygun(c):
        return False
    if yontem_al(c) == 'outlook':
        return True  # Outlook masaüstü — ek bilgi gerekmez (açık ve girişli olmalı)
    return all(c.get(k) for k in ('smtp_host', 'smtp_port', 'gonderen'))


def gonderim_saati():
    """(saat, dakika) — config'ten, yoksa 17:00. scheduler bunu okur."""
    c = config_yukle() or {}
    raw = str(c.get('gonderim_saati', '17:00'))
    try:
        sa, dk = raw.split(':')
        return int(sa), int(dk)
    except Exception:
        return 17, 0


def durum_ozeti():
    """Dashboard göstergesi için maskelenmiş durum (şifre ASLA dönmez)."""
    c = config_yukle()
    if not c:
        h = config_hatasi()
        # dosya var ama bozuk -> "tanimli degil" demek yaniltici olur
        return {'config_var': bool(h), 'etkin': False, 'config_hatasi': h}
    import socket
    try:
        bu_host = socket.gethostname()
    except Exception:
        bu_host = ''
    return {
        'config_var': True,
        'etkin': bool(etkin()),
        # Host kilidi gönderimi SESSIZCE durdurabilir (config'te sadece_host yanlış
        # yazılırsa). Panelde nedeni görünsün diye dönüyoruz.
        'sadece_host': (c.get('sadece_host') or ''),
        'bu_host': bu_host,
        'host_engeli': bool(c.get('etkin')) and not host_uygun(c),
        'yontem': yontem_al(c),
        'smtp_host': c.get('smtp_host', ''),
        'smtp_port': c.get('smtp_port', ''),
        'gonderen': c.get('gonderen', ''),
        'gonderim_saati': c.get('gonderim_saati', '17:00'),
        'kapsam_lokasyon': c.get('kapsam_lokasyon', 'HEPSI'),
        # Kaç ayrı mail gidecek (TK1/TK2 ayrımı panelde görünsün)
        'kapsamlar': kapsam_lokasyonlari(c),
        # Tesis seçmediği için rapor ALMAYAN aktif alıcılar — panel uyarır
        'raporsuz': raporsuz_alicilar(),
    }


# ─────────────────────────────────────────────────────────────
# ALICILAR (mail_alicilari tablosu)
# ─────────────────────────────────────────────────────────────
def _hepsi_mi(lokasyon):
    """Bu kapsam 'tüm tesisler' mi? (None / '' / 'HEPSI')"""
    return not lokasyon or str(lokasyon).strip().upper() == 'HEPSI'


TESISLER = ('TK2', 'TK1')


def kapsam_lokasyonlari(c=None):
    """Rapor hangi tesis(ler) için üretilecek — HER BİRİ AYRI MAİL.

    RAPOR HER ZAMAN TESİS BAZLI AYRIDIR (kullanıcı 2026-08-26). Birleşik tek
    rapor seçeneği KALDIRILDI: alıcı bazlı TK1/TK2 seçimi gelince birleşik mod
    hem seçimi anlamsız kılıyor hem de sessiz bir tuzak oluşturuyordu —
    config'te eski 'HEPSI' değeri kalınca panelde kutucuklar ayarlanmış olsa
    bile herkese tek birleşik rapor gidiyordu (kullanıcı: "tek kişi seçmeme
    rağmen 6 kişiye gönderildi").

    mail_config.json 'kapsam_lokasyon' artık yalnız KAPSAMI DARALTIR:
      yok / "AYRI" / "HEPSI" → ['TK2', 'TK1'] (ikisi de, AYRI mail)
      "TK1"                  → yalnız TK1
      ["TK2","TK1"]          → açıkça listelenen tesisler
    """
    c = c or config_yukle() or {}
    d = c.get('kapsam_lokasyon', 'AYRI')
    if isinstance(d, (list, tuple)):
        liste = [str(x).strip().upper() for x in d if str(x).strip()]
        liste = [x for x in liste if x in TESISLER]
        return liste or list(TESISLER)
    d = str(d or '').strip().upper()
    if d in TESISLER:
        return [d]
    return list(TESISLER)


def _db(conn=None):
    if conn is not None:
        return conn, False
    c = sqlite3.connect(DB_PATH, timeout=20.0)
    c.row_factory = sqlite3.Row
    return c, True


def _alici_lokasyonlari(deger):
    """Alıcının seçtiği tesisler. BOŞ = HEPSİ (bkz. database migration notu)."""
    return [x.strip().upper() for x in str(deger or '').split(',') if x.strip()]


def aktif_alicilar(conn=None, lokasyon=None):
    """aktif=1 olan alıcı e-postaları. lokasyon verilirse YALNIZ o tesisi
    SEÇMİŞ olanlar.

    KUTUCUK BELİRLEYİCİDİR (kullanıcı 2026-08-26): hiç tesis seçmemiş alıcıya
    rapor GİTMEZ. Önce "boş = hepsi" varsayılmıştı; kullanıcı tek kişiyi
    işaretleyip diğerlerinin kutucuklarını boşalttığında mail yine 6 kişiye
    gitti. Ekranda işaretsiz duran bir kutunun "hepsi seçili" anlamına gelmesi
    yanıltıcı — işaretsiz kutu 'hayır' demektir.
    Rapor almayan alıcılar panelde AÇIKÇA uyarı olarak gösterilir, sessizce
    listeden düşmüş gibi durmasınlar.

    "1 kişi 2 raporu da seçerse 2 ayrı excel gitsin": gönderim tesis başına bir
    kez döner, iki tesisi de seçen kişi iki turda da listede olur → iki ayrı
    mail, iki ayrı Excel."""
    c, kapat = _db(conn)
    try:
        rows = c.execute(
            "SELECT email, COALESCE(lokasyonlar,'') AS lokasyonlar "
            "FROM mail_alicilari WHERE COALESCE(aktif,1)=1 ORDER BY email"
        ).fetchall()
    except Exception:
        # Kolon henüz eklenmemiş (eski DB) → tesis süzmesi yapmadan devam et
        rows = c.execute(
            "SELECT email, '' AS lokasyonlar FROM mail_alicilari "
            "WHERE COALESCE(aktif,1)=1 ORDER BY email").fetchall()
    finally:
        if kapat:
            c.close()
    if not lokasyon or _hepsi_mi(lokasyon):
        return [r['email'] for r in rows]
    hedef = str(lokasyon).upper()
    return [r['email'] for r in rows
            if hedef in _alici_lokasyonlari(r['lokasyonlar'])]


def raporsuz_alicilar(conn=None):
    """Aktif ama HİÇBİR tesis seçmemiş alıcılar — bunlara rapor gitmez.

    Panel bunu uyarı olarak gösterir: kutucuğu boş bırakmak artık 'hepsi'
    değil 'hiçbiri' demek; sessiz kalırsa kişi raporu almayı kestiğini
    fark etmez."""
    c, kapat = _db(conn)
    try:
        rows = c.execute(
            "SELECT email, COALESCE(lokasyonlar,'') AS lokasyonlar "
            "FROM mail_alicilari WHERE COALESCE(aktif,1)=1 ORDER BY email").fetchall()
    except Exception:
        return []
    finally:
        if kapat:
            c.close()
    return [r['email'] for r in rows if not _alici_lokasyonlari(r['lokasyonlar'])]


# ─────────────────────────────────────────────────────────────
# VERİ + EXCEL
# ─────────────────────────────────────────────────────────────
def gunluk_veri(tarih=None, lokasyon=None):
    """O günün üretim satırları: (tesis, bolum, operator, referans, adet).
    lokasyon verilirse yalnız o tesis; verilmezse (None/HEPSI) tüm tesisler.
    """
    if tarih is None:
        tarih = date.today().isoformat()
    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    conn.row_factory = sqlite3.Row
    try:
        sql = """
            SELECT
                COALESCE(v.lokasyon, 'TK2')      AS tesis,
                COALESCE(v.bolum, 'kaynak')      AS bolum,
                v.operator_adi                   AS operator,
                u.referans_kodu                  AS referans,
                SUM(u.ok_adet)                   AS adet
            FROM uretim_kayitlari u
            JOIN vardiyalar v ON v.id = u.vardiya_id
            WHERE v.tarih = ?
              AND u.referans_kodu IS NOT NULL AND u.referans_kodu != ''
        """
        params = [tarih]
        if lokasyon and lokasyon.upper() != 'HEPSI':
            sql += " AND COALESCE(v.lokasyon,'TK2') = ?"
            params.append(lokasyon.upper())
        # Hat (robot_no) da kırılıma girer: tel üretiminde aynı operatör aynı gün
        # farklı makinelerde çalışabilir ('Kapama 1' ve 'Kapama 3') — satırlar
        # birleşmesin, kimin hangi makinede ne yaptığı görünsün.
        sql += """
            GROUP BY tesis, bolum, v.operator_adi, u.referans_kodu, v.robot_no
            HAVING SUM(u.ok_adet) > 0
            ORDER BY tesis, bolum, v.operator_adi, u.referans_kodu
        """
        # ── TEL ÜRETİMİ: TÜM OPERATÖRLER RAPORDA GÖRÜNÜR (2026-08-04) ─────────
        # Kullanıcı: "raporda diğer operatörlerin de ne yaptığını görelim."
        # Ara adımlar SÜZÜLMEZ — çoklu sayım artık REFERANS KODUNDAN ayrışıyor:
        # ara operasyonlar adım ekiyle kaydediliyor ('93.TK.464 KESIM'), bitmiş
        # ürün base kodla ('93.TK.464'). Kesimi yapan operatörün 200 adet kesim
        # yaptığı raporda görünür ama ürün toplamına karışmaz — farklı koddur.
        return conn.execute(sql, params).fetchall()
    finally:
        conn.close()


# Bölüm kimlik renkleri — dashboard paletiyle (--b-*) AYNI değerler.
_BOLUM_RENK = {'kaynak': '#1D4ED8', 'montaj': '#07734F', 'metal': '#8F6100',
               'isleme': '#3F6212', 'lazer': '#0E7490', 'pres': '#9D174D',
               'plastik': '#A21CAF', 'tel': '#B45309'}


def _bolum_kirilimi(tarih, lokasyon=None):
    """Tesis+bölüm bazlı özet: [{tesis, bolum, satir, adet}] — adede göre azalan.
    Mail gövdesindeki "nerede ne üretildi" tablosu için (2026-07-30)."""
    ozet = {}
    for r in gunluk_veri(tarih, lokasyon):
        a = dict(r) if not isinstance(r, dict) else r
        k = (a['tesis'], a['bolum'])
        d = ozet.setdefault(k, {'tesis': a['tesis'], 'bolum': a['bolum'], 'satir': 0, 'adet': 0})
        d['satir'] += 1
        d['adet'] += int(a['adet'] or 0)
    return sorted(ozet.values(), key=lambda x: -x['adet'])


# ─────────────────────────────────────────────────────────────
# ANALİZ BÖLÜMÜ (kullanıcı 2026-08-21: "her gün seçili kişilere rapor gitsin")
# ─────────────────────────────────────────────────────────────
# analiz.py yerel motoru bulguları çıkarır; ai_config.json hazırsa Claude bir de
# yönetici özeti yazar. HER İKİSİ DE OPSİYONELDİR: analiz patlarsa mail yine
# gider (sadece bu bölüm boş kalır) — günlük rapor akışı asla analize bağlanmaz.

def analiz_topla(tarih, lokasyon=None, yorum=True):
    """(ozet, bulgular, yorum_metni, yorum_hata).

    YORUM_HATA NEDEN DÖNÜYOR (kullanıcı 2026-08-27: "dün akşam raporda yönetici
    özeti yoktu, test mailinde geldi — neden?"): AI çağrısı bilinçli olarak maili
    ASLA engellemez; ama hata yalnız sunucu konsoluna yazılınca bölümün neden
    kaybolduğu görünmez kalıyordu. Sebep artık maile de düşer."""
    try:
        import analiz as _an
    except Exception as e:
        print(f'[MAIL] analiz modulu yuklenemedi: {e}')
        return None, [], ''
    # 'HEPSI' BİR LOKASYON ADI DEĞİL (kullanıcı 2026-08-26 hata bildirimi):
    # analiz.vardiya_metrikleri lokasyonu "COALESCE(lokasyon,'TK2') = ?" ile
    # süzüyor; 'HEPSI' geçilince HİÇBİR vardiya eşleşmiyor ve mailin analiz
    # bölümü "bu dönem için sisteme yansımış herhangi bir vardiya bulunmamaktadır"
    # diyordu — aynı mailin üretim tablosunda 34 satır veri dururken.
    # gunluk_veri bu ayrımı zaten yapıyordu, analiz yolu yapmıyordu.
    _lok = None if _hepsi_mi(lokasyon) else str(lokasyon).upper()
    try:
        a = _an.gunluk_ozet(tarih, _lok)
    except Exception as e:
        print(f'[MAIL] analiz hatasi: {e}')
        return None, [], '', f'analiz çalışmadı: {e}'
    metin, yorum_hata = '', ''
    if yorum:
        try:
            y, hata = _an.yorum_uret(a)
            if y:
                metin = y.get('metin') or ''
            elif hata:
                yorum_hata = hata
                print(f'[MAIL] AI yorumu atlandi: {hata}')
        except Exception as e:
            yorum_hata = f'{type(e).__name__}: {e}'
            print(f'[MAIL] AI yorumu hatasi: {e}')
    return a.get('ozet'), a.get('bulgular') or [], metin, yorum_hata


SIDDET_RENK_MAIL = {'kritik': '#DC2626', 'uyari': '#D97706', 'bilgi': '#4E4C63'}
SIDDET_AD_MAIL = {'kritik': 'KRİTİK', 'uyari': 'UYARI', 'bilgi': 'BİLGİ'}


def _html_govde(tarih_tr, satir, toplam, kirilim, analiz_ozet=None,
                bulgular=None, yorum='', tesis='', yorum_hata=''):
    """Mailin HTML gövdesi — Cofle Forge paleti (2026-07-30 kullanıcı isteği:
    "gunluk uretim raporu tasarimi ... yeni tasarimimiza gore").

    E-posta istemcisi kısıtları: <style> bloğu ve harici font YOK (Outlook
    masaüstü çoğunu atar) → her kural INLINE, yerleşim TABLO ile. Genişlik 640px
    ve hücreler yüzdeli — dar ekranda da okunur. Düz metin sürümü de gönderilir
    (multipart/alternative), HTML kapalı istemciler onu görür."""
    mor, koyu, gri, cizgi = '#6D28D9', '#14122B', '#4E4C63', '#E1E1EA'

    def hucre(baslik, deger, renk):
        return (f'<td width="50%" style="padding:0 6px" valign="top">'
                f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
                f'style="background:#FFFFFF;border:1px solid {cizgi};border-radius:10px"><tr>'
                f'<td style="padding:11px 13px;font-family:Segoe UI,Arial,sans-serif">'
                f'<div style="font-size:10px;font-weight:700;color:{gri};'
                f'text-transform:uppercase;letter-spacing:.4px">{baslik}</div>'
                f'<div style="font-size:20px;font-weight:800;color:{renk};padding-top:3px">'
                f'{deger}</div></td></tr></table></td>')

    if satir:
        # TK1: tel adetleri proses adımı — 'Toplam işlem' + açıklama (2026-08-27)
        _t_etiket = 'Toplam işlem' if tesis == 'TK1' else 'Toplam üretim'
        kartlar = ('<tr>' + hucre(_t_etiket, f'{toplam:,}'.replace(',', '.'), mor)
                   + hucre('Kayıt satırı', f'{satir}', koyu) + '</tr>')
        if tesis == 'TK1':
            kartlar += ('<tr><td colspan="2" style="padding:2px 14px 8px;'
                        'font-family:Segoe UI,Arial,sans-serif;font-size:10.5px;'
                        'color:#8A8798;font-style:italic">Tel üretimi tüm proses '
                        'adımlarıyla dahildir — rakam bitmiş ürün adedi değildir.</td></tr>')
        satirlar = ''
        for k in kirilim:
            ad = BOLUM_AD.get(k['bolum'], k['bolum'])
            renk = _BOLUM_RENK.get(k['bolum'], mor)
            satirlar += (
                f'<tr>'
                f'<td style="padding:8px 13px;border-bottom:1px solid {cizgi};'
                f'font-family:Segoe UI,Arial,sans-serif;font-size:12px;color:{koyu}">'
                f'<span style="display:inline-block;width:3px;height:13px;background:{renk};'
                f'vertical-align:-2px;margin-right:9px"></span>'
                f'<b>{ad}</b> <span style="color:{gri}">· {k["tesis"]}</span></td>'
                f'<td align="right" style="padding:10px 14px;border-bottom:1px solid {cizgi};'
                f'font-family:Segoe UI,Arial,sans-serif;font-size:11px;color:{gri}">'
                f'{k["satir"]} satır</td>'
                f'<td align="right" style="padding:10px 14px;border-bottom:1px solid {cizgi};'
                f'font-family:Segoe UI,Arial,sans-serif;font-size:13px;font-weight:800;color:{koyu}">'
                f'{k["adet"]:,}'.replace(',', '.') + '</td></tr>')
        tablo = ('' if not kirilim else
                 f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
                 f'style="background:#FFFFFF;border:1px solid {cizgi};border-radius:10px;'
                 f'margin-top:18px"><tr><td colspan="3" style="padding:13px 14px;'
                 f'border-bottom:1px solid {cizgi};font-family:Segoe UI,Arial,sans-serif;'
                 f'font-size:10px;font-weight:700;color:{gri};text-transform:uppercase;'
                 f'letter-spacing:.4px">Bölüm bazında</td></tr>{satirlar}</table>')
        icerik = (
            f'<p style="margin:0 0 14px;font-family:Segoe UI,Arial,sans-serif;font-size:13px;'
            f'color:{koyu}">Merhaba,<br><b>{tarih_tr}</b> tarihli günlük üretim raporu ektedir.</p>'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="margin:0 -6px">{kartlar}</table>{tablo}'
            f'<p style="margin:16px 0 0;font-family:Segoe UI,Arial,sans-serif;font-size:12px;'
            f'color:{gri}">Operatör ve referans kırılımı ekteki Excel dosyasındadır.</p>')
    else:
        icerik = (
            f'<p style="margin:0 0 12px;font-family:Segoe UI,Arial,sans-serif;font-size:13px;'
            f'color:{koyu}">Merhaba,</p>'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="background:#FFFFFF;border:1px solid {cizgi};border-radius:10px"><tr>'
            f'<td style="padding:18px;font-family:Segoe UI,Arial,sans-serif;font-size:14px;'
            f'color:{koyu};text-align:center;font-size:13px">'
            f'<b>{tarih_tr}</b> tarihinde sisteme kayıtlı üretim bulunmamaktadır.</td>'
            f'</tr></table>')

    # ── ANALİZ BÖLÜMÜ ───────────────────────────────────────────────────────
    # Üretim tablosunun ALTINA eklenir: önce ne üretildiği, sonra neyin ters
    # gittiği. Bulgu yoksa bölüm hiç basılmaz (boş kutu gürültüdür).
    analiz_html = ''
    if analiz_ozet and (bulgular or yorum or yorum_hata):
        satirlar = ''.join(
            f'<tr><td style="padding:7px 0;border-bottom:1px solid {cizgi};'
            f'font-family:Segoe UI,Arial,sans-serif;font-size:12px;color:{koyu};'
            f'line-height:1.5" valign="top">'
            f'<span style="display:inline-block;background:'
            f'{SIDDET_RENK_MAIL.get(b["siddet"], gri)}18;color:'
            f'{SIDDET_RENK_MAIL.get(b["siddet"], gri)};font-size:9px;font-weight:700;'
            f'padding:1px 6px;border-radius:4px;letter-spacing:.4px">'
            f'{SIDDET_AD_MAIL.get(b["siddet"], "")}</span> '
            f'<b>{b["baslik"]}</b><br>'
            f'<span style="color:{gri};font-size:11px">{b["detay"]}</span></td></tr>'
            for b in (bulgular or [])[:6])
        yorum_html = ''
        if not yorum and yorum_hata:
            yorum_html = (
                f'<tr><td style="padding:8px 14px;font-family:Segoe UI,Arial,sans-serif;'
                f'font-size:11px;color:#8A8798;font-style:italic">Yönetici özeti bu '
                f'koşuda üretilemedi: {yorum_hata}</td></tr>'
            )
        if yorum:
            yorum_html = (
                f'<tr><td style="padding:12px 14px;background:#F5F3FF;'
                f'border:1px solid #DDD6FE;border-radius:10px;'
                f'font-family:Segoe UI,Arial,sans-serif;font-size:12px;color:{koyu};'
                f'line-height:1.65;white-space:pre-wrap">'
                f'<div style="font-weight:800;color:{mor};padding-bottom:5px">'
                f'Yönetici Özeti</div>{yorum}</td></tr>'
                f'<tr><td style="height:12px"></td></tr>')
        analiz_html = (
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="background:#FFFFFF;border:1px solid {cizgi};border-radius:10px;'
            f'margin-top:14px"><tr><td style="padding:14px 16px">'
            f'<div style="font-family:Segoe UI,Arial,sans-serif;font-size:13px;'
            f'font-weight:800;color:{koyu};padding-bottom:4px">Otomatik Analiz</div>'
            f'<div style="font-family:Segoe UI,Arial,sans-serif;font-size:11px;'
            f'color:{gri};padding-bottom:10px">'
            f'OEE %{analiz_ozet.get("oee", 0)} · Kullanılabilirlik '
            f'%{analiz_ozet.get("availability", 0)} · Performans '
            f'%{min(analiz_ozet.get("performance", 0), 100)} · Kalite '
            f'%{analiz_ozet.get("quality", 0)} · Plansız duruş '
            f'{analiz_ozet.get("plansiz_dk", 0)} dk</div>'
            f'<table width="100%" cellpadding="0" cellspacing="0" border="0">'
            f'{yorum_html}{satirlar}</table>'
            f'</td></tr></table>')
        icerik = icerik + analiz_html

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8">'
        f'<meta name="viewport" content="width=device-width,initial-scale=1">'
        f'<meta name="color-scheme" content="light only">'
        f'<meta name="supported-color-schemes" content="light"></head>'
        f'<body style="margin:0;padding:0;background:#ECECF1">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        f'style="background:#ECECF1;padding:24px 12px"><tr><td align="center">'
        f'<table width="640" cellpadding="0" cellspacing="0" border="0" '
        f'style="width:100%;max-width:640px">'
        # başlık şeridi — logo gradyanı (#7C3AED → #A78BFA); gradyan desteklemeyen
        # istemcide düz mor kalır, metin beyaz olduğu için kontrast korunur
        f'<tr><td style="background:{mor};background-image:linear-gradient(90deg,#7C3AED,#A78BFA);'
        f'border-radius:12px 12px 0 0;padding:15px 18px">'
        f'<div style="font-family:Segoe UI,Arial,sans-serif;font-size:16px;font-weight:800;'
        f'color:#FFFFFF;letter-spacing:.5px">COFLE FORGE</div>'
        f'<div style="font-family:Segoe UI,Arial,sans-serif;font-size:11px;color:#EDE4FE;'
        # TESİS BAŞLIKTA (2026-08-26): TK1 ve TK2 ayrı mail gidiyor, açan kişi
        # hangi tesisin raporuna baktığını ilk satırda görsün.
        f'padding-top:3px">Günlük Üretim Raporu'
        + (f' · {tesis}' if tesis else '') + f' · {tarih_tr}</div></td></tr>'
        f'<tr><td style="background:#ECECF1;padding:20px 4px 4px">{icerik}</td></tr>'
        f'<tr><td style="padding:16px 4px 0;font-family:Segoe UI,Arial,sans-serif;'
        f'font-size:10px;color:{gri};border-top:1px solid {cizgi};margin-top:10px">'
        f'Bu e-posta Cofle Forge üretim takip sistemi tarafından otomatik gönderilmiştir.'
        f'</td></tr></table></td></tr></table></body></html>')


def excel_olustur(tarih=None, lokasyon=None):
    """Günlük üretim Excel'i üretir. (dosya_yolu, satir_sayisi, toplam_adet) döner."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if tarih is None:
        tarih = date.today().isoformat()
    rows = gunluk_veri(tarih, lokasyon)

    os.makedirs(RAPOR_DIR, exist_ok=True)
    # TESİS DOSYA ADINDA (2026-08-26): TK1 ve TK2 raporları ayrı maille
    # gidiyor; aynı ada yazsalardı ikinci rapor birincinin üstüne biner ve
    # ekler karışırdı (aynı dakikada üretilirler).
    _ek_ad = '' if _hepsi_mi(lokasyon) else f'_{str(lokasyon).upper()}'
    dosya = os.path.join(RAPOR_DIR, f'{tarih}_UretimRaporu{_ek_ad}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Günlük Üretim'

    # 'Proses Adımı' (2026-08-18): tel üretiminde aynı ürün birden çok adımdan
    # geçiyor ve her adım kendi kod ekiyle ayrı satır oluyor. Adım artık koddan
    # ayıklanıp KENDİ KOLONUNA yazılır — okuyan "aynı referans neden iki kez
    # yazılmış" diye düşünmez, satırların farklı OPERASYONLAR olduğunu görür.
    basliklar = ['Tarih', 'Tesis', 'Bölüm', 'Operatör', 'Referans Kodu',
                 'Proses Adımı', 'Adet']
    ws.append(basliklar)

    # Başlık stili (koyu lacivert — Cofle kimliği)
    baslik_fill = PatternFill('solid', fgColor='1E293B')
    baslik_font = Font(bold=True, color='FFFFFF', size=11)
    ince = Side(style='thin', color='D0D7E2')
    kenar = Border(left=ince, right=ince, top=ince, bottom=ince)
    for i in range(1, len(basliklar) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = baslik_fill
        c.font = baslik_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = kenar

    tarih_tr = _tarih_tr(tarih)
    # TEL DE TOPLAMA GİRER (kullanıcı 2026-08-27 — 2026-08-18 kararını değiştirdi):
    # eskiden tel satırları çift-sayım yanılgısı ("100 kapama + 100 son montaj =
    # 200 ürün") yüzünden genel toplamdan dışlanıyordu ve TK1 toplamı gerçek işin
    # çoğunu göstermiyordu. Yeni karar: rakam DAHİL, ama adı dürüst — tel varsa
    # etiket 'TOPLAM İŞLEM' olur ve bitmiş ürün adedi olmadığı açıkça yazılır;
    # adım bazlı alt toplamlar da durur (hangi adım ne üretti oradan okunur).
    toplam_adet = 0
    tel_adim_toplam = {}
    for r in rows:
        adet = int(r['adet'] or 0)
        tel_mi = (r['bolum'] or '') == 'tel'
        adim = tel_koddan_adim(r['referans']) if tel_mi else None
        if tel_mi:
            tel_adim_toplam[adim or '—'] = tel_adim_toplam.get(adim or '—', 0) + adet
        toplam_adet += adet
        ws.append([
            tarih_tr,
            r['tesis'],
            BOLUM_AD.get(r['bolum'], r['bolum']),
            r['operator'] or '',
            # Tel'de kod ADIMSIZ yazılır: aynı ürünün satırları aynı kodda görünsün,
            # hangi adım olduğu yan kolonda dursun.
            tel_ek_ayikla(r['referans']) if tel_mi else r['referans'],
            adim or '',
            adet,
        ])

    # Gövde hücre kenarları + Adet sağa yasla
    for satir in range(2, ws.max_row + 1):
        for sutun in range(1, len(basliklar) + 1):
            hc = ws.cell(row=satir, column=sutun)
            hc.border = kenar
            if sutun == 7:
                hc.alignment = Alignment(horizontal='right')

    # Toplam satır(lar)ı
    if rows:
        def _toplam_satiri(etiket, deger, kalin=True):
            r = ws.max_row + 1
            e = ws.cell(row=r, column=6, value=etiket)
            e.font = Font(bold=kalin)
            e.alignment = Alignment(horizontal='right')
            d = ws.cell(row=r, column=7, value=deger)
            d.font = Font(bold=kalin)
            d.alignment = Alignment(horizontal='right')

        _toplam_satiri('TOPLAM İŞLEM' if tel_adim_toplam else 'TOPLAM', toplam_adet)
        if tel_adim_toplam:
            not_r = ws.max_row + 2
            nc = ws.cell(row=not_r, column=1,
                         value='TEL ÜRETİMİ — adım bazlı kırılım (aynı ürün birden '
                               'çok adımdan geçer; üstteki toplam bitmiş ürün adedi '
                               'DEĞİL, yapılan toplam iştir):')
            nc.font = Font(bold=True, italic=True)
            for adim, deger in tel_adim_toplam.items():
                _toplam_satiri(adim, deger, kalin=False)

    # Kolon genişlikleri
    for kol, gen in zip('ABCDEFG', (12, 8, 18, 22, 20, 14, 10)):
        ws.column_dimensions[kol].width = gen
    ws.freeze_panes = 'A2'

    wb.save(dosya)
    _eski_raporlari_temizle()
    return dosya, len(rows), toplam_adet


def _eski_raporlari_temizle():
    if not os.path.isdir(RAPOR_DIR):
        return
    dosyalar = sorted(
        (f for f in os.listdir(RAPOR_DIR) if f.endswith('.xlsx')),
        reverse=True
    )
    for eski in dosyalar[RAPOR_GUN_LIMIT:]:
        try:
            os.remove(os.path.join(RAPOR_DIR, eski))
        except Exception:
            pass


def _tarih_tr(iso):
    """'2026-07-10' → '10.07.2026'."""
    try:
        p = iso.split('-')
        return f'{p[2]}.{p[1]}.{p[0]}'
    except Exception:
        return iso


# ─────────────────────────────────────────────────────────────
# GÖNDERİM
# ─────────────────────────────────────────────────────────────
def _outlook_gonder(alicilar, konu, govde, ek_yol, html=None):
    """Açık Outlook masaüstü uygulaması üzerinden gönderir (win32com COM) — ŞİFRE GEREKMEZ.
    Şart: Outlook aynı Windows oturumunda çalışıyor ve giriş yapılmış olmalı.
    Not: COM çağrısı arka planda çalıştığından pythoncom.CoInitialize() zorunlu."""
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    try:
        ol = win32com.client.Dispatch('Outlook.Application')
        mail = ol.CreateItem(0)  # 0 = olMailItem
        mail.To = '; '.join(alicilar)
        mail.Subject = konu
        if html:
            mail.HTMLBody = html   # Outlook HTMLBody atanınca Body'yi yok sayar
        else:
            mail.Body = govde
        if ek_yol and os.path.exists(ek_yol):
            mail.Attachments.Add(os.path.abspath(ek_yol))
        mail.Send()
    finally:
        pythoncom.CoUninitialize()


# ── GÖNDERİM İZİ (kullanıcı 2026-08-26) ──────────────────────────────
# "mail gönderildi dedi ama Outlook'ta göremedim." SMTP 250 yanıtı 'kabul
# ettim' demektir, 'kullanıcının kutusuna düştü' demek DEĞİLDİR — arada spam
# filtresi, yönlendirme kuralı, hatta MX taşınması varken eski posta kutusuna
# bakıyor olmak var. Sunucunun son yanıtındaki KUYRUK KİMLİĞİ bu boşluğu
# kapatır: Google Admin > Reporting > Email Log Search'te o kimlikle mesajın
# nereye teslim edildiği birebir görülür.
class _IzliSMTP(smtplib.SMTP):
    son_yanit = None

    def getreply(self):
        kod, mesaj = super().getreply()
        self.son_yanit = (kod, mesaj)
        return kod, mesaj


class _IzliSMTPSSL(smtplib.SMTP_SSL):
    son_yanit = None

    def getreply(self):
        kod, mesaj = super().getreply()
        self.son_yanit = (kod, mesaj)
        return kod, mesaj


def _yanit_metni(srv):
    """Sunucunun son yanıtı — kuyruk kimliği burada geçer. Okunamazsa ''."""
    try:
        kod, mesaj = srv.son_yanit or (None, None)
        if kod is None:
            return ''
        if isinstance(mesaj, bytes):
            mesaj = mesaj.decode('utf-8', 'replace')
        return f'{kod} {" ".join(str(mesaj).split())}'
    except Exception:
        return ''


def _sifre_temizle(host, sifre):
    """Şifreyi giriş için hazırlar.

    GOOGLE UYGULAMA ŞİFRESİ BOŞLUKLU GÖSTERİLİR: Google ekranda 16 haneyi
    "abcd efgh ijkl mnop" diye dörtlü gruplar hâlinde yazar. Kopyalayıp
    yapıştıranın elinde boşluklu bir metin kalıyor ve giriş '535 BadCredentials'
    ile reddediliyor — sebebi görünmediği için "port mu yanlış" diye aranıyor
    (kullanıcı 2026-08-26 tam bunu sordu). Google boşluğu şifrenin parçası
    saymaz; burada temizlenir.
    Diğer sunucularda YALNIZ baş/son boşluk alınır: oralarda iç boşluk gerçek
    bir karakter olabilir, sessizce silmek girişi bozardı."""
    s = str(sifre or '')
    _h = str(host or '').lower()
    if 'gmail' in _h or 'google' in _h:
        return ''.join(s.split())
    return s.strip()


def _giris(srv, host, kullanici, sifre):
    """SMTP login — hata mesajını ANLAŞILIR hâle getirir.

    Gmail'e geçişte (2026-08-26) en sık takılınan yer: hesabın normal şifresi
    SMTP'de ÇALIŞMAZ, 'Uygulama Şifresi' (App Password) gerekir ve o da ancak
    2 adımlı doğrulama açıkken üretilebilir. Ham hata ('535 5.7.8 Username and
    Password not accepted') panelde 'şifre yanlış' gibi okunup saatler
    kaybettirir; sebebi ve çözümü burada yazılır."""
    _h = str(host or '').lower()
    sifre = _sifre_temizle(host, sifre)
    try:
        srv.login(kullanici, sifre)
    except smtplib.SMTPAuthenticationError as e:
        if 'gmail' in _h or 'google' in _h:
            # UZUNLUK İPUCU: uygulama şifresi TAM 16 hane. Farklıysa büyük
            # ihtimalle hesabın normal şifresi yazılmış — en sık hata bu.
            _n = len(sifre)
            _ipucu = ('Girilen şifre 16 hane DEĞİL (%d hane) — büyük ihtimalle hesabın '
                      'normal şifresi yazılmış. ' % _n) if _n != 16 else (
                     'Şifre 16 hane, biçim doğru görünüyor — şifre yanlış/iptal edilmiş '
                     'olabilir ya da Workspace yöneticisi uygulama şifrelerini kapatmış '
                     'olabilir. ')
            raise RuntimeError(
                f'Gmail girişi reddedildi ({kullanici}). '
                f'PORT/SUNUCU AYARI SORUN DEĞİL — Google sunucusu yanıt verdi, '
                f'reddedilen şey kimlik bilgisi. {_ipucu}'
                f'Yapılacak: Google Hesabı > Güvenlik > 2 Adımlı Doğrulama AÇIK olmalı, '
                f'sonra "Uygulama şifreleri"nden 16 haneli şifre üretip '
                f'mail_config.json > "sifre" alanına yazın (boşluklar önemsiz, '
                f'sistem temizler). Workspace hesabında bu seçenek yoksa yönetici '
                f'engellemiştir; alternatif smtp-relay.gmail.com ile IP izinli röledir. '
                f'Sunucu yanıtı: {e}') from e
        raise RuntimeError(f'SMTP girişi reddedildi ({kullanici}). Kullanıcı adı/şifre '
                           f'veya sunucu ayarı hatalı. Sunucu yanıtı: {e}') from e


def _smtp_gonder(cfg, alicilar, konu, govde, ek_yol, html=None):
    """Tek SMTP oturumunda maili kurar ve gönderir. Hata fırlatır (çağıran yakalar)."""
    msg = MIMEMultipart()
    gonderen = cfg['gonderen']
    msg['From'] = formataddr((cfg.get('gonderen_ad', 'Cofle Forge'), gonderen))
    msg['To'] = ', '.join(alicilar)
    msg['Subject'] = konu
    # multipart/alternative: HTML gosteremeyen istemci duz metni gorur.
    # Ek VARSA dis kabuk multipart/mixed kalmali, alternatif ic ice girer.
    if html:
        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText(govde, 'plain', 'utf-8'))
        alt.attach(MIMEText(html, 'html', 'utf-8'))
        msg.attach(alt)
    else:
        msg.attach(MIMEText(govde, 'plain', 'utf-8'))

    if ek_yol and os.path.exists(ek_yol):
        with open(ek_yol, 'rb') as f:
            ek = MIMEApplication(f.read(), _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            ek.add_header('Content-Disposition', 'attachment', filename=os.path.basename(ek_yol))
            msg.attach(ek)

    host = cfg['smtp_host']
    port = int(cfg['smtp_port'])
    kullanici = cfg.get('kullanici') or gonderen
    sifre = cfg.get('sifre', '')
    mod = str(cfg.get('guvenlik', 'starttls')).lower()  # starttls | ssl | none

    if mod == 'ssl':
        ctx = ssl.create_default_context()
        with _IzliSMTPSSL(host, port, timeout=30, context=ctx) as srv:
            if sifre:
                _giris(srv, host, kullanici, sifre)
            red = srv.sendmail(gonderen, alicilar, msg.as_string())
            iz = _yanit_metni(srv)
    else:
        with _IzliSMTP(host, port, timeout=30) as srv:
            srv.ehlo()
            if mod == 'starttls':
                srv.starttls(context=ssl.create_default_context())
                srv.ehlo()
            if sifre:
                _giris(srv, host, kullanici, sifre)
            red = srv.sendmail(gonderen, alicilar, msg.as_string())
            iz = _yanit_metni(srv)
    # REDDEDİLEN ALICI SESSİZ KALMASIN: sendmail bir kısmı reddedilse bile
    # istisna atmaz (hepsi reddedilirse atar). Eskiden bu durumda ekranda
    # 'gönderildi' yazıyordu ve o adrese mail hiç gitmiyordu.
    if red:
        print(f'[MAIL] REDDEDILEN ALICI: {red}')
    print(f'[MAIL] SMTP kabul etti -> {", ".join(alicilar)} | sunucu yaniti: {iz or "-"}')
    return {'gonderilen': [a for a in alicilar if a not in (red or {})],
            'reddedilen': {k: str(v) for k, v in (red or {}).items()},
            'smtp_yanit': iz}


def gunluk_mail_gonder(tarih=None, zorla_alicilar=None):
    """Günlük raporu üretip alıcılara gönderir. scheduler ve 'şimdi gönder' bunu çağırır.
    zorla_alicilar verilirse (test) DB yerine o listeye gönderir.
    Döner: {basarili, mesaj, ...}
    """
    cfg = config_yukle()
    if not cfg or not cfg.get('etkin'):
        return {'basarili': False, 'mesaj': 'Mail yapılandırılmamış (mail_config.json yok/etkin değil).', 'atlandi': True}
    yontem = yontem_al(cfg)
    if yontem == 'smtp' and not all(cfg.get(k) for k in ('smtp_host', 'smtp_port', 'gonderen')):
        return {'basarili': False, 'mesaj': 'SMTP config eksik: smtp_host / smtp_port / gonderen zorunlu.'}

    if tarih is None:
        tarih = date.today().isoformat()

    if not zorla_alicilar and not aktif_alicilar():
        return {'basarili': False, 'mesaj': 'Aktif alıcı yok — dashboard\'dan e-posta ekleyin.', 'atlandi': True}

    def _alicilar(lok):
        # Test maili SEÇİMİ AŞAR: kurulum doğrulanırken tesis kutucuğu engel olmasın.
        return zorla_alicilar if zorla_alicilar else aktif_alicilar(lokasyon=lok)

    # TESİS BAŞINA AYRI MAİL (2026-08-26). Tek kapsam varsa döngü bir kez döner
    # ve sonuç sözlüğü eskisiyle AYNI şekilde (düz) döner — panelin 'şimdi
    # gönder' ekranı ve testler bozulmasın.
    kapsamlar = kapsam_lokasyonlari(cfg)
    if len(kapsamlar) > 1:
        sonuclar = [_tek_rapor_gonder(cfg, tarih, lok, _alicilar(lok), yontem, zorla_alicilar)
                    for lok in kapsamlar]
        gonderilen = [r for r in sonuclar if r.get('basarili') and not r.get('atlandi')]
        hatali = [r for r in sonuclar if not r.get('basarili')]
        return {
            'basarili': not hatali,
            'coklu': True,
            'kapsamlar': kapsamlar,
            'sonuclar': dict(zip(kapsamlar, sonuclar)),
            # Kapsamlar farklı alıcı kümelerine gidiyor — TEKİL kişi sayısı
            # (iki raporu da alan biri bir kez sayılır).
            'alici_sayisi': len({a for r in sonuclar for a in (r.get('alicilar') or [])}),
            'satir': sum(r.get('satir', 0) or 0 for r in sonuclar),
            'toplam_adet': sum(r.get('toplam_adet', 0) or 0 for r in sonuclar),
            'mesaj': ' | '.join(f'{k}: {r.get("mesaj", "")}' for k, r in zip(kapsamlar, sonuclar)),
            'atlandi': (not gonderilen and not hatali),
        }
    return _tek_rapor_gonder(cfg, tarih, kapsamlar[0], _alicilar(kapsamlar[0]),
                             yontem, zorla_alicilar)


def _tek_rapor_gonder(cfg, tarih, lokasyon, alicilar, yontem, zorla_alicilar=None):
    """TEK bir kapsam için raporu üretir ve gönderir. Sonuç sözlüğü döner.

    Gövde eskiden gunluk_mail_gonder içindeydi; TK1/TK2 ayrımı için buraya
    alındı — iki tesis için iki kez, ama AYNI kodla çalışsın."""
    if not alicilar:
        _k = '' if _hepsi_mi(lokasyon) else f' ({str(lokasyon).upper()})'
        return {'basarili': True, 'atlandi': True, 'satir': 0, 'toplam_adet': 0,
                'lokasyon': lokasyon,
                'mesaj': f'Bu tesisi{_k} seçen aktif alıcı yok — mail gönderilmedi.'}
    try:
        dosya, satir, toplam = excel_olustur(tarih, lokasyon)
    except Exception as e:
        return {'basarili': False, 'mesaj': f'Excel oluşturulamadı: {e}'}

    # ── ÜRETİMSİZ GÜNDE MAIL GÖNDERİLMEZ (kullanıcı 2026-08-24) ──────────
    # Hafta sonu/tatilde herkese "üretim bulunmamaktadır" maili gidiyordu; bu
    # maillerin okunmaması, günün birinde GERÇEK raporun da atlanmasına yol açar.
    # İSTİSNA: test maili (zorla_alicilar) — SMTP kurulumu üretimsiz günde de
    # doğrulanabilmeli. mail_config.json'a "bos_gun_gonder": true yazılırsa eski
    # davranış geri gelir.
    if not satir and not zorla_alicilar and not cfg.get('bos_gun_gonder'):
        _k = '' if _hepsi_mi(lokasyon) else f' ({str(lokasyon).upper()})'
        print(f'[MAIL] {tarih}{_k}: uretim yok — mail gonderilmedi')
        return {'basarili': True, 'atlandi': True, 'satir': 0, 'toplam_adet': 0,
                'lokasyon': lokasyon,
                'mesaj': f'{_tarih_tr(tarih)}{_k} tarihinde üretim yok — mail gönderilmedi.'}

    tarih_tr = _tarih_tr(tarih)
    # Tesis KONUDA: iki rapor aynı anda geldiğinde hangisinin hangisi olduğu
    # gelen kutusunda açmadan görünsün.
    _tesis_et = '' if _hepsi_mi(lokasyon) else f' — {str(lokasyon).upper()}'
    konu = f'Cofle Forge — Günlük Üretim Raporu{_tesis_et} ({tarih_tr})'
    kirilim = _bolum_kirilimi(tarih, lokasyon) if satir else []
    if satir:
        govde = (
            f'Merhaba,\n\n'
            f'{tarih_tr} tarihli günlük üretim raporu ektedir.\n\n'
            f'Özet:\n'
            f'  • Kayıt satırı : {satir}\n'
            # TK1'de tel adetleri PROSES ADIMI (bitmiş ürün değil) — toplamın
            # çoğunu onlar oluşturuyor, ad dürüst olsun (kullanıcı 2026-08-27).
            + (f'  • Toplam işlem : {toplam:,} adet (tel tüm adımlarıyla dahil — '
               f'bitmiş ürün adedi değildir)\n\n'
               if str(lokasyon).upper() == 'TK1' else
               f'  • Toplam üretim: {toplam:,} adet\n\n')
            + (''.join(f'  • {k["tesis"]} · {BOLUM_AD.get(k["bolum"], k["bolum"])}: '
                       f'{k["adet"]:,} adet\n' for k in kirilim) + '\n' if kirilim else '')
            + f'Detaylar ekteki Excel dosyasındadır.\n\n'
            f'Bu e-posta Cofle Forge tarafından otomatik gönderilmiştir.'
        ).replace(',', '.')
    else:
        govde = (
            f'Merhaba,\n\n'
            f'{tarih_tr} tarihinde sisteme kayıtlı üretim bulunmamaktadır.\n\n'
            f'Bu e-posta Cofle Forge tarafından otomatik gönderilmiştir.'
        )
    # Analiz bölümü — hatası maili ENGELLEMEZ (analiz_topla kendi içinde yutar)
    an_ozet, an_bulgular, an_yorum, an_yorum_hata = (
        analiz_topla(tarih, lokasyon) if satir else (None, [], '', ''))
    if an_bulgular or an_yorum or an_yorum_hata:
        govde += '\n\n' + '-' * 52 + '\nOTOMATİK ANALİZ\n' + '-' * 52 + '\n'
        # Özet üretilemediyse SEBEBİYLE söyle — sessiz kaybolunca "dün vardı
        # bugün yok" bilmecesi doğuyordu (bkz. analiz_topla açıklaması).
        if not an_yorum and an_yorum_hata:
            govde += f'(Yönetici özeti bu koşuda üretilemedi: {an_yorum_hata})\n\n'
        if an_ozet:
            govde += (f'OEE %{an_ozet.get("oee", 0)} · Plansız duruş '
                      f'{an_ozet.get("plansiz_dk", 0)} dk · Hurda '
                      f'%{an_ozet.get("hurda_oran", 0)}\n\n')
        if an_yorum:
            govde += an_yorum + '\n\n'
        for b in an_bulgular[:6]:
            govde += f'  [{SIDDET_AD_MAIL.get(b["siddet"], "")}] {b["baslik"]}\n'
    html = _html_govde(tarih_tr, satir, toplam, kirilim, an_ozet, an_bulgular, an_yorum,
                       tesis=('' if _hepsi_mi(lokasyon) else str(lokasyon).upper()),
                       yorum_hata=an_yorum_hata)

    try:
        ek = dosya if satir else None
        if yontem == 'outlook':
            _outlook_gonder(alicilar, konu, govde, ek, html)
            iz = {}
        else:
            iz = _smtp_gonder(cfg, alicilar, konu, govde, ek, html) or {}
    except Exception as e:
        return {'basarili': False, 'mesaj': f'Gönderim hatası: {e}', 'alici_sayisi': len(alicilar)}

    print(f'[MAIL] {tarih}{_tesis_et} raporu gonderildi -> {len(alicilar)} alici, '
          f'{satir} satir, {toplam} adet')
    # Mesajda ADRESLER de yazar: 'gönderildi' deyip nereye gittiğini
    # söylememek, mail gelmediğinde yanlış yerde aramaya yol açıyordu.
    _kime = ', '.join(alicilar[:4]) + (f' (+{len(alicilar) - 4})' if len(alicilar) > 4 else '')
    _red = iz.get('reddedilen') or {}
    _mesaj = (f'{len(alicilar)} alıcıya gönderildi ({satir} satır, {toplam} adet) → {_kime}.'
              + (f' REDDEDİLEN: {", ".join(_red)}' if _red else '')
              + (f' [sunucu: {iz.get("smtp_yanit")}]' if iz.get('smtp_yanit') else ''))
    return {
        'basarili': True,
        'lokasyon': lokasyon,
        'mesaj': _mesaj,
        'alicilar': alicilar,
        'reddedilen': _red,
        'smtp_yanit': iz.get('smtp_yanit', ''),
        'gonderen': cfg.get('gonderen', ''),
        'alici_sayisi': len(alicilar),
        'satir': satir,
        'toplam_adet': toplam,
        'dosya': os.path.basename(dosya),
    }


def test_gonder(email, tarih=None):
    """Tek adrese test maili — kurulumu doğrulamak için."""
    return gunluk_mail_gonder(tarih=tarih, zorla_alicilar=[email])


if __name__ == '__main__':
    # Elle test: python mail_raporu.py [YYYY-MM-DD]
    import sys
    t = sys.argv[1] if len(sys.argv) > 1 else None
    print('Config etkin mi:', etkin())
    if t or True:
        d, n, top = excel_olustur(t)
        print(f'Excel: {d}  ({n} satır, {top} adet)')
