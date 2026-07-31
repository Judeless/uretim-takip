# -*- coding: utf-8 -*-
"""Kaynak planındaki referansların ÜRETİLEBİLİR ADEDİNİ ERP'den hesaplar (2026-07-30).

NEDEN: Planlama her hafta bir "Kaynak ihtiyaçları" dosyası çıkarıyor; referanslar
önem sırasına dizili. Launch talimatı vermek için her kod AS400'de 07>10>01
ekranında tek tek açılıp 9+Enter ile ürün ağacı görüntüleniyor ve 1. seviye alt
parçaların 01D/CF2 stoğu gözle kontrol ediliyor. 230 satır için bu iş saatler
sürüyor ve hata kaldırmıyor.

BU SCRIPT aynı kontrolü toplu yapar:
  1. Plan dosyasındaki "Kaynak kodu" sütununu SIRASIYLA okur (sıra = öncelik)
  2. Her kod için ERP'den 1. SEVİYE ürün ağacını çeker      (TKC0301F.BSPEF2)
  3. Alt parçaların depo stoklarını çeker                    (TKC0300F.BSMAF0)
  4. Üretilebilir adet = min(alt parça stoğu / birim miktar) — en kısıtlayan parça
  5. Excel raporu yazar: özet + parça kırılımı

STOK KURALI: kullanıcının elle yaptığı kontrolle aynı — yalnız 01D ve CF2.
Diğer depolar (01W, REP, MDT...) rapora bilgi olarak yazılır ama hesaba GİRMEZ.

NEGATİF STOK: ERP'de eksi bakiye olabiliyor (girilmemiş hareket, ters kayıt).
Eksi bakiye "o kadar üretilebilir" demek değildir → 0 sayılır ve işaretlenir.

KULLANIM:
    python kaynak_plan_kontrol.py
    python kaynak_plan_kontrol.py --plan "Q:\\...\\Kaynak ihtiyaçları 260729.xlsb"
    python kaynak_plan_kontrol.py --limit 20        (ilk 20 kod — hızlı deneme)
"""
import argparse
import math
import os
import re
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

KOK = os.path.dirname(os.path.abspath(__file__))
VARSAYILAN_PLAN = r"Q:\UretimPlanlama\Yarımamul Üretim Planları\Kaynak ihtiyaçları 260729.xlsb"
SAYILAN_DEPOLAR = ('01D', 'CF2')          # kullanıcının kuralı
GOSTERILEN_DEPOLAR = ('01D', 'CF2', '01W', 'REP', 'MDT', 'MK2', 'MT2')
GUVENLI_KOD = re.compile(r'^[A-Za-z0-9./\- ]{3,21}$')


def _satirlari_getir(yol):
    """Dosyayı uzantısına göre açar ve satırları {kolon_index: değer} olarak verir.

    .xlsb  → pyxlsb (planlamanın çıkardığı özgün biçim)
    .xlsx  → openpyxl (zaten kurulu; pyxlsb kurulamadığında kaçış yolu —
             kullanıcı dosyayı Excel'de "Farklı Kaydet > .xlsx" ile verebilir)
    pyxlsb yoksa hata mesajı NE YAPILACAĞINI söyler (2026-07-31: sunucuda modül
    kurulu olmadığı için ham "No module named 'pyxlsb'" görünüyordu)."""
    uzanti = os.path.splitext(yol)[1].lower()
    if uzanti == '.xlsb':
        try:
            from pyxlsb import open_workbook
        except ImportError:
            raise RuntimeError(
                "Bu sunucuda .xlsb okuyucusu (pyxlsb) kurulu değil. İki çözüm: "
                "(1) sunucuda 'pip install pyxlsb' çalıştırıp servisi yeniden başlatın, "
                "(2) ya da dosyayı Excel'de 'Farklı Kaydet > Excel Çalışma Kitabı (.xlsx)' "
                "ile kaydedip .xlsx olarak yükleyin — o biçim ek kurulum istemez.")
        with open_workbook(yol) as wb:
            sayfa = 'Kaynak' if 'Kaynak' in wb.sheets else wb.sheets[0]
            with wb.get_sheet(sayfa) as s:
                for satir in s.rows():
                    yield {c.c: c.v for c in satir}
        return
    if uzanti in ('.xlsx', '.xlsm'):
        import openpyxl
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
        try:
            ad = 'Kaynak' if 'Kaynak' in wb.sheetnames else wb.sheetnames[0]
            for satir in wb[ad].iter_rows(values_only=True):
                yield {i: v for i, v in enumerate(satir)}
        finally:
            wb.close()
        return
    raise RuntimeError(f'Desteklenmeyen dosya biçimi: {uzanti or "(uzantısız)"} — .xlsb veya .xlsx olmalı')


def plan_oku(yol, limit=None):
    """Plan dosyasının 'Kaynak' sayfasını okur. [{sira, kaynak_kod, urun, ustler}]
    Satır sırası = öncelik; korunur."""
    satirlar, gorulen = [], set()

    def _say(h, idx):
        try:
            return float(h.get(idx))
        except (TypeError, ValueError):
            return 0.0

    for i, h in enumerate(_satirlari_getir(yol)):
        if i == 0:
            continue                              # başlık satırı
        kod = str(h.get(5) or '').strip()         # F = Kaynak kodu
        if not kod or kod in gorulen:
            continue
        gorulen.add(kod)
        satirlar.append({
            'sira': len(satirlar) + 1,
            'kaynak_kod': kod,
            'urun': str(h.get(1) or '').strip(),
            'ustler': [str(h.get(k) or '').strip() for k in (2, 3, 4) if str(h.get(k) or '').strip()],
            # ── İHTİYAÇ ÖLÇÜTÜ (kullanıcı 2026-07-31, sütunları kendisi verdi) ──
            #   BO (66) "Toplam 6 haft iht"  = ihtiyaç adedi
            #   BV (73) "Launch"             = mevcutta üretime alınan adet
            #   BQ (68) "toplam stok"
            # Kural: üretime alınan ihtiyaçtan azsa KALANI üretmek gerekir ve
            # eldeki stok da ihtiyaçtan düşülür:
            #     GEREKEN = max(0, BO − BV − toplam stok)
            # NOT: dosyadaki hazır "6 hafta için fark" (BT/71) sütunu KULLANILMIYOR;
            # o sütun negatif stoğu yok sayıyor ve 12 satırda bu formülden
            # ayrışıyor (ör. BO=100, stok=−44 → BT 100 der, bu kural 144).
            'acik_launch': _say(h, 73),          # BV — üretime alınan
            'bakiye': _say(h, 15),
            'toplam_stok': _say(h, 68),          # BQ
            'iht_2h': _say(h, 20),
            'iht_6h': _say(h, 66),               # BO — 6 haftalık ihtiyaç
            'kontrol6': (None if h.get(74) is None else bool(h.get(74))),
            'gun_stok': _say(h, 75),
            # Planlamanın elle yazdığı alt parçalar — ERP ağacıyla kıyaslanır
            'plan_parcalar': [str(h.get(k) or '').strip()
                              for k in range(6, 13) if str(h.get(k) or '').strip()],
        })
        son = satirlar[-1]
        son['gereken'] = max(0.0, (son['iht_6h'] or 0) - (son['acik_launch'] or 0)
                            - (son['toplam_stok'] or 0))
        son['stok_eksi'] = (son['toplam_stok'] or 0) < 0
        if limit and len(satirlar) >= limit:
            break
    return satirlar


def erp_baglan():
    d = os.path.join(KOK, 'as400')
    if d not in sys.path:
        sys.path.insert(0, d)
    import launch_cek
    return launch_cek.baglan(timeout=60)


def urun_agaci(cn, kodlar):
    """1. seviye açılım: {ust_kod: [(alt_kod, birim_miktar, birim)]}"""
    agac = {}
    kodlar = [k for k in kodlar if GUVENLI_KOD.match(k)]
    cu = cn.cursor()
    for i in range(0, len(kodlar), 50):
        grup = kodlar[i:i + 50]
        yer = ','.join('?' * len(grup))
        cu.execute(
            f"SELECT TRIM(Y2ARTI), TRIM(Y2RICD), Y2IVQT, Y2IVUM "
            f"FROM TKC0301F.BSPEF2 WHERE TRIM(Y2ARTI) IN ({yer})", grup)
        for ust, alt, miktar, um in cu.fetchall():
            if not (ust and alt):
                continue
            agac.setdefault(ust, []).append((alt, float(miktar or 0), str(um or '').strip()))
    return agac


def stoklar(cn, kodlar):
    """{artikel: {depo: stok}} — tüm depolar (hesapta yalnız SAYILAN_DEPOLAR)."""
    st = {}
    kodlar = [k for k in kodlar if GUVENLI_KOD.match(k)]
    cu = cn.cursor()
    for i in range(0, len(kodlar), 50):
        grup = kodlar[i:i + 50]
        yer = ','.join('?' * len(grup))
        cu.execute(
            f"SELECT TRIM(S0ARTI), TRIM(S0MGCD), SUM(S0GIAD) FROM TKC0300F.BSMAF0 "
            f"WHERE S0SOCI='01' AND S0ARTI IN ({yer}) GROUP BY TRIM(S0ARTI), TRIM(S0MGCD)", grup)
        for art, depo, q in cu.fetchall():
            st.setdefault(art, {})[depo] = float(q or 0)
    return st


def hesapla(satirlar, agac, stok):
    """Her plan satırına üretilebilir adet ve kısıtlayan parçayı ekler."""
    for s in satirlar:
        parcalar = agac.get(s['kaynak_kod']) or []
        s['parcalar'] = []
        if not parcalar:
            s['durum'] = 'AGAC YOK'
            s['uretilebilir'] = None
            s['kisitlayan'] = ''
            continue
        kapasiteler = []
        for alt, birim, um in sorted(parcalar):
            depolar = stok.get(alt, {})
            ham = sum(depolar.get(d, 0) for d in SAYILAN_DEPOLAR)
            eldeki = max(0.0, ham)          # eksi bakiye "üretilebilir" değildir
            if birim > 0:
                kap = int(math.floor(eldeki / birim))
                kapasiteler.append((kap, alt))
            else:
                kap = None                  # birim tanımsız → kısıt sayma, işaretle
            s['parcalar'].append({
                'kod': alt, 'birim': birim, 'um': um,
                'stok_sayilan': ham, 'eksi_bakiye': ham < 0,
                'kapasite': kap,
                'depolar': {d: depolar[d] for d in GOSTERILEN_DEPOLAR if depolar.get(d)},
            })
        if not kapasiteler:
            s['durum'] = 'BIRIM YOK'
            s['uretilebilir'] = None
            s['kisitlayan'] = ''
            continue
        en_dusuk = min(kapasiteler)
        s['uretilebilir'] = en_dusuk[0]
        s['kisitlayan'] = en_dusuk[1]
        s['durum'] = 'URETILEBILIR' if en_dusuk[0] > 0 else 'STOK YOK'
        s['eksi_var'] = any(p['eksi_bakiye'] for p in s['parcalar'])

    # ── KARAR: planlamanın launch ihtiyacı + malzeme durumu ──────────────
    # Kullanıcının elle yaptığı iş tam olarak bu: "launch gerekiyor mu?" (plan
    # dosyasındaki Launch sütunu) ile "malzemesi var mı?" (AS400'de ağaç+stok)
    # sorularını birleştirip talimat vermek. İkisi burada birleşir.
    for s in satirlar:
        ihtiyac = s.get('gereken') or 0
        u = s.get('uretilebilir')
        # ERP ağacı planlamanın listesini kapsıyor mu? (kapsamıyorsa ağaç
        # değişmiş olabilir — karar elle doğrulanmalı)
        erp_kodlar = {p['kod'] for p in s.get('parcalar', [])}
        s['agac_farki'] = sorted(set(s.get('plan_parcalar') or []) - erp_kodlar)
        if ihtiyac <= 0:
            # 6 haftalık açık yok ya da açığı zaten mevcut launch'lar kapatıyor.
            s['karar'] = 'GEREK YOK'
        elif u is None:
            s['karar'] = 'ELLE BAK'
        elif u >= ihtiyac:
            s['karar'] = 'TALIMAT VER'
        elif u > 0:
            s['karar'] = 'KISMI'
        else:
            s['karar'] = 'MALZEME YOK'
    return satirlar


def excel_yaz(satirlar, cikti):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()

    ince = Side(style='thin', color='D0D7E2')
    kenar = Border(left=ince, right=ince, top=ince, bottom=ince)
    bas_fill = PatternFill('solid', fgColor='6D28D9')
    bas_font = Font(bold=True, color='FFFFFF', size=11)

    def basliklar(ws, kolonlar, genislikler):
        ws.append(kolonlar)
        for i, g in enumerate(genislikler, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g
            c = ws.cell(row=1, column=i)
            c.fill, c.font, c.border = bas_fill, bas_font, kenar
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = 'A2'

    KARAR_RENK = {'TALIMAT VER': 'D6EBDD', 'KISMI': 'F7EFDD',
                  'MALZEME YOK': 'F7E6E4', 'ELLE BAK': 'EDEAF7', 'GEREK YOK': 'F3F3F8'}

    def _satir_yaz(ws, s):
        kis_stok = ''
        for p in s.get('parcalar', []):
            if p['kod'] == s.get('kisitlayan'):
                kis_stok = p['stok_sayilan']
                break
        notlar = []
        if s.get('agac_farki'):
            notlar.append('ERP ağacı plandakinden farklı: ' + ', '.join(s['agac_farki'][:3]))
        if s.get('eksi_var'):
            notlar.append('bazı parçalarda EKSİ bakiye (0 sayıldı)')
        if s['durum'] == 'AGAC YOK':
            notlar.append("ERP'de ürün ağacı yok — satın alma parçası olabilir")
        if s['durum'] == 'BIRIM YOK':
            notlar.append('ağaçta birim miktar tanımsız')
        ws.append([s['sira'], s['kaynak_kod'], s['urun'],
                   s.get('iht_6h') or 0, s.get('acik_launch') or 0, s.get('toplam_stok') or 0,
                   s.get('gereken') or 0,
                   s['uretilebilir'] if s['uretilebilir'] is not None else '—',
                   s.get('karar', ''), s.get('kisitlayan', ''), kis_stok,
                   len(s.get('parcalar', [])), ' · '.join(notlar)])
        r = ws.max_row
        f = KARAR_RENK.get(s.get('karar'))
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = kenar
            if f:
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=f)
        ws.cell(row=r, column=7).font = Font(bold=True)
        ws.cell(row=r, column=8).font = Font(bold=True)
        ws.cell(row=r, column=9).font = Font(bold=True)

    BASLIK = ['Sıra', 'Kaynak Kodu', 'Ürün', '6 Hf İhtiyaç (BO)', 'Üretimde (BV)',
              'Toplam Stok', 'GEREKEN', 'Üretilebilir', 'KARAR', 'Kısıtlayan Parça',
              'Kısıtlayan Stok', 'Parça Sayısı', 'Not']
    GENIS = [6, 22, 18, 15, 13, 12, 11, 13, 15, 22, 15, 12, 46]

    # ── 1. SAYFA: yalnız AKSİYON gerekenler (launch ihtiyacı > 0) ──
    ws = wb.active
    ws.title = 'Aksiyon'
    basliklar(ws, BASLIK, GENIS)
    sira_onem = {'TALIMAT VER': 0, 'KISMI': 1, 'MALZEME YOK': 2, 'ELLE BAK': 3}
    aksiyon = [s for s in satirlar if (s.get('gereken') or 0) > 0]
    for s in sorted(aksiyon, key=lambda x: (sira_onem.get(x.get('karar'), 9), x['sira'])):
        _satir_yaz(ws, s)

    # ── 2. SAYFA: planın tamamı, öncelik sırası korunmuş ──
    ws_t = wb.create_sheet('Tüm Plan')
    basliklar(ws_t, BASLIK, GENIS)
    for s in satirlar:
        _satir_yaz(ws_t, s)

    # ── PARÇA KIRILIMI ──
    ws2 = wb.create_sheet('Parça Kırılımı')
    basliklar(ws2, ['Sıra', 'Kaynak Kodu', 'Alt Parça', 'Birim Miktar', 'Birim',
                    '01D', 'CF2', '01D+CF2', 'Bu parçadan kaç adet', 'Kısıt mı?', 'Diğer depolar'],
              [6, 22, 22, 12, 7, 11, 11, 11, 20, 10, 30])
    for s in satirlar:
        for p in s.get('parcalar', []):
            diger = ' '.join(f'{d}:{v:g}' for d, v in p['depolar'].items()
                             if d not in SAYILAN_DEPOLAR)
            ws2.append([s['sira'], s['kaynak_kod'], p['kod'], p['birim'], p['um'],
                        p['depolar'].get('01D', 0), p['depolar'].get('CF2', 0),
                        p['stok_sayilan'],
                        p['kapasite'] if p['kapasite'] is not None else '—',
                        'KISIT' if p['kod'] == s.get('kisitlayan') else '',
                        diger])
            r = ws2.max_row
            for c in range(1, 12):
                ws2.cell(row=r, column=c).border = kenar
            if p['eksi_bakiye']:
                for c in (6, 7, 8):
                    ws2.cell(row=r, column=c).fill = PatternFill('solid', fgColor='F7E6E4')
            if p['kod'] == s.get('kisitlayan'):
                ws2.cell(row=r, column=10).font = Font(bold=True, color='B4231C')

    # ── AÇIKLAMA ──
    ws3 = wb.create_sheet('Nasıl okunur')
    for satir in [
        ['Kaynak plan kontrolü — otomatik üretilebilirlik hesabı'],
        [f'Üretim tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'],
        [],
        ['Bu rapor, AS400 07>10>01 ekranında her kodu tek tek açıp (9+Enter) 1. seviye'],
        ['alt parçaların stoğuna bakma işini otomatikleştirir.'],
        [],
        ['Üretilebilir', 'En kısıtlayan alt parçanın izin verdiği adet:'],
        ['', 'min( alt parça 01D+CF2 stoğu / ürün ağacındaki birim miktar )'],
        ['Stok kuralı', 'Yalnız 01D ve CF2 depoları sayılır (elle yapılan kontrolle aynı).'],
        ['', 'Diğer depolar (01W, REP, MDT...) bilgi olarak gösterilir, hesaba girmez.'],
        ['Eksi bakiye', "ERP'de eksi stok olabiliyor; 0 sayılır ve satır kırmızı işaretlenir."],
        ['AGAC YOK', "ERP'de 1. seviye ürün ağacı bulunamadı — satın alma parçası olabilir,"],
        ['', 'bu satırı elle kontrol edin.'],
        ['Sıra', 'Plan dosyasındaki öncelik sırası korunmuştur.'],
        [],
        ['DİKKAT: hesap yalnız STOK kısıtına bakar. Açık sipariş, rezervasyon, kalite'],
        ['bloğu ve kapasite dikkate alınmaz — launch kararında bunları ayrıca değerlendirin.'],
    ]:
        ws3.append(satir)
    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 92
    ws3['A1'].font = Font(bold=True, size=13, color='6D28D9')

    wb.save(cikti)
    return cikti


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--plan', default=VARSAYILAN_PLAN)
    ap.add_argument('--limit', type=int, default=None)
    ap.add_argument('--cikti', default=None)
    a = ap.parse_args()

    if not os.path.exists(a.plan):
        print(f'HATA: plan dosyası bulunamadı: {a.plan}')
        return 1
    print(f'Plan okunuyor: {os.path.basename(a.plan)}')
    satirlar = plan_oku(a.plan, a.limit)
    print(f'  {len(satirlar)} kaynak kodu (öncelik sırasıyla)')

    print('ERP bağlanıyor...')
    cn = erp_baglan()
    try:
        print('  ürün ağaçları çekiliyor...')
        agac = urun_agaci(cn, [s['kaynak_kod'] for s in satirlar])
        alt_kodlar = sorted({alt for lst in agac.values() for alt, _, _ in lst})
        print(f'  {len(agac)} kodun ağacı bulundu · {len(alt_kodlar)} farklı alt parça')
        print('  stoklar çekiliyor...')
        stok = stoklar(cn, alt_kodlar)
        print(f'  {len(stok)} parçanın stok kaydı var')
    finally:
        cn.close()

    hesapla(satirlar, agac, stok)
    say = {}
    for s in satirlar:
        say[s['karar']] = say.get(s['karar'], 0) + 1
    aksiyon = [s for s in satirlar if (s.get('gereken') or 0) > 0]
    print(f'\n6 haftalık açığı olan: {len(aksiyon)} / {len(satirlar)} satır')
    print('KARAR: ' + ' · '.join(f'{k}: {v}' for k, v in sorted(say.items())))

    ver = [s for s in aksiyon if s['karar'] == 'TALIMAT VER']
    print(f'\n>>> TALİMAT VERİLEBİLİR ({len(ver)}) — malzemesi tam:')
    for s in ver:
        print(f"   {s['sira']:>3}. {s['kaynak_kod']:<20} gereken {s['gereken']:>7.0f} · "
              f"üretilebilir {s['uretilebilir']:>7}")
    kis = [s for s in aksiyon if s['karar'] == 'KISMI']
    if kis:
        print(f'\n>>> KISMİ ({len(kis)}) — malzeme yetmiyor:')
        for s in kis[:12]:
            print(f"   {s['sira']:>3}. {s['kaynak_kod']:<20} gereken {s['gereken']:>7.0f} · "
                  f"üretilebilir {s['uretilebilir']:>7}  (kısıt: {s['kisitlayan']})")
    yok = [s for s in aksiyon if s['karar'] == 'MALZEME YOK']
    if yok:
        print(f'\n>>> MALZEME YOK ({len(yok)}):')
        for s in yok[:12]:
            print(f"   {s['sira']:>3}. {s['kaynak_kod']:<20} gereken {s['gereken']:>7.0f} · "
                  f"kısıt: {s['kisitlayan']}")
    fark = [s for s in aksiyon if s.get('agac_farki')]
    if fark:
        print(f'\n>>> ERP AĞACI PLANDAN FARKLI ({len(fark)}) — kararı elle doğrulayın:')
        for s in fark[:8]:
            print(f"   {s['sira']:>3}. {s['kaynak_kod']:<20} planda var, ERP ağacında yok: "
                  f"{', '.join(s['agac_farki'][:3])}")

    cikti = a.cikti or os.path.join(
        os.path.expanduser('~'), 'OneDrive', 'Masaüstü',
        f'Kaynak_Uretilebilirlik_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    try:
        excel_yaz(satirlar, cikti)
        print(f'\nRapor: {cikti}')
    except PermissionError:
        print(f'\nHATA: dosya açık olabilir → {cikti}')
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
